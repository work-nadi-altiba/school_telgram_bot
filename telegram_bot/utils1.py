#!/usr/bin/env python3

import requests
import json
from pygments import highlight
from pygments.lexers import JsonLexer
from pygments.formatters import TerminalFormatter
from openpyxl import Workbook , load_workbook
from docxtpl import DocxTemplate
from docx2pdf import convert
import subprocess
import os 
import glob
from odf.opendocument import load
from odf import text, table
from odf.table import Table ,TableCell ,TableRow
from odf.namespaces import DRAWNS, STYLENS, SVGNS, TEXTNS
from odf.draw import CustomShape
from odf.style import Style, TextProperties
from num2words import num2words
from odf.text import P, Span
import hijri_converter
import fitz
import webcolors
import ezodf
import shutil
import PyPDF2
import PyPDF4
import os
from openpyxl import load_workbook
from openpyxl.styles import Font
import random
import re
import itertools
import openpyxl
import tempfile
import zipfile
from PyPDF4 import PdfFileMerger
import datetime
from dateutil.relativedelta import relativedelta
import calendar 
import locale
from itertools import product
import pdb
import wfuzz
from tqdm import tqdm
from pprint import pprint
from requests.packages.urllib3.exceptions import InsecureRequestWarning
from itertools import groupby
import traceback
import pandas as pd
from loguru import logger
from setting import *
import time 
import ijson 
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.firefox.service import Service
from selenium.webdriver.firefox.options import Options
from webdriver_manager.firefox import GeckoDriverManager

requests.packages.urllib3.disable_warnings(InsecureRequestWarning)

# Global variables should under her please
secondery_students = []

# Globals for debugging purposes
open_emis_core_marks = []
grouped_list = []

# New code should be under here please
def remove_and_pass_to_function(password , certs=False , main_marks=False , side_marks=False , term2 = True):
    password = password.replace('  Ø´','').replace('  Ø¬','').replace('  Ø±','')
    try : 
        username = password.split('/')[0].strip()
        password = password.split('/')[1].strip()
    except:
        username ,password =[password.strip()]*2
    
    try:
        logger.info(f"username:{username} ----> password :{password}")
        if certs:
            create_from_certs_template_wrapper(username , password , term2=term2)
        elif main_marks:
            create_e_side_marks_doc(username , password )
        elif side_marks:
            create_two_terms_side_marks_doc(username , password )
    except Exception as e:
        
        print("\033[91m There is error in \n{}/{}\033[00m" .format(username , password))
        # print(username , password)
        # traceback.print_exc()
        logger.error(f"username:{username} ----> password :{password}"+'\n'+traceback.format_exc())

def diverse_bulk_work(passwords ,term2=True):
    passwords = passwords.split('\n')
    for password in passwords : 
        if '  Ø´' in password:
            remove_and_pass_to_function(password ,certs=True, term2=term2)
        
        if '  Ø±' in password:
            remove_and_pass_to_function(password ,main_marks=True, term2=term2)
        
        if '  Ø¬' in password:
            remove_and_pass_to_function(password ,side_marks=True, term2=term2)

def convert_if_whole(value):
    # Check if the fractional part is zero
    if not isinstance(value , str):
        if value % 1 == 0:
            # Convert the float to an integer
            return int(value)
        else:
            # Return the float as is
            return value
    else:
        return value

def create_certs_from_docx_template(students_statistics_assesment_data ,absent_list_with_names, term2=False ,template='./templet_files/cartoon 1-10 FC_modified_windows.docx' , outdir='./send_folder/'):
    term = 1 if term2 else 0
    context = {}
    statistic_data = students_statistics_assesment_data['students_info']
    assessments_data_groups = students_statistics_assesment_data['assessments_data']
    
    for group in assessments_data_groups:
        if not any("Ø¹Ø´Ø±" in i['student_grade_name'] for i in group) :

            dir_name = students_statistics_assesment_data['teacher_incharge_name'] +'-'+ group[0]['student_class_name_letter']
            os.makedirs(f'{outdir}{dir_name}', exist_ok=True)
            outdir = f'{outdir}{dir_name}/'
            
            group = sort_dictionary_list_based_on(group ,simple=False,item_in_list=term)
            
            for counter , group_item in enumerate(group , start=2):
                
                wanted_id = group_item['student_id']                
                student_statistic_info = [i for i in statistic_data if i['student_id'] == wanted_id ][0]
                absent_count = [i for i in absent_list_with_names if i[0] == wanted_id ]
                
                context['stname']= group_item['student__full_name']
                context['clas']= group_item['student_class_name_letter']
                context['nat']= group_item['student_nat']
                context['id']= group_item['student_nat_id']
                context['bdp']= group_item['student_birth_place']+' '+group_item['student_birth_date']
                context['deyanh']= student_statistic_info['religion']
                context['adres']= student_statistic_info['address']
                context['school']= group_item['student_school_name']
                context['sadres']=  students_statistics_assesment_data['school_address']
                context['mod']= students_statistics_assesment_data['school_directorate']
                context['lewa']= students_statistics_assesment_data['school_bridge']
                context['soltah']=  'ÙˆØ²Ø§Ø±Ø© Ø§Ù„ØªØ±Ø¨ÙŠØ© Ùˆ Ø§Ù„ØªØ¹Ù„ÙŠÙ…'
                context['hatef']= students_statistics_assesment_data['school_phone_number']
                context['schoolid']= students_statistics_assesment_data['school_national_id']
                context['day1']= ''
                context['date1']= ''
                context['day2']= ''
                context['date2']= ''
                context['day3']= ''
                context['date3']= ''
                context['day4']= ''
                context['date4']= ''
                context['sy1']= ''
                context['day5']= ''
                context['date5']= ''

                context['y1']= students_statistics_assesment_data['academic_year_1']
                context['y2' ]= students_statistics_assesment_data['academic_year_2']
                context['stname']= group_item['student__full_name']
                context['clas_only']= group_item['student_grade_name']

                islam_subject = [value for key ,value in group_item['subject_sums'].items() if 'Ø³Ù„Ø§Ù…ÙŠØ©' in key]
                arabic_subject = [value for key ,value in group_item['subject_sums'].items() if 'Ø¹Ø±Ø¨ÙŠØ©' in key]
                english_subject = [value for key ,value in group_item['subject_sums'].items() if 'Ø¬Ù„ÙŠØ²ÙŠØ©' in key]
                math_subject = [value for key ,value in group_item['subject_sums'].items() if 'Ø±ÙŠØ§Ø¶ÙŠØ§Øª' in key]
                social_subjects = [value for key ,value in group_item['subject_sums'].items() if 'Ø§Ø¬ØªÙ…Ø§Ø¹ÙŠØ© Ùˆ Ø§Ù„ÙˆØ·Ù†ÙŠØ©' in key]
                science_subjects = [value for key ,value in group_item['subject_sums'].items() if 'Ø§Ù„Ø¹Ù„ÙˆÙ…' in key]
                art_subject = [value for key ,value in group_item['subject_sums'].items() if 'Ø§Ù„ÙÙ†ÙŠØ© ÙˆØ§Ù„Ù…ÙˆØ³' in key]
                sport_subject = [value for key ,value in group_item['subject_sums'].items() if 'Ø±ÙŠØ§Ø¶ÙŠØ©' in key]
                vocational_subject = [value for key ,value in group_item['subject_sums'].items() if 'Ù…Ù‡Ù†ÙŠØ©' in key]
                computer_subject = [value for key ,value in group_item['subject_sums'].items() if 'Ø­Ø§Ø³ÙˆØ¨' in key]
                financial_subject = [value for key ,value in group_item['subject_sums'].items() if 'Ù…Ø§Ù„ÙŠØ©' in key]
                franch_subject = [value for key ,value in group_item['subject_sums'].items() if 'ÙØ±Ù†Ø³ÙŠØ©' in key]
                christian_subject = [value for key ,value in group_item['subject_sums'].items() if 'Ø§Ù„Ø¯ÙŠÙ† Ø§Ù„Ù…Ø³ÙŠØ­ÙŠ' in key]
                # --------------------------------------
                # Ø¯ÙŠÙ†
                if 'Ø«Ø§Ù…Ù†' in group_item['student_grade_name'] or 'ØªØ§Ø³Ø¹' in group_item['student_grade_name'] or 'Ø¹Ø§Ø´Ø±' in group_item['student_grade_name']:
                    islam_subject_maxMark = 200
                    context['m11']= islam_subject_maxMark /2
                    context['m12']= islam_subject_maxMark
                else:
                    islam_subject_maxMark = 100
                    context['m11']= islam_subject_maxMark /2
                    context['m12']= islam_subject_maxMark
                context['m13']= islam_subject[0][0] if islam_subject and len(islam_subject[0]) != 0 else ''
                context['m14']= islam_subject[0][1] if term2 and islam_subject and len(islam_subject[0]) != 0 else ''
                context['m15']=(islam_subject[0][0] + islam_subject[0][1])/2 if term2 and islam_subject and len(islam_subject[0]) != 0 else ''
                # --------------------------------------
                # Ø¹Ø±Ø¨ÙŠ 
                if 'Ø«Ø§Ù…Ù†' in group_item['student_grade_name'] or 'ØªØ§Ø³Ø¹' in group_item['student_grade_name'] or 'Ø¹Ø§Ø´Ø±' in group_item['student_grade_name']:
                    arabic_subject_maxMark = 300
                    context['m21']= arabic_subject_maxMark / 2
                    context['m22']= arabic_subject_maxMark
                else:
                    arabic_subject_maxMark = 100
                    context['m21']= arabic_subject_maxMark / 2
                    context['m22']= arabic_subject_maxMark
                context['m23']=arabic_subject[0][0] if arabic_subject and len(arabic_subject[0]) != 0 else ''
                context['m24']=arabic_subject[0][1] if term2 and arabic_subject and len(arabic_subject[0]) != 0 else ''
                context['m25']=(arabic_subject[0][0] + arabic_subject[0][1])/2 if term2 and arabic_subject and len(arabic_subject[0]) != 0 else ''
                # --------------------------------------
                # Ø§Ù†Ø¬Ù„ÙŠØ²ÙŠ
                if 'Ø«Ø§Ù…Ù†' in group_item['student_grade_name'] or 'ØªØ§Ø³Ø¹' in group_item['student_grade_name'] or 'Ø¹Ø§Ø´Ø±' in group_item['student_grade_name']:
                    english_subject_maxMark = 200
                    context['m31']= english_subject_maxMark / 2
                    context['m32']= english_subject_maxMark
                else:
                    english_subject_maxMark = 100
                    context['m31']= english_subject_maxMark / 2
                    context['m32']= english_subject_maxMark
                context['m33']= english_subject[0][0] if english_subject and len(english_subject[0]) != 0 else ''
                context['m34']= english_subject[0][1] if term2 and english_subject and len(english_subject[0]) != 0 else ''
                context['m35']= (english_subject[0][0] + english_subject[0][1])/2 if term2 and english_subject and len(english_subject[0]) != 0 else ''
                # --------------------------------------
                # Ø±ÙŠØ§Ø¶ÙŠØ§Øª

                if 'Ø«Ø§Ù…Ù†' in group_item['student_grade_name'] or 'ØªØ§Ø³Ø¹' in group_item['student_grade_name'] or 'Ø¹Ø§Ø´Ø±' in group_item['student_grade_name']:
                    math_subject_maxMark = 200
                    context['m41']=  math_subject_maxMark / 2
                    context['m42']=  math_subject_maxMark
                else:
                    math_subject_maxMark = 100
                    context['m41']=  math_subject_maxMark / 2
                    context['m42']=  math_subject_maxMark
                context['m43']= math_subject[0][0] if math_subject and len(math_subject[0]) != 0 else ''
                context['m44']= math_subject[0][1] if term2 and math_subject and len(math_subject[0]) != 0 else ''
                context['m45']= (math_subject[0][0] + math_subject[0][1])/2 if term2 and math_subject and len(math_subject[0]) != 0 else ''
                # --------------------------------------
                # Ø§Ø¬ØªÙ…Ø§Ø¹ÙŠØ§Øª
                if 'Ø«Ø§Ù…Ù†' in group_item['student_grade_name'] or 'ØªØ§Ø³Ø¹' in group_item['student_grade_name'] or 'Ø¹Ø§Ø´Ø±' in group_item['student_grade_name']:
                    social_subjects_maxMark = 200
                    context['m51']= social_subjects_maxMark / 2
                    context['m52']= social_subjects_maxMark
                    context['m53']= int(social_subjects[0][0]*(2/3)) if social_subjects and len(social_subjects[0]) != 0 else ''
                    context['m54']= int(social_subjects[0][1]*(2/3)) if term2 and social_subjects and len(social_subjects[0]) != 0 else ''
                    context['m55']= round((((social_subjects[0][0] + social_subjects[0][1])/2)*(2/3)),1) if term2 and social_subjects and len(social_subjects[0]) != 0 else ''
                elif 'Ø³Ø§Ø¯Ø³' in group_item['student_grade_name']:
                    social_subjects_maxMark = 100                
                    context['m51']=( social_subjects_maxMark / 2)
                    context['m52']=( social_subjects_maxMark)
                    context['m53']= int(social_subjects[0][0]/3) if social_subjects and len(social_subjects[0]) != 0 else ''
                    context['m54']= int(social_subjects[0][1]/3) if term2 and social_subjects and len(social_subjects[0]) != 0 else ''
                    context['m55']= round(((social_subjects[0][0] + social_subjects[0][1])/2)/3 , 1) if term2 and social_subjects and len(social_subjects[0]) != 0 else ''
                else:
                    social_subjects_maxMark = 100
                    context['m51']= social_subjects_maxMark / 2
                    context['m52']= social_subjects_maxMark
                    context['m53']= int(social_subjects[0][0]) if social_subjects and len(social_subjects[0]) != 0 else ''
                    context['m54']= int(social_subjects[0][1]) if term2 and social_subjects and len(social_subjects[0]) != 0 else ''
                    context['m55']= (social_subjects[0][0] + social_subjects[0][1])/2 if term2 and social_subjects and len(social_subjects[0]) != 0 else ''

                # --------------------------------------
                # Ø¹Ù„ÙˆÙ…
                if 'Ø«Ø§Ù…Ù†' in group_item['student_grade_name'] :
                    science_subjects_maxMark = 200
                    context['m61']= science_subjects_maxMark / 2
                    context['m62']= science_subjects_maxMark
                elif 'ØªØ§Ø³Ø¹' in group_item['student_grade_name'] or 'Ø¹Ø§Ø´Ø±' in group_item['student_grade_name']:
                    science_subjects_maxMark = 400
                    context['m61']= science_subjects_maxMark / 2
                    context['m62']= science_subjects_maxMark
                else:
                    science_subjects_maxMark = 100
                    context['m61']= science_subjects_maxMark / 2
                    context['m62']= science_subjects_maxMark
                context['m63']= science_subjects[0][0] if science_subjects and len(science_subjects[0]) != 0 else ''
                context['m64']= science_subjects[0][1] if term2 and science_subjects and len(science_subjects[0]) != 0 else ''
                context['m65']= (science_subjects[0][0] + science_subjects[0][1])/2 if term2 and science_subjects and len(science_subjects[0]) != 0 else ''

                # --------------------------------------
                # ÙÙ†

                context['m71']= 50 if art_subject and len(art_subject[0]) != 0 else ''
                context['m72']= 100 if art_subject and len(art_subject[0]) != 0 else ''
                context['m73']= art_subject[0][0] if art_subject and len(art_subject[0]) != 0 else ''
                context['m74']= art_subject[0][1] if term2 and art_subject and len(art_subject[0]) != 0 else ''
                context['m75']= (art_subject[0][0] + art_subject[0][1])/2 if term2 and art_subject and len(art_subject[0]) != 0 else ''
                # --------------------------------------
                # Ø±ÙŠØ§Ø¶Ø©
                context['m81']= 50 if sport_subject and len(sport_subject[0]) != 0 else ''
                context['m82']= 100 if sport_subject and len(sport_subject[0]) != 0 else ''
                context['m83']= sport_subject[0][0] if sport_subject and len(sport_subject[0]) != 0 else ''
                context['m84']= sport_subject[0][1] if term2 and sport_subject and len(sport_subject[0]) != 0 else ''
                context['m85']= (sport_subject[0][0] + sport_subject[0][1])/2 if term2 and sport_subject and len(sport_subject[0]) != 0 else ''
                # --------------------------------------
                # Ù…Ù‡Ù†ÙŠ
                context['m91']= 50 if vocational_subject and len(vocational_subject[0]) != 0 else ''
                context['m92']= 100 if vocational_subject and len(vocational_subject[0]) != 0 else ''
                context['m93']= vocational_subject[0][0] if vocational_subject and len(vocational_subject[0]) != 0 else ''
                context['m94']= vocational_subject[0][1] if term2 and vocational_subject and len(vocational_subject[0]) != 0 else ''
                context['m95']= (vocational_subject[0][0] + vocational_subject[0][1])/2 if term2 and vocational_subject and len(vocational_subject[0]) != 0 else ''

                # --------------------------------------
                # Ø­Ø§Ø³ÙˆØ¨
                context['m101']= 50 if computer_subject and len(computer_subject[0]) != 0 else ''
                context['m102']= 100 if computer_subject and len(computer_subject[0]) != 0 else ''
                context['m103']= computer_subject[0][0] if computer_subject and len(computer_subject[0]) != 0 else ''
                context['m104']= computer_subject[0][1] if term2 and computer_subject and len(computer_subject[0]) != 0 else ''
                context['m105']= (computer_subject[0][0] + computer_subject[0][1])/2 if term2 and computer_subject and len(computer_subject[0]) != 0 else ''
                # --------------------------------------
                # Ù…Ø§Ù„ÙŠØ©
                context['m111']= 50 if financial_subject and len(financial_subject[0]) != 0 else ''
                context['m112']= 100 if financial_subject and len(financial_subject[0]) != 0 else ''
                context['m113']= financial_subject[0][0] if financial_subject and len(financial_subject[0]) != 0 else ''
                context['m114']= financial_subject[0][1] if term2 and financial_subject and len(financial_subject[0]) != 0 else ''
                context['m115']= (financial_subject[0][0] + financial_subject[0][1])/2 if term2 and financial_subject and len(financial_subject[0]) != 0 else ''
                # --------------------------------------
                # ÙØ±Ù†Ø³ÙŠ
                context['m121']= 50 if franch_subject and len(franch_subject[0]) != 0 else ''
                context['m122']= 100 if franch_subject and len(franch_subject[0]) != 0 else ''
                context['m123']= franch_subject[0][0] if franch_subject and len(franch_subject[0]) != 0 else ''
                context['m124']= franch_subject[0][1] if term2 and franch_subject and len(franch_subject[0]) != 0 else ''
                context['m125']= (franch_subject[0][0] + franch_subject[0][1])/2 if term2 and franch_subject and len(franch_subject[0]) != 0 else ''
                # --------------------------------------
                # Ù…Ø³ÙŠØ­ÙŠ

                context['m131']= 50 if christian_subject and len(christian_subject[0]) != 0 else ''
                context['m132']= 100 if christian_subject and len(christian_subject[0]) != 0 else ''
                context['m133']= christian_subject[0][0] if christian_subject and len(christian_subject[0]) != 0 else ''
                context['m134']= christian_subject[0][1] if term2 and christian_subject and len(christian_subject[0]) != 0 else ''
                context['m135']= (christian_subject[0][0] + christian_subject[0][1])/2 if term2 and christian_subject and len(christian_subject[0]) != 0 else ''

                counter = 0
                for subject_name  ,S1_S2 in group_item['subject_sums'].items():
                    average = (S1_S2[0]+S1_S2[1])/2
                    # print( subject_name, S1_S2)
                    if 'Ø³Ù„Ø§Ù…ÙŠ' in subject_name and average < islam_subject_maxMark / 2 : 
                        counter+=1
                    elif "Ø¹Ø±Ø¨ÙŠØ©"  in subject_name and average < arabic_subject_maxMark / 2 : 
                        counter+=1
                    elif "Ù†Ø¬Ù„ÙŠØ²ÙŠ"  in subject_name and average < english_subject_maxMark / 2 : 
                        counter+=1
                    elif "Ø±ÙŠØ§Ø¶ÙŠØ§Øª"  in subject_name and average < math_subject_maxMark / 2 : 
                        counter+=1
                    elif "Ø¬ØªÙ…Ø§Ø¹ÙŠØ©"  in subject_name and average < social_subjects_maxMark / 2 : 
                        counter+=1
                    elif "Ø¹Ù„ÙˆÙ…"  in subject_name and average < science_subjects_maxMark / 2 : 
                        counter+=1
                    elif  average < 50: 
                        counter+=1
                    # Ø·Ø±ÙŠÙ‚Ø© Ø·Ø¨Ø§Ø¹Ø© Ø§Ù„Ø±Ù‚Ù… ØµØ­ÙŠØ­ Ø§Ø°Ø§ ÙƒØ§Ù† Ø¨Ø¯ÙˆÙ† Ø§Ø¹Ø´Ø§Ø± 
                    # print(subject_name , int((S1_S2[0]+S1_S2[1])/2) if str((S1_S2[0]+S1_S2[1])/2).split('.')[1] == '0' else (S1_S2[0]+S1_S2[1])/2 )
                    
                # print(counter)
                if counter > 4 : 
                    print('ÙŠØ¨Ù‚Ù‰ ÙÙŠ ØµÙÙ‡')
                    result = 2
                elif counter == 0 :
                    print("Ù†Ø§Ø¬Ø­")
                    result = 0
                else :     
                    print('Ù…ÙƒÙ…Ù„')
                    result = 1

                # context['p'] = pass , context['f'] = failed , context['s'] = stay in his class    
                result_symbol = ['p' , 'f' , 's']
                
                context[result_symbol[0]] = 'â˜'
                context[result_symbol[1]] = 'â˜'
                context[result_symbol[2]] = 'â˜'
                if term2 :
                    # Ø§Ù„Ù…Ø¹Ø¯Ù„ Ø§Ù„Ù…Ø¦ÙˆÙŠ Ø¨Ø§Ù„Ø±Ù‚Ø§Ù… 
                    context['av']= group_item['t1+t2+year_avarage'][2]
                    #Ø§Ù„Ù†ØªÙŠØ¬Ø© 
                    context[result_symbol[result]] = 'ğŸ—¹'
                else:
                    #Ø§Ù„Ù…Ø¹Ø¯Ù„ Ø§Ù„Ù…Ø¦ÙˆÙŠ Ø¨Ø§Ù„Ø±Ù‚Ø§Ù… 
                    context['av']=  group_item['t1+t2+year_avarage'][term]
                    if result == 2 : # Ø§Ø°Ø§ ÙƒØ§Ù† Ù…ÙƒÙ…Ù„ ÙÙŠ ØµÙÙ‡ Ø§Ù„ÙØµÙ„ Ø§Ù„Ø§ÙˆÙ„ Ø®Ù„ÙŠÙ‡Ø§ Ø§Ù„Ù‡ Ø¨Ø³ Ø±Ø§Ø³Ø¨ Ù„Ø§Ù†Ù‡ Ø¨Ø¬ÙˆØ² Ø§Ù„ÙØµÙ„ Ø§Ù„Ø«Ø§Ù†ÙŠ ÙŠØªØ­Ø³Ù† 
                        context[1] =  'ğŸ—¹'
                    else:    
                        context[result] = 'ğŸ—¹'

                absent_as_string = absent_count[0][1] if len(absent_count) else 0 
                
                context['days']= ''
                context['abs0']= f'     {absent_as_string}'
                context['abs1']= ''
                context['abs2']= ''
                # context['av']= group_item[]
                context['murabi']= students_statistics_assesment_data['teacher_incharge_name']
                context['modeer']= students_statistics_assesment_data['principle_name']
                context =  {key: convert_if_whole(value) for  key , value in context.items()}
                fill_doc(template , context , outdir+f"{group_item['student__full_name']}.docx" )
                context.clear()

def create_from_certs_template_wrapper(username , password ,term2=False , just_teacher_class=True ,curr_year = None , skip_art_sport=True):
    """
    Retrieves student information, statistics, and marks, then generates colored certificates in OpenDocument Spreadsheet (ODS) format.

    The function performs the following steps:
    1. Establishes a session using the provided username and password.
    2. Retrieves student information and subject marks using the established session.
    3. Groups students based on their information.
    4. Adds subject sum dictionary to the grouped list.
    5. Calculates and adds averages to the grouped list.
    6. Saves the dictionary containing assessment data and statistics to a JSON file.
    7. Generates colored certificates using the saved data.

    Example Usage:
    ```python
    create_coloured_certs_wrapper(username, password, term2=False)
    ```

    Parameters:
    - username (str): The username for authentication.
    - password (str): The password for authentication.
    - term2 (bool): Flag indicating whether to consider term 2. Default is False.

    Returns:
    - None

    Side Effect:
    - Saves colored certificates in the current working directory.

    """
    session = requests.Session()
    auth = get_auth(username , password)
    student_info_marks = get_students_info_subjectsMarks( username , password ,just_teacher_class=just_teacher_class, session=session,curr_year = curr_year , auth = auth)
    if just_teacher_class:
        student_ids = [i['student_id'] for i in student_info_marks]
    else:
        student_ids = [i['student_id'] for i in get_school_students_ids(auth)]
    students_statistics_assesment_data = get_student_statistic_info(username,password , student_ids= student_ids, session=session , auth=auth)
    dic_list4 = student_info_marks
    listo=get_school_absent_days_with_studentID_and_countOfAbsentDays_and_classID_and_className(auth)
    absent_list_with_names=get_absens_studentName_and_countOfDays_and_studentID(dic_list4,listo)
    grouped_list = group_students(dic_list4 )
    
    
    add_subject_sum_dictionary(grouped_list)

    add_averages_to_group_list(grouped_list ,skip_art_sport=skip_art_sport)
    students_statistics_assesment_data['assessments_data'] = grouped_list
    
    create_certs_from_docx_template(students_statistics_assesment_data,absent_list_with_names , term2=term2)

def is_json_response(response):
    try:
        response.json()  # Try to parse the response as JSON
        return True
    except ValueError:
        # If JSON parsing fails, it's not JSON
        return False

def insert_to_two_terms_side_marks_doc(classes_data , template_sheet_or_file=None):
    """
    Insert marks data into the E-side marks document.

    Parameters:
        title (str): Title of the worksheet.
        class_name (str): Name of the class.
        assessments_json (dict): Assessments JSON data.
        assessments (dict): Assessments data.
        secondary_students (list): List of secondary students.
        necessary_data_dict (dict): Necessary data dictionary.
        counter (int): Counter value.
        template_sheet_or_file: Template sheet or file.
    """
    for class_data in classes_data:
        # copy the worksheet
        sheet_copy = template_sheet_or_file.copy_worksheet(template_sheet_or_file['(1)'])
        marks_and_name = []

        # rename the new worksheet
        #FIXME: remove try and catch clous here also 
        try:
            class_name = get_class_short(class_data['class_name'].split('=')[0])
            subject_name = class_data['class_name'].split('=')[1]
        except:
            class_name = get_class_short(class_data['class_name'])
            subject_name = class_data['subject_name']
        sheet_title = subject_name +'-'+ class_name
        
        sheet_copy.title = sheet_title
        sheet_copy.sheet_view.rightToLeft = True
        sheet_copy.print_title_cols = 'A:R' # the first two cols
        sheet_copy.print_title_rows = '1:6' # the first row
        
        
        class_ ,section = class_data['class_name'].split('=')[0].split('-')
        
        sheet_copy.cell(row=2, column=2).value = class_
        sheet_copy.cell(row=2, column=7).value = section
        sheet_copy.cell(row=2, column=14).value = subject_name
        
        
        # Write data to the worksheet and calculate the sum of some columns in each row
        for row_number, dataFrame in enumerate(class_data['students_data'], start=7):
            sheet_copy.cell(row=row_number, column=2).value = dataFrame['name']
            sheet_copy.cell(row=row_number, column=4).value = dataFrame['term1']['assessment1']
            sheet_copy.cell(row=row_number, column=5).value = dataFrame['term1']['assessment2']
            sheet_copy.cell(row=row_number, column=6).value = dataFrame['term1']['assessment3']
            sheet_copy.cell(row=row_number, column=7).value = dataFrame['term1']['assessment4']
            sheet_copy.cell(row=row_number, column=12).value = dataFrame['term2']['assessment1']
            sheet_copy.cell(row=row_number, column=13).value = dataFrame['term2']['assessment2']
            sheet_copy.cell(row=row_number, column=14).value = dataFrame['term2']['assessment3']
            sheet_copy.cell(row=row_number, column=15).value = dataFrame['term2']['assessment4']
            # Set the font for the data rows
            # for cell in sheet_copy[row_number]:
            #     cell.font = data_font    

def create_two_terms_side_marks_doc(username=None , password=None ,classes_data=None,template='./templet_files/Mark Book H.xlsx' ,outdir='./send_folder' ,student_status_ids = [1], period_id = None , empty_marks = False , session=None):
    """
    The function `create_e_side_marks_doc` creates a document with e-side marks using a specified
    template and saves it in a specified output directory.
    
    :param username: The username is the username of the user who wants to create the document. It is
    used for authentication purposes
    :param password: The password is a string that represents the password for the user's account
    :param template: The template parameter is the path to the Excel file that will be used as a
    template for creating the document. It should be in the format './templet_files/e_side_marks.xlsx',
    defaults to ./templet_files/e_side_marks.xlsx (optional)
    :param outdir: The `outdir` parameter specifies the directory where the generated document will be
    saved, defaults to ./send_folder (optional)
    :param session: The `session` parameter is an optional parameter that can be used to pass an
    existing session object. This can be useful if you want to reuse an existing session for
    authentication or other purposes. If no session object is provided, a new session will be created
    """
    
    if username is not None and password is not None : 
        auth = get_auth(username , password )
        if period_id is None :
            period_id = get_curr_period(auth,session=session)['data'][0]['id']
        user = user_info(auth , username,session=session)
        userInfo = user['data'][0]
        user_id , user_name = userInfo['id'] ,f"{userInfo['first_name']} {userInfo['middle_name']} {userInfo['last_name']} - {str(username)}"  
        # years = get_curr_period(auth)
        school_data = inst_name(auth,session=session)['data'][0]
        inst_id = school_data['Institutions']['id']
        school_name = school_data['Institutions']['name']
        school_name_id = f'{school_name}={inst_id}'
        school_id=inst_id

        baldah = make_request(auth=auth , url=f'https://emis.moe.gov.jo/openemis-core/restful/Institution-Institutions.json?_limit=1&id={inst_id}&_contain=InstitutionLands.CustomFieldValues',session=session)['data'][0]['address'].split('-')[0]
        # grades = make_request(auth=auth , url='https://emis.moe.gov.jo/openemis-core/restful/Education.EducationGrades?_limit=0')
        school_place_data= make_request(auth=auth , url=f'https://emis.moe.gov.jo/openemis-core/restful/Institution-Institutions.json?_limit=1&id={inst_id}&_contain=InstitutionLands.CustomFieldValues', session=session)['data'][0]
        indcator_of_private_techers_sector=school_place_data['institution_sector_id']
        if indcator_of_private_techers_sector == 12 : 
            area_data = get_AreaAdministrativeLevels(auth, session=session)['data']
            area_chain_list = find_area_chain(school_place_data['area_administrative_id'], area_data).split(' - ')
            modeeriah_v2=area_chain_list[1]
            modeeriah=f'Ø§Ù„ØªØ¹Ù„ÙŠÙ… Ø§Ù„Ø®Ø§Øµ / {modeeriah_v2}'
        else:
            modeeriah = inst_area(auth , session=session)['data'][0]['Areas']['name']
            modeeriah=f'{modeeriah}'
        school_year = get_curr_period(auth,session=session)['data']
        hejri1 = str(hijri_converter.convert.Gregorian(school_year[0]['start_year'], 1, 1).to_hijri().year)
        hejri2 =  str(hijri_converter.convert.Gregorian(school_year[0]['end_year'], 1, 1).to_hijri().year)
        melady1 = str(school_year[0]['start_year'])
        melady2 = str(school_year[0]['end_year'])
        teacher = user['data'][0]['name'].split(' ')[0]+' '+user['data'][0]['name'].split(' ')[-1]
        
        assessment_periods = make_request(auth =auth,url=f'https://emis.moe.gov.jo/openemis-core/restful/v2/Assessment-AssessmentPeriods.json?_limit=0' , session=session)
        # Ù…Ø§ Ø¨Ø¹Ø±Ù ÙƒÙŠÙ Ø³ÙˆÙŠØªÙ‡Ø§ Ù„ÙƒÙ† Ø²Ø¨Ø·Øª 
        classes_id_2 =[lst for lst in get_teacher_classes_v2(auth ,inst_id, user_id, period_id)['data'] if lst]
        assessments_period_data = []
        grades_info = get_grade_info(auth,period_id,session=session)
        
        teacher_load_marks_data = get_marks_v2(auth , inst_id, period_id, classes_id_2 , grades_info , assessment_periods,session,student_status_ids=student_status_ids, empty_marks=empty_marks)
    else:
        student_details = classes_data
        school_name = student_details['custom_shapes']['school']
        # modified_classes =student_details['custom_shapes']['classes']
        teacher = student_details['custom_shapes']['teacher'] 
        melady2 = student_details['custom_shapes']['melady1']
        melady1 = student_details['custom_shapes']['melady2']
        modeeriah = student_details['custom_shapes']['modeeriah']
        teacher_load_marks_data = student_details['file_data']
    
    sofof , mawad = [] , []
    for students_data_list in teacher_load_marks_data:
        sofof.append(get_class_short(students_data_list['class_name'].split('=')[0]).replace(' ' , ''))
        # FIXME: ÙŠØ¬Ø¨ Ø§Ù† Ø§Ø¬Ø¯ Ø·Ø±ÙŠÙ‚Ø© ØºÙŠØ± Ø·Ø±ÙŠÙ‚Ø© ØªØ±Ø§ÙŠ Ù„ÙƒÙŠ Ø§Ø­Ø¶Ø± Ø§Ø³Ù… Ø§Ù„Ù…Ø§Ø¯Ø©
        try:
            mawad.append(students_data_list['class_name'].split('=')[1])
        except:
            mawad.append(students_data_list['subject_name'])
    sofof = ','.join(list(set(sofof)))
    mawad = ','.join(list(set(mawad)))
    
    # load the existing workbook
    existing_wb = load_workbook(template)
    
    insert_to_two_terms_side_marks_doc(teacher_load_marks_data , template_sheet_or_file=existing_wb)

    existing_wb.remove(existing_wb['(1)'])
    existing_wb.remove(existing_wb['(2)'])
    existing_wb.remove(existing_wb['(3)'])
    existing_wb.remove(existing_wb['(4)'])
    existing_wb.remove(existing_wb['(5)'])
    
    # Access the new sheet by name
    info_sheet = existing_wb["ØºÙ„Ø§Ù"]

    # Write data to the new sheet
    info_sheet["b1"] = f' Ù…Ø¯ÙŠØ±ÙŠØ© Ø§Ù„ØªØ±Ø¨ÙŠØ© Ùˆ Ø§Ù„ØªØ¹Ù„ÙŠÙ… ÙÙŠ {modeeriah}'
    info_sheet["b2"] = school_name
    info_sheet["b5"] = sofof
    info_sheet["b7"] = mawad
    info_sheet["b10"] = teacher
    info_sheet["b12"] = melady1 +'/'+melady2

    # save the modified workbook
    existing_wb.save(f'{outdir}/Ø¯ÙØªØ± Ø¹Ù„Ø§Ù…Ø§Øª Ø¬Ø§Ù†Ø¨ÙŠ -{teacher}.xlsx')

def get_cookie_as_string(usern , passd , url='https://emis.moe.gov.jo/openemis-core/'):
    '''
    Ø¯Ø§Ù„Ø© ØªÙ‚ÙˆÙ… Ø¨Ø§Ø­Ø¶Ø§Ø± Ø§Ù„ÙƒÙˆÙƒÙŠ Ù„ÙƒÙŠ ÙŠØªÙ… Ø§Ø³ØªØ¹Ù…Ø§Ù„Ù‡Ø§ Ø¹Ù†Ø¯Ù…Ø§ ØªÙƒÙˆÙ† Ø§Ù„ÙˆØ§Ø¬Ù‡Ø© Ø§Ù„Ø±Ø³ÙˆÙ…ÙŠØ© Ù…Ø¹Ø·Ù„Ø©
    '''
    options = Options()
    # options.add_argument("--headless")

    driver = webdriver.Firefox(service=Service(GeckoDriverManager().install()), options=options)


    # # Set up the WebDriver
    # driver = webdriver.Firefox(service=Service(GeckoDriverManager().install()))

    try : 
        # Open a webpage
        driver.get(url=url)
        
        current_url = driver.current_url

        # Locate the username and password fields
        username_field = driver.find_element(By.ID, "username")  # Replace "username" with the actual ID of the username field
        password_field = driver.find_element(By.ID, "password")  # Replace "password" with the actual ID of the password field
        

        # Enter the username and password
        username_field.send_keys(usern)  # Replace with your actual username
        password_field.send_keys(passd)  # Replace with your actual password

        # Locate the login button and click it
        login_button = driver.find_element(By.NAME, "submit")  # Replace "loginButton" with the actual ID of the login button
        login_button.click()


        wait = WebDriverWait(driver, 600 )
        wait.until(EC.url_changes(current_url))

        cookies = driver.get_cookies()
        result = {item['name']: item['value'] for item in cookies}
        string_cookie = f"System={result['System']};csrfToken={result['csrfToken']};AWSALB={result['AWSALB']};AWSALBCORS={result['AWSALBCORS']};PHPSESSID={result['PHPSESSID']}"
        print(string_cookie)
    except:
        # Close the driver
        driver.close()
    # Close the driver
    driver.close()
    
    return string_cookie

def is_valid_date(date_str):
    try:
        # Attempt to parse the date string
        datetime.strptime(date_str, '%Y-%m-%d')
        return True
    except ValueError:
        # ValueError will be raised for invalid dates
        return False

def fill_student_absent_doc_wrapper_with_absent_filled(username, password ,template='./templet_files/plus_st_abs_A4.ods' , outdir='./send_folder/' ,teacher_full_name=False , context =None):
    """
    Fills the student absent notebook document template with data and saves it.

    Parameters:
    - username (str): The username for authentication.
    - password (str): The password for authentication.
    - template (str): Path to the ODS template file (default: './templet_files/new_empty_absence_notebook_doc_white_cover.ods').
    - outdir (str): Directory to save the filled document (default: './send_folder/').
    - teacher_full_name (bool): Flag to include teacher's full name in the document (default: False).

    Example Usage:
    ```python
    fill_student_absent_doc_wrapper('your_username', 'your_password', teacher_full_name=True)
    ```

    Note:
    - This function fetches student statistical information using the provided credentials.
    - It then uses the data to fill the specified ODS template with student details and saves the filled document.
    - The filled document is saved in the specified output directory.

    """
    if context is None :
        context = {2: 'Y69=AP123', 1: 'A69=V123', 4: 'Y128=AP182', 3: 'A128=V182', 6: 'Y186=AP240', 5: 'A186=V240', 8: 'Y244=AP298', 7: 'A244=V298', 10: 'Y302=AP356', 9: 'A302=V356', 12: 'Y360=AP414', 11: 'A360=V414', 14: 'Y418=AP472', 13: 'A418=V472', 16: 'Y476=AP530', 15: 'A476=V530', 18: 'Y534=AP588', 17: 'A534=V588', 20: 'Y592=AP646', 19: 'A592=V646', 22: 'Y650=AP704', 21: 'A650=V704', 24: 'Y708=AP762', 23: 'A708=V762', 26: 'Y766=AP820', 25: 'A766=V820'}
    auth = get_auth(username , password)
    student_details = get_student_statistic_info(username,password,teacher_full_name=teacher_full_name)
    
    required_data = get_required_data_to_enter_absent(auth)
    absent_data_list = get_class_absent_days_with_id(auth , required_data=required_data)
    fill_student_absent_doc_name_days_cover(student_details , template , outdir , context = context ,absent_data_list=absent_data_list)

def read_large_json_files(files_list,item ='item.institution_class_id'):
    data = []
    for file in files_list:
        praser = ijson.parse(open(file, 'r'))
        praser2 = ijson.items(praser, item)
        items = [i for i in praser2]
        data.extend(items)
    return data

def find_assessment_above_max_for_one_student(student_data , max_assess_dict , class_name , subject_name):
    above_max = []
    mark_dict = {}
    for term in ['term1' , 'term2']:
        for assess in student_data[term]:
            student_mark = student_data[term][assess]
            term_name_in_arabic = " Ø§Ù„ÙØµÙ„ Ø§Ù„Ø§ÙˆÙ„ /" if term =='term1' else ' Ø§Ù„ÙØµÙ„ Ø§Ù„Ø«Ø§Ù†ÙŠ /'
            assess_name_in_arabic = assess.replace('assessment' , 'Ø§Ù„ØªÙ‚ÙˆÙŠÙ… ').replace('1' , 'Ø§Ù„Ø§ÙˆÙ„').replace('2' , 'Ø§Ù„Ø«Ø§Ù†ÙŠ').replace('3' , 'Ø§Ù„Ø«Ø§Ù„Ø«').replace('4' , 'Ø§Ù„Ø±Ø§Ø¨Ø¹')
            mark_dict =  { 'class_name':class_name ,
                            'subject_name':subject_name ,
                            'term&assess':term_name_in_arabic+assess_name_in_arabic  ,
                            'name':student_data['name'] ,
                            'id' : student_data['id'] 
                        }
            if student_mark == '':
                mark_dict['mark'] = 'ÙØ§Ø±Øº'
                above_max.append(mark_dict)
            elif student_mark > max_assess_dict[assess] :
                mark_dict['mark'] = student_mark
                above_max.append(mark_dict)
    return above_max

def find_above_max_mark_for_assessments(excel_output , max_assessments_dictionaries , skip_empty = True):
    '''
    function usage : 
        auth = get_auth(9891009452 , 9891009452)
        session = requests.Session()
        max_marks_for_classes_based_on_subject_id = create_max_of_dictionaries(excel_output , auth , session)
    '''
    above_max_dictionaries = []
    for class_data in excel_output['file_data'] :
        class_info = class_data['class_name'].split('=')
        class_name = class_info[0]
        subject_name = class_info[1]
        class_institution_id = class_info[2]
        subject_id = class_info[3]
        max_assess_dict = max_assessments_dictionaries[int(class_institution_id)][int(subject_id)]['max_marks_dictionary']
        for student in class_data['students_data']:
            marks_above_max = find_assessment_above_max_for_one_student(student , max_assess_dict ,class_name , subject_name)
            above_max_dictionaries.extend(marks_above_max)
            
    if skip_empty:
        return [i for i in above_max_dictionaries if i['mark'] != 'ÙØ§Ø±Øº']
    else:
        return above_max_dictionaries

def get_all_assessments(auth , session=None):
    assessment_periods = make_request(auth =auth,url=f'https://emis.moe.gov.jo/openemis-core/restful/v2/Assessment-AssessmentPeriods.json?_limit=0' , session=session)
    return assessment_periods

def get_all_AssessmentItemsGradingTypes(auth ,session=None):
    get_max_mark = f"https://emis.moe.gov.jo/openemis-core/restful/v2/Assessment-AssessmentItemsGradingTypes.json?_contain=EducationSubjects,AssessmentGradingTypes.GradingOptions&_limit=0"
    data = make_request(auth=auth , url =get_max_mark , session=session)
    data = data['data']
    return data 

def create_max_marks_assessments_dictionary (data ,assessment_id , education_id , sorted_class_assessments_list):
    sorted_class_assessments_list = sorted_class_assessments_list[0:4]
    max_marks_assessments_dictionary = {'assessment1':'' , 'assessment2':'' , 'assessment3':'' , 'assessment4':''}
    targeted_assessments_with_max = {
                            i['assessment_period_id']: i['assessment_grading_type']['grading_options'][0]['max'] 
                                for i in data 
                                    if i['assessment_id'] == assessment_id 
                                        and
                                        i['education_subject_id'] == education_id 
                            }
    
    for assessment_id ,dictionary_key in  zip(sorted_class_assessments_list , max_marks_assessments_dictionary.keys()) :
        max_marks_assessments_dictionary[dictionary_key] = targeted_assessments_with_max[int(assessment_id)]
    return max_marks_assessments_dictionary

def create_max_of_dictionaries(e_side_document_output , auth , session=None):
    max_of_assessments = get_all_AssessmentItemsGradingTypes(auth , session=session)
    assessment_periods = get_all_assessments(auth , session=session)    
    max_marks_dictionary = {}
    for item in e_side_document_output['file_data']:
        class_information = item['class_name'].split('=')
        institution_class_id , subject_education_id = int(class_information[2]) , int(class_information[3]) 
        class_data = e_side_document_output['required_data_for_mrks_enter'][institution_class_id]
        assessment_grade_id = class_data['assessment_grade_id']
        assessments_period_ids = [int(i) for i in class_data['assessments_period_ids']]
        sorted_class_assessments_list = offline_sort_assessement_period_ids_v2(assessments_period_ids , assessment_periods , True)[0:4]
        maxes_dictionary = create_max_marks_assessments_dictionary(max_of_assessments , assessment_grade_id , subject_education_id , sorted_class_assessments_list) 
        max_marks_dictionary[institution_class_id] = {  subject_education_id : { 'max_marks_dictionary':maxes_dictionary } }
    return max_marks_dictionary

def get_absens_studentName_and_countOfDays_and_studentID(dic_list4,listo):
    """ØªÙ‚ÙˆÙ… Ù‡Ø°Ù‡ Ø§Ù„Ø¯Ø§Ù„Ù‡ Ø¨Ø£Ø±Ø¬Ø§Ø¹ Ù„ÙŠØ³Øª Ù…ÙƒÙˆÙ†Ù‡ Ù…Ù† Ø§Ø³Ù…Ø§Ø¡ Ø§Ù„Ø·Ù„Ø§Ø¨ ÙˆØ¹Ø¯Ø¯ Ø§ÙŠØ§Ù… ØºÙŠØ§Ø¨Ù‡Ù… ÙˆØ§Ù„Ø±Ù‚Ù… Ø§Ù„ØªØ¹Ø±ÙŠÙÙŠ Ø§Ù„Ø®Ø§Øµ Ø¨Ø§Ù„Ø·Ø§Ù„Ø¨ 

    Args:
        dic_list4 (dic_list):create_cert_wrapper Ø¯ÙƒØ´Ù†ÙŠØ±ÙŠ Ù„ÙŠØ³Øª Ø§Ù„Ù…ÙˆØ¬ÙˆØ¯Ù‡ ÙÙŠ Ø¯Ø§Ù„Ø© 
        listo (_list_):get_school_absent_days_with_studentID_and_countOfAbsentDays_and_classID_and_className Ø§Ù„Ù„ÙŠØ³Øª Ø§Ù„ØªÙŠ ØªÙ‚ÙˆÙ… Ø¨Ø§Ø±Ø¬Ø§Ø¹Ù‡Ø§ Ø¯Ø§Ù„Ø© 

    Returns:
        list: Ù„ÙŠØ³Øª Ù…ÙƒÙˆÙ†Ù‡ Ù…Ù† Ø§Ø³Ù… Ø§Ù„Ø·Ø§Ù„Ø¨ ÙˆØ¹Ø¯Ø¯ Ø§ÙŠØ§Ù… Ø§Ù„ØºÙŠØ§Ø¨ ÙˆØ§Ù„Ø§ÙŠØ¯ÙŠ Ø§Ù„Ø®Ø§Øµ Ø¨Ù‡ 
    """    
    student_ids=[[item['student_id'],item["student__full_name"]] for item in dic_list4]
    absens_stu=[]
    for i in student_ids :
        for s in listo :
            if i[0] == s[0]:
                absens_stu.append([i[0], s[1],s[3],i[1]])
    
    return absens_stu

def get_school_absent_days_with_studentID_and_countOfAbsentDays_and_classID_and_className(auth , required_data = None):
    """_ØªÙ‚ÙˆÙ… Ù‡Ø°Ù‡ Ø§Ù„Ø¯Ø§Ù„Ù‡ ÙÙŠ Ø¬Ù„Ø¨ ØºÙŠØ§Ø¨ Ø§Ù„Ù…Ø¯Ø±Ø³Ù‡ ÙƒØ§Ù…Ù„Ø§ Ù…Ù† Ø®Ù„Ø§Ù„ Ø§Ù„Ø§ÙˆØ«Ù†ÙŠØªÙƒØ´Ù† _

    Args:
        auth (_type_): _ Ø§Ù„Ø§ÙˆØ«Ù†ØªÙŠÙƒØ´Ù† Ø§Ù„Ù…Ø·Ù„ÙˆØ¨_
        required_data (_type_, optional): Ø¹Ù†ØµØ± Ø§Ø®ØªÙŠØ§Ø±ÙŠ ÙŠÙØ¶Ù„ ØªØ±ÙƒÙ‡ ÙƒÙ…Ø§ Ù‡ÙˆÙ‡ 

    Returns:
        _list_: _Ù‚Ø§Ø¦Ù…Ù‡ ØªØ¹ÙˆØ¯ Ø¨Ù‡Ø§ Ø§Ù„Ø¯Ø§Ù„Ù‡ ØªØ­ØªÙˆÙŠ Ø¹Ù„Ù‰ Ø§Ù„Ø§Ø¯ÙŠ Ø§Ù„Ø®Ø§Øµ Ø¨Ø§Ù„Ø·Ø§Ù„Ø¨ Ùˆ ÙƒÙ… Ø¹Ø¯Ø¯ Ø§ÙŠØ§Ù… Ø§Ù„ØºÙŠØ§Ø¨ ÙˆØ§Ù„Ø§ÙŠØ¯ÙŠ Ø§Ù„Ø®Ø§Øµ Ø¨Ø§Ù„ØµÙ ÙˆØ§Ø³Ù… Ø§Ù„ØµÙ_
    """    
    
    if not required_data:
        required_data = get_required_data_to_enter_absent(auth)
    
    institution_id=required_data['institution_id']   
    education_grade_id=required_data['education_grade_id']
    academic_period_id=required_data['academic_period_id']
    
    url = f'https://emis.moe.gov.jo/openemis-core/restful/Institution.StudentAbsencesPeriodDetails?institution_id={institution_id}&education_grade_id={education_grade_id}&academic_period_id={academic_period_id}&_limit=0&_fields=student_id,institution_id,academic_period_id,institution_class_id,education_grade_id,date,period,comment,absence_type_id'
    absent_data = make_request(auth=auth , url=url)
    absent_data = [[i['student_id'] , i['date'],i['institution_class_id']] for i in absent_data['data']]
    #########################Ø¹Ø²Ù„ Ø§Ù„ØµÙÙˆÙ Ù…Ø¹ Ø§Ù„Ø§Ø³Ù…Ø§Ø¡ ###########################
    class_names = get_classes_ids_with_names_dict(auth)
    data = absent_data
    # Dictionary to store class names
    class_names_dict = {
        class_id: class_name for class_id, class_name in class_names.items()
    }
    # List to store the updated data with class names
    updated_data = []
    # Iterate through the data
    for item in data:
        student_id, _, class_id = item
        # Check if the class ID is in the dictionary
        if class_id in class_names_dict:
            # Add class name to the data
            class_name = class_names_dict[class_id]
            updated_item = [student_id, _, class_id, class_name]
            updated_data.append(updated_item)
        else:
            # If the class ID is not in the dictionary, add a default value or handle as needed
            updated_data.append(item)
    listo = []

    for i in updated_data:
        student_id, date, class_id, class_name = i

        # Check if the student ID is not in the listo
        found = False
        for item in listo:
            if item[0] == student_id:
                # If the student ID is already in the listo, increment the count
                item[1] += 1
                found = True
                break

        # If the student ID is not found, add it to the listo with count 1
        if not found:
                listo.append([student_id, 1,i[2],i[3]])
    return listo

def get_teacher_class_students(auth , institution_class_id , inst_id=None , curr_year=None , just_id_and_name_and_empty_marks =True ,student_status_ids=[1] ,session=None):
    """
    Retrieves secondary students enrolled in a specific institution class.

    Parameters:
    - auth (str): Authorization token.
    - institution_class_id (int): Identifier for the institution class.
    - inst_id (int, optional): Identifier for the institution (default is None).
    - curr_year (int, optional): Current academic year (default is None).
    - session (requests.Session, optional): Session object for making HTTP requests (default is None).

    Returns:
    - dict: Dictionary containing data about enrolled secondary students.
    """    
    id_and_name_dic_list = []
    secondery_students =  get_school_students_ids(auth, inst_id=inst_id , curr_year=curr_year ,student_status_ids=student_status_ids , session=session)
    data = [i for i in secondery_students if i['institution_class_id'] == int(institution_class_id) and i['student_status_id'] in student_status_ids]
    data = {'data': data , 'total': len(data)}
    
    enrolled = [i for i in data['data'] if i['student_status_id'] in student_status_ids]
    data = {'data': enrolled , 'total': len(enrolled)}
    if just_id_and_name_and_empty_marks:
        dic = {'id':'' ,'name': '','term1':{ 'assessment1': '' ,'assessment2': '' , 'assessment3': '' , 'assessment4': ''} ,'term2':{ 'assessment1': '' ,'assessment2': '' , 'assessment3': '' , 'assessment4': ''} ,'assessments_periods_ides':[]}
        
        for item in data['data']:
            dic['id'] = item['student_id']
            dic['name'] = item['user']['name']
            id_and_name_dic_list.append(dic)
            dic = {'id':'' ,'name': '','term1':{ 'assessment1': '' ,'assessment2': '' , 'assessment3': '' , 'assessment4': ''} ,'term2':{ 'assessment1': '' ,'assessment2': '' , 'assessment3': '' , 'assessment4': ''} ,'assessments_periods_ides':[]}
        return id_and_name_dic_list
    else:
        return data

def get_teacher_homeroom_class(auth,institution_id,period_id):
    url = f'https://emis.moe.gov.jo/openemis-core/restful/v2/Institution-InstitutionClasses?institution_id={institution_id}&academic_period_id={period_id}&_limit=0'
    return make_request(url=url , auth =auth)

def get_subjects_dictionary_list_from_the_site(auth ,session):
    url = GET_SUBJCTS_DATA
    
    return make_request(auth=auth , url=url , session=session)

def fill_student_absent_A4_doc_wrapper(username, password ,template='./templet_files/plus_st_abs_A4.ods' , outdir='./send_folder/' ,teacher_full_name=False , context =None):
    """
    Fills the student absent notebook document template with data and saves it.

    Parameters:
    - username (str): The username for authentication.
    - password (str): The password for authentication.
    - template (str): Path to the ODS template file (default: './templet_files/new_empty_absence_notebook_doc_white_cover.ods').
    - outdir (str): Directory to save the filled document (default: './send_folder/').
    - teacher_full_name (bool): Flag to include teacher's full name in the document (default: False).

    Example Usage:
    ```python
    fill_student_absent_doc_wrapper('your_username', 'your_password', teacher_full_name=True)
    ```

    Note:
    - This function fetches student statistical information using the provided credentials.
    - It then uses the data to fill the specified ODS template with student details and saves the filled document.
    - The filled document is saved in the specified output directory.

    """
    if context is None :
        context = {2: 'Y69=AP123', 1: 'A69=V123', 4: 'Y128=AP182', 3: 'A128=V182', 6: 'Y186=AP240', 5: 'A186=V240', 8: 'Y244=AP298', 7: 'A244=V298', 10: 'Y302=AP356', 9: 'A302=V356', 12: 'Y360=AP414', 11: 'A360=V414', 14: 'Y418=AP472', 13: 'A418=V472', 16: 'Y476=AP530', 15: 'A476=V530', 18: 'Y534=AP588', 17: 'A534=V588', 20: 'Y592=AP646', 19: 'A592=V646', 22: 'Y650=AP704', 21: 'A650=V704', 24: 'Y708=AP762', 23: 'A708=V762', 26: 'Y766=AP820', 25: 'A766=V820'}
    student_details = get_student_statistic_info(username,password,teacher_full_name=teacher_full_name)
    fill_student_absent_doc_name_days_cover(student_details , template , outdir , context = context )

def setup_logging(log_file_path: str):
    log_directory = os.path.join(os.getcwd(), "logs")
    os.makedirs(log_directory, exist_ok=True)
    log_file = os.path.join(log_directory, log_file_path)
    
    logger.remove()  # Remove existing handlers
    logger.add(log_file, rotation="500 MB", compression="zip", backtrace=True, diagnose=True, format="{time} | {function} | {level} | {message}")

def log_info(message: str):
    logger.info(message)

def log_warning(message: str):
    logger.warning(message)

def log_error(message: str):
    logger.error(message)

def log_exception(message: str):
    logger.exception(message)

def get_school_classed_and_unclassed_students(auth,session=None):
    inst_id = inst_name(auth)['data'][0]['Institutions']['id']
    curr_year = get_curr_period(auth)['data'][0]['id']
    unclassed_ss = [
                i 
                for i in make_request(session=session ,auth=auth,url=f'https://emis.moe.gov.jo/openemis-core/restful/v2/Institution.Students?_limit=0&_finder=Users.address_area_id,Users.birthplace_area_id,Users.gender_id,Users.date_of_birth,Users.date_of_death,Users.nationality_id,Users.identity_number,Users.external_reference,Users.status&institution_id={inst_id}&academic_period_id={curr_year}&_contain=Users')['data']
                    
                    if i['student_status_id'] == 1
                ]
    classed_ss = [
                    i 
                    for i in make_request(auth=auth, url=f'https://emis.moe.gov.jo/openemis-core/restful/v2/Institution-InstitutionClassStudents.json?_limit=0&_finder=Users.address_area_id,Users.birthplace_area_id,Users.gender_id,Users.date_of_birth,Users.date_of_death,Users.nationality_id,Users.identity_number,Users.external_reference,Users.status&institution_id={inst_id}&academic_period_id={curr_year}&_contain=Users')['data'] 
                    
                        if i['student_status_id'] == 1 
                    ]
    return {
        'unclassed_ss' : unclassed_ss ,
        'classed_ss' : classed_ss ,
        }

def turn_classed_and_unclassed_students_to_diclist(data):
    classed_ss = [
                    {
                        'student_id' :i['student_id'] , 
                        'student_openemis_no' :i['user']['openemis_no']  , 
                        'identity_number' :i['user']['identity_number'] , 
                        'full_name' :i['user']['name'],
                        'institution_class_id' : i['institution_class_id'],
                        'grade_id' : i['education_grade_id']
                    } for i in data['classed_ss']
                ]

    unclassed_ss = [
                    {
                        'student_id' :i['student_id'] , 
                        'student_openemis_no' :i['user']['openemis_no'] , 
                        'identity_number' :i['user']['identity_number'] , 
                        'full_name' :i['user']['name'],
                        'grade_id' : i['education_grade_id']
                    } for i in data['unclassed_ss']
                ]
    return  classed_ss, unclassed_ss

def get_classes_ids_with_names_dict(auth=None , classes_data=None , session = None ):
    if classes_data is None :
        inst_id = inst_name(auth)['data'][0]['Institutions']['id']
        period_id = get_curr_period(auth)['data'][0]['id']
        student_classess = make_request(auth=auth, url=f'https://emis.moe.gov.jo/openemis-core/restful/v2/Institution-InstitutionClassStudents.json?institution_id={inst_id}&academic_period_id={period_id}&_contain=Users&_limit=0', session=session)['data']
        class_names_dic = {i['institution_class_id'] :{'education_grade_id': i['education_grade_id']} for i in student_classess}
        classes = [i for i in class_names_dic]
        classes_str = ','.join([f'institution_class_id:{i}' for i in classes])
        url = f"https://emis.moe.gov.jo/openemis-core/restful/Institution.InstitutionClassSubjects?status=1&_contain=InstitutionSubjects,InstitutionClasses&_limit=0&_orWhere={classes_str}"
        classes_data = make_request(url=url,auth=auth,session=session)['data']
        return { i['institution_class']['id'] : i['institution_class']['name'] for i in classes_data}
    
    return { i['institution_class']['id'] : i['institution_class']['name'] for i in classes_data}

def get_education_grade_id_with_grade_name_dic(auth=None , grades_data=None):
    if grades_data is None :
        grades_data = get_grade_info(auth)
    pattern = r'.* Ù„Ù„ØµÙ'
    replacement = 'Ø§Ù„ØµÙ'
    returned_dict = { i['education_grade_id'] : re.sub(pattern, replacement, i['name'])  for i in grades_data}
    returned_dict[0] = 'Ø¨Ø¯ÙˆÙ† ØµÙ'
    return returned_dict

def create_excel_for_school_students_with_class_status(auth):
    classes_dictionary = get_classes_ids_with_names_dict(auth=auth)
    grades_dictionary = get_education_grade_id_with_grade_name_dic(auth=auth)


    classed_ss , unclassed_ss  = turn_classed_and_unclassed_students_to_diclist (get_school_classed_and_unclassed_students(auth))

    unique_classed_ss ,unique_unclassed_ss = [dict(t) for t in {tuple(d.items()) for d in classed_ss}] , [dict(t) for t in {tuple(d.items()) for d in unclassed_ss}]

    students = []

    for unclassed_student in unique_unclassed_ss:
        find_classed_student = [i for i in unique_classed_ss if i['student_id'] == unclassed_student['student_id']]    
        
        if len(find_classed_student):
            find_classed_student[0]['student_class_status'] = 'Ù…Ø´Ø¹Ø¨'
            find_classed_student[0]['grade_name'] =  grades_dictionary[find_classed_student[0]['grade_id']] 
            find_classed_student[0]['institution_class_name'] =  classes_dictionary[find_classed_student[0]['institution_class_id']]        
            students.append(find_classed_student[0])
        else:
            unclassed_student['grade_name'] =  grades_dictionary[unclassed_student['grade_id']] 
            unclassed_student['student_class_status'] = 'ØºÙŠØ± Ù…Ø´Ø¹Ø¨'
            students.append(unclassed_student)
    create_excel_from_data(students , 'send_folder/Ø§Ù„Ø·Ù„Ø§Ø¨ Ø¨Ø§Ù„Ø´Ø¹Ø¨.xlsx')

def divide_teacher_load(classes):
    pages = 0
    divided_lists = []
    current_list = []

    for _class in classes:
        _class_size = len(_class['students_data'])
        if _class_size > 25:
            pages += 4
        else:
            pages += 2
            
        if pages == 46:
            pages = 0
            divided_lists.append(current_list)
            if _class_size > 25:
                pages += 4
            else:
                pages += 2
            current_list = [_class]
        elif pages > 46:
            pages = 0
            divided_lists.append(current_list)
            if _class_size > 25:
                pages += 4
            else:
                pages += 2
            current_list = [_class]
        else:
            current_list.append(_class)
    # If there are remaining classes in current_list, add it to divided_lists
    if current_list:
        divided_lists.append(current_list)
        
    return divided_lists

def fill_official_marks_functions_wrapper_v2(username=None , password=None , outdir='./send_folder' , templet_file = './templet_files/official_marks_doc_a3_two_face_white_cover.ods',A3_context=True ,A4_context=True ,e_side_notebook_data=None ,empty_marks=False,divded_dfter_to_primary_and_secnedry=False, do_not_delete_send_folder=False,session = None, default=True):
    
    if A3_context :
        if default : 
            context = {'46': 'A6:A30', '4': 'A39:A63', '3': 'L6:L30', '45': 'L39:L63', '44': 'A71:A95', '6': 'A103:A127', '5': 'L71:L95', '43': 'L103:L127', '42': 'A135:A159', '8': 'A167:A191', '7': 'L135:L159', '41': 'L167:L191', '40': 'A199:A223', '10': 'A231:A255', '9': 'L199:L223', '39': 'L231:L255', '38': 'A263:A287', '12': 'A295:A319', '11': 'L263:L287', '37': 'L295:L319', '36': 'A327:A351', '14': 'A359:A383', '13': 'L327:L351', '35': 'L359:L383', '34': 'A391:A415', '16': 'A423:A447', '15': 'L391:L415', '33': 'L423:L447', '32': 'A455:A479', '18': 'A487:A511', '17': 'L455:L479', '31': 'L487:L511', '30': 'A519:A543', '20': 'A551:A575', '19': 'L519:L543', '29': 'L551:L575', '28': 'A583:A607', '22': 'A615:A639', '21': 'L583:L607', '27': 'L615:L639', '26': 'A647:A671', '24': 'A679:A703', '23': 'L647:L671', '25': 'L679:L703' ,'sheet_47': 'Ø§Ù„ØºÙ„Ø§Ù Ø§Ù„Ø¯Ø§Ø®Ù„ÙŠ' ,'sheet_47_plot': 'L37:L61'}
        else:
            context = context
    elif A4_context :
        if default:
            context = {'4': 'A39:A63','5': 'L39:L63','6': 'A71:A95','7': 'L71:L95','8': 'A103:A127','9': 'L103:L127','10': 'A135:A159','11': 'L135:L159','12': 'A167:A191','13': 'L167:L191','14': 'A199:A223','15': 'L199:L223','16': 'A231:A255','17': 'L231:L255','18': 'A263:A287','19': 'L263:L287','20': 'A295:A319','21': 'L295:L319','22': 'A327:A351','23': 'L327:L351','24': 'A359:A383','25': 'L359:L383','26': 'A391:A415','27': 'L391:L415','28': 'A423:A447','29': 'L423:L447','30': 'A455:A479','31': 'L455:L479','32':  'A487:A511','33':  'L487:L511','34':  'A519:A543','35':  'L519:L543','36':  'A551:A575','37':  'L551:L575','38':  'A583:A607','39':  'L583:L607','40':  'A615:A639','41':  'L615:L639','42':  'A647:A671','43':  'L647:L671','44':  'A679:A703','45':  'L679:L703' ,'46':  'A711:A735' ,'47':  'L711:L735'}
        else:
            context = A4_context
            
    if A3_context :
        paper_type = 'A3'
    else: 
        paper_type = 'A4'

    # ods_file = f'{ods_name}{ods_num}.ods'
    
    
    if (username is not None and password is not None ):
        auth = get_auth(username , password)
        period_id = get_curr_period(auth , session=session)['data'][0]['id']
        inst_id = inst_name(auth, session=session)['data'][0]['Institutions']['id']
        user_id = user_info(auth , username, session=session)['data'][0]['id']
        
        user = user_info(auth , username, session=session)
        school_name = inst_name(auth, session=session)['data'][0]['Institutions']['name']

        baldah = make_request(auth=auth , url=f'https://emis.moe.gov.jo/openemis-core/restful/Institution-Institutions.json?_limit=1&id={inst_id}&_contain=InstitutionLands.CustomFieldValues', session=session)['data'][0]['address'].split('-')[0]
        # grades= make_request(auth=auth , url='https://emis.moe.gov.jo/openemis-core/restful/Education.EducationGrades?_limit=0')
        
        school_place_data= make_request(auth=auth , url=f'https://emis.moe.gov.jo/openemis-core/restful/Institution-Institutions.json?_limit=1&id={inst_id}&_contain=InstitutionLands.CustomFieldValues', session=session)['data'][0]
        indcator_of_private_techers_sector=school_place_data['institution_sector_id']

        if indcator_of_private_techers_sector == 12 : 
            area_data = get_AreaAdministrativeLevels(auth, session=session)['data']
            area_chain_list = find_area_chain(school_place_data['area_administrative_id'], area_data).split(' - ')
            modeeriah_v2=area_chain_list[1]
            modeeriah=f'Ø§Ù„ØªØ¹Ù„ÙŠÙ… Ø§Ù„Ø®Ø§Øµ / {modeeriah_v2}'
        else:
            modeeriah = inst_area(auth, session=session)['data'][0]['Areas']['name']
            modeeriah=f'{modeeriah}'
                    

        school_year = get_curr_period(auth, session=session)['data']
        hejri1 = str(hijri_converter.convert.Gregorian(school_year[0]['start_year'], 1, 1).to_hijri().year)
        hejri2 =  str(hijri_converter.convert.Gregorian(school_year[0]['end_year'], 1, 1).to_hijri().year)
        melady1 = str(school_year[0]['start_year'])
        melady2 = str(school_year[0]['end_year'])
        teacher = f"{user['data'][0]['first_name']}  {user['data'][0]['last_name']}"
        
        
        classes_id_2 =[lst for lst in get_teacher_classes_v2(auth, inst_id , user_id ,period_id ,session=session)['data'] if lst]
        assessment_periods = make_request(auth =auth,url=f'https://emis.moe.gov.jo/openemis-core/restful/v2/Assessment-AssessmentPeriods.json?_limit=0' , session=session)
        grades_info = get_grade_info(auth)
        students_data_lists = get_marks_v2(auth ,inst_id , period_id , classes_id_2 , grades_info ,assessment_periods , session=session , empty_marks=empty_marks)
    else: 
        students_data_lists = e_side_notebook_data
        modeeriah = e_side_notebook_data['custom_shapes']['modeeriah']
        hejri1 = e_side_notebook_data['custom_shapes']['hejri1']
        hejri2 = e_side_notebook_data['custom_shapes']['hejri2']
        melady1 = e_side_notebook_data['custom_shapes']['melady1']
        melady2 = e_side_notebook_data['custom_shapes']['melady2']
        baldah = e_side_notebook_data['custom_shapes']['baldah']
        school_name = e_side_notebook_data['custom_shapes']['school']
        teacher = e_side_notebook_data['custom_shapes']['teacher']
        period_id = e_side_notebook_data['custom_shapes']['period_id']
        
    devided_teacher_load_list = divide_teacher_load(students_data_lists['file_data'])
    print('hi')

    
    custom_shapes = {
                    'modeeriah':modeeriah,
                    'hejri1': hejri1,
                    'hejri2': hejri2,
                    'melady1': melady1,
                    'melady2': melady2,
                    'baldah': baldah,
                    'school': school_name,
                    'teacher': teacher,
                    'modeeriah_20_2': f' {modeeriah}',
                    'hejri_20_1': hejri1,
                    'hejri_20_2': hejri2,
                    'melady_20_1': melady1,
                    'melady_20_2': melady2,
                    'hejri_20_5': hejri1,
                    'hejri_20_6': hejri2,
                    'melady_20_7': melady1,
                    'melady_20_8': melady2,        
                    'baldah_20_2': baldah,
                    'school_20_2': school_name,
                    'teacher_20_2': teacher,
                    'modeeriah_20_1': f'{modeeriah}',
                    'hejri1': hejri1,
                    'hejri2': hejri2,
                    'melady1': melady1,
                    'melady2': melady2,
                    'baldah_20_1': baldah,
                    'school_20_1': school_name,
                    'teacher_20_1': teacher,
                    'period_id': period_id
                    }
    
    primary_classes,other_classes=extract_primary_and_other_classes(devided_teacher_load_list)
    if divded_dfter_to_primary_and_secnedry : 
        
        if len(primary_classes) > 0 :
            templet_file='./templet_files/official_marks_document_from_grade_1-3_white_cover.ods'
            for counter , section in enumerate(devided_teacher_load_list, start=1 ):
                modified_classes = []
                primary_classes = ['Ø§Ù„ØµÙ Ø§Ù„Ø£ÙˆÙ„','Ø§Ù„ØµÙ Ø§Ù„Ø«Ø§Ù†ÙŠ','Ø§Ù„ØµÙ Ø§Ù„Ø«Ø§Ù„Ø«',]
                mawad = [i['subject_name'] for i in section]
                classes = [i['class_name'] for i in section]
                all_classes = [i['class_name'] for i in section]
                classes = [class_name for class_name in all_classes if any(primary_class in class_name for primary_class in primary_classes)]
                all_class_names = classes
                unique_class_names = set(all_class_names)
                unique_class_names_list = list(unique_class_names)
                filtered_basedOnPrimary_section = [class_data for class_data in section if any(primary_class in class_data['class_name'] for primary_class in primary_classes)]
                section=filtered_basedOnPrimary_section
                for i in unique_class_names_list: 
                    if '-' not in i:
                        i = ' '.join(i.split(' ')[0:-1])+'-'+i.split(' ')[-1]
                    modified_classes.append(get_class_short(i))
                modified_classes = ' ØŒ '.join(modified_classes)
                mawad = sorted(set(mawad))
                mawad = ' ØŒ '.join(mawad)

                custom_shapes['mawad'] = mawad
                custom_shapes['classes'] = modified_classes
                custom_shapes['classes_20_2'] = modified_classes
                custom_shapes['mawad_20_2'] = mawad
                custom_shapes['classes_20_1'] = modified_classes
                custom_shapes['mawad_20_1'] = mawad
                
                copy_ods_file(templet_file , f'{outdir}/{teacher}_Ø¬_{counter}.ods')
                fill_official_marks_v2(students_data_lists=section , ods_file=f'{outdir}/{teacher}_Ø¬_{counter}.ods' ,context=context, session=session)
                fill_custom_shape(doc= f'{outdir}/{teacher}_Ø¬_{counter}.ods' ,sheet_name= 'Ø§Ù„ØºÙ„Ø§Ù Ø§Ù„Ø¯Ø§Ø®Ù„ÙŠ' , custom_shape_values= custom_shapes , outfile=f'{outdir}/modified.ods')
                fill_custom_shape(doc=f'{outdir}/modified.ods', sheet_name='Ø§Ù„ØºÙ„Ø§Ù Ø§Ù„Ø§Ø²Ø±Ù‚', custom_shape_values=custom_shapes, outfile=f"{outdir}/final_{counter}")
                os.system(f'soffice --headless --convert-to pdf:writer_pdf_Export --outdir {outdir} {outdir}/final_{counter}')
                os.rename(f"{outdir}/final_{counter}", f"{outdir}/ Ø¯ÙØªØ± _Ø¹Ù„Ø§Ù…Ø§Øª_{teacher}_Ø¬Ø²Ø¡_{counter}_Ø§Ù„ØµÙÙˆÙ Ø§Ù„Ø§Ø¨ØªØ¯Ø§Ø¦ÙŠ{paper_type}.ods")
                os.rename(f"{outdir}/final_{counter}.pdf", f"{outdir}/Ø¯ÙØªØ± _Ø¹Ù„Ø§Ù…Ø§Øª_{teacher}_Ø¬Ø²Ø¡_{counter}_Ø§Ù„ØµÙÙˆÙ Ø§Ù„Ø§Ø¨ØªØ¯Ø§Ø¦ÙŠ{paper_type}.pdf")
            
        
        if len(other_classes) > 0 :
            
            for counter , section in enumerate(devided_teacher_load_list, start=1 ):
                modified_classes = []
                mawad = [i['subject_name'] for i in section]
                classes = [i['class_name'] for i in section]
                primary_classes = ['Ø§Ù„ØµÙ Ø§Ù„Ø£ÙˆÙ„','Ø§Ù„ØµÙ Ø§Ù„Ø«Ø§Ù†ÙŠ','Ø§Ù„ØµÙ Ø§Ù„Ø«Ø§Ù„Ø«',]
                all_classes = [i['class_name'] for i in section]
                classes = [class_name for class_name in all_classes if all(primary_class not in class_name for primary_class in primary_classes)]
                all_class_names = classes
                unique_class_names = set(all_class_names)
                unique_class_names_list = list(unique_class_names)
                other_classes = [class_data for class_data in section if all(primary_class not in class_data['class_name'] for primary_class in primary_classes)]
                section=other_classes
                for i in unique_class_names_list: 
                    if '-' not in i:
                        i = ' '.join(i.split(' ')[0:-1])+'-'+i.split(' ')[-1]
                    modified_classes.append(get_class_short(i))
                modified_classes = ' ØŒ '.join(modified_classes)
                mawad = sorted(set(mawad))
                unique_mawad_names=set(mawad)
                unique_mawad_names_list=list(unique_mawad_names)
                mawad=unique_mawad_names_list
                mawad = ' ØŒ '.join(mawad)
                

                custom_shapes['mawad'] = mawad
                custom_shapes['classes'] = modified_classes
                custom_shapes['classes_20_2'] = modified_classes
                custom_shapes['mawad_20_2'] = mawad
                custom_shapes['classes_20_1'] = modified_classes
                custom_shapes['mawad_20_1'] = mawad
                
                copy_ods_file(templet_file , f'{outdir}/{teacher}_Ø¬_{counter}.ods')
                fill_official_marks_v2(students_data_lists=section , ods_file=f'{outdir}/{teacher}_Ø¬_{counter}.ods' ,context=A3_context, session=session)
                fill_custom_shape(doc= f'{outdir}/{teacher}_Ø¬_{counter}.ods' ,sheet_name= 'Ø§Ù„ØºÙ„Ø§Ù Ø§Ù„Ø¯Ø§Ø®Ù„ÙŠ' , custom_shape_values= custom_shapes , outfile=f'{outdir}/modified.ods')
                fill_custom_shape(doc=f'{outdir}/modified.ods', sheet_name='Ø§Ù„ØºÙ„Ø§Ù Ø§Ù„Ø§Ø²Ø±Ù‚', custom_shape_values=custom_shapes, outfile=f"{outdir}/final_{counter}")
                os.system(f'soffice --headless --convert-to pdf:writer_pdf_Export --outdir {outdir} {outdir}/final_{counter}')
                os.rename(f"{outdir}/final_{counter}", f"{outdir}/Ø¯ÙØªØ± _Ø¹Ù„Ø§Ù…Ø§Øª_{teacher}_Ø¬Ø²Ø¡_{counter}_{paper_type}.ods")
                os.rename(f"{outdir}/final_{counter}.pdf", f"{outdir}/Ø¯ÙØªØ± _Ø¹Ù„Ø§Ù…Ø§Øª_{teacher}_Ø¬Ø²Ø¡_{counter}_{paper_type}.pdf")
    else :
        for counter , section in enumerate(devided_teacher_load_list, start=1 ):
                modified_classes = []
                if (username is not None and password is not None ):
                    mawad = [i['subject_name'] for i in section]
                    classes = [i['class_name'] for i in section]
                else:
                    classes = [i['class_name'].split('=')[0] for i in section]
                    mawad = [i['class_name'].split('=')[1] for i in section]
                
                all_class_names = classes
                unique_class_names = set(all_class_names)
                unique_class_names_list = list(unique_class_names)
                for i in unique_class_names_list: 
                    if '-' not in i:
                        i = ' '.join(i.split(' ')[0:-1])+'-'+i.split(' ')[-1]
                    modified_classes.append(get_class_short(i))
                modified_classes = ' ØŒ '.join(modified_classes)
                mawad = sorted(set(mawad))
                unique_mawad_names=set(mawad)
                unique_mawad_names_list=list(unique_mawad_names)
                mawad=unique_mawad_names_list
                mawad = ' ØŒ '.join(mawad)
                custom_shapes['mawad'] = mawad
                custom_shapes['classes'] = modified_classes
                custom_shapes['classes_20_2'] = modified_classes
                custom_shapes['mawad_20_2'] = mawad
                custom_shapes['classes_20_1'] = modified_classes
                custom_shapes['mawad_20_1'] = mawad
                
                copy_ods_file(templet_file , f'{outdir}/{teacher}_Ø¬_{counter}.ods')
                fill_official_marks_v2(students_data_lists=section , ods_file=f'{outdir}/{teacher}_Ø¬_{counter}.ods' ,context=context, session=session)
                fill_custom_shape(doc= f'{outdir}/{teacher}_Ø¬_{counter}.ods' ,sheet_name= 'Ø§Ù„ØºÙ„Ø§Ù Ø§Ù„Ø¯Ø§Ø®Ù„ÙŠ' , custom_shape_values= custom_shapes , outfile=f'{outdir}/modified.ods')
                fill_custom_shape(doc=f'{outdir}/modified.ods', sheet_name='Ø§Ù„ØºÙ„Ø§Ù Ø§Ù„Ø§Ø²Ø±Ù‚', custom_shape_values=custom_shapes, outfile=f"{outdir}/final_{counter}")
                os.system(f'soffice --headless --convert-to pdf:writer_pdf_Export --outdir {outdir} {outdir}/final_{counter}')
                os.rename(f"{outdir}/final_{counter}", f"{outdir}/Ø¯ÙØªØ± _Ø¹Ù„Ø§Ù…Ø§Øª_{teacher}_Ø¬Ø²Ø¡_{counter}_{paper_type}.ods")
                os.rename(f"{outdir}/final_{counter}.pdf", f"{outdir}/Ø¯ÙØªØ± _Ø¹Ù„Ø§Ù…Ø§Øª_{teacher}_Ø¬Ø²Ø¡_{counter}_{paper_type}.pdf")
    
    if not do_not_delete_send_folder :    
        delete_files_except(
                            [
                                i for i in os.listdir("./send_folder") 
                                            if "Ø¯ÙØªØ± _Ø¹Ù„Ø§Ù…Ø§Øª" in i
                            ]
                            , outdir)

def get_marks_v2(auth=None , inst_id=None , period_id=None , classes_id_2=None ,grades_info=None , assessment_periods=None , session=None,student_status_ids=[1] ,empty_marks=False):
    """
    Retrieves marks data for the specified classes and periods.    
    
    # Example usage:
    assessments_period_data = get_marks(
                                            auth, 
                                            inst_id,
                                            period_id, 
                                            classes_id_2,
                                            grades_info,
                                            insert_function=insert_to_e_side_marks_doc,
                                            existing_wb=your_existing_workbook,
                                            session=session
                                        )
                
                
    assessments_period_data_text = '\\\\'.join([str(list(dictionary.items())[0][0]) + ',' + ','.join(str(i) for i in list(dictionary.items())[0][1]) for dictionary in assessments_period_data])

    Parameters:
        auth (dict): Authentication data. Defaults to None.
        inst_id (str): Institution ID. Defaults to None.
        period_id (str): Period ID. Defaults to None.
        classes_id_2 (list): List of dictionaries containing class information. Defaults to None.
        grades_info (dict): Grades information. Defaults to None.
        assessment_periods (dict): Assessment periods information. Defaults to None.
        session: Session data. Defaults to None.

    Returns:
        list: List of dictionaries containing marks data for each class.    
    """
    classes_id_3, classes_data_and_marks= [] ,[]
    global secondery_students
    
    for class_info in classes_id_2:
        classes_id_3.append([{'institution_class_id': class_info['institution_class_id'] ,
                                'sub_name': class_info['institution_subject']['name'],
                                'class_name': class_info['institution_class']['name'] ,
                                'subject_id': class_info['institution_subject']['education_subject_id'] ,
                                'education_grade_id':class_info['institution_subject']['education_grade_id'],
                                'institution_subject_id': class_info['institution_subject_id']}])

    for v in range(len(classes_id_2)):
        # id institution_class_id
        institution_class_id = classes_id_3[v][0]['institution_class_id']
        
        # subject name 
        subject_name = classes_id_3[v][0]['sub_name'].replace("\\", "_")
        
        # class name
        class_name = classes_id_3[v][0]['class_name']
        
        # subject id 
        subject_id = classes_id_3[v][0]['subject_id']
        
        # institution subject id 
        institution_subject_id = classes_id_3[v][0]['institution_subject_id']
        
        # education grade id
        education_grade_id = classes_id_3[v][0]['education_grade_id']
        
        title = f'{class_name}={subject_name}={institution_class_id}={subject_id}'.replace('/', '~')
        
        try:
            # assessment id 
            assessment_id = offline_get_assessment_id_from_grade_id(education_grade_id ,grades_info)
            
            # print( institution_class_id ,subject_name,class_name,subject_id , institution_subject_id ,sep='\n')
            
            assessments_json = make_request(auth=auth , url=f'https://emis.moe.gov.jo/openemis-core/restful/v2/Institution-InstitutionSubjectStudents.json?_finder=StudentResults[institution_id:{inst_id};institution_class_id:{institution_class_id};assessment_id:{assessment_id};academic_period_id:{period_id};institution_subject_id:{institution_subject_id};education_grade_id:{education_grade_id}]&_limit=0&_contain=EducationSubjects',session=session)
        
            if 'Ø¹Ø´Ø±' in class_name :
                id_name_marks = get_secondery_students(auth,institution_class_id,inst_id=inst_id , curr_year=period_id ,student_status_ids=student_status_ids,session=session)
            else:
                id_name_marks = get_marks_and_names_dictionary_list(class_name , assessment_periods ,assessments_json ,empty_marks=empty_marks)
            classes_data_and_marks.append(
                                            {
                                                'title':title ,
                                                'students_data' :id_name_marks,
                                                'institution_class_id' :institution_class_id,
                                                'subject_name' :subject_name,
                                                'class_name' :class_name,
                                                'subject_id' :subject_id,
                                                'institution_subject_id' :institution_subject_id,
                                                'education_grade_id' :education_grade_id,
                                                'assessment_id':assessment_id
                                            }
                                        )
        except:
            print(title)
            pass
    return classes_data_and_marks

def fill_official_marks_v2(username=None, password=None , ods_file=None ,students_data_lists=None, context={} ,session=None ):
    """
    Fills the official marks document.

    Parameters:
        username (str): Username for authentication.
        password (str): Password for authentication.
        ods_file (str): Path to the ODS file.
        students_data_lists (list): List of dictionaries containing students' data. Defaults to None.
        context (dict): Contextual data. Defaults to {}.
        e_side_notebook_data: Data from e-Side notebook. Defaults to None.
        session: Session data. Defaults to None.

    Returns:
        dict: Custom shapes data.
        list: List of dictionaries containing students' data.
    """    
    context = context 
    page = 4
    name_counter = 1
    if username is not None and password is not None:
        auth = get_auth(username , password)
        period_id = get_curr_period(auth , session=session)['data'][0]['id']
        inst_id = inst_name(auth, session=session)['data'][0]['Institutions']['id']
        user_id = user_info(auth , username, session=session)['data'][0]['id']
        
        user = user_info(auth , username, session=session)
        school_name = inst_name(auth, session=session)['data'][0]['Institutions']['name']
        baldah = make_request(auth=auth , url=FILL_OFFICIAL_MARKS_DOC_WRAPPER_OFFLINE_URL_1.format(inst_id=inst_id), session=session)['data'][0]['address'].split('-')[0]
        # grades= make_request(auth=auth , url=FILL_OFFICIAL_MARKS_DOC_WRAPPER_OFFLINE_URL_2)
        modeeriah = inst_area(auth, session=session)['data'][0]['Areas']['name']
        school_year = get_curr_period(auth, session=session)['data']
        hejri1 = str(hijri_converter.convert.Gregorian(school_year[0]['start_year'], 1, 1).to_hijri().year)
        hejri2 =  str(hijri_converter.convert.Gregorian(school_year[0]['end_year'], 1, 1).to_hijri().year)
        melady1 = str(school_year[0]['start_year'])
        melady2 = str(school_year[0]['end_year'])
        teacher = user['data'][0]['name'].split(' ')[0]+' '+user['data'][0]['name'].split(' ')[-1]
        
        
        classes_id_2 =[lst for lst in get_teacher_classes_v2(auth, inst_id , user_id ,period_id ,session=session)['data'] if lst]
        assessment_periods = make_request(auth =auth,url=f'https://emis.moe.gov.jo/openemis-core/restful/v2/Assessment-AssessmentPeriods.json?_limit=0' , session=session)
        grades_info = get_grade_info(auth)
        students_data_lists = get_marks_v2(auth ,inst_id , period_id , classes_id_2 , grades_info ,assessment_periods , session=session)
    
    doc = ezodf.opendoc(ods_file)
    
    sheet_name = 'sheet'
    sheet = doc.sheets[sheet_name]

    for students_data_list in students_data_lists:
        
        # ['Ø§Ù„ØµÙ Ø§Ù„Ø³Ø§Ø¨Ø¹', 'Ø£', 'Ø§Ù„Ù„ØºØ© Ø§Ù„Ø§Ù†Ø¬Ù„ÙŠØ²ÙŠØ©', '786118']
        
        if username is None and password is None:        
            class_data = students_data_list['class_name'].split('=')
        else: 
            class_data = students_data_list['title'].split('=')[0:2]

        class_name = class_data[0].replace('Ø§Ù„ØµÙ ' , '').split('-')[0]
        class_char = class_data[0].split('-')[1]
        sub_name = class_data[1]
        
        sheet[f"D{int(context[str(page)].split(':')[0][1:])-5 }"].set_value(f' Ø§Ù„ØµÙ: {class_name}')
        sheet[f"I{int(context[str(page)].split(':')[0][1:])-5 }"].set_value(f'Ø§Ù„Ø´Ø¹Ø¨Ø© (   {class_char}    )')
        if page == 46:
            sheet_47 = doc.sheets[context['sheet_47']]
            sheet[f"O{int(context[str('sheet_47_plot')].split(':')[0][1:])-5}"].set_value(sub_name)
        else:
            sheet[f"O{int(context[str(page+1)].split(':')[0][1:])-5}"].set_value(sub_name)
                            

        for counter,student_info in enumerate(students_data_list['students_data'], start=1):
            if counter >= 26:
                page += 2
                counter = 1
                
                sheet[f"D{int(context[str(page)].split(':')[0][1:])-5}"].set_value(f' Ø§Ù„ØµÙ: {class_name}')
                sheet[f"I{int(context[str(page)].split(':')[0][1:])-5}"].set_value(f'Ø§Ù„Ø´Ø¹Ø¨Ø© (   {class_char}    )')
                # sheet[f"O{int(context[str(page+1)].split(':')[0][1:])-5}"].set_value(sub_name)
                #    Ø§Ù„Ù…Ø§Ø¯Ø© Ø§Ù„Ø¯Ø±Ø§Ø³ÙŠØ©
                
                # {'id': 3824166, 'name': 'Ù†ÙˆØ±Ø§Ù„Ø¯ÙŠÙ† Ù…Ø­Ù…ÙˆØ¯ Ø±Ø§Ø¶ÙŠ Ø§Ù„Ø¯ØºÙŠÙ…Ø§Øª', 'term1': {'assessment1': 9, 'assessment2': 10, 'assessment3': 11, 'assessment4': 20}}
                
                for student_info in students_data_list['students_data'][25:] :
                    row_idx = counter + int(context[str(page)].split(':')[0][1:]) - 1  # compute the row index based on the counter
                    sheet[f"A{row_idx}"].set_value(name_counter)
                    sheet[f"B{row_idx}"].set_value(student_info['name'])
                    if 'term1' in student_info and 'assessment1' in student_info['term1'] and 'assessment2' in student_info['term1'] and 'assessment3' in student_info['term1'] and 'assessment4' in student_info['term1']:
                        sheet[f"D{row_idx}"].set_value(student_info['term1']['assessment1'])
                        sheet[f"E{row_idx}"].set_value(student_info['term1']['assessment2'])
                        sheet[f"F{row_idx}"].set_value(student_info['term1']['assessment3'])
                        sheet[f"G{row_idx}"].set_value(student_info['term1']['assessment4'])
                    if 'term2' in student_info:
                        try :
                            if page == 46:
                                sheet_47 = doc.sheets[context['sheet_47']]
                                row_idx2 = counter + int(context['sheet_47_plot'].split(':')[0][1:]) - 1
                                sheet_47[f"O{int(context['sheet_47_plot'].split(':')[0][1:])-5}"].set_value(sub_name)
                                sheet_47[f"L{row_idx2}"].set_value(student_info['term2']['assessment1']) 
                                sheet_47[f"M{row_idx2}"].set_value(student_info['term2']['assessment2']) 
                                sheet_47[f"N{row_idx2}"].set_value(student_info['term2']['assessment3'])
                                sheet_47[f"O{row_idx2}"].set_value(student_info['term2']['assessment4'])
                            else:
                                raise ValueError("page is not forty six")                                
                        except:
                            sheet[f"O{int(context[str(page+1)].split(':')[0][1:])-5}"].set_value(sub_name)
                            row_idx2 = counter + int(context[str(page+1)].split(':')[0][1:]) - 1  # compute the row index based on the counter 
                            sheet[f"L{row_idx2}"].set_value(student_info['term2']['assessment1'])
                            sheet[f"M{row_idx2}"].set_value(student_info['term2']['assessment2'])
                            sheet[f"N{row_idx2}"].set_value(student_info['term2']['assessment3'])
                            sheet[f"O{row_idx2}"].set_value(student_info['term2']['assessment4'])
                    counter += 1
                    name_counter += 1              
                break
            row_idx = counter + int(context[str(page)].split(':')[0][1:]) - 1  # compute the row index based on the counter
            sheet[f"A{row_idx}"].set_value(name_counter)
            sheet[f"B{row_idx}"].set_value(student_info['name'])
            if 'term1' in student_info and 'assessment1' in student_info['term1'] and 'assessment2' in student_info['term1'] and 'assessment3' in student_info['term1'] and 'assessment4' in student_info['term1']:
                sheet[f"D{row_idx}"].set_value(student_info['term1']['assessment1']) 
                sheet[f"E{row_idx}"].set_value(student_info['term1']['assessment2']) 
                sheet[f"F{row_idx}"].set_value(student_info['term1']['assessment3'])
                sheet[f"G{row_idx}"].set_value(student_info['term1']['assessment4'])
            if 'term2' in student_info:
                try :
                    if 46 == page:
                        sheet_47 = doc.sheets[context['sheet_47']]
                        row_idx2 = counter + int(context['sheet_47_plot'].split(':')[0][1:]) - 1
                        sheet_47[f"O{int(context['sheet_47_plot'].split(':')[0][1:])-5}"].set_value(sub_name)
                        sheet_47[f"L{row_idx2}"].set_value(student_info['term2']['assessment1']) 
                        sheet_47[f"M{row_idx2}"].set_value(student_info['term2']['assessment2']) 
                        sheet_47[f"N{row_idx2}"].set_value(student_info['term2']['assessment3'])
                        sheet_47[f"O{row_idx2}"].set_value(student_info['term2']['assessment4'])
                    else:
                        raise ValueError("page is not forty six")
                except:
                    sheet[f"O{int(context[str(page+1)].split(':')[0][1:])-5}"].set_value(sub_name)
                    row_idx2 = counter + int(context[str(page+1)].split(':')[0][1:]) - 1  # compute the row index based on the counter 
                    sheet[f"L{row_idx2}"].set_value(student_info['term2']['assessment1'])
                    sheet[f"M{row_idx2}"].set_value(student_info['term2']['assessment2'])
                    sheet[f"N{row_idx2}"].set_value(student_info['term2']['assessment3'])
                    sheet[f"O{row_idx2}"].set_value(student_info['term2']['assessment4'])
            name_counter += 1 
        name_counter = 1
        page += 2

    # FIXME: make the customshapes crop _20_ to the rest of the key in the custom_shapes
    # Iterate through the cells of the sheet and fill in the values you want
    doc.save()

    #     # FIXME: make the customshapes crop _20_ to the rest of the key in the custom_shapes

    modified_classes = []
    if (username is not None and password is not None ):
        mawad = [i['subject_name'] for i in students_data_lists]
        classes = [i['class_name'] for i in students_data_lists]
    else:
        classes = [i['class_name'].split('=')[0] for i in students_data_lists]
        mawad = [i['class_name'].split('=')[1] for i in students_data_lists]
    for i in classes: 
        if '-' not in i:
            i = ' '.join(i.split(' ')[0:-1])+'-'+i.split(' ')[-1]
        modified_classes.append(get_class_short(i))
    modified_classes = ' ØŒ '.join(modified_classes)
    mawad = sorted(set(mawad))
    mawad = ' ØŒ '.join(mawad)

    if username is not None and password is not None:
        custom_shapes = {
            'modeeriah': f'{modeeriah}',
            'hejri1': hejri1,
            'hejri2': hejri2,
            'melady1': melady1,
            'melady2': melady2,
            'baldah': baldah,
            'school': school_name,
            'classes': modified_classes,
            'mawad': mawad,
            'teacher': teacher,
            'modeeriah_20_2': f'{modeeriah}',
            'hejri_20_1': hejri1,
            'hejri_20_2': hejri2,
            'melady_20_1': melady1,
            'melady_20_2': melady2,
            'hejri_20_5': hejri1,
            'hejri_20_6': hejri2,
            'melady_20_7': melady1,
            'melady_20_8': melady2,        
            'baldah_20_2': baldah,
            'school_20_2': school_name,
            'classes_20_2': modified_classes,
            'mawad_20_2': mawad,
            'teacher_20_2': teacher,
            'modeeriah_20_1': f'{modeeriah}',
            'hejri1': hejri1,
            'hejri2': hejri2,
            'melady1': melady1,
            'melady2': melady2,
            'baldah_20_1': baldah,
            'school_20_1': school_name,
            'classes_20_1': modified_classes,
            'mawad_20_1': mawad,
            'teacher_20_1': teacher,
            'period_id': period_id
        }
        #     # # FIXME: make the customshapes crop _20_ to the rest of the key in the custom_shapes
        #     # # Iterate through the cells of the sheet and fill in the values you want
        #     # doc.save()


        return custom_shapes , students_data_lists

def fill_official_marks_wrapper_v2(username , password , ods_name='send', outdir='./send_folder' ,ods_num=1 , templet_file = './templet_files/official_marks_doc_a3_two_face_white_cover.ods', color="#ffffff"):
    """
    Fills the official marks wrapper document.

    Parameters:
        username (str): Username for authentication.
        password (str): Password for authentication.
        ods_name (str): Name of the ODS file. Defaults to 'send'.
        outdir (str): Output directory path. Defaults to './send_folder'.
        ods_num (int): Number of the ODS file. Defaults to 1.
        templet_file (str): Path to the template ODS file. Defaults to './templet_files/official_marks_doc_a3_two_face_white_cover.ods'.
        color (str): Color for margins. Defaults to "#ffffff".
    """    
    ods_file = f'{ods_name}{ods_num}.ods'
    copy_ods_file(templet_file , f'{outdir}/{ods_file}')
    
    custom_shapes = fill_official_marks_a3_two_face_doc2(username= username, password= password , ods_file=f'{outdir}/{ods_file}')
    fill_custom_shape(doc= f'{outdir}/{ods_file}' ,sheet_name= 'Ø§Ù„ØºÙ„Ø§Ù Ø§Ù„Ø¯Ø§Ø®Ù„ÙŠ' , custom_shape_values= custom_shapes , outfile=f'{outdir}/modified.ods')
    fill_custom_shape(doc=f'{outdir}/modified.ods', sheet_name='Ø§Ù„ØºÙ„Ø§Ù Ø§Ù„Ø§Ø²Ø±Ù‚', custom_shape_values=custom_shapes, outfile=f'{outdir}/final_'+ods_file)
    os.system(f'soffice --headless --convert-to pdf:writer_pdf_Export --outdir {outdir} {outdir}/final_{ods_file} ')
    add_margins(f"{outdir}/final_{ods_name}{ods_num}.pdf", f"{outdir}/output_file.pdf",top_rec=30, bottom_rec=80, left_rec=1, right_rec=120, color_name=color)
    add_margins(f"{outdir}/output_file.pdf", f"{outdir}/{custom_shapes['teacher']}.pdf",page=1 , top_rec=60, bottom_rec=80, left_rec=1, right_rec=120, color_name=color)
    split_A3_pages(f"{outdir}/output_file.pdf" , outdir)
    reorder_official_marks_to_A4(f"{outdir}/output.pdf" , f"{outdir}/reordered.pdf")

    add_margins(f"{outdir}/reordered.pdf", f"{outdir}/output_file.pdf",top_rec=60, bottom_rec=50, left_rec=20, right_rec=20, color_name=color)
    add_margins(f"{outdir}/output_file.pdf", f"{outdir}/output_file1.pdf",page=1 , top_rec=100, bottom_rec=80, left_rec=90, right_rec=120, color_name=color)
    add_margins(f"{outdir}/output_file1.pdf", f"{outdir}/output_file2.pdf",page=50 , top_rec=100, bottom_rec=80, left_rec=70, right_rec=60, color_name=color)
    add_margins(f"{outdir}/output_file2.pdf", f"{outdir}/{custom_shapes['teacher']}_A4.pdf",page=51 , top_rec=100, bottom_rec=80, left_rec=90, right_rec=120, color_name=color)
    delete_files_except([f"{custom_shapes['teacher']}.pdf",f"{custom_shapes['teacher']}_A4.pdf",f'final_{ods_file}'], outdir)

def insert_to_side_marks_document_with_marks(title, class_name , assessments_json, assessments ,secandary_students , names_only = False , term = 1 , necessary_data_dict=None , outdir='./send_folder/',counter=None ,template_sheet_or_file=None):
    """
    Inserts marks data into a side marks document.

    Parameters:
        title (str): Title of the document.
        class_name (str): Name of the class.
        assessments_json (dict): JSON data containing assessments.
        assessments (list): List of assessments.
        secandary_students (list): List of secondary students.
        names_only (bool): Flag indicating whether to include only names or marks as well. Defaults to False.
        term (int): Term number. Defaults to 1.
        necessary_data_dict (dict): Dictionary containing necessary data. Defaults to None.
        outdir (str): Output directory path. Defaults to './send_folder/'.
        counter (int): Counter value. Defaults to None.
        template_sheet_or_file: Template sheet or file. Defaults to None.
    """
    context = {}
    marks_and_name = get_marks_and_names_dictionary_list(class_name , assessments ,assessments_json)

    melady1  = necessary_data_dict['melady1'] 
    melady2  = necessary_data_dict['melady2'] 
    userInfo  = necessary_data_dict['userInfo'] 
    modeeriah  = necessary_data_dict['modeeriah'] 
    school_name  = necessary_data_dict['school_name'] 
    
    if 'Ø¹Ø´Ø±' in class_name :
        counter2 = 0
        for item in marks_and_name :
            context[f'name{counter2}'] = item['name']
            counter2+=1 
    else:
        counter2 = 0
        for item in marks_and_name :
            context[f'name{counter2}'] = item['name']
            if not names_only :
                assessments = [
                            item[f'term{term}']['assessment1'],
                            item[f'term{term}']['assessment2'],
                            item[f'term{term}']['assessment3'],
                            item[f'term{term}']['assessment4']
                            ]
                context[f'A1_{counter2}'] = item[f'term{term}']['assessment1']
                context[f'A2_{counter2}'] = item[f'term{term}']['assessment2']
                context[f'A3_{counter2}'] = item[f'term{term}']['assessment3']
                context[f'A4_{counter2}'] = item[f'term{term}']['assessment4']
                SUM = sum(int(assessment) if assessment != '' else 0 for assessment in assessments)                    
                context[f'S_{counter2}'] = SUM if SUM !=0 else ''
                total = item[f'term{term}']['assessment3']

                try :                    
                    variables = [random.randint(3, min(total, 5)) for _ in range(3) if total > 0]
                    variables.append(total - sum(variables))
                    context[f'M1_{counter2}'] ,context[f'M2_{counter2}'] ,context[f'M3_{counter2}'] ,context[f'M4_{counter2}'] = variables
                except : 
                    context[f'M1_{counter2}'] ,context[f'M2_{counter2}'] ,context[f'M3_{counter2}'] ,context[f'M4_{counter2}'] =['']*4
            counter2+=1 
    context['teacher'] = userInfo['first_name']+' '+ userInfo['middle_name'] +' '+ userInfo['last_name']
    context[f'class_name'] = class_name
    context[f'term'] = 'Ø§Ù„Ø£ÙˆÙ„' if term == 1 else 'Ø§Ù„Ø«Ø§Ù†ÙŠ'
    context['school'] = school_name
    context['directory'] = modeeriah
    context['y1'] = melady1
    context['y2'] = melady2
    context['sub'] = title.split('=')[1]
    fill_doc(template_sheet_or_file , context , outdir+f'send{counter2}.docx' )
    context.clear()
    generate_pdf(outdir+f'send{counter2}.docx' , outdir ,counter2)
    delete_pdf_page(outdir+f'send{counter2}.pdf', outdir+f'SEND{counter2}.pdf', 1)
    delete_file(outdir+f'send{counter2}.pdf')

def get_marks_and_names_dictionary_list(class_name , assessment_periods ,assessments_json,empty_marks=False):
    """
    Extracts marks and names dictionary list from assessment data.

    Parameters:
        class_name (str): Name of the class.
        assessment_periods (dict): Dictionary containing assessment periods data.
        assessments_json (dict): JSON data containing assessments.

    Returns:
        list: List of dictionaries containing marks and names data.
    """    
    grouped_students = {key: list(group) for key, group in groupby(assessments_json['data'], key=lambda x: x['student_id'])}
    marks_and_names = []
    
    dic = {'id':'' ,'name': '','term1':{ 'assessment1': '' ,'assessment2': '' , 'assessment3': '' , 'assessment4': ''} ,'term2':{ 'assessment1': '' ,'assessment2': '' , 'assessment3': '' , 'assessment4': ''} ,'assessments_periods_ides':[]}
    
    for id , values in grouped_students.items():
        dic['id'] = id
        dic['name'] = values[0]['the_student_name']

        if 'Ø¹Ø´Ø±' not in class_name  :
            values = offline_sort_assessement_period_ids_v2( values , assessment_periods)
            dic['assessments_periods_ides'] = [int(x) for x in [i['assessment_period_id'] for i in values ] if x is not None]
            dic['term1']['assessment1'] = float(values[0]["mark"]) if values[0]["mark"] is not None and not empty_marks else ''
            dic['term1']['assessment2'] = float(values[1]["mark"]) if values[1]["mark"] is not None and not empty_marks else ''
            dic['term1']['assessment3'] = float(values[2]["mark"]) if values[2]["mark"] is not None and not empty_marks else ''
            dic['term1']['assessment4'] = float(values[3]["mark"]) if values[3]["mark"] is not None and not empty_marks else ''
            dic['term2']['assessment1'] = float(values[4]["mark"]) if values[4]["mark"] is not None and not empty_marks else ''
            dic['term2']['assessment2'] = float(values[5]["mark"]) if values[5]["mark"] is not None and not empty_marks else ''
            dic['term2']['assessment3'] = float(values[6]["mark"]) if values[6]["mark"] is not None and not empty_marks else ''
            dic['term2']['assessment4'] = float(values[7]["mark"]) if values[7]["mark"] is not None and not empty_marks else ''
        marks_and_names.append(dic)
        dic = {'id':'' ,'name': '','term1':{ 'assessment1': '' ,'assessment2': '' , 'assessment3': '' , 'assessment4': ''} ,'term2':{ 'assessment1': '' ,'assessment2': '' , 'assessment3': '' , 'assessment4': ''} ,'assessments_periods_ides':[]}
    
    marks_and_names = [d for d in marks_and_names if d['name'] != '']
    marks_and_names = sorted(marks_and_names, key=lambda x: x['name'])
    
    return marks_and_names

def get_teacher_classes_v2(auth,ins_id,staff_id,academic_period,session=None):
    """
    Get the classes taught by a teacher for a specific academic period.
    
    Parameters:
        auth (str): Authentication token.
        ins_id (int): Institution ID.
        staff_id (int): Staff ID of the teacher.
        academic_period (int): Academic period ID.
        session (object, optional): Requests session object. Defaults to None.
    
    Returns:
        dict: Teacher's classes data.
    """    
    subjects  = get_teacher_classes1(auth ,ins_id, staff_id, academic_period,session=session)
    subject_ids = sorted([i['institution_subject_id'] for i in subjects['data']])
    joined_subjects = ','.join(f'institution_subject_id:{id}' for id in subject_ids)
    url = f"https://emis.moe.gov.jo/openemis-core/restful/Institution.InstitutionClassSubjects?status=1&_contain=InstitutionSubjects,InstitutionClasses&_limit=0&_orWhere={joined_subjects}"
    return make_request(url,auth)

def offline_sort_assessement_period_ids_v2(marks_data ,assessments_periods , assessments_periods_ara_just_digits = False):
    """
    Offline sorting of assessment IDs based on their codes.

    Parameters:
    - assessment_id (str): Identifier for the assessment.
    - marks_data (list): List of dictionaries containing marks data.
    - assessments (dict): Dictionary containing assessment data.

    Returns:
    - list: Sorted list of dictionaries based on assessment codes.
    """    
    sorted_values =[]
    assessments_periods_dictionary = get_assessment_periods_dictionary_offline(assessments_periods)
    assessments_codes = {f'S{i}A{x}' : { 'term': "Ø§Ù„ÙØµÙ„ Ø§Ù„"+num2words(i,lang='ar', to='ordinal_num'), 'assessment_name':"Ø§Ù„ØªÙ‚ÙˆÙŠÙ… Ø§Ù„"+num2words(x,lang='ar', to='ordinal_num')} for i in [1,2] for x in [1,2,3,4]}
    for code in assessments_codes:
        # target_id = str([i['id'] for i in assessments_periods if code in i['code']][0])
        if assessments_periods_ara_just_digits: 
            target_value = [i for i in marks_data 
                                if code in assessments_periods_dictionary[i]['code'] ]
        else:
            target_value = [i for i in marks_data 
                                    if code in assessments_periods_dictionary[int(i['assessment_period_id'])]['code'] ]
        # Add code to each dictionary in target_value
        if not assessments_periods_ara_just_digits: 
            for item in target_value:
                # Add your code here
                item['code'] = code
        sorted_values.extend(target_value)
    # Check if the length is less than 8
    while len(sorted_values) < 8:
        # Add dictionaries with the value {'mark': None}
        if assessments_periods_ara_just_digits: 
            sorted_values.append(None)
        else : 
            sorted_values.append({'mark': None , 'assessment_period_id': None})
        
    return sorted_values

def get_assessment_periods_dictionary_offline(assessments_periods ):
    """
    Create a dictionary of assessment periods from offline data.

    Parameters:
        assessments_periods (dict): Assessment periods data.

    Returns:
        dict: Dictionary of assessment periods.
    """    
    return {
            i['id']:{
                    'code' : i['code'],
                    'name' : i['name'],
                    'academic_term' : i['academic_term'],
                    } 
            for i in assessments_periods['data']
            }

def insert_to_e_side_marks_doc(classes_data , template_sheet_or_file=None):
    """
    Insert marks data into the E-side marks document.

    Parameters:
        title (str): Title of the worksheet.
        class_name (str): Name of the class.
        assessments_json (dict): Assessments JSON data.
        assessments (dict): Assessments data.
        secondary_students (list): List of secondary students.
        necessary_data_dict (dict): Necessary data dictionary.
        counter (int): Counter value.
        template_sheet_or_file: Template sheet or file.
    """
    for class_data in classes_data:
        # copy the worksheet
        sheet_copy = template_sheet_or_file.copy_worksheet(template_sheet_or_file.active)
        marks_and_name = []

        # rename the new worksheet
        sheet_copy.title = f"{class_data['title'].split('=')[0]}={class_data['title'].split('=')[1]}"
        # sheet_copy.title = class_data['title']
        sheet_copy.sheet_view.rightToLeft = True
        
        # marks_and_name = get_marks_and_names_dictionary_list(class_name , assessments ,assessments_json)
        # marks_and_name = []
        data_font = Font(name='Arial', size=16, bold=False)
        
        
        sheet_copy.cell(row=1, column=39).value = f"{class_data['institution_class_id']}={class_data['subject_id']}"
        
        # print([d['name'] for d in class_data['students_data'] if d['name'] != ''])
        
        # class_data = {f'{institution_class_id}-{assessment_id}-{education_grade_id}' : '' if len(marks_and_name) == 0 else marks_and_name[0]['assessments_periods_ides']}
        # Write data to the worksheet and calculate the sum of some columns in each row
        for row_number, dataFrame in enumerate(class_data['students_data'], start=3):
            sheet_copy.cell(row=row_number, column=1).value = row_number-2
            sheet_copy.cell(row=row_number, column=2).value = dataFrame['id']
            sheet_copy.cell(row=row_number, column=3).value = dataFrame['name']
            sheet_copy.cell(row=row_number, column=4).value = dataFrame['term1']['assessment1']
            sheet_copy.cell(row=row_number, column=5).value = dataFrame['term1']['assessment2']
            sheet_copy.cell(row=row_number, column=6).value = dataFrame['term1']['assessment3']
            sheet_copy.cell(row=row_number, column=7).value = dataFrame['term1']['assessment4']
            sheet_copy.cell(row=row_number, column=8).value = f'=SUM(D{row_number}:G{row_number})'
            sheet_copy.cell(row=row_number, column=9).value = dataFrame['term2']['assessment1']
            sheet_copy.cell(row=row_number, column=10).value = dataFrame['term2']['assessment2']
            sheet_copy.cell(row=row_number, column=11).value = dataFrame['term2']['assessment3']
            sheet_copy.cell(row=row_number, column=12).value = dataFrame['term2']['assessment4']
            sheet_copy.cell(row=row_number, column=13).value = f'=SUM(I{row_number}:L{row_number})'
            sheet_copy.cell(row=row_number, column=14).value = f'=SUM(H{row_number},M{row_number})/2'
            # Set the font for the data rows
            for cell in sheet_copy[row_number]:
                cell.font = data_font    

def get_marks(auth=None , inst_id=None , period_id=None , classes_id_2=None ,grades_info=None , assessments = None , insert_function=None , existing_wb=None ,necessary_data_dict=None, session=None , template_sheet_or_file = None):
    """
    Get marks data for specified classes and subjects.
    
    # Example usage:
    assessments_period_data = get_marks(
                                            auth, 
                                            inst_id,
                                            period_id, 
                                            classes_id_2,
                                            grades_info,
                                            insert_function=insert_to_e_side_marks_doc,
                                            existing_wb=your_existing_workbook,
                                            session=session
                                        )
    assessments_period_data_text = '\\\\'.join([str(list(dictionary.items())[0][0]) + ',' + ','.join(str(i) for i in list(dictionary.items())[0][1]) for dictionary in assessments_period_data])

    Parameters:
        auth (str): Authentication token.
        inst_id (str): Institution ID.
        period_id (str): Period ID.
        classes_id_2 (list): List of dictionaries containing class information.
        grades_info (dict): Grades information.
        assessments (dict): Assessments data.
        insert_function (function): Function to insert marks data.
        existing_wb: Existing workbook.
        necessary_data_dict (dict): Necessary data dictionary.
        session: Session object.
        template_sheet_or_file: Template sheet or file.

    Returns:
        list: List of assessment period data.
    """
    classes_id_3,assessments_period_data ,secandary_students= [] ,[], []
    global secondery_students
    
    for class_info in classes_id_2:
        classes_id_3.append([{'institution_class_id': class_info['institution_class_id'] ,
                                'sub_name': class_info['institution_subject']['name'],
                                'class_name': class_info['institution_class']['name'] ,
                                'subject_id': class_info['institution_subject']['education_subject_id'] ,
                                'education_grade_id':class_info['institution_subject']['education_grade_id'],
                                'institution_subject_id': class_info['institution_subject_id']}])

    for v in range(len(classes_id_2)):
        # id institution_class_id
        institution_class_id = classes_id_3[v][0]['institution_class_id']
        # subject name 
        subject_name = classes_id_3[v][0]['sub_name'].replace("\\", "_")
        # class name
        class_name = classes_id_3[v][0]['class_name']
        # subject id 
        subject_id = classes_id_3[v][0]['subject_id']
        # institution subject id 
        institution_subject_id = classes_id_3[v][0]['institution_subject_id']
        # education grade id
        education_grade_id = classes_id_3[v][0]['education_grade_id']
        assessment_id = offline_get_assessment_id_from_grade_id(education_grade_id ,grades_info)
        
        print( institution_class_id ,subject_name,class_name,subject_id , institution_subject_id ,sep='\n')

        
        assessments_json = make_request(auth=auth , url=f'https://emis.moe.gov.jo/openemis-core/restful/v2/Institution-InstitutionSubjectStudents.json?_finder=StudentResults[institution_id:{inst_id};institution_class_id:{institution_class_id};assessment_id:{assessment_id};academic_period_id:{period_id};institution_subject_id:{institution_subject_id};education_grade_id:{education_grade_id}]&_limit=0&_contain=EducationSubjects',session=session)
        
        title = f'{class_name}={subject_name}={institution_subject_id}={subject_id}'.replace('/', '~')

        if 'Ø¹Ø´Ø±' in class_name :
            secandary_students = get_secondery_students(auth,institution_class_id,inst_id=inst_id , curr_year=period_id ,just_id_and_name_and_empty_marks=True,session=session)

        # if not touched by the user in the parameters 
        if not necessary_data_dict: 
            necessary_data_dict = {
                                    'assessment_id' : assessment_id ,
                                    'education_grade_i' : education_grade_id, 
                                    'institution_class_id' : institution_class_id ,
                                }
                        # insert_to_e_side_marks_doc(title , class_name ,marks_sheet_copy, assessments_json,)
        returned_data = insert_function(title, class_name , assessments_json, assessments,secandary_students ,necessary_data_dict=necessary_data_dict , counter = v , template_sheet_or_file=template_sheet_or_file)
        # Pass the insert_function as an argument to insert_to_e_side_marks_doc
        assessments_period_data.append(returned_data)
        
    return assessments_period_data

def teachers_marks_upload_percentage_wrapper_version_2(auth , first_term =False,second_term = False, both_terms=False ,inst_id=None , inst_nat=None , session=None ,curr_year =None, template='./templet_files/ÙƒØ´Ù Ù†Ø³Ø¨Ø© Ø§Ù„Ø§Ø¯Ø®Ø§Ù„ Ù…Ø¹Ø¯Ù„ Ù†Ø³Ø®Ø©_2.xlsx' ,outdir='./send_folder/' ,student_status_list=[1]):
    """
    Generate a report containing the upload percentages of teachers' marks and empty marks.

    Parameters:
        auth (str): Authentication token.
        first_term (bool): Flag indicating the first term. Default is False.
        second_term (bool): Flag indicating the second term. Default is False.
        both_terms (bool): Flag indicating both terms. Default is False.
        inst_id (str): Institution ID. Default is None.
        inst_nat (str): Institution nationality. Default is None.
        session: Session object. Default is None.
        template (str): Path to the template file. Default is './templet_files/ÙƒØ´Ù Ù†Ø³Ø¨Ø© Ø§Ù„Ø§Ø¯Ø®Ø§Ù„ Ù…Ø¹Ø¯Ù„ Ù†Ø³Ø®Ø©_2.xlsx'.
        outdir (str): Path to the output directory. Default is './send_folder/'.
    """    
    if curr_year is None :
        curr_year = get_curr_period(auth=auth,session=session)['data'][0]['id']
    
    if inst_id is None and inst_nat is None : 
        inst_id = inst_name(auth ,session=session)['data'][0]['Institutions']['id']

    data_dict = get_marks_upload_percentages_v2(auth , inst_id,curr_year,first_term =first_term,second_term = second_term, both_terms=both_terms , session=session , student_status_list=student_status_list)
    
    empty_marks_list = data_dict['school_percentage']['row_empty_marks']
    classes_data = data_dict['classes_percentages']
    teachers_percentages_data = data_dict['teachers_percentages']
    teachers_names = [i for i in data_dict['teachers_percentages']]
    teachers_names.sort()

    existing_WorkBook = load_workbook(template)
    school_percentage_WorkSheet = existing_WorkBook['Ù†Ø³Ø¨ Ø§Ù„Ø§Ø¯Ø®Ø§Ù„']
    school_percentage_WorkSheet['B3'] = int(data_dict['school_percentage']['percentage'])
    
    # Insert teachers percentages in the first page which is the active page
    for counter , teacher in enumerate(teachers_names , start=10):
        # Insert name
        school_percentage_WorkSheet[f'A{counter}'] = teacher
        # Insert percentage
        school_percentage_WorkSheet[f'B{counter}'] = teachers_percentages_data[teacher]['percentage']
        # Insert inserted marks number
        school_percentage_WorkSheet[f'C{counter}'] = teachers_percentages_data[teacher]['inserted_marks']
        # Insert empty marks number
        school_percentage_WorkSheet[f'D{counter}'] = teachers_percentages_data[teacher]['empty_marks']
    
    # Insert classes subjects marks percentage (for each class)
    for class_id in classes_data:
        class_sheet_copy = existing_WorkBook.copy_worksheet(existing_WorkBook['Sheet1'])
        class_data = classes_data[class_id]
        class_sheet_copy['N1'] = class_data['name']
        class_sheet_copy.title = class_data['name']
        class_sheet_copy.sheet_view.rightToLeft = True
        class_sheet_copy.sheet_view.rightToLeft = True
        class_subjects_ids = class_data['subjects_percentage']
        
        for counter , subject_id in enumerate(class_subjects_ids , start=5):
            subject_data = class_subjects_ids[subject_id]
            assessments_periods_percentages = subject_data['subject_marks_percentage']
            # subject name
            class_sheet_copy[f'B{counter}'] = subject_data['name']
            # first semester marks percentages
            class_sheet_copy[f'C{counter}'] = assessments_periods_percentages['S1A1']
            class_sheet_copy[f'D{counter}'] = assessments_periods_percentages['S1A2']
            class_sheet_copy[f'E{counter}'] = assessments_periods_percentages['S1A3']
            class_sheet_copy[f'F{counter}'] = assessments_periods_percentages['S1A4']
            # second semester marks percentages
            class_sheet_copy[f'G{counter}'] = assessments_periods_percentages['S2A1']
            class_sheet_copy[f'H{counter}'] = assessments_periods_percentages['S2A2']
            class_sheet_copy[f'I{counter}'] = assessments_periods_percentages['S2A3']
            class_sheet_copy[f'J{counter}'] = assessments_periods_percentages['S2A4']
            
            class_sheet_copy[f'K{counter}'] = subject_data['subject_teacher']
    
    empty_marks_sheet = existing_WorkBook.create_sheet("Ø§Ù„Ø¹Ù„Ø§Ù…Ø§Øª Ø§Ù„ÙØ§Ø±ØºØ©")
    empty_marks_sheet.sheet_view.rightToLeft = True
    
    
    header = [
                'student_id',
                'Ø§Ø³Ù… Ø§Ù„Ø·Ø§Ù„Ø¨',
                'Ø­Ø§Ù„Ø© Ø§Ù„Ø·Ø§Ù„Ø¨',
                'status_id',
                'Ø§Ø³Ù… Ø§Ù„Ù…Ø¹Ù„Ù…',
                'Ø±Ù‚Ù… Ø§Ù„ØµÙ',
                'Ø§Ù„ØµÙ+Ø§Ù„Ø´Ø¹Ø¨Ø©',
                'Ø±Ù‚Ù… Ø§Ù„Ù…Ø§Ø¯Ø©',
                'Ø§Ø³Ù… Ø§Ù„Ù…Ø§Ø¯Ø©',
                'Ø§Ù„ØªÙ‚ÙˆÙŠÙ…',
                'Ø§Ù„ÙØµÙ„',
                'code',
                'Ø§Ù„Ø¹Ù„Ø§Ù…Ø©'
            ]
    empty_marks_sheet.append(header)
    for data in empty_marks_list:
        values = [data.get(key, '') for key in header]
        empty_marks_sheet.append(values)    
    
    
    existing_WorkBook.remove(existing_WorkBook['Sheet1'])
    existing_WorkBook.save( outdir + f'Ù†Ø³Ø¨ Ø§Ù„Ø¹Ù„Ø§Ù…Ø§Øª Ù„Ù„Ø§Ø³Ø§Ø³ÙŠ.xlsx')

def calculate_percentage(part, whole):
    """
    Calculate the percentage.

    Parameters:
        part (float): Part value.
        whole (float): Whole value.

    Returns:
        float: Calculated percentage.
    """    
    if whole == 0:
        return 0
    return (part / whole) * 100

def inserted_marks_percentage_from_dataframes_variable_v2(marks , with_entered_and_not_marks=False , row_empty_marks=False):
    """
    Calculate the percentage of inserted marks from a list of marks.

    Parameters:
        marks (list): List of marks data.
        with_entered_and_not_marks (bool): Flag to include details of entered and empty marks. Default is False.
        row_empty_marks (bool): Flag to include the list of rows with empty marks. Default is False.

    Returns:
        float or dict: If with_entered_and_not_marks is False, returns the inserted marks percentage as float.
                        If with_entered_and_not_marks is True, returns a dictionary with percentage, inserted marks,
                        empty marks count, and optionally the list of rows with empty marks.
    """

    empty_marks = [mark for mark in marks if not isinstance(mark['Ø§Ù„Ø¹Ù„Ø§Ù…Ø©'], str)]
    inserted_marks = abs(len(empty_marks)-len(marks))
    inserted_marks_percentage = calculate_percentage(inserted_marks ,len(marks) )
    
    if with_entered_and_not_marks:
        if row_empty_marks:
            return {'percentage': inserted_marks_percentage,
                    'inserted_marks': inserted_marks,
                    'empty_marks': len(empty_marks),
                    'row_empty_marks': empty_marks
                    }
        else:
            return {'percentage': inserted_marks_percentage,
                    'inserted_marks': inserted_marks,
                    'empty_marks': len(empty_marks),
                    }            
    else:
        return inserted_marks_percentage

def create_fuzz_list(inst_id, period_id ,class_data_dic):
    """
    Create a fuzz list based on institution ID, academic period ID, and class data dictionary.

    Parameters:
        inst_id (int): The ID of the institution.
        period_id (int): The ID of the academic period.
        class_data_dic (dict): Dictionary containing class data.

    Returns:
        list: A list of fuzz strings.
    """
    _fuzz_list = []
    for class_id in class_data_dic:
        name = class_data_dic[class_id]['name']
        if 'Ø¹Ø´Ø±' not in name :
            class_subjects = class_data_dic[class_id]['subjects']
            assessment_id = class_data_dic[class_id]['assessment_id']
            education_grade_id = class_data_dic[class_id]['education_grade_id']    
            for subject in class_subjects :
                _fuzz_list.append(f'institution_id:{inst_id};institution_class_id:{class_id};assessment_id:{assessment_id};academic_period_id:{period_id};institution_subject_id:{subject["id"]};education_grade_id:{education_grade_id}')
    return _fuzz_list

def wfuzz_function_can_return_data(url,_fuzz_list , headers , body_postdata , method='POST' , proxies = None):
    """ 
    Ø¯Ø§Ù„Ø© Ø§Ø³ØªØ®Ø¯Ù…Ù‡Ø§ Ù„Ø§Ø±Ø³Ø§Ù„ Ø·Ù„Ø¨ Ø¨ÙˆØ³Øª Ø¨Ø´ÙƒÙ„ Ø³Ø±ÙŠØ¹ ØŒ Ùˆ Ø¨Ø¥Ù…ÙƒØ§Ù†Ù‡Ø§ Ø§Ù„Ø¹ÙˆØ¯Ø© Ø¨Ø¨ÙŠØ§Ù†Ø§Øª Ù…Ø¹ÙŠÙ†Ø© Ø§Ù„Ù…Ø·ÙˆØ± ÙŠØ­Ø¯Ø¯Ù‡Ø§

    Args:
        _fuzz_list (list): Ù‚Ø§Ø¦Ù…Ø© ÙÙŠ Ø¨ÙŠØ§Ù†Ø§Øª Ø§Ù„Ø·Ù„Ø§Ø¨ Ø§Ù„Ù…Ø±Ø§Ø¯ Ø§Ø¯Ø®Ø§Ù„Ù‡Ø§
        headers (tuple-list): Ø±Ø§Ø³ÙŠØ§Øª Ø§Ù„Ø·Ù„Ø¨ Ø§Ùˆ Ø§Ù„Ø±ÙƒÙˆÙŠØ³Øª
        body_postdata (str): Ø¬Ø³Ù… Ø§Ù„Ø¨ÙˆØ³Øª Ø¯Ø§ØªØ§
        method (str, optional): Ø·Ø±ÙŠÙ‚Ø© Ø§Ù„Ø·Ù„Ø¨. Defaults to 'POST'.

    Returns:
        any : ØªØ¹ÙˆØ¯ Ø¨Ù‚Ø§Ø¦Ù…Ø© Ø§Ù„Ø·Ù„Ø¨Ø§Øª ØºÙŠØ± Ø§Ù„Ù†Ø§Ø¬Ø­Ø©
    """    
    auth_header = [i[1] for i in headers if i[0] == 'Authorization'][0]
    if 'csrfToken' in auth_header or 'PHPSESSID' in auth_header or 'System' in auth_header:
        value = [i[1] for i in headers if i[0] == 'Authorization'][0]
        headers = [i for i in headers if i[0] != 'Authorization']
        headers.append(('Cookie' , value))
        
    unsuccessful_requests=[]
    _data=[]
    with tqdm(total=len(_fuzz_list), bar_format='{postfix[0]} {n_fmt}/{total_fmt}',
            postfix=["scraped schools", {"value": 0}]) as t:
            s = wfuzz.get_payloads([_fuzz_list])
            for idx , r in enumerate(s.fuzz(
                            url=url ,
                            # hc=[404] , 
                            # payloads=[("list",_fuzz_list)] ,
                            headers=headers ,
                            postdata = body_postdata ,
                            proxies= proxies ,
                            method= method,
                            # concurrent=100,
                            scanmode=True,
                            req_delay=1000000
                            ),start =1):
                    
                t.postfix[1]["value"] = idx
                t.update()
                try:
                    if r.history.code == 200 :
                        dict_content = json.loads(r.content)
                        _data.append(dict_content)
                except:
                    # if len(dict_content['data']):
                    print ('there is error at fuzz item : ' + r.description)
                    # there
                    pass
            #     print(r)
            #     print(r.content)
            #     print(r.history.code) # ÙƒÙˆØ¯ Ø§Ù„Ø±ÙƒÙˆÙŠØ³Øª
                if r.history.code != 200 :
                    if r.history.code in [i for i in range(499 , 600)]:
                        time.sleep(5)
                        print(r.history.code ,r.content)
                    unsuccessful_requests.append(r.description)                        
    return [unsuccessful_requests , _data]

def get_school_marks_version_2(auth , inst_id , period_id , _class_data_dic):
    """
    Retrieves school marks for a specific institution, academic period, and class data dictionary.

    Parameters:
        auth (dict): Authentication information.
        inst_id (int): The ID of the institution.
        period_id (int): The ID of the academic period.
        _class_data_dic (dict): Dictionary containing class data.

    Returns:
        list: A list of school marks.
    """
    _school_marks = []
    url = url=f'https://emis.moe.gov.jo/openemis-core/restful/v2/Institution-InstitutionSubjectStudents.json?_finder=StudentResults[FUZZ]&_limit=0'
    headers = [("User-Agent" , "python-requests/2.28.1"),("Accept-Encoding" , "gzip, deflate"),("Accept" , "*/*"),("Connection" , "close"),("Authorization" , f"{auth}"),("ControllerAction" , "Results"),("Content-Type" , "application/json")]

    _fuzz_list = create_fuzz_list(inst_id , period_id ,_class_data_dic)
    unsuccessful_requests , _data_list = wfuzz_function_can_return_data(url ,_fuzz_list,headers,body_postdata=None,method='GET')

    while len(unsuccessful_requests) != 0:
        requests  = wfuzz_function_can_return_data(url ,unsuccessful_requests,headers,body_postdata=None,method='GET')
        unsuccessful_requests = requests[0]
        _data_list.extend(requests[1])
    
    for i in _data_list:
        if len(i):
            try :
                if len(i['data']):
                    _school_marks.extend(i['data'])
            except:
                pass
    
    return _school_marks

def get_school_classes_and_students_with_classes(auth ,inst_id , period_id , session=None):
    """
    Retrieves information about school classes and students enrolled in those classes.

    Parameters:
        auth (dict): Authentication information.
        inst_id (int): The ID of the institution.
        period_id (int): The ID of the academic period.
        session (object, optional): Session information. Defaults to None.

    Returns:
        tuple: A tuple containing dictionaries for class names with their associated information and students with their associated data.
    """    
    grades_info = get_grade_info(auth , period_id)
    student_classess = make_request(auth=auth, url=f'https://emis.moe.gov.jo/openemis-core/restful/v2/Institution-InstitutionClassStudents.json?institution_id={inst_id}&academic_period_id={period_id}&_contain=Users&_limit=0', session=session)['data']
    class_names_dic = {i['institution_class_id'] :{'education_grade_id': i['education_grade_id']} for i in student_classess}
    students_with_data_dic = {i['student_id']:{'full_name':i['user']['name'] ,'status_id':i['student_status_id'] ,'class_id':i['institution_class_id']} for i in student_classess }
    classes = [i for i in class_names_dic]
    classes_str = ','.join([f'institution_class_id:{i}' for i in classes])
    url = f"https://emis.moe.gov.jo/openemis-core/restful/Institution.InstitutionClassSubjects?status=1&_contain=InstitutionSubjects,InstitutionClasses&_limit=0&_orWhere={classes_str}"
    classes_data = make_request(url=url,auth=auth,session=session)['data']
    for i in classes_data:
        class_names_dic[i['institution_class_id']]['name'] = i['institution_class']['name']
    for clas in class_names_dic:
        class_names_dic[clas]['assessment_id'] = offline_get_assessment_id_from_grade_id(class_names_dic[clas]['education_grade_id'] ,grades_info)
    return class_names_dic , students_with_data_dic

def get_marks_upload_percentages_v2(auth , inst_id , period_id ,first_term =False,second_term = False, both_terms=False, student_status_list = [1],subject_search_name_wanted_index = [2,3,5],session=None):
    """
    Retrieves marks upload percentages for students, teachers, and classes.

    Parameters:
        auth (dict): Authentication information.
        inst_id (int): The ID of the institution.
        period_id (int): The ID of the academic period.
        first_term (bool, optional): Whether to consider the first term. Defaults to False.
        second_term (bool, optional): Whether to consider the second term. Defaults to False.
        both_terms (bool, optional): Whether to consider both terms. Defaults to False.
        student_status_list (list, optional): List of student status IDs. Defaults to [1].
        subject_search_name_wanted_index (list, optional): List of indices for subject search names. Defaults to [2, 3, 5].
        session (object, optional): Session information. Defaults to None.

    Returns:
        dict: Dictionary containing school, teachers, and classes percentages along with data frames.
    """
    # function variables here 
    techers_percentages ,teachers_empty_marks,data_frames , subject_ids ,terms_list = {}, [], [], [] , []
    assessments_codes = {f'S{i}A{x}' : { 'term': "Ø§Ù„ÙØµÙ„ Ø§Ù„"+num2words(i,lang='ar', to='ordinal_num'), 'assessment_name':"Ø§Ù„ØªÙ‚ÙˆÙŠÙ… Ø§Ù„"+num2words(x,lang='ar', to='ordinal_num')} for i in [1,2] for x in [1,2,3,4]}
    search_names =['Ø±ÙŠØ§Ø¶ÙŠØ©', 'Ù†Ø´Ø§Ø·', 'Ù…Ø³ÙŠØ­ÙŠØ©', 'ÙÙ†', 'ÙØ±Ù†Ø³']
    
    search_names = [search_names[abs(i - 1)] for i in subject_search_name_wanted_index]
    unique_names = {}
    
    # Ø§Ø°Ø§ Ù„Ù… ÙŠØ®ØªØ± Ø§Ù„Ù…Ø³ØªØ®Ø¯Ù… Ø§Ù„ÙØµÙ„ Ø§Ù„Ø§ÙˆÙ„ Ø§Ø°Ø§ Ø§Ø®ØªØ± Ø§Ù„ÙØµÙ„ Ø§Ù„Ø«Ø§Ù†ÙŠ 
    # Ùˆ Ø§Ø°Ø§ Ù„Ù… ÙŠØ®ØªØ± Ø§Ù„ÙØµÙ„ Ø§Ù„Ø§ÙˆÙ„ ÙˆÙ„Ø§ Ø§Ù„Ø«Ø§Ù†ÙŠ 
    # Ø§Ø°Ø§ ÙˆØ§Ø®ØªØ§Ø± Ø§Ù„ÙØµÙ„ÙŠÙ† Ø§Ø°Ø§ Ø§Ø¸Ù‡Ø± Ù„Ù‡ Ù†ØªØ§Ø¦Ø¬ Ø§Ù„ÙØµÙ„ÙŠÙ†
    if first_term:
        terms_list = [i for i in assessments_codes if 'S1' in i]
        terms_list.extend('Assess')
    elif second_term:
        terms_list = [i for i in assessments_codes if 'S2' in i]
        terms_list.extend('Assess2')
    elif both_terms:
        terms_list = [i for i in assessments_codes]
        terms_list.extend(['Assess1' , 'Assess2'])

    # get the marks that the teachers uploaded on the emis site 
    # get the classes and the students 
    class_data_dic , students_with_data_dic = get_school_classes_and_students_with_classes(auth ,inst_id , period_id,session=session)
    
    # add subjects to the class dictionary variable which is class_data_dic
    class_data_with_subjects_dictionary = add_subjects_to_class_data_dic(auth,inst_id , period_id,class_data_dic,session=session)
    
    open_emis_core_marks = get_school_marks_version_2(auth,inst_id , period_id,class_data_dic)

    # get the teachers or staff data (what the subjects they teach and the class names)
    SubjectStaff_data = make_request(auth=auth , url=f'https://emis.moe.gov.jo/openemis-core/restful/v2/Institution-InstitutionSubjectStaff.json?institution_id={inst_id}&academic_period_id={period_id}&_contain=Users,InstitutionSubjects&_limit=0',session=session)['data']
    
    # get the assessment periods dictionary 
    assessment_periods_dictionary = get_assessment_periods_dictionary(auth)
    
    # map the followings 
    # teachers load  
    # subjects for each teacher  
    # the teacher with subjects
    staff_load_mapping = {
                        x['staff_id'] : {
                            'name': x['user']['name'],
                            'teacher_subjects':
                                [
                                    {
                                        'subject_class_id' :i['institution_subject']['id'] ,
                                        'subject_name' :i['institution_subject']['name'] ,
                                        'subject_grade_id' :i['institution_subject']['education_grade_id'],
                                        'subject_id' :i['institution_subject']['education_subject_id'] ,
                                    
                                    } for i in SubjectStaff_data if x['staff_id'] == i['staff_id']
                                ]
                            }
                        for x in SubjectStaff_data
                            if x['end_date'] is None
                        }
    subject_mapping_for_teachers = {
                                    i['id'] : { 
                                            'name': i['name'] , 
                                            'class_id': class_id ,
                                            'class_name' : class_data_dic[class_id]['name'] ,
                                            'education_subject_id': i['education_subject_id']
                                            }    
                                    for class_id in class_data_with_subjects_dictionary 
                                    for i in class_data_with_subjects_dictionary[class_id]['subjects']
                                    }
    teacher_with_subject_mapping = {
                                        i['subject_class_id'] : { 
                                                'teacher_name': staff_load_mapping[teacher_id]['name'] , 
                                                'education_subject_name': i['subject_name'],
                                                'education_subject_id': i['subject_id']
                                                }    
                                        for teacher_id in staff_load_mapping 
                                        for i in staff_load_mapping[teacher_id]['teacher_subjects']
                                    }
    class_subject_teacher_mapping = get_class_subject_teacher_mapping_dictionary( class_data_with_subjects_dictionary , subject_mapping_for_teachers , teacher_with_subject_mapping)
    
    # Create data_frames for these porposes :-
    # 1) writing the resulted marks in excel file  
    # 2)to get the percentages for the school ,and teachers , classes
    for student in students_with_data_dic :
        # FIXME: make execulding sacendary students option in the function
        if 'Ø¹Ø´Ø±' not in class_data_dic[students_with_data_dic[student]['class_id']]['name'] :
            #  Ø§Ø¨Ø­Ø« Ø¹Ù† Ø§Ù„Ø·Ø§Ù„Ø¨ ØµØ§Ø­Ø¨ Ø§Ù„Ø±Ù‚Ù… Ø§Ù„ØªØ¹Ø±ÙŠÙÙŠ 
            student_marks = [i for i in open_emis_core_marks if i['student_id']==student]
            # Ø§Ø¨Ø­Ø« ÙÙŠ ÙƒÙ„ Ø§Ù„Ù…ÙˆØ§Ø¯ Ø§Ù„ØªØ§Ù„ÙŠØ© 
            for subject in class_data_dic[students_with_data_dic[student]['class_id']]['subjects'] :
                subject_marks = [i for i in student_marks if i['education_subject_id'] ==int(subject['education_subject_id'])]

                # Ùˆ ØªØ­Ù‚Ù‚ Ù…Ù† ÙˆØ¬ÙˆØ¯ ÙƒÙˆØ¯ Ø§Ù„ØªÙ‚ÙˆÙŠÙ…Ø§Øª Ø§Ù„Ø«Ù…Ø§Ù†ÙŠØ© Ù„Ù„ÙØµÙ„ Ø§Ù„Ø§ÙˆÙ„ Ùˆ Ø§Ù„ÙØµÙ„ Ø§Ù„Ø«Ø§Ù†ÙŠ Ùˆ Ø§Ø°Ø§ Ù„Ù… ØªØ¬Ø¯ Ø§Ø±ØµØ¯ ÙØ§Ø±Øº Ù„Ù„Ø¹Ù„Ø§Ù…Ø© 
                for mark in subject_marks :
                    # Ø§Ù…Ø§ Ø§Ø°Ø§ ÙˆØ¬Ø¯Øª ÙØ§Ø±ØµØ¯ Ø¹Ù„Ø§Ù…Ø© Ø§Ù„Ø·Ø§Ù„Ø¨ Ø§Ù„Ø­Ù‚ÙŠÙ‚ÙŠØ©
                    student_class = students_with_data_dic[mark['student_id']]['class_id']
                    assessment_period_data = assessment_periods_dictionary[int(mark['assessment_period_id'])]
                    data_frames.append({
                        'student_id': mark['student_id'] ,
                        'Ø§Ø³Ù… Ø§Ù„Ø·Ø§Ù„Ø¨': students_with_data_dic[mark['student_id']]['full_name'],
                        'Ø­Ø§Ù„Ø© Ø§Ù„Ø·Ø§Ù„Ø¨': 'Ù…Ù„ØªØ­Ù‚' if mark['student_status_id'] in student_status_list else 'ØºÙŠØ± Ø°Ù„Ùƒ',
                        'status_id' : mark['student_status_id'] ,
                        'Ø§Ø³Ù… Ø§Ù„Ù…Ø¹Ù„Ù…':class_subject_teacher_mapping[student_class][mark['education_subject_id']]['teacher_name'] ,
                        'Ø±Ù‚Ù… Ø§Ù„ØµÙ' : student_class,
                        'Ø§Ù„ØµÙ+Ø§Ù„Ø´Ø¹Ø¨Ø©': class_subject_teacher_mapping[student_class][mark['education_subject_id']]['class_name'],
                        'Ø±Ù‚Ù… Ø§Ù„Ù…Ø§Ø¯Ø©' : mark['education_subject_id'],
                        'Ø§Ø³Ù… Ø§Ù„Ù…Ø§Ø¯Ø©':class_subject_teacher_mapping[student_class][mark['education_subject_id']]['name'],
                        'Ø§Ù„ØªÙ‚ÙˆÙŠÙ…': assessment_period_data['name'],
                        'Ø§Ù„ÙØµÙ„': assessment_period_data['academic_term'],
                        'code':assessment_period_data['code'],
                        'Ø§Ù„Ø¹Ù„Ø§Ù…Ø©': mark['mark']
                    })

    # Code that i wrote to skip some subjects from the percentages
    subjects_dictionary_list = [
                                    {
                                        'name':values['name'] ,
                                        'education_subject_id':values['education_subject_id']
                                    } for values in subject_mapping_for_teachers.values()
                                ]
        # Create a Wiktionary to track unique names and their IDs
    sorted_unique_data = sorted({item['education_subject_id']: item for item in subjects_dictionary_list}.values(), key=lambda x: x['education_subject_id'])
    # Iterate through the data and update the dictionary
    for item in sorted_unique_data:
        stripped_name = item['name'].strip()
        if stripped_name in unique_names:
            unique_names[stripped_name].append(item['education_subject_id'])
        else:
            unique_names[stripped_name] = [item['education_subject_id']]
    for name, ids in unique_names.items():
        if any(search_name in name for search_name in search_names):
            ids = [int(i) for i in ids]
            subject_ids.extend(ids)


    # get the teachers percentage and the uploaded marks and unuploaded marks
    teachers_names = list(set(i['Ø§Ø³Ù… Ø§Ù„Ù…Ø¹Ù„Ù…'] for i in data_frames))
    teacher_marks = {
                        i: {'row_marks':[
                                        {
                                            'student_id': x['student_id'],
                                            'Ø§Ø³Ù… Ø§Ù„Ø·Ø§Ù„Ø¨': x['Ø§Ø³Ù… Ø§Ù„Ø·Ø§Ù„Ø¨'],
                                            'Ø­Ø§Ù„Ø© Ø§Ù„Ø·Ø§Ù„Ø¨': x['Ø­Ø§Ù„Ø© Ø§Ù„Ø·Ø§Ù„Ø¨'],
                                            'status_id' : x['status_id' ],
                                            'Ø§Ø³Ù… Ø§Ù„Ù…Ø¹Ù„Ù…': x['Ø§Ø³Ù… Ø§Ù„Ù…Ø¹Ù„Ù…'],
                                            'Ø±Ù‚Ù… Ø§Ù„ØµÙ' : x['Ø±Ù‚Ù… Ø§Ù„ØµÙ' ],
                                            'Ø§Ù„ØµÙ+Ø§Ù„Ø´Ø¹Ø¨Ø©': x['Ø§Ù„ØµÙ+Ø§Ù„Ø´Ø¹Ø¨Ø©'],
                                            'Ø±Ù‚Ù… Ø§Ù„Ù…Ø§Ø¯Ø©' : x['Ø±Ù‚Ù… Ø§Ù„Ù…Ø§Ø¯Ø©' ],
                                            'Ø§Ø³Ù… Ø§Ù„Ù…Ø§Ø¯Ø©': x['Ø§Ø³Ù… Ø§Ù„Ù…Ø§Ø¯Ø©'],
                                            'Ø§Ù„ØªÙ‚ÙˆÙŠÙ…': x['Ø§Ù„ØªÙ‚ÙˆÙŠÙ…'],
                                            'Ø§Ù„ÙØµÙ„': x['Ø§Ù„ÙØµÙ„'],
                                            'code': x['code'],
                                            'Ø§Ù„Ø¹Ù„Ø§Ù…Ø©': x['Ø§Ù„Ø¹Ù„Ø§Ù…Ø©']
                                        } for x in data_frames 
                                            if x['Ø§Ø³Ù… Ø§Ù„Ù…Ø¹Ù„Ù…'] == i
                                            # FIXME: try to find other way to get the term percentage and student status
                                            and
                                            any(term_item in x['code'] for term_item in terms_list)
                                            and
                                            x['status_id'] in student_status_list
                                        ]}
                        for i in teachers_names 
                    }

    for teacher in teacher_marks :
        marks = teacher_marks[teacher]['row_marks']
        # print(teacher ,inserted_marks_percentage_from_dataframes_variable(marks))
        teachers_empty_marks.extend(inserted_marks_percentage_from_dataframes_variable_v2(marks ,True, True)['row_empty_marks'])
        techers_percentages[teacher] = inserted_marks_percentage_from_dataframes_variable_v2(marks ,True)

    classes_with_subjects_percentage = class_data_dic

    for class_number in class_data_dic:
        class_subjects_dict = {
                                int(i['education_subject_id']) : {
                                                                    'name' :
                                                                        i['name']
                                                                    }
                                    for i in class_data_dic[class_number]['subjects']
                            }
        class_subjects_dict = dict(sorted(class_subjects_dict.items(), key=lambda item: item[0]))
        class_marks = [i for i in data_frames if i['Ø±Ù‚Ù… Ø§Ù„ØµÙ'] == class_number and "ØºÙŠØ± Ø°Ù„Ùƒ" not in i['Ø­Ø§Ù„Ø© Ø§Ù„Ø·Ø§Ù„Ø¨'] ]
        for subject_id in class_subjects_dict:
            subject_data_list = [i for i in class_marks if i['Ø±Ù‚Ù… Ø§Ù„Ù…Ø§Ø¯Ø©'] == subject_id]
            assessments_dict = {i:'' for i in assessments_codes}        
            for assessment in assessments_codes:
                # FIXME: make it compare the all cases of academic term with all possible strings
                assessment_marks = [i for i in subject_data_list if assessment in i['code'] ]
                assessments_dict[assessment] = inserted_marks_percentage_from_dataframes_variable_v2(assessment_marks)
                
            class_subjects_dict[subject_id]['subject_marks_percentage']  = assessments_dict
            class_subjects_dict[subject_id]['subject_teacher'] = class_subject_teacher_mapping[class_number][subject_id]['teacher_name']
            
        classes_with_subjects_percentage[class_number]['subjects_percentage'] = class_subjects_dict
        class_subjects_dict = {}

    school_marks = [
                    {
                        'student_id': x['student_id'],
                        'Ø§Ø³Ù… Ø§Ù„Ø·Ø§Ù„Ø¨': x['Ø§Ø³Ù… Ø§Ù„Ø·Ø§Ù„Ø¨'],
                        'Ø­Ø§Ù„Ø© Ø§Ù„Ø·Ø§Ù„Ø¨': x['Ø­Ø§Ù„Ø© Ø§Ù„Ø·Ø§Ù„Ø¨'],
                        'status_id' : x['status_id' ],
                        'Ø§Ø³Ù… Ø§Ù„Ù…Ø¹Ù„Ù…': x['Ø§Ø³Ù… Ø§Ù„Ù…Ø¹Ù„Ù…'],
                        'Ø±Ù‚Ù… Ø§Ù„ØµÙ' : x['Ø±Ù‚Ù… Ø§Ù„ØµÙ' ],
                        'Ø§Ù„ØµÙ+Ø§Ù„Ø´Ø¹Ø¨Ø©': x['Ø§Ù„ØµÙ+Ø§Ù„Ø´Ø¹Ø¨Ø©'],
                        'Ø±Ù‚Ù… Ø§Ù„Ù…Ø§Ø¯Ø©' : x['Ø±Ù‚Ù… Ø§Ù„Ù…Ø§Ø¯Ø©' ],
                        'Ø§Ø³Ù… Ø§Ù„Ù…Ø§Ø¯Ø©': x['Ø§Ø³Ù… Ø§Ù„Ù…Ø§Ø¯Ø©'],
                        'Ø§Ù„ØªÙ‚ÙˆÙŠÙ…': x['Ø§Ù„ØªÙ‚ÙˆÙŠÙ…'],
                        'Ø§Ù„ÙØµÙ„': x['Ø§Ù„ÙØµÙ„'],
                        'code': x['code'],
                        'Ø§Ù„Ø¹Ù„Ø§Ù…Ø©': x['Ø§Ù„Ø¹Ù„Ø§Ù…Ø©']
                    }
                    for x in data_frames 
                        if any(term_item in x['code'] for term_item in terms_list)
                        and
                        x['status_id'] in student_status_list
                        and
                        x['Ø±Ù‚Ù… Ø§Ù„Ù…Ø§Ø¯Ø©'] not in subject_ids
                    ]
    return {
            'school_percentage' : inserted_marks_percentage_from_dataframes_variable_v2(school_marks , True ,True),
            'teachers_percentages' :techers_percentages,
            'classes_percentages' : classes_with_subjects_percentage,
            'data_frames' : data_frames
            }

def get_class_subjects(auth , class_id , assessment_id , academic_period_id , institution_id , session=None):
    """
    Retrieves subjects for a specific class.

    Parameters:
        auth (dict): Authentication information.
        class_id (int): The ID of the class.
        assessment_id (int): The ID of the assessment.
        academic_period_id (int): The ID of the academic period.
        institution_id (int): The ID of the institution.
        session (object, optional): Session information. Defaults to None.

    Returns:
        list: A list of dictionaries containing information about the subjects for the specified class.
    """    
    url = GET_SUB_INFO_URL.format(class_id=class_id,assessment_id=assessment_id,academic_period_id=academic_period_id,institution_id=institution_id)
    return make_request(auth=auth , url=url , session=session)['data']

def get_class_subject_teacher_mapping_dictionary(_class_data_with_subjects_dictionary , _subject_mapping_for_teachers ,_teacher_with_subject_mapping):
    """
    Creates a dictionary mapping classes to their subjects and teachers.

    Parameters:
        _class_data_with_subjects_dictionary (dict): Dictionary containing class data with subjects.
        _subject_mapping_for_teachers (dict): Mapping of subjects to teachers.
        _teacher_with_subject_mapping (dict): Mapping of teachers to subjects.

    Returns:
        dict: A dictionary where keys are class IDs and values are dictionaries containing subject IDs mapped to their names, class IDs, class names, and teacher names.
    """    
    _class_subject_teacher_mapping = {}
    _class_subjects_linked_to_teacher =[]
    for class_id in _class_data_with_subjects_dictionary:
        _class_subjects_linked_to_teacher.clear()
        class__subject_ids =[i['id'] for i in _class_data_with_subjects_dictionary[class_id]['subjects']]
        for education_subject_id in class__subject_ids:
            try:
                _class_subjects_linked_to_teacher.append(_teacher_with_subject_mapping[int(education_subject_id)])
            except KeyError:
                subject_data = _subject_mapping_for_teachers[education_subject_id]
                _class_subjects_linked_to_teacher.append({
                                                            'education_subject_name': subject_data['name'],
                                                            'class_id': subject_data['class_id'],
                                                            'class_name': subject_data['class_name'],
                                                            'education_subject_id': subject_data['education_subject_id'],
                                                            'teacher_name': 'Ø¨Ø¯ÙˆÙ† Ù…Ø¹Ù„Ù…',
                                                        })

        
        _class_subject_teacher_mapping[class_id] = {
                                                    int(i['education_subject_id']):
                                                                                    {
                                                                                        'name': i['education_subject_name'],
                                                                                        'class_id': class_id,
                                                                                        'class_name': _class_data_with_subjects_dictionary[class_id]['name'],
                                                                                        'teacher_name': i['teacher_name'],
                                                                                    }
                                                                                        for i in _class_subjects_linked_to_teacher
                                                    }
    return _class_subject_teacher_mapping

def get_subject_dictionary(_class_data_with_subjects_dictionary):
    """
    Extracts a dictionary of subjects from a dictionary containing class data with subjects.
    
    # Example flatten list of lists
    matrix = [
        [1, 2, 3],
        [4, 5, 6],
        [7, 8, 9]
    ]
    numbers =[]
    for row in matrix:
        for col in row:
            numbers.append(col)
    print(numbers)
    # or 
    # for row in matrix
    # for col in row
    # and what we want is the col
    print([col for row in matrix for col in row])


    Parameters:
        _class_data_with_subjects_dictionary (dict): Dictionary containing class data with subjects.

    Returns:
        dict: A dictionary where keys are education_subject_ids and values are subject names.
    """
    return {
            int(subject_data['education_subject_id']): subject_data['name'] 
                for class_id in _class_data_with_subjects_dictionary 
                for subject_data in _class_data_with_subjects_dictionary[class_id]['subjects']
            }

def get_assessment_periods_list(auth , session=None):
    """
    Retrieves a list of assessment periods.

    Parameters:
        auth (dict): Authentication information.
        session (object, optional): Session information. Defaults to None.

    Returns:
        list: A list of dictionaries containing information about assessment periods.
    """    
    assessment_periods = make_request(auth =auth,url=f'https://emis.moe.gov.jo/openemis-core/restful/v2/Assessment-AssessmentPeriods.json?_limit=0' , session=session)
    return assessment_periods['data']

def get_assessment_periods_dictionary(auth ):
    """
    Retrieves assessment periods and creates a dictionary with their information.

    Parameters:
        auth (dict): Authentication information.

    Returns:
        dict: A dictionary containing assessment period IDs as keys and dictionaries containing the following keys:
            - 'code': The code of the assessment period.
            - 'name': The name of the assessment period.
            - 'academic_term': The academic term of the assessment period.
    """    
    return {
            i['id']:{
                    'code' : i['code'],
                    'name' : i['name'],
                    'academic_term' : i['academic_term'],
                    } 
            for i in get_assessment_periods_list(auth)
            }

def add_subjects_to_class_data_dic(auth , inst_id ,period_id , _class_data_dic ,session=None):
    """
    Adds subjects to the class_data_dic.

    Parameters:
        auth (dict): Authentication information.
        inst_id (int): Institution ID.
        period_id (int): Period ID.
        _class_data_dic (dict): Dictionary containing class data.
        session (object, optional): Session information. Defaults to None.

    Returns:
        dict: The updated class_data_dic with subjects added.
    """
    # add subjects to class_data_dic
    for class_ in _class_data_dic:
        class_subject_data = get_class_subjects(auth ,class_ , _class_data_dic[class_]['assessment_id'] ,period_id, inst_id ,session=session)
        _class_data_dic[class_]['subjects'] = [i['InstitutionSubjects'] for i in class_subject_data]
    return _class_data_dic

def get_school_marks(auth , inst_id , period_id , limit =1000,session = None):
    """
    Creates an Excel file containing student marks.

    Parameters:
        data_frames (list of dict): A list of dictionaries representing student marks.
        excel_file_name (str, optional): The name of the Excel file to be created. Defaults to 'Ø¹Ù„Ø§Ù…Ø§Øª Ø§Ù„Ø·Ù„Ø§Ø¨ Ø§Ù„Ø¯Ù‚ÙŠÙ‚Ø©.xlsx'.
    """
    start_page = 1
    school_marks = []
    url = f'https://emis.moe.gov.jo/openemis-core/restful/Assessment.AssessmentItemResults?_fields=created_user_id,AssessmentGradingOptions.name,AssessmentGradingOptions.min,AssessmentGradingOptions.max,EducationSubjects.name,EducationSubjects.code,AssessmentPeriods.code,AssessmentPeriods.name,AssessmentPeriods.academic_term,marks,assessment_grading_option_id,student_id,assessment_id,education_subject_id,education_grade_id,assessment_period_id,institution_classes_id&academic_period_id={period_id}&_contain=AssessmentPeriods,AssessmentGradingOptions,EducationSubjects&institution_id={inst_id}'+'&_limit=1'
    total = make_request(auth=auth,url=url, session=session)['total']

    # +2 because of the range in python 
    end_page = int(total/limit)+2
    pages = [i for i in  range(start_page , end_page)]
    headers = [("User-Agent" , "python-requests/2.28.1"),("Accept-Encoding" , "gzip, deflate"),("Accept" , "*/*"),("Connection" , "close"),("Authorization" , f"{auth}"),("ControllerAction" , "Results"),("Content-Type" , "application/json")]
    url = url + f'&_limit={limit}&_page=FUZZ'

    unsuccessful_requests , data_list = wfuzz_function_can_return_data(url ,pages,headers,body_postdata=None,method='GET')

    while len(unsuccessful_requests) != 0:
        requests  = wfuzz_function_can_return_data(url ,unsuccessful_requests,headers,body_postdata=None,method='GET')
        unsuccessful_requests = requests[0]
        data_list.append(requests[1])


    for i in data_list:
        if len(i['data']):
            school_marks.extend(i['data'])
    
    return school_marks

def create_excel_from_data(data_frames , excel_file_name = 'Ø¹Ù„Ø§Ù…Ø§Øª Ø§Ù„Ø·Ù„Ø§Ø¨ Ø§Ù„Ø¯Ù‚ÙŠÙ‚Ø©.xlsx' ,rtl=True):
    """
    Calculates the percentage of a part compared to a whole.

    Parameters:
        part (int or float): The part of the whole.
        whole (int or float): The total value representing the whole.

    Returns:
        float: The percentage of the part compared to the whole. If whole is 0, returns 0.
    """    
    # Convert the list of dictionaries to a pandas DataFrame
    df = pd.DataFrame(data_frames)
    # Write the DataFrame to an Excel file
    df.to_excel(excel_file_name, index=False)
    
    if rtl:
        # Load the workbook
        workbook = load_workbook(excel_file_name)

        # Set right-to-left direction
        for sheetname in workbook.sheetnames:
            sheet = workbook[sheetname]
            sheet.sheet_view.rightToLeft = True

        # Save the Excel file
        workbook.save(excel_file_name)

    print(f"Excel file '{excel_file_name}' has been created.")

def calculate_percentage(part, whole):
    if whole == 0:
        return 0
    return (part / whole) * 100

def inserted_marks_percentage_from_dataframes_variable(marks , with_entered_and_not_marks=False):
    """
    Calculates the percentage of inserted marks from a list of marks.

    Parameters:
        marks (list): A list of dictionaries representing marks, where each dictionary contains the keys 'Ø§Ù„Ø¹Ù„Ø§Ù…Ø©' representing the mark.
        with_entered_and_not_marks (bool, optional): If True, returns additional information including the percentage of inserted marks, the number of inserted marks, and the list of empty marks. Defaults to False.

    Returns:
        float or dict: If with_entered_and_not_marks is False, returns the percentage of inserted marks as a float. If True, returns a dictionary containing the following keys:
            - 'percentage': The percentage of inserted marks.
            - 'inserted_marks': The number of inserted marks.
            - 'empty_marks': A list of dictionaries representing marks with no integer value.
    """    
    empty_marks = [mark for mark in marks if not isinstance(mark['Ø§Ù„Ø¹Ù„Ø§Ù…Ø©'], int)]
    inserted_marks = abs(len(empty_marks)-len(marks))
    inserted_marks_percentage = calculate_percentage(inserted_marks ,len(marks) )
    
    if with_entered_and_not_marks:
        return {'percentage': inserted_marks_percentage,
                'inserted_marks': inserted_marks,
                'empty_marks': empty_marks
                }
    else:
        return inserted_marks_percentage

def get_marks_upload_percentages(auth , inst_id , period_id ,first_term =False,second_term = False, both_terms=False, student_status_list = [1],subject_search_name_wanted_index = [2,3,5],session=None):
    """
    Retrieves and calculates percentages related to marks uploaded by teachers.

    Parameters:
        auth (dict): Authentication information.
        inst_id (int): Institution ID.
        period_id (int): Period ID.
        first_term (bool, optional): Flag indicating consideration of the first term. Defaults to False.
        second_term (bool, optional): Flag indicating consideration of the second term. Defaults to False.
        both_terms (bool, optional): Flag indicating consideration of both terms. Defaults to False.
        student_status_list (list, optional): List of student statuses to include. Defaults to [1].
        subject_search_name_wanted_index (list, optional): Indices indicating which subjects to include. Defaults to [2, 3, 5].
        session (object, optional): Session information. Defaults to None.

    Returns:
        dict: A dictionary containing the following keys:
            - 'school_percentage': Percentage of uploaded marks for the entire school.
            - 'teachers_percentages': Dictionary containing percentages for each teacher.
            - 'classes_percentages': Dictionary containing percentages for each class and subject.
            - 'data_frames': List of dictionaries representing data frames.
    """
    # function variables here 
    techers_percentages ,data_frames , subject_ids ,terms_list = {}, [], [], []
    assessments_codes = {f'S{i}A{x}' : { 'term': "Ø§Ù„ÙØµÙ„ Ø§Ù„"+num2words(i,lang='ar', to='ordinal_num'), 'assessment_name':"Ø§Ù„ØªÙ‚ÙˆÙŠÙ… Ø§Ù„"+num2words(x,lang='ar', to='ordinal_num')} for i in [1,2] for x in [1,2,3,4]}
    search_names =['Ø±ÙŠØ§Ø¶ÙŠØ©', 'Ù†Ø´Ø§Ø·', 'Ù…Ø³ÙŠØ­ÙŠØ©', 'ÙÙ†', 'ÙØ±Ù†Ø³']
    
    search_names = [search_names[abs(i - 1)] for i in subject_search_name_wanted_index]
    unique_names = {}
    
    # Ø§Ø°Ø§ Ù„Ù… ÙŠØ®ØªØ± Ø§Ù„Ù…Ø³ØªØ®Ø¯Ù… Ø§Ù„ÙØµÙ„ Ø§Ù„Ø§ÙˆÙ„ Ø§Ø°Ø§ Ø§Ø®ØªØ± Ø§Ù„ÙØµÙ„ Ø§Ù„Ø«Ø§Ù†ÙŠ 
    # Ùˆ Ø§Ø°Ø§ Ù„Ù… ÙŠØ®ØªØ± Ø§Ù„ÙØµÙ„ Ø§Ù„Ø§ÙˆÙ„ ÙˆÙ„Ø§ Ø§Ù„Ø«Ø§Ù†ÙŠ 
    # Ø§Ø°Ø§ ÙˆØ§Ø®ØªØ§Ø± Ø§Ù„ÙØµÙ„ÙŠÙ† Ø§Ø°Ø§ Ø§Ø¸Ù‡Ø± Ù„Ù‡ Ù†ØªØ§Ø¦Ø¬ Ø§Ù„ÙØµÙ„ÙŠÙ†
    if first_term:
        terms_list = [i for i in assessments_codes if 'S1' in i]
        terms_list.append('Assess')
    elif second_term:
        terms_list = [i for i in assessments_codes if 'S2' in i]
        terms_list.append('Assess2')
    elif both_terms:
        terms_list = [i for i in assessments_codes]
        terms_list.append('Assess1' , 'Assess2')

    # get the marks that the teachers uploaded on the emis site 
    # get the classes and the students 
    open_emis_core_marks = get_school_marks(auth,inst_id , period_id,session=session)
    class_data_dic , students_with_data_dic = get_school_classes_and_students_with_classes(auth ,inst_id , period_id,session=session)

    # add subjects to the class dictionary variable which is class_data_dic
    for class_ in class_data_dic:
        class_subject_data = get_class_subjects(auth ,class_ , class_data_dic[class_]['assessment_id'] ,period_id, inst_id,session=session)
        class_data_dic[class_]['subjects'] = [i['InstitutionSubjects'] for i in class_subject_data]

    # get the teachers or staff data (what the subjects they teach and the class names)
    SubjectStaff_data = make_request(auth=auth , url=f'https://emis.moe.gov.jo/openemis-core/restful/v2/Institution-InstitutionSubjectStaff.json?institution_id={inst_id}&academic_period_id={period_id}&_contain=Users,InstitutionSubjects&_limit=0',session=session)['data']

    # map the followings 
    # teachers load  
    # subjects for each teacher  
    # the teacher with subjects
    staff_load_mapping = {
                        x['staff_id'] : {
                            'name': x['user']['name'],
                            'teacher_subjects':
                                [
                                    [
                                        i['institution_subject']['id'] ,
                                        i['institution_subject']['name'] ,
                                        i['institution_subject']['education_grade_id'],
                                        i['institution_subject']['education_subject_id'] ,
                                    
                                    ] for i in SubjectStaff_data if x['staff_id'] == i['staff_id']
                                ]
                            }
                        for x in SubjectStaff_data
                            if x['end_date'] is None
                        }
    subject_mapping_for_teachers = {
                                    i['id'] : { 
                                            'name': i['name'] , 
                                            'class_name': class_id ,
                                            'class_id' : class_data_dic[class_id]['name'] ,
                                            'education_subject_id': i['education_subject_id']
                                            }    
                                    for class_id in class_data_dic 
                                    for i in class_data_dic[class_id]['subjects']
                                    }
    teacher_with_subject_mapping = {
                                        i[0] : { 
                                                'teacher_name': staff_load_mapping[teacher_id]['name'] , 
                                                'education_subject_name': i[1]
                                                }    
                                        for teacher_id in staff_load_mapping 
                                        for i in staff_load_mapping[teacher_id]['teacher_subjects']
                                    }

    # Create data_frames for these porposes :-
    # 1) writing the resulted marks in excel file  
    # 2)to get the percentages for the school ,and teachers , classes
    for student in students_with_data_dic :
        # FIXME: make execulding sacendary students option in the function
        if 'Ø¹Ø´Ø±' not in class_data_dic[students_with_data_dic[student]['class_id']]['name'] :
            #  Ø§Ø¨Ø­Ø« Ø¹Ù† Ø§Ù„Ø·Ø§Ù„Ø¨ ØµØ§Ø­Ø¨ Ø§Ù„Ø±Ù‚Ù… Ø§Ù„ØªØ¹Ø±ÙŠÙÙŠ 
            student_marks = [i for i in open_emis_core_marks if i['student_id']==student]
            # Ø§Ø¨Ø­Ø« ÙÙŠ ÙƒÙ„ Ø§Ù„Ù…ÙˆØ§Ø¯ Ø§Ù„ØªØ§Ù„ÙŠØ© 
            for subject in class_data_dic[students_with_data_dic[student]['class_id']]['subjects'] :
                subject_marks = [i for i in student_marks if i['education_subject_id'] ==int(subject['education_subject_id'])]
                
                missing_codes = set(i for i in assessments_codes) - set(item['assessment_period']['code'][-4:] for item in subject_marks)
                try:
                    teacher_name =teacher_with_subject_mapping[int(subject['id'])]['teacher_name']
                except:
                    teacher_name = 'Ù„Ø§ ÙŠÙˆØ¬Ø¯ Ù…Ø¹Ù„Ù…'
                    
                # Ùˆ ØªØ­Ù‚Ù‚ Ù…Ù† ÙˆØ¬ÙˆØ¯ ÙƒÙˆØ¯ Ø§Ù„ØªÙ‚ÙˆÙŠÙ…Ø§Øª Ø§Ù„Ø«Ù…Ø§Ù†ÙŠØ© Ù„Ù„ÙØµÙ„ Ø§Ù„Ø§ÙˆÙ„ Ùˆ Ø§Ù„ÙØµÙ„ Ø§Ù„Ø«Ø§Ù†ÙŠ Ùˆ Ø§Ø°Ø§ Ù„Ù… ØªØ¬Ø¯ Ø§Ø±ØµØ¯ ÙØ§Ø±Øº Ù„Ù„Ø¹Ù„Ø§Ù…Ø© 
                for mark in subject_marks :
                    # Ø§Ù…Ø§ Ø§Ø°Ø§ ÙˆØ¬Ø¯Øª ÙØ§Ø±ØµØ¯ Ø¹Ù„Ø§Ù…Ø© Ø§Ù„Ø·Ø§Ù„Ø¨ Ø§Ù„Ø­Ù‚ÙŠÙ‚ÙŠØ©
                    code = mark['assessment_period']['code'][-4:]
                    data_frames.append({
                        'Ø§Ø³Ù… Ø§Ù„Ø·Ø§Ù„Ø¨': students_with_data_dic[mark['student_id']]['full_name'],
                        'Ø­Ø§Ù„Ø© Ø§Ù„Ø·Ø§Ù„Ø¨': 'Ù…Ù„ØªØ­Ù‚' if students_with_data_dic[mark['student_id']]['status_id'] in student_status_list else 'ØºÙŠØ± Ø°Ù„Ùƒ',
                        'status_id' : students_with_data_dic[mark['student_id']]['status_id'],
                        'Ø§Ø³Ù… Ø§Ù„Ù…Ø¹Ù„Ù…': teacher_name ,
                        'Ø±Ù‚Ù… Ø§Ù„ØµÙ' : students_with_data_dic[student]['class_id'],
                        'Ø§Ù„ØµÙ+Ø§Ù„Ø´Ø¹Ø¨Ø©': class_data_dic[mark['institution_classes_id']]['name'],
                        'Ø±Ù‚Ù… Ø§Ù„Ù…Ø§Ø¯Ø©' : mark['education_subject_id'],
                        'Ø§Ø³Ù… Ø§Ù„Ù…Ø§Ø¯Ø©': mark['education_subject']['name'],
                        'Ø§Ù„ØªÙ‚ÙˆÙŠÙ…': mark['assessment_period']['name'],
                        'Ø§Ù„ÙØµÙ„': mark['assessment_period']['academic_term'],
                        'code':code,
                        'Ø§Ù„Ø¹Ù„Ø§Ù…Ø©': mark['marks']
                    })
                if len(missing_codes): 
                    for code in missing_codes:
                        # Ø±ØµØ¯ ØµÙØ± ÙÙŠ ÙƒÙ„ Ø§Ù„ØªÙ‚ÙˆÙŠÙ…Ø§Øª Ø§Ù„ØªÙŠ Ù„Ø§ ÙŠÙˆØ¬Ø¯ Ø¨Ù‡Ø§ ÙƒÙˆØ¯ Ø§Ù„ØªÙ‚ÙˆÙŠÙ…Ø§Øª Ø§Ù„Ù…ÙˆØ¬ÙˆØ¯Ø©
                        data_frames.append({
                            'Ø§Ø³Ù… Ø§Ù„Ø·Ø§Ù„Ø¨': students_with_data_dic[student]['full_name'],
                            'Ø­Ø§Ù„Ø© Ø§Ù„Ø·Ø§Ù„Ø¨': 'Ù…Ù„ØªØ­Ù‚' if students_with_data_dic[student]['status_id'] in student_status_list else 'ØºÙŠØ± Ø°Ù„Ùƒ',
                            'status_id' : students_with_data_dic[student]['status_id'],
                            'Ø§Ø³Ù… Ø§Ù„Ù…Ø¹Ù„Ù…': teacher_name ,
                            'Ø±Ù‚Ù… Ø§Ù„ØµÙ' : students_with_data_dic[student]['class_id'],
                            'Ø§Ù„ØµÙ+Ø§Ù„Ø´Ø¹Ø¨Ø©': class_data_dic[students_with_data_dic[student]['class_id']]['name'],
                            'Ø±Ù‚Ù… Ø§Ù„Ù…Ø§Ø¯Ø©' : int(subject['education_subject_id']),
                            'Ø§Ø³Ù… Ø§Ù„Ù…Ø§Ø¯Ø©': subject['name'] ,
                            'Ø§Ù„ØªÙ‚ÙˆÙŠÙ…': assessments_codes[code]['assessment_name'],
                            'Ø§Ù„ÙØµÙ„': assessments_codes[code]['term'],
                            'code':code,
                            'Ø§Ù„Ø¹Ù„Ø§Ù…Ø©': 'ÙØ§Ø±Øº'
                        })

    # Code that i wrote to skip some subjects from the percentages
    subjects_dictionary_list = [
                                    {
                                        'name':values['name'] ,
                                        'education_subject_id':values['education_subject_id']
                                    } for values in subject_mapping_for_teachers.values()
                                ]
        # Create a Wiktionary to track unique names and their IDs
    sorted_unique_data = sorted({item['education_subject_id']: item for item in subjects_dictionary_list}.values(), key=lambda x: x['education_subject_id'])
        # Iterate through the data and update the dictionary
    for item in sorted_unique_data:
        stripped_name = item['name'].strip()
        if stripped_name in unique_names:
            unique_names[stripped_name].append(item['education_subject_id'])
        else:
            unique_names[stripped_name] = [item['education_subject_id']]
    for name, ids in unique_names.items():
        if any(search_name in name for search_name in search_names):
            ids = [int(i) for i in ids]
            subject_ids.extend(ids)


    # get the teachers percentage and the uploaded marks and unuploaded marks
    teachers_names = list(set(i['Ø§Ø³Ù… Ø§Ù„Ù…Ø¹Ù„Ù…'] for i in data_frames))
    teacher_marks = {
        i: {'row_marks':[
                        {'Ø§Ù„Ø¹Ù„Ø§Ù…Ø©' :x['Ø§Ù„Ø¹Ù„Ø§Ù…Ø©'] , 'code' : x['code']} for x in data_frames 
                            if x['Ø§Ø³Ù… Ø§Ù„Ù…Ø¹Ù„Ù…'] == i
                            # FIXME: try to find other way to get the term percentage and student status
                            and
                            x['code'] in terms_list
                            and
                            x['status_id'] in student_status_list
                        ]}
        for i in teachers_names 
    }

    for teacher in teacher_marks :
        marks = teacher_marks[teacher]['row_marks']
        # print(teacher ,inserted_marks_percentage_from_dataframes_variable(marks))
        techers_percentages[teacher] = inserted_marks_percentage_from_dataframes_variable(marks ,True)

    classes_with_subjects_percentage = class_data_dic

    for class_number in class_data_dic:
        class_subjects_dict = {
                                int(i['education_subject_id']) : {
                                                                    'name' :
                                                                        i['name']
                                                                    }
                                    for i in class_data_dic[class_number]['subjects']
                            }
        class_subjects_dict = dict(sorted(class_subjects_dict.items(), key=lambda item: item[0]))
        class_marks = [i for i in data_frames if i['Ø±Ù‚Ù… Ø§Ù„ØµÙ'] == class_number and "ØºÙŠØ± Ø°Ù„Ùƒ" not in i['Ø­Ø§Ù„Ø© Ø§Ù„Ø·Ø§Ù„Ø¨'] ]
        for subject_id in class_subjects_dict:
            subject_data_list = [i for i in class_marks if i['Ø±Ù‚Ù… Ø§Ù„Ù…Ø§Ø¯Ø©'] == subject_id]
            # subject_marks = [i['Ø§Ù„Ø¹Ù„Ø§Ù…Ø©'] for i in subject_data_list]
            if len(subject_data_list):
                teacher_name = subject_data_list[0]['Ø§Ø³Ù… Ø§Ù„Ù…Ø¹Ù„Ù…']
            # [i for i in assessments_codes]
            assessments_dict = {i:'' for i in assessments_codes}        
            for assessment in assessments_codes:
                assessment_marks = [i for i in subject_data_list if i['code'] == assessment]
                assessments_dict[assessment] = inserted_marks_percentage_from_dataframes_variable(assessment_marks)
                
            class_subjects_dict[subject_id]['subject_marks_percentage']  = assessments_dict
            class_subjects_dict[subject_id]['subject_teacher'] = teacher_name
            
        classes_with_subjects_percentage[class_number]['subjects_percentage'] = class_subjects_dict
        class_subjects_dict = {}
        teacher_name = ''

    # for class_number in classes_with_subjects_percentage:
    #     print( class_number , classes_with_subjects_percentage[class_number]['name'] )
    #     pprint(classes_with_subjects_percentage[class_number]['subjects_percentage'])

    school_marks = [
                    {'Ø§Ù„Ø¹Ù„Ø§Ù…Ø©' :x['Ø§Ù„Ø¹Ù„Ø§Ù…Ø©'] , 'code' : x['code']} 
                    for x in data_frames 
                        if x['code'] in terms_list
                        and
                        x['status_id'] in student_status_list
                        and
                        x['Ø±Ù‚Ù… Ø§Ù„Ù…Ø§Ø¯Ø©'] not in subject_ids
                    ]
    return {
            'school_percentage' : inserted_marks_percentage_from_dataframes_variable(school_marks , True),
            'teachers_percentages' :techers_percentages,
            'classes_percentages' : classes_with_subjects_percentage,
            'data_frames' : data_frames
            }

def get_secondery_students(auth , institution_class_id , inst_id=None , curr_year=None , just_id_and_name_and_empty_marks =True,student_status_ids=[1],session=None):
    """
    Retrieves secondary students enrolled in a specific institution class.

    Parameters:
    - auth (str): Authorization token.
    - institution_class_id (int): Identifier for the institution class.
    - inst_id (int, optional): Identifier for the institution (default is None).
    - curr_year (int, optional): Current academic year (default is None).
    - session (requests.Session, optional): Session object for making HTTP requests (default is None).

    Returns:
    - dict: Dictionary containing data about enrolled secondary students.
    """    
    global secondery_students 
    id_and_name_dic_list = []
    if not len(secondery_students):
        secondery_students =  get_school_students_ids(auth, inst_id=inst_id , curr_year=curr_year ,student_status_ids=student_status_ids , session=session)
    data = [i for i in secondery_students if i['institution_class_id'] == int(institution_class_id) and i['student_status_id'] in student_status_ids]
    data = {'data': data , 'total': len(data)}
    
    enrolled = [i for i in data['data'] if i['student_status_id'] in student_status_ids]
    data = {'data': enrolled , 'total': len(enrolled)}
    if just_id_and_name_and_empty_marks:
        dic = {'id':'' ,'name': '','term1':{ 'assessment1': '' ,'assessment2': '' , 'assessment3': '' , 'assessment4': ''} ,'term2':{ 'assessment1': '' ,'assessment2': '' , 'assessment3': '' , 'assessment4': ''} ,'assessments_periods_ides':[]}
        
        for item in data['data']:
            dic['id'] = item['student_id']
            dic['name'] = item['user']['name']
            id_and_name_dic_list.append(dic)
            dic = {'id':'' ,'name': '','term1':{ 'assessment1': '' ,'assessment2': '' , 'assessment3': '' , 'assessment4': ''} ,'term2':{ 'assessment1': '' ,'assessment2': '' , 'assessment3': '' , 'assessment4': ''} ,'assessments_periods_ides':[]}
        return id_and_name_dic_list
    else:
        return data

def offline_sort_assessement_ids(assessment_id ,marks_data ,assessments):
    """
    Offline sorting of assessment IDs based on their codes.

    Parameters:
    - assessment_id (str): Identifier for the assessment.
    - marks_data (list): List of dictionaries containing marks data.
    - assessments (dict): Dictionary containing assessment data.

    Returns:
    - list: Sorted list of dictionaries based on assessment codes.
    """    
    sorted_values = []
    codes = sorted([i['code'][-4:] for i in assessments['data'] if i['assessment_id'] == assessment_id])
    assessments = [i for i in assessments['data'] if i['assessment_id'] == assessment_id]
    for code in codes:
        target_id = str([i['id'] for i in assessments if code in i['code']][0])
        target_value = [i for i in marks_data if i['assessment_period_id'] == target_id]
        # Add code to each dictionary in target_value
        for item in target_value:
            # Add your code here
            item['code'] = code
        sorted_values.extend(target_value)
    # Check if the length is less than 8
    while len(sorted_values) < 8:
        # Add dictionaries with the value {'mark': None}
        sorted_values.append({'mark': None , 'assessment_period_id': None})
    return sorted_values

def sort_assessement_ids(auth ,assessment_id ,marks_data , session=None):
    """
    Sorting of assessment IDs based on their codes.

    Parameters:
    - auth (str): Authorization token.
    - assessment_id (str): Identifier for the assessment.
    - marks_data (list): List of dictionaries containing marks data.
    - session (requests.Session, optional): Session object for making HTTP requests (default is None).

    Returns:
    - list: Sorted list of dictionaries based on assessment codes.
    """
    assessments= make_request(auth =auth,url=f'https://emis.moe.gov.jo/openemis-core/restful/v2/Assessment-AssessmentPeriods.json?assessment_id={assessment_id}&_limit=0' , session=session)
    sorted_values = []
    codes = sorted([i['code'][-4:] for i in assessments['data']])
    for code in codes:
        target_id = str([i['id'] for i in assessments['data'] if code in i['code']][0])
        target_value = [i for i in marks_data if i['assessment_period_id'] == target_id]
        # Add code to each dictionary in target_value
        for item in target_value:
            # Add your code here
            item['code'] = code
        sorted_values.extend(target_value)
    # Check if the length is less than 8
    while len(sorted_values) < 8:
        # Add dictionaries with the value {'mark': None}
        sorted_values.append({'mark': None , 'assessment_period_id': None})
    return sorted_values

def offline_get_assessment_id_from_grade_id(grade_id , grades_info):
    """
    Offline retrieval of assessment ID based on the education grade ID.

    Parameters:
    - grade_id (int): Identifier for the education grade.
    - grades_info (list): List of dictionaries containing grades information.

    Returns:
    - str: Identifier for the assessment corresponding to the education grade ID.
    """    
    return [d['id'] for d in grades_info if d.get('education_grade_id') == grade_id][0]

def mark_all_students_as_present(auth ,term_days_dates ,r_data = None , proxies = None):
    """
    Marks all students as present on specified dates.

    Parameters:
    - auth (str): Authentication token for making requests.
    - term_days_dates (list): List of dates to mark all students as present.
    - r_data (dict): Dictionary containing required data, fetched using get_required_data_to_enter_absent.
    - proxies (dict): Optional. Dictionary of proxy settings for requests.

    Returns:
    None
    """
    session = requests.Session()
    if not r_data:
        r_data = get_required_data_to_enter_absent(auth,session=session)
    institution_id = r_data['institution_id']
    institution_class_id = r_data['institution_class_id']
    education_grade_id = r_data['education_grade_id']
    academic_period_id = r_data['academic_period_id']
    
    url = f'https://emis.moe.gov.jo/openemis-core/restful/v2/Institution-StudentAttendances.json?_finder=classStudentsWithAbsenceSave[institution_id:{institution_id};institution_class_id:{institution_class_id};education_grade_id:{education_grade_id};academic_period_id:{academic_period_id};attendance_period_id:1;day_id:FUZZ;week_id:undefined;week_start_day:undefined;week_end_day:undefined;subject_id:0]&_limit=0'

    headers = [("User-Agent" , "python-requests/2.28.1"),("Accept-Encoding" , "gzip, deflate"),("Accept" , "*/*"),("Connection" , "close"),("Authorization" , f"{auth}"),("ControllerAction" , "StudentAttendances"),("Content-Type" , "application/json")]


    unsuccessful_requests = wfuzz_function(url , term_days_dates,headers,None,method='GET',proxies = proxies)

    while len(unsuccessful_requests) != 0:
        unsuccessful_requests = wfuzz_function(url , unsuccessful_requests,headers,None,method='GET',proxies = proxies)

    print("All requests were successful!")

def mark_students_absent_in_dates(auth ,students_id_with_names, absent_days_list ,institution_id , institution_class_id , education_grade_id , academic_period_id , year1 , year2 , helper=False ,proxies = None):
    """
    Marks students absent on specified dates.

    Parameters:
    - auth (str): Authentication token for making requests.
    - students_id_with_names (list): List containing tuples of (student_name, student_id).
    - absent_days_list (list): List of dates in "student_id/day/month" format for marking absent.
    - institution_id (int): ID of the institution.
    - institution_class_id (int): ID of the institution class.
    - education_grade_id (int): ID of the education grade.
    - academic_period_id (int): ID of the academic period.
    - year1 (int): Starting year of the academic period.
    - year2 (int): Ending year of the academic period.
    - helper (bool): If True, prints student names and dates for each iteration.
    - proxies (dict): Optional. Dictionary of proxy settings for requests.

    usage:
        required_data = get_required_data_to_enter_absent(auth)
        # for i in f:
        #     print(f"{i} = required_data['{i}']",end=', ')
        mark_students_absent_in_dates(auth , id_with_names, absent_days_list, institution_id = required_data['institution_id'], institution_class_id = required_data['institution_class_id'], education_grade_id = required_data['education_grade_id'], academic_period_id = required_data['academic_period_id'], year1 = required_data['year1'], year2 = required_data['year2'] )

    Returns:
    None
    """
    students_dictionary = {}
    for idx, (name, student_id) in enumerate(students_id_with_names, start=1) : 
        students_dictionary[idx]= [student_id,name ]

    absent_days_list = [day.split('/') for day in absent_days_list]

    # students_dictionary
    absent_data = []

    for date in absent_days_list:
        student_name = str(students_dictionary[int(date[0])][0])
        student_id = str(students_dictionary[int(date[0])][1])
        year = year1 if int(date[-1]) in [9,10,11,12] else year2
        
        for day in date[1:-1]:
            date_str = f"{year}-{date[-1].zfill(2)}-{day.zfill(2)}"
            if helper:
                print ( student_name  , date_str)
            item = json.dumps({"student_id": student_id,
                                "institution_id": institution_id,
                                "academic_period_id": academic_period_id,
                                "institution_class_id": institution_class_id,
                                "absence_type_id": "2",
                                "student_absence_reason_id": None,
                                "comment": None,
                                "period": 1,
                                "date": date_str,
                                "subject_id": 0,
                                "education_grade_id": education_grade_id}).replace('}','')
            absent_data.append(item)

    # 'student_id': 7388854,
    # 'date': '2022-09-19',

    # {"student_id": 7388854, "institution_id": 2600, "academic_period_id": 13, "institution_class_id": 786118, "absence_type_id": "2", "student_absence_reason_id": null, "comment": null, "period": 1, "date": "2022-09-19", "subject_id": 0, "education_grade_id": 275,
    if helper:
        pprint(absent_data[0])


    headers = [("User-Agent" , "python-requests/2.28.1"),("Accept-Encoding" , "gzip, deflate"),("Accept" , "*/*"),("Connection" , "close"),("Authorization" , f"{auth}"),("ControllerAction" , "StudentAttendances"),("Content-Type" , "application/json")]

    body_postdata = json.dumps({ "action_type": "default"}).replace('{',' FUZZ ,')


    url = 'https://emis.moe.gov.jo/openemis-core/restful/v2/Institution-StudentAbsencesPeriodDetails.json?_limit=0'

    unsuccessful_requests = wfuzz_function(url , absent_data , headers ,body_postdata ,method='POST', proxies = proxies)

    while len(unsuccessful_requests) != 0:
        unsuccessful_requests = wfuzz_function(url , unsuccessful_requests , headers ,body_postdata ,method='POST',proxies = proxies)

    print("All requests were successful!")

def fill_students_absent_in_dates_wrapper(auth ,students_absent_multiline_string=None , random_values = False ,start_date_str = None , end_date_str = None , skip_dates_list = None , required_data=None, proxies = None):
    """
    Fills attendance for students based on provided absent dates.

    Parameters:
    - auth (str): Authentication token for making requests.
    - students_absent_multiline_string (str): Multiline string containing absent dates in the specified format.
    - random_values (bool): If True, generates random absent dates for students within the specified date range.
    - start_date_str (str): Start date of the attendance period in "Y-m-d" format (for random_values).
    - end_date_str (str): End date of the attendance period in "Y-m-d" format (for random_values).
    - skip_dates_list (list): List of dates to skip during the attendance period in "Y-m-d" format (for random_values).
    - required_data (dict): Optional. Required data for entering absent days, retrieved by get_required_data_to_enter_absent.
    - proxies (dict): Optional. Dictionary of proxy settings for requests.

    Returns:
    None
    """
    if not required_data:
        required_data = get_required_data_to_enter_absent(auth)
    id_with_names = get_names_for_absent_purposes(auth)
    if students_absent_multiline_string :
        absent_days_list = extract_absent_dates_from_text(students_absent_multiline_string)
    elif random_values :
        days_list = get_period_days_dates(start_date_str, end_date_str, skip_dates_list, skip_weekend=True)
        absent_days_list = create_random_absent_list(days_list , id_with_names)
    else:
        print('give me absent_multiline_string value or choose random_values')
    mark_students_absent_in_dates(auth , id_with_names, absent_days_list, institution_id = required_data['institution_id'], institution_class_id = required_data['institution_class_id'], education_grade_id = required_data['education_grade_id'], academic_period_id = required_data['academic_period_id'], year1 = required_data['year1'], year2 = required_data['year2'] ,proxies = proxies)

def fill_all_students_present_wrapper(auth , start_date_str , end_date_str , skip_dates_list ,required_data=None ,proxies = None):
    """
    Fills attendance for all students as present within a specified date range.

    Parameters:
    - auth (str): Authentication token for making requests.
    - start_date_str (str): Start date of the attendance period in "Y-m-d" format.
    - end_date_str (str): End date of the attendance period in "Y-m-d" format.
    - skip_dates_list (list): List of dates to skip during the attendance period in "Y-m-d" format.
    - required_data (dict): Optional. Required data for entering absent days, retrieved by get_required_data_to_enter_absent.
    - proxies (dict): Optional. Dictionary of proxy settings for requests.

    Returns:
    None
    """
    days_list = get_period_days_dates(start_date_str, end_date_str, skip_dates_list, skip_weekend=True)
    mark_all_students_as_present(auth , days_list , required_data ,proxies = proxies)

def erase_students_absent_dates(auth ,required_data=None ,helper=False,proxies = None):
    """
    Erases the absent dates for students in a specific class.

    Parameters:
    - auth (str): Authentication token for making requests.
    - required_data (dict): Optional. Required data for entering absent days, retrieved by get_required_data_to_enter_absent.
    - helper (bool): Optional. If True, prints the structure of the first absent data item.
    - proxies (dict): Optional. Dictionary of proxy settings for requests.

    Returns:
    None
    """    
    if not required_data:
        required_data = get_required_data_to_enter_absent(auth)
    absent_data_list = get_class_absent_days_with_id(auth ,required_data=required_data)
    
    institution_id=required_data['institution_id']
    institution_class_id=required_data['institution_class_id']
    education_grade_id=required_data['education_grade_id']
    academic_period_id=required_data['academic_period_id']
    
    absent_data = []    
    
    for date_item in absent_data_list:
        student_id = str(date_item[0])
        date_str = date_item[1]

        item = json.dumps({"student_id": student_id,
                            "institution_id": institution_id,
                            "academic_period_id": academic_period_id,
                            "institution_class_id": institution_class_id,
                            "absence_type_id": "0",
                            "student_absence_reason_id": None,
                            "comment": None,
                            "period": 1,
                            "date": date_str,
                            "subject_id": 0,
                            "education_grade_id": education_grade_id}).replace('}','')
        absent_data.append(item)

    # 'student_id': 7388854,
    # 'date': '2022-09-19',

    # {"student_id": 7388854, "institution_id": 2600, "academic_period_id": 13, "institution_class_id": 786118, "absence_type_id": "2", "student_absence_reason_id": null, "comment": null, "period": 1, "date": "2022-09-19", "subject_id": 0, "education_grade_id": 275,
    if helper:
        pprint(absent_data[0])


    headers = [("User-Agent" , "python-requests/2.28.1"),("Accept-Encoding" , "gzip, deflate"),("Accept" , "*/*"),("Connection" , "close"),("Authorization" , f"{auth}"),("ControllerAction" , "StudentAttendances"),("Content-Type" , "application/json")]

    body_postdata = json.dumps({ "action_type": "default"}).replace('{',' FUZZ ,')


    url = 'https://emis.moe.gov.jo/openemis-core/restful/v2/Institution-StudentAbsencesPeriodDetails.json?_limit=0'

    unsuccessful_requests = wfuzz_function(url , absent_data , headers ,body_postdata ,method='POST',proxies = proxies)

    while len(unsuccessful_requests) != 0:
        unsuccessful_requests = wfuzz_function(url , unsuccessful_requests , headers ,body_postdata ,method='POST',proxies = proxies)

    print("All requests were successful!")

def get_class_absent_days_with_id(auth ,simple_list=True , required_data = None):
    """
    Gets the list of absent days for each student in a specific class with their IDs.

    Parameters:
    - auth (str): Authentication token for making requests.
    - simple_list (bool): Optional. If True, returns a simplified list containing student IDs and absent dates.
    - required_data (dict): Optional. Required data for entering absent days, retrieved by get_required_data_to_enter_absent.

    Returns:
    - list: List of absent days for each student in the specified class with their IDs.
    """    
    if not required_data:
        required_data = get_required_data_to_enter_absent(auth)
    
    institution_id=required_data['institution_id']
    institution_class_id=required_data['institution_class_id']
    education_grade_id=required_data['education_grade_id']
    academic_period_id=required_data['academic_period_id']
    
    url = f'https://emis.moe.gov.jo/openemis-core/restful/Institution.StudentAbsencesPeriodDetails?institution_id={institution_id}&institution_class_id={institution_class_id}&education_grade_id={education_grade_id}&academic_period_id={academic_period_id}&_limit=0&_fields=student_id,institution_id,academic_period_id,institution_class_id,education_grade_id,date,period,comment,absence_type_id'
    absent_data = make_request(auth=auth , url=url)
    if simple_list:
        absent_data = [[i['student_id'] , i['date']] for i in absent_data['data']]
    return absent_data

def create_random_absent_list(dates_list ,id_with_names):
    """
    Generates a random list of absent students for given dates and student IDs.

    Parameters:
    - dates_list (list): List of dates in "MM-DD" format.
    - id_with_names (list): List of student IDs with corresponding names.

    Returns:
    - list: Random list of absent students with formatted dates.
    """    
    number_of_the_students = len(id_with_names)
    random_absent = []
    for date in dates_list:
        month , day = date.split('-')[1] , date.split('-')[2] 
        number_of_absency = random.randint(1, 8)
        random_students = []
        for _ in range(number_of_absency):
            random_student_index = random.randint(1, number_of_the_students)
            random_students.append(f'{random_student_index}/{day}/{month}')
        # print(random_students)
        random_absent.extend(random_students)
    return random_absent

def extract_absent_dates_from_text(text , helper=False):
    """
    Ù†Øµ Ø§Ù„ØºÙŠØ§Ø¨ ÙŠØ¬Ø¨ Ø§Ù† ÙŠØªØ¨Ø¹ Ø§Ù„Ù‚ÙˆØ§Ø¹Ø¯ Ø§Ù„Ø§ØªÙŠØ© :
    Ø§Ù† ØºÙŠØ§Ø¨ ÙƒÙ„ Ø·Ø§Ù„Ø¨ ÙÙŠ Ø³Ø·Ø± Ø­Ø³Ø¨ Ø±Ù‚Ù…Ù‡ Ø¹Ù„Ù‰ Ø§Ù„Ø§Ø³Ù…Ø§Ø¡ Ø§Ù„ØªÙŠ ØªØ³ØªØ®Ø±Ø¬Ù‡Ø§ Ø¯Ø§Ù„Ø© get_names_for_absent_purposes
    Ø§Ù„ØºÙŠØ§Ø¨ ÙŠÙƒÙˆÙ† ÙÙŠÙ‡ Ø§Ù„ÙŠÙˆÙ… Ø«Ù… Ø§Ù„Ø´Ù‡Ø± 
    Ø§Ø°Ø§ ÙƒØ§Ù† Ø§Ù„ØºÙŠØ§Ø¨ Ø¨ÙŠÙ† Ø§ÙŠØ§Ù… Ù…ØªÙˆØ§ØµÙ„Ø© Ø§Ø³ØªØ¹Ù…Ù„ Ø¹Ù„Ø§Ù…Ø© Ø§Ù„Ù†Ø§Ù‚Øµ Ø¨ÙŠÙ† Ø§Ù„ÙŠÙˆÙ… Ø§Ù„Ø§ÙˆÙ„ Ùˆ Ø§Ù„ÙŠÙˆÙ… Ø§Ù„Ø§Ø®ÙŠØ±
    Ø§Ø°Ø§ ÙƒØ§Ù†Øª Ø§Ù„Ø§ÙŠØ§Ù… Ù…ØªÙØ±Ù‚Ø© Ø§Ø³ØªØ¹Ù…Ù„ Ø¨ÙŠÙ† ÙƒÙ„ ÙŠÙˆÙ… Ùˆ ÙŠÙˆÙ… Ø¹Ù„Ø§Ù…Ø© Ø§Ù„ÙØ§ØµÙ„Ø© 
    ÙƒØ§Ù„Ù…Ø«Ø§Ù„ Ø§Ù„ØªØ§Ù„ÙŠ:
        8/11 5,7/10 17-19/10 23,25/10
    Ø§Ù„Ù…Ø®Ø±Ø¬Ø§Øª Ø³ØªÙƒÙˆÙ†:
    1    /    24    /     9
    Ø§Ù„Ø´Ù‡Ø±      Ø§Ù„ÙŠÙˆÙ…      Ø±Ù‚Ù…
    Ø§Ù„Ø·Ø§Ù„Ø¨
    Ø§Ù„Ù…ØªØ³Ù„Ø³Ù„       

    Extracts absent dates from a text following specific rules:
    - Each student's absence is in a separate line, indexed by their number.
    - The absence format is day/month.
    - For consecutive days, use the hyphen (-) between the first and last day.
    - For separate days, use a comma (,) between each day.
    
    Example:
    8/11 5,7/10 17-19/10 23,25/10
    
    Outputs:
    1/24/9
    - Day   - Month   - Student Number
    - Student Serial
    
    Parameters:
    - text (str): The absent text following the specified rules.
    - helper (bool): Optional. If True, prints additional information for debugging.

    Returns:
    - list: List of formatted absent dates.     
    """
    absent_days_list = []
    absent_string_list = text.split('\n')
    for idx ,item in enumerate(absent_string_list ,start=1):
        if helper:
            print(idx)
        dates = item.split(' ')
        for date in dates :
            if '-' in date:
                start , end = date.split('/')[0].split('-')
                month = date.split('/')[1]
                e = [f"{idx}/{i}/{month}" for i in range(int(start), int(end)+1)] 
                absent_days_list.extend(e)
                if helper:
                    print(' '.join(e),end=' ')              
            elif ',' in date :
                days ,month= date.split('/')
                splitted_days = days.split(',')
                d = [f"{idx}/{i}/{month}" for i in splitted_days]
                absent_days_list.extend(d)
                if helper:
                    print(' '.join(d) ,end=' ')
            else :
                absent_days_list.append(f'{idx}/{date}')
                if helper:
                    print(f'{idx}/{date}' ,end=' ')
        if helper:            
            print('\n'+'-'*50)
    return absent_days_list

def get_period_days_dates(start_date_str, end_date_str, skip_dates_list=[], skip_weekend=True):
    """
    Generates a list of days within a specified time period, excluding skipped dates and weekends.

    Parameters:
    - start_date_str (str): Start date in "YYYY-MM-DD" format.
    - end_date_str (str): End date in "YYYY-MM-DD" format.
    - skip_dates_list (list): Optional. List of dates to skip.
    - skip_weekend (bool): Optional. If True, skips weekends.

    Returns:
    - list: List of formatted days within the specified period.

    Example Usage:
    ```
    Ø§Ø¹Ø·Ø§Ø¡ Ø§Ù„Ø§ÙŠØ§Ù… Ø§Ù„Ù…ÙˆØ¬ÙˆØ¯Ø© ÙÙŠ ÙØªØ±Ø© Ù…Ù† Ø§Ù„Ø²Ù…Ù† Ùˆ Ø§Ø³ØªØ«Ù†Ø§Ø¡ Ø§ÙŠØ§Ù… Ø§Ù„Ø¹Ø·Ù„ Ø§Ù„Ø±Ø³Ù…ÙŠØ© Ø¨ØªÙˆØ¬ÙŠÙ‡ Ø§Ù„Ù…Ø³ØªØ®Ø¯Ù…
    start_date_str = "2023-08-20"
    end_date_str = "2023-11-12"
    skip_dates_list = ["2023-09-27"]  # Specify dates to skip in "Y-m-d" format
    
    result_dates = get_period_days_dates(start_date_str, end_date_str, skip_dates_list, skip_weekend=True)

    # len(result_dates)
    print('\n'.join(result_dates))
    # print(result_dates)
    ```
    """
    start_date = datetime.datetime.strptime(start_date_str, "%Y-%m-%d")
    end_date = datetime.datetime.strptime(end_date_str, "%Y-%m-%d")
    skip_dates_list = [datetime.datetime.strptime(date_str, "%Y-%m-%d").date() for date_str in skip_dates_list]
    result_dates = []
    current_date = start_date
    delta = datetime.timedelta(days=1)

    while current_date <= end_date:
        if current_date.date() not in skip_dates_list:
            if not (skip_weekend and current_date.weekday() in [4, 5]):
                result_dates.append(current_date.strftime("%Y-%m-%d"))
        
        current_date += delta

    return result_dates

def contains_else_number_or_slash(text):
    """
    Checks if a text contains characters other than digits, "/", and whitespace.

    Parameters:
    - text (str): Input text to check.

    Returns:
    - bool: True if special characters are present, False otherwise.
    """    
    # Define a regular expression pattern to match characters other than "/", digits, and whitespace
    pattern = re.compile(r'[^/\d\s]')
    
    # Search for the pattern in the text
    match = pattern.search(text)
    
    # Return True if a match is found (i.e., special characters are present), False otherwise
    return match is not None

def intended_for_pytest_for_the_absent_text(absent_days_list):
    """
    Prints items in a list that contain characters other than digits, "/", and whitespace.

    Parameters:
    - absent_days_list (list): List of absent days.

    Returns:
    - None
    """    
    for i in absent_days_list:
        if contains_else_number_or_slash(i):
            print(i)
    # get the monthes of the proccessed text 
    # Ù„Ø§Ø­Ø¶Ø§Ø± Ø§Ù„Ø§Ø´Ù‡Ø± Ø§Ù„ØªÙŠ ØªØ­ØªØ§Ø¬ Ø§Ù„Ù‰ ØªØ¹Ø¯ÙŠÙ„ Ø§Ùˆ Ø§Ù„Ù…Ø®ØªÙ„ÙØ© 
    # set([i.split('/')[2] for i in l])

def get_names_for_absent_purposes(auth , session=None , current= True):
    """
    Retrieves a list of student names and IDs for absent purposes based on institution, class, and academic period.

    Parameters:
    - auth (str): Authentication token.
    - session: Optional. Requests Session object.

    Returns:
    - list: Sorted list of tuples containing student IDs and names.
    """        
    status = [1] if current else [1,2,3,4,5,6]
    d = get_required_data_to_enter_absent(auth=auth, session=session)
    institution_id = d['institution_id']
    institution_class_id = d['institution_class_id']
    academic_period_id = d['academic_period_id']
    url = f"https://emis.moe.gov.jo/openemis-core/restful/v2/Institution.InstitutionSubjectStudents?_fields=student_id,student_status_id,Users.id,Users.username,Users.openemis_no,Users.first_name,Users.middle_name,Users.third_name,Users.last_name,Users.address,Users.address_area_id,Users.birthplace_area_id,Users.gender_id,Users.date_of_birth,Users.date_of_death,Users.nationality_id,Users.identity_type_id,Users.identity_number,Users.external_reference,Users.status,Users.is_guardian&_limit=0&academic_period_id={academic_period_id}&institution_class_id={institution_class_id}&institution_id={institution_id}&_contain=Users"
    students_with_ids = make_request(url=url,auth=auth,session=session)
    u_names_with_ids = set([(i['student_id'] ,i['user']['name']) for i in students_with_ids['data'] if i['student_status_id'] in status ])

    sorted_list = sorted(u_names_with_ids, key=lambda x: x[1])
    
    return sorted_list

def get_required_data_to_enter_absent(auth , session=None):
    """
    Retrieves the required data to enter student absences on the EMIS system.

    This function fetches essential data needed to enter student absences, including
    information about the current academic period, home room class, and education grade.

    Parameters:
    - auth (str): The authentication token for accessing the API.
    - session (requests.Session): An optional session object to reuse existing connections.

    Returns:
    - dict: A dictionary containing the following keys:
      - 'institution_id': The ID of the institution.
      - 'institution_class_id': The ID of the home room class.
      - 'education_grade_id': The ID of the education grade.
      - 'academic_period_id': The ID of the current academic period.
      - 'year1': The start year of the academic period.
      - 'year2': The end year of the academic period.

    Example Usage:
    ```python
    auth_token = get_auth('your_username', 'your_password')
    required_data = get_required_data_to_enter_absent(auth_token)
    ```
    """    
    url = 'https://emis.moe.gov.jo/openemis-core/restful/v2/Institution-InstitutionClasses'
    classes = make_request(url=url , auth=auth , session=session)
    curr_per = get_curr_period(auth=auth, session=session)['data'][0]
    academic_period_id =curr_per['id']
    home_room_class = [i for i in classes['data'] if i['academic_period_id'] ==academic_period_id][0]

    institution_class_id = home_room_class['id']
    institution_id = home_room_class['institution_id']

    url = f'https://emis.moe.gov.jo/openemis-core/restful/v2/Institution-InstitutionClasses.json?_finder=gradesByInstitutionAndAcademicPeriodAndInstitutionClass[institution_id:{institution_id};academic_period_id:{academic_period_id};institution_class_id:{institution_class_id}]&_limit=0'

    education_grade_id = make_request(url=url , auth=auth, session=session)['data'][0]['id']

    return {
            'institution_id':institution_id,
            'institution_class_id':institution_class_id,
            'education_grade_id':education_grade_id,
            'academic_period_id':academic_period_id,
            'year1': curr_per['start_year'],
            'year2': curr_per['end_year']
            }

def bulk_e_side_note_marks(passwords):
    """
    Generate E-Side Note marks documents for multiple users.

    This function iterates over a list of username/password combinations and generates
    E-Side Note marks documents for each user.

    Parameters:
    - passwords (str): A string containing username/password combinations separated by '\n'.

    Returns:
    None

    Example Usage:
    ```python
    bulk_e_side_note_marks('username1/password1\nusername2/password2\n...')
    ```

    """    
    session = requests.Session()
    for p in passwords.split('\n'):
        # print(p,'-------<>')
        try : 
            username = p.split('/')[0]
            password = p.split('/')[1]
        except:
            username ,password =[p]*2
        # print(username , password)
        # FIXME: ØµÙ„Ø­ Ù…Ø´ÙƒÙ„Ø© Ø§Ù„Ø³ÙŠØ´ÙŠÙ† ÙÙŠ Ø§Ù„Ø±ÙŠÙƒÙŠÙˆØ³Øª
        # session = requests.Session()
        try:
            logger.info(f"username:{username} ----> password :{password}")
            create_e_side_marks_doc(username , password ,session=session)
        except Exception as e:
            
            print("\033[91m There is error in \n{}/{}\033[00m" .format(username , password))
            # print(username , password)
            # traceback.print_exc()
            logger.error(f"username:{username} ----> password :{password}"+'\n'+traceback.format_exc())
        # if not get_auth(username , password): 
        #     print(username , password)

def read_all_xlsx_in_folder(directory_path='./send_folder'):
    """
    Reads all Excel files (ODS or XLSX) in the specified folder and returns a list of dictionaries.

    This function scans the specified directory for Excel files and reads their contents. It distinguishes
    between ODS (OpenDocument Spreadsheet) and XLSX (Excel) files and uses appropriate parsers accordingly.
    The resulting data is stored in a list of dictionaries.

    Parameters:
    - directory_path (str): The path to the folder containing Excel files. Default is './send_folder'.

    Returns:
    - list: A list of dictionaries, each containing the data read from an Excel file.

    Example Usage:
    ```python
    dic_list = read_all_xlsx_in_folder(directory_path='./send_folder')
    ```
    """    
    dic_list = []
    for item in os.listdir(directory_path):
        if not os.path.isdir(f'{directory_path}/{item}'):
            if item.lower().endswith('.ods'):
                item_path = os.path.join(directory_path, item)  
                dic_list.append(Read_E_Side_Note_Marks_ods(file_path=item_path))    
            else:
                item_path = os.path.join(directory_path, item)  
                dic_list.append(Read_E_Side_Note_Marks_xlsx(file_path=item_path))
    return dic_list

def convert_to_marks_offline_from_send_folder(A3_context , A4_context , directory_path='./send_folder',do_not_delete_send_folder=True , template='./templet_files/official_marks_doc_a3_two_face_white_cover.ods' , color ="#8cd6e6" , default_value = True  ):
    """
    Converts data from multiple Excel files in a specified folder to official marks documents.

    This function reads Excel files from the specified directory and converts the data
    into official marks documents using a provided template. The resulting documents are saved
    in the same directory with the prefix "official_marks_" added to their original file names.

    Parameters:
    - directory_path (str): The path to the folder containing Excel files. Default is './send_folder'.
    - do_not_delete_send_folder (bool): If True, the 'send_folder' directory will not be deleted after processing.
                                        Default is True.
    - template (str): The path to the template file (ODS format) used for generating official marks documents.
                     Default is './templet_files/official_marks_doc_a3_two_face_white_cover.ods'.
    - color (str): The background color for cells in the generated documents. Default is "#8cd6e6".

    Example Usage:
    ```python
    convert_to_marks_offline_from_send_folder(directory_path='./send_folder', do_not_delete_send_folder=True,
                                              template='./templet_files/official_marks_doc_a3_two_face_white_cover.ods',
                                              color="#8cd6e6")
    ```
    """
    dic_list = read_all_xlsx_in_folder(directory_path)
    for file_content in dic_list:
        fill_official_marks_functions_wrapper_v2( e_side_notebook_data=file_content ,do_not_delete_send_folder=do_not_delete_send_folder, templet_file=template , divded_dfter_to_primary_and_secnedry=False, default = default_value , A3_context = A3_context , A4_context = A4_context)
        # fill_official_marks_doc_wrapper_offline(file_content , do_not_delete_send_folder=do_not_delete_send_folder , templet_file=template ,color=color)

def fill_student_absent_doc_wrapper(username, password ,template='./templet_files/new_empty_absence_notebook_doc_white_cover.ods' , outdir='./send_folder/' ,teacher_full_name=False , context =None):
    """
    Fills the student absent notebook document template with data and saves it.

    Parameters:
    - username (str): The username for authentication.
    - password (str): The password for authentication.
    - template (str): Path to the ODS template file (default: './templet_files/new_empty_absence_notebook_doc_white_cover.ods').
    - outdir (str): Directory to save the filled document (default: './send_folder/').
    - teacher_full_name (bool): Flag to include teacher's full name in the document (default: False).

    Example Usage:
    ```python
    fill_student_absent_doc_wrapper('your_username', 'your_password', teacher_full_name=True)
    ```

    Note:
    - This function fetches student statistical information using the provided credentials.
    - It then uses the data to fill the specified ODS template with student details and saves the filled document.
    - The filled document is saved in the specified output directory.

    """
    student_details = get_student_statistic_info(username,password,teacher_full_name=teacher_full_name)
    fill_student_absent_doc_name_days_cover(student_details , template , outdir , context = context )

def vacancies_dictionary2Html(dict_list , outdir='./send_folder/'):
    """
    Generates an HTML table from dictionaries and saves it to a file.

    The function combines data from two dictionaries (dict_list1 and dict_list2)
    into an HTML table and saves the resulting HTML to a file named "ØªØ´ÙƒÙŠÙ„Ø§Øª.html".

    Note:
    - The data is formatted using the Jinja2 templating engine.
    - The table includes columns for school name, school position, teachers, and classes.
    - The generated HTML is saved to a file, and a confirmation message is printed.

    Example Usage:
    ```python
    vacancies_dictionary2Html()
    ```

    """
    from jinja2 import Template
    # from mydicts import dict_list1 ,dict_list2

    table_data = dict_list


    # Define the HTML table code as a string
    table_template = """
    <style>
    table {
    border-collapse: collapse;
    width: 100%;
    margin-left: auto;
    margin-right: 0;
    font-size: 20px;
    font-family: 'Times New Roman', Times, serif;
    }

    th, td {
    border: 1px solid black;
    padding: 8px;
    }

    </style>

    <table dir="rtl">
    <thead>
    <tr>
    <th style="text-align: right;">Ø§Ø³Ù… Ø§Ù„Ù…Ø¯Ø±Ø³Ø©</th>
    <th style="text-align: right;">Ø§Ù†ØµØ¨Ø© Ø§Ù„Ù…Ø¯Ø±Ø³Ø©</th>
    <th style="text-align: right;">Ø§Ù„Ù…Ø¹Ù„Ù…ÙŠÙ†</th>
    <th style="text-align: right;">Ø§Ù„ØµÙÙˆÙ</th>
    </tr>
    </thead>
    <tbody>
    {% for item in data %}
    <tr>
    <td style="text-align: right;">{{ item['school_name'] }}</td>
    <td style="text-align: right;">{{ item['school_load'] | replace("\n", "<br>") }}</td>
    <td style="text-align: right;">{{ item['teachers'] | replace("\n", "<br>") }}</td>
    <td style="text-align: right;">{{ item['classes'] | replace("\n", "<br>") }}</td>
    </tr>
    {% endfor %}
    </tbody>
    </table>
    """

    # format the data into the table template
    table_html = Template(table_template).render(data=table_data)

    html = f"""
    <html lang="ar">
    <head>
        <meta charset="UTF-8">
        <meta http-equiv="X-UA-Compatible" content="IE=edge">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Ø§Ù„Ø´ÙˆØ§ØºØ±</title>
    </head>
    <body>
    {table_html}
    </body>
    </html>
    """



    # Specify the file path and name
    file_path = outdir + "ØªØ´ÙƒÙŠÙ„Ø§Øª.html"

    # Open the file in write mode
    with open(file_path, "w") as file:
        # Write the content to the file
        file.write(html)

    # Confirmation message
    print(f"Content saved to {file_path}.")

def tor_code():
    """
    Ø¯Ø§Ù„Ø© Ù„Ù…ØªØµÙØ­ ØªÙˆØ± ÙƒØªØ¨ØªÙ‡Ø§ Ù„ÙƒÙŠ Ø§ØªÙ…ÙƒÙ† Ù…Ù† Ù…Ø¹Ø§Ù„Ø¬Ø© Ù…Ø´ÙƒÙ„Ø© Ø§Ù„Ø³ÙŠØ±ÙØ± Ø§Ù„Ø°ÙŠ ÙŠØ­ØªØ§Ø¬ Ù…Ù†ÙŠ Ø§Ù† ÙŠÙƒÙˆÙ† Ø¹Ù†ÙˆØ§Ù† Ø¬Ù‡Ø§Ø²ÙŠ Ø§Ù…Ø±ÙŠÙƒÙŠ
    """
    import stem.process
    from stem.util import term
    from stem import Signal
    from stem.control import Controller
    import subprocess
    import requests 

    SOCKS_PORT = 9050
    # CTL_SOCKS_PORT = 9051

    def get_tor_nodes_with_prefix(prefix , CTL_SOCKS_PORT = 9051):

        with Controller.from_port(port=CTL_SOCKS_PORT) as controller:
            controller.authenticate(password="MyStr0n9P#D")  # Authenticate with the Tor control port

            # Retrieve the list of Tor nodes
            relay_list = controller.get_network_statuses()

            nodes_with_prefix = []

            for relay in relay_list:
                if relay.flags and 'Exit' in relay.flags:  # Filter exit relays
                    if relay.address and relay.address.startswith(prefix):  # Filter IP address prefix
                        nodes_with_prefix.append(relay)

            return nodes_with_prefix

    # Example usage
    # nodes = get_tor_nodes_with_prefix('23.')
    # for node in nodes:
    #     print("Fingerprint:", node.fingerprint)
    #     print("IP Address:", node.address)

    with subprocess.Popen(['tor'], stdout=subprocess.PIPE, universal_newlines=True) as tor:
        bootstrapped = False
        for line in tor.stdout:
            if not bootstrapped and 'Bootstrapped 100%' not in line:
                continue
            elif not bootstrapped:
                bootstrapped = True
            print(line.strip())
            nodes = get_tor_nodes_with_prefix('23.')
            tor.terminate()
            break
    print(nodes)
    # Start an instance of Tor configured to only exit through Russia. This prints
    # Tor's bootstrap information as it starts. Note that this likely will not
    # work if you have another Tor instance running.

    def print_bootstrap_lines(line):
        if "Bootstrapped " in line:
            print(term.format(line, term.Color.BLUE))


        print(term.format("Starting Tor:\n", term.Attr.BOLD))

    for node in nodes :
        tor_process = stem.process.launch_tor_with_config(
        config = {
            'SocksPort': str(SOCKS_PORT),
            'ExitNodes': node.fingerprint ,
        },
        init_msg_handler = print_bootstrap_lines,
        )

        print(term.format("\nChecking our endpoint:\n", term.Attr.BOLD))
        
        print(term.format(node.address,term.Color.GREEN))

        try:
            auth = get_auth(9971055725,9971055725 , proxies={"http": "socks5://127.0.0.1:9050", "https": "socks5://127.0.0.1:9050"})
            print(term.format( auth, term.Color.BLUE))
        except:
            print(term.format('error occured' , term.Color.RED))

        tor_process.kill()  # stops tor

def get_year_days_dates(start_date=None , end_date=None , skip_start_date=None , skip_end_date=None):
    """
    Generates a list of dates representing the school year days within a specified range, excluding weekends and specified skip dates.

    Parameters:
    - start_date (str or None): Start date of the school year (format: "YYYY-MM-DD") or None to use the default (2022-08-30).
    - end_date (str or None): End date of the school year (format: "YYYY-MM-DD") or None to use the default (2023-06-30).
    - skip_start_date (str or None): Start date of any skip period (format: "YYYY-MM-DD") or None to use the default (2023-01-01).
    - skip_end_date (str or None): End date of any skip period (format: "YYYY-MM-DD") or None to use the default (2023-02-05).

    Returns:
    - list: A list of date strings in the format "YYYY-MM-DD" representing the school year days.

    Example Usage:
    ```python
    # Using default dates
    year_days = get_year_days_dates()

    # Specifying custom dates and skip periods
    custom_year_days = get_year_days_dates(start_date="2022-09-01", end_date="2023-06-01", skip_start_date="2023-01-01", skip_end_date="2023-02-05")
    ```

    Note:
    - The function excludes weekends (Friday and Saturday).
    - The default start_date is 2022-08-30, and the default end_date is 2023-06-30.
    - The default skip period is from 2023-01-01 to 2023-02-05.

    """
    present_days = []
    start_date = datetime.date(2022, 8, 30) if not start_date else datetime.strptime(start_date, "%Y-%m-%d")
    end_date = datetime.date(2023, 6, 30) if not end_date else datetime.strptime(end_date , "%Y-%m-%d")
    skip_start_date = datetime.date(2023, 1, 1) if not skip_start_date else datetime.strptime(skip_start_date , "%Y-%m-%d")
    skip_end_date = datetime.date(2023, 2, 5) if not skip_end_date else datetime.strptime(skip_end_date , "%Y-%m-%d")

    current_date = start_date
    while current_date <= end_date:
        if current_date < skip_start_date or current_date > skip_end_date:
            if current_date.weekday() not in [4, 5]:  # Exclude Friday (4) and Saturday (5)
                present_days.append(current_date.strftime("%Y-%m-%d"))
        current_date += datetime.timedelta(days=1)

    return present_days

def wfuzz_function(url, fuzz_list,headers,body_postdata,method='POST',proxies = None , timout_req_delay = 1000000, threshhold=True):
    """Ø¯Ø§Ù„Ø© Ø§Ø³ØªØ®Ø¯Ù…Ù‡Ø§ Ù„Ø§Ø±Ø³Ø§Ù„ Ø·Ù„Ø¨ Ø¨ÙˆØ³Øª Ø¨Ø´ÙƒÙ„ Ø³Ø±ÙŠØ¹

    Args:
        fuzz_list (list): Ù‚Ø§Ø¦Ù…Ø© ÙÙŠ Ø¨ÙŠØ§Ù†Ø§Øª Ø§Ù„Ø·Ù„Ø§Ø¨ Ø§Ù„Ù…Ø±Ø§Ø¯ Ø§Ø¯Ø®Ø§Ù„Ù‡Ø§
        headers (tuple-list): Ø±Ø§Ø³ÙŠØ§Øª Ø§Ù„Ø·Ù„Ø¨ Ø§Ùˆ Ø§Ù„Ø±ÙƒÙˆÙŠØ³Øª
        body_postdata (str): Ø¬Ø³Ù… Ø§Ù„Ø¨ÙˆØ³Øª Ø¯Ø§ØªØ§
        method (str, optional): Ø·Ø±ÙŠÙ‚Ø© Ø§Ù„Ø·Ù„Ø¨. Defaults to 'POST'.
        proxies (tuple_list) : [("127.0.0.1","8080","HTTP")]

    Returns:
        any : ØªØ¹ÙˆØ¯ Ø¨Ù‚Ø§Ø¦Ù…Ø© Ø§Ù„Ø·Ù„Ø¨Ø§Øª ØºÙŠØ± Ø§Ù„Ù†Ø§Ø¬Ø­Ø©
    """    
    auth_header = [i[1] for i in headers if i[0] == 'Authorization'][0]
    if 'csrfToken' in auth_header or 'PHPSESSID' in auth_header or 'System' in auth_header:
        value = [i[1] for i in headers if i[0] == 'Authorization'][0]
        headers = [i for i in headers if i[0] != 'Authorization']
        headers.append(('Cookie' , value))
    unsuccessful_requests=[]
    with tqdm(total=len(fuzz_list), bar_format='{postfix[0]} {n_fmt}/{total_fmt}',
            postfix=["uploaded mark", {"value": 0}]) as t:
            s = wfuzz.get_payloads([fuzz_list])
            for idx , r in enumerate(s.fuzz(
                            url=url ,
                            # hc=[404] , 
                            # payloads=[("list",fuzz_list)] ,
                            headers=headers ,
                            postdata = body_postdata ,
                            proxies= proxies ,
                            method= method,
                            delay=.5,
                            req_delay= timout_req_delay
                            ),start =1):
                    
                t.postfix[1]["value"] = idx
                t.update()    
            #     print(r)
            #     print(r.content)
            #     print(r.history.code) # ÙƒÙˆØ¯ Ø§Ù„Ø±ÙƒÙˆÙŠØ³Øª
                if r.history.code != 200 :
                    # if threshhold and r.history.code in [i for i in range(300 , 600)]:
                        # print(r.history.code ,r.content)                    
                        # time.sleep(30)
                    unsuccessful_requests.append(r.description)
    return unsuccessful_requests

def upload_marks_optimized(username , password , classess_data , empty = False):
    """
    Uploads student marks to the EMIS system for the specified classes and assessments.

    Parameters:
    - username (str): The username for authentication.
    - password (str): The password for authentication.
    - classess_data (dict): Dictionary containing class data, including student information and marks.
    - empty (bool): If True, marks will be uploaded as empty (default is False).

    Returns:
    None

    Example Usage:
    ```python
    file_name = 'Ø¹Ù„ÙŠ Ø§Ù„Ù…Ø­Ø§Ù…ÙŠØ¯-9901024120(6).ods'
    student_details = Read_E_Side_Note_Marks_ods('./'+file_name)
    fuzz_list = upload_marks_optimized(9901024120 , 9901024120 , student_details ,empty=False)
    ```

    Note:
    This function utilizes the EMIS API to upload student marks. It is crucial to handle authentication properly and ensure that the API responses are in JSON format without errors.

    **Important:**
    - This function is powerful and requires careful consideration and modification of certain aspects.
    - The response body from the request should be in JSON format and should not contain errors.
    - Modify the while loop from a `for` loop to a `while` loop with a maximum of five iterations, ensuring code repetition within it does not exceed the maximum limit."""
    fuzz_postdata_list = []
    session = requests.Session()
    auth = get_auth(username , password)
    period_id = classess_data['custom_shapes']['period_id']
    school_id = classess_data['custom_shapes']['school_id']
    # term1_assessment_codes = ['S1A1', 'S1A2', 'S1A3', 'S1A4']
    # term2_assessment_codes = ['S2A1', 'S2A2', 'S2A3', 'S2A4']
    assessment_codes = ['S1A1', 'S1A2', 'S1A3', 'S1A4' , 'S2A1', 'S2A2', 'S2A3', 'S2A4']
    assessment_code_dic = {'S1A1': {'term' :'term1' , 'assess' : 'assessment1'},
                            'S1A2': {'term' :'term1' , 'assess' : 'assessment2'},
                            'S1A3': {'term' :'term1' , 'assess' : 'assessment3'},
                            'S1A4': {'term' :'term1' , 'assess' : 'assessment4'},
                            'S2A1': {'term' :'term2' , 'assess' : 'assessment1'},
                            'S2A2': {'term' :'term2' , 'assess' : 'assessment2'},
                            'S2A3': {'term' :'term2' , 'assess' : 'assessment3'},
                            'S2A4': {'term' :'term2' , 'assess' : 'assessment4'}}
    
    assessments_periods_data = classess_data['required_data_for_mrks_enter']
    for class_data in classess_data['file_data']:
        class_id = class_data['class_name'].split('=')[2] 
        class_subject = class_data['class_name'].split('=')[3]
        class_name = class_data['class_name'].split('=')[0]
        if 'Ø¹Ø´Ø±' not in class_name : 
            students_marks_ids = class_data['students_data']
            assessment_grade_id = assessments_periods_data[int(class_id)]['assessment_grade_id']
            grade_id = assessments_periods_data[int(class_id)]['grade_id']
            assessment_periods = get_editable_assessments(auth,username,assessment_grade_id,class_subject,session=session)
            # assessment_ids = assessments_periods_data[int(class_id)]['assessments_period_ids']
            # s1a1, s1a2, s1a3, s1a4, s2a1, s2a2, s2a3, s2a4 = [assessment_ids[i] if i < len(assessment_ids) else None for i in range(8)]
            for student_info in students_marks_ids:
                for code in assessment_codes:
                    if len([i for i in assessment_periods if code in i['code']]) != 0:
                        assessment_period_id = [i for i in assessment_periods if code in i['code']][0]['AssesId']
                        term = assessment_code_dic[code]['term']
                        assess = assessment_code_dic[code]['assess']
                        term_marks = student_info[term]
                        mark = '' if empty else term_marks.get(assess)
                        fuzz_postdata = {
                                'marks': '' if mark== '' else str("{:.2f}".format(float(mark))),
                                'assessment_id': assessment_grade_id,
                                'education_subject_id': class_subject,
                                'education_grade_id': grade_id,
                                'institution_classes_id': class_id,
                                'student_id': student_info['id'],
                                'assessment_period_id': assessment_period_id,
                                'action_type': 'default'
                            }
                        fuzz_postdata_list.append(json.dumps(fuzz_postdata).replace('{','').replace('}',''))
                    
    body_postdata = json.dumps({
            'assessment_grading_option_id': 8,
            'institution_id': school_id,
            'academic_period_id': period_id,
            'student_status_id': 1,
            'action_type': 'default'}).replace('}',', FUZZ }')

    headers = [("User-Agent" , "python-requests/2.28.1"),("Accept-Encoding" , "gzip, deflate"),("Accept" , "*/*"),("Connection" , "close"),("Authorization" , f"{auth}"),("ControllerAction" , "Results"),("Content-Type" , "application/json")]
    
    url = ENTER_MARK_URL
    
    unsuccessful_requests = wfuzz_function(url , fuzz_postdata_list,headers,body_postdata )

    while len(unsuccessful_requests) != 0:
        unsuccessful_requests = wfuzz_function(url ,unsuccessful_requests,headers,body_postdata)

    print("All requests were successful!")

def read_json_file(file_path):
    """
    Reads a JSON file and returns its content as a Python dictionary.

    Parameters:
    - file_path (str): The path to the JSON file.

    Returns:
    - dict: The content of the JSON file as a dictionary.

    Example Usage:
    ```python
    data = read_json_file('./input_data/data.json')
    print(data)
    ```

    """    
    with open(file_path, 'r', encoding='utf-8') as file:
        dictionary = json.load(file)
    return dictionary

def save_dictionary_to_json_file(dictionary, file_path='./send_folder/output.json', indent=None):
    """
    Saves a Python dictionary to a JSON file.

    Parameters:
    - dictionary (dict): The dictionary to be saved.
    - file_path (str): The path to the JSON file. Default is './send_folder/output.json'.
    - indent (int or None): The number of spaces to use for indentation. If None, the JSON is compact. Default is None.

    Returns:
    - None

    Example Usage:
    ```python
    save_dictionary_to_json_file(my_dict, file_path='./output_data/data.json', indent=4)
    ```
    """    
    with open(file_path, 'w', encoding='utf-8') as file:
        json.dump(dictionary, file, indent=indent, ensure_ascii=False)

def create_coloured_certs_wrapper(username , password ,term2=False , just_teacher_class=True ,curr_year = None):
    """
    Retrieves student information, statistics, and marks, then generates colored certificates in OpenDocument Spreadsheet (ODS) format.

    The function performs the following steps:
    1. Establishes a session using the provided username and password.
    2. Retrieves student information and subject marks using the established session.
    3. Groups students based on their information.
    4. Adds subject sum dictionary to the grouped list.
    5. Calculates and adds averages to the grouped list.
    6. Saves the dictionary containing assessment data and statistics to a JSON file.
    7. Generates colored certificates using the saved data.

    Example Usage:
    ```python
    create_coloured_certs_wrapper(username, password, term2=False)
    ```

    Parameters:
    - username (str): The username for authentication.
    - password (str): The password for authentication.
    - term2 (bool): Flag indicating whether to consider term 2. Default is False.

    Returns:
    - None

    Side Effect:
    - Saves colored certificates in the current working directory.

    """
    session = requests.Session()
    auth = get_auth(username , password)
    student_info_marks = get_students_info_subjectsMarks( username , password ,just_teacher_class=just_teacher_class, session=session,curr_year = curr_year)
    students_statistics_assesment_data = get_student_statistic_info(username,password , student_ids=[i['student_id'] for i in get_school_students_ids(auth)] , session=session)
    dic_list4 = student_info_marks
    listo=get_school_absent_days_with_studentID_and_countOfAbsentDays_and_classID_and_className(get_auth(username,password))
    absent_list_with_names=get_absens_studentName_and_countOfDays_and_studentID(dic_list4,listo)
    grouped_list = group_students(dic_list4 )
    
    
    add_subject_sum_dictionary(grouped_list)

    add_averages_to_group_list(grouped_list ,skip_art_sport=False)
    students_statistics_assesment_data['assessments_data'] = grouped_list
    
    save_dictionary_to_json_file(dictionary=students_statistics_assesment_data)
    
    create_coloured_certs_ods(students_statistics_assesment_data,absent_list_with_names , term2=term2)

def convert_files_to_pdf(outdir,pages_range):
    """Ø¯Ø§Ù„Ù‡ ØªÙ‚ÙˆÙ… Ø¨ØªØ­ÙˆÙŠÙ„ Ø§Ù„Ù…Ù„ÙØ§Øª ÙÙŠ Ù…Ø¬Ù„Ø¯ Ø§Ù„Ù‰ ØµÙŠØºØ© pdf 

    Args:
        outdir (string): Ø§Ø³Ù… Ø§Ù„Ù…Ø¬Ù„Ø¯ Ø§Ù„Ø°ÙŠ ØªØ±ÙŠØ¯ ØªØ­ÙˆÙŠÙ„ Ø§Ù„Ù…Ù„ÙØ§Øª Ù…Ù†Ù‡
    """    
    files = os.listdir(outdir)

    for file in files:
        if not file.endswith(".json"):
            if not pages_range:
                subprocess.run(['soffice', '--headless', '--convert-to', 'pdf:writer_pdf_Export', '--outdir', outdir, '--page-ranges', '1-2', f'{outdir}/{file}'])
            else:
                subprocess.run(['soffice', '--headless', '--convert-to', 'pdf:writer_pdf_Export', '--outdir', outdir, f'{outdir}/{file}'])

def column_index_from_string(column_string):
    """Converts a column letter to a column index."""
    index = 0
    for i, char in enumerate(column_string):
        index = index * 26 + (ord(char.upper()) - ord('A')) + 1
    return index

def get_column_letter(column_index):
    """
    Convert a column index to the corresponding column letter.
    """
    div = column_index +1
    column_letter = ""
    while div > 0:
        modulo = (div - 1) % 26
        column_letter = chr(65 + modulo) + column_letter
        div = (div - modulo) // 26
    return column_letter

def strategy_ladder ( ):
    """
    Processes and fills a strategy ladder template with assessment data for each student.

    Assumes an existing strategy ladder template at './ods_strategy_ladder/SL-5,6,7.ods'.
    Reads student data from an Excel file obtained using the Read_E_Side_Note_Marks_ods function.
    Generates random numbers for specific assessment cells based on provided ranges.

    Example Usage:
    ```python
    strategy_ladder()
    ```

    Note:
    Ensure that the file specified by Read_E_Side_Note_Marks_ods('./'+file_name) returns the expected student details.

    Parameters:
    - None

    Returns:
    - None

    Side Effect:
    - Saves individual strategy ladder sheets for each class in the './send_folder' directory.

    """    
    doc = ezodf.opendoc('./ods_strategy_ladder/SL-5,6,7.ods')
    sheet = doc.sheets[0] # Assuming you want to work with the first sheet

    # Access cell values
    cell = sheet['AM1']
    cells = cell.value.split('-')
    firstCell, secondCell, beforeLastCell = cells[0].split('/')[0], cells[0].split('/')[1], cells[0].split('/')[-2]
    thirdCell = column_index_from_string(secondCell)+1
    arb_fill_cells = cells[1].split('.')

    # function parameter variables
    outdir = './send_folder'
    file_name = 'Ø¹Ù…Ø±Ùˆ Ø§Ù„Ù…Ø·Ø§Ø±Ù†Ø©-9981048186(2).ods'
    student_details = Read_E_Side_Note_Marks_ods('./'+file_name)
    term = 2


    # Iterate over student data
    for students_data_list in student_details['file_data']:
        class_name = students_data_list['class_name'].split('=')[0].replace('Ø§Ù„ØµÙ', '')
        
        for row_idx, student_info in enumerate(students_data_list['students_data'], start=9):
            assessments = [
                student_info[f'term{term}']['assessment1'],
                student_info[f'term{term}']['assessment2'],
                student_info[f'term{term}']['assessment3'],
                student_info[f'term{term}']['assessment4']
                ]
            name_conter = row_idx - 8
            sheet[row_idx ,0].set_value(name_conter)
            sheet[row_idx ,1].set_value(student_info['name'])
            sheet[row_idx ,column_index_from_string(firstCell)].set_value(student_info[f'term{term}']['assessment1'])
            sheet[row_idx ,column_index_from_string(secondCell)].set_value(student_info[f'term{term}']['assessment2'])
            sheet[row_idx ,thirdCell].set_value(student_info[f'term{term}']['assessment3'])
            sheet[row_idx ,column_index_from_string(beforeLastCell)].set_value(student_info[f'term{term}']['assessment4'])
            sheet[row_idx ,column_index_from_string(beforeLastCell)+1].set_value(sum(int(assessment) if assessment != '' else 0 for assessment in assessments) )

            Assess1_generator = RandomNumberGenerator(student_info[f'term{term}']['assessment1'], RandomNumberGenerator.convert_to_ranges(arb_fill_cells)).generate_numbers_with_sum()
            Assess1_generator = Assess1_generator if Assess1_generator is not None else [''] * len(arb_fill_cells)
            Assess1_start_column = column_index_from_string(firstCell) - len(arb_fill_cells)
            for column_idx, value in enumerate(Assess1_generator, start=Assess1_start_column):
                sheet[row_idx ,column_idx].set_value(value)

            Assess2_generator = RandomNumberGenerator(student_info[f'term{term}']['assessment2'], RandomNumberGenerator.convert_to_ranges(arb_fill_cells)).generate_numbers_with_sum() 
            Assess2_generator = Assess2_generator if Assess2_generator is not None else [''] * len(arb_fill_cells)
            Assess2_start = column_index_from_string(secondCell) - len(arb_fill_cells)
            for column_idx, value in enumerate(Assess2_generator, start=Assess2_start):
                sheet[row_idx ,column_idx].set_value(value)

            Assess4_generator = RandomNumberGenerator(student_info[f'term{term}']['assessment4'], RandomNumberGenerator.convert_to_ranges([int(i) * 2 for i in arb_fill_cells])).generate_numbers_with_sum() 
            Assess4_generator = Assess4_generator if Assess4_generator is not None else [''] * len(arb_fill_cells)
            Assess4_start = column_index_from_string(beforeLastCell) - len(arb_fill_cells)
            for column_idx, value in enumerate(Assess4_generator, start=Assess4_start):
                sheet[row_idx ,column_idx].set_value(value)
            
            
        doc.saveas(f'{outdir}/{class_name}.ods')    

class RandomNumberGenerator:
    """A class for generating random numbers with a specified sum and individual ranges.

    Example Usage:
    ```python
    total_sum = 18
    numbers = [3, 2, 4, 5, 1, 3]
    ranges = RandomNumberGenerator.convert_to_ranges(numbers)  # ranges = [(0, 3), (0, 3), (0, 5), (0, 5), (0, 2), (0, 2)]
    
    generator = RandomNumberGenerator(total_sum, ranges)
    result = generator.generate_numbers_with_sum()
    print(result)
    ```

    Attributes:
    - total_sum (int): The desired sum of the generated numbers.
    - ranges (list): A list of tuples representing the minimum and maximum values for each number.

    Methods:
    - `generate_numbers(self)`: Generates a list of random numbers based on the specified ranges.
    - `check_sum(self, numbers)`: Checks if the sum of the given list of numbers is equal to the specified total_sum.
    - `generate_numbers_with_sum(self)`: Generates a set of numbers that satisfy the specified sum and ranges.

    Static Methods:
    - `convert_to_ranges(numbers)`: Converts a list of numbers into a list of ranges represented as tuples.

    Example Usage:
    ```python
    total_sum = 18
    numbers = [3, 2, 4, 5, 1, 3]
    ranges = RandomNumberGenerator.convert_to_ranges(numbers)
    
    generator = RandomNumberGenerator(total_sum, ranges)
    result = generator.generate_numbers_with_sum()
    print(result)
    ```"""
    def __init__(self, total_sum, ranges):
        self.total_sum = total_sum
        self.ranges = ranges
    
    def generate_numbers(self):
        numbers = [random.randint(minimum, maximum) for minimum, maximum in self.ranges]
        return numbers
    
    def check_sum(self, numbers):
        return sum(numbers) == self.total_sum
    
    def generate_numbers_with_sum(self):
        for numbers in product(*(range(minimum, maximum + 1) for minimum, maximum in self.ranges)):
            if self.check_sum(numbers):
                return numbers
    
    @staticmethod
    def convert_to_ranges(numbers):
        ranges = [(1, int(number)) for number in numbers]
        return ranges

def fill_student_absent_doc_name_days_cover(student_details , ods_file, outdir ,context = None ,absent_data_list = None):
    """
    Fill an OpenDocument Spreadsheet (ODS) file with student information and generate corresponding documents.

    Parameters:
    - student_details (dict): A dictionary containing information about students.
    - ods_file (str): The path to the input ODS file.
    - outdir (str): The directory where the generated documents will be saved.

    Returns:
    None

    This function opens an ODS file, fills it with student details, and then saves the modified file.
    It generates additional documents with custom shapes based on the provided information.
    The generated documents include student details, attendance data, and cover information.

    The 'student_details' dictionary should have the following keys:
    - 'students_info': A list of dictionaries containing individual student information.
    - 'class_name': The class name.
    - 'year_code': The academic year code.
    - 'start_date': The start date of the academic year.
    - 'school_bridge': The school bridge information.
    - 'school_name_code': The school name and code.
    - 'teacher_incharge_name': The name of the teacher in charge.

    The 'ods_file' parameter is the path to the input ODS file that will be modified.

    The 'outdir' parameter specifies the directory where the generated documents will be saved.
    The generated documents include ODS files and corresponding PDF files.

    Example Usage:
    fill_student_absent_doc_name_days_cover(student_details, 'input.ods', 'output_directory')
    """    
    doc = ezodf.opendoc(ods_file)
        
    sheet_name = 'Sheet1'
    sheet = doc.sheets[sheet_name]

    students_data_lists = student_details['students_info']
    class_name = student_details['class_name']
    if context is None :
        context = {27 : 'Y69=AP123', 2 : 'A69=V123' ,3 : 'Y128=AP182', 26 : 'A128=V182' ,25 : 'Y186=AP240', 4 : 'A186=V240' ,5 : 'Y244=AP298', 24 : 'A244=V298' ,
                        23 : 'Y302=AP356', 6 : 'A302=V356' ,7 : 'Y360=AP414', 22 : 'A360=V414' ,21 : 'Y418=AP472', 8 : 'A418=V472' ,9 : 'Y476=AP530', 20 : 'A476=V530' ,
                        19 : 'Y534=AP588', 10 : 'A534=V588' ,11 : 'Y592=AP646', 18 : 'A592=V646' ,17 : 'Y650=AP704', 12 : 'A650=V704' ,13 : 'Y708=AP762', 16 : 'A708=V762' ,
                        15 : 'Y766=AP820', 14 : 'A766=V820' }

    year1 , year2 = student_details['year_code'].split('-')
    for i in range(183,820,58):
        sheet[f"E{i}"].set_value(f'{class_name.replace("Ø§Ù„ØµÙ" ,"")}')
        sheet[f"AN{i}"].set_value(f'{year2} / {year1}')

    for counter,student_info in enumerate(students_data_lists, start=0):
        
        # row_idx = counter + int(context[str(page)].split(':')[0][1:]) - 1  # compute the row index based on the counter
        if context is not None:
            row_idx = counter + 69
            row_idx2 = counter + 128
        else:
            row_idx = counter + int(context[1].split('=')[0][1:])
            row_idx2 = counter + int(context[2].split('=')[0][1:])
        birth_data = student_info['birth_date'].split('-')
        years, months, days = calculate_age(student_info['birth_date'],student_details['start_date'] )
        
        sheet[f"G{row_idx}"].set_value(student_info['first_name'])
        sheet[f"I{row_idx}"].set_value(student_info['second_name'])
        sheet[f"K{row_idx}"].set_value(student_info['third_name'])
        sheet[f"M{row_idx}"].set_value(student_info['last_name'])
        sheet[f"O{row_idx}"].set_value(int(birth_data[2]))
        sheet[f"P{row_idx}"].set_value(int(birth_data[1]))
        sheet[f"Q{row_idx}"].set_value(birth_data[0])
        sheet[f"S{row_idx}"].set_value(student_info['birthPlace_area'])
        sheet[f"U{row_idx}"].set_value(student_info['nationality'])
        sheet[f"Y{row_idx2}"].set_value(student_info['religion'])
        
        sheet[f"AA{row_idx2}"].set_value(days)
        sheet[f"AB{row_idx2}"].set_value(months)
        sheet[f"AC{row_idx2}"].set_value(years)
        
        sheet[f"AH{row_idx2}"].set_value(student_info['guardian_name'])
        sheet[f"AJ{row_idx2}"].set_value(student_info['guardian_student_relationship'])
        sheet[f"AL{row_idx2}"].set_value(student_info['guardian_employment'])
        sheet[f"AN{row_idx2}"].set_value(student_info['guardian_phone_number'])
        sheet[f"AO{row_idx2}"].set_value(student_info['address'])
        sheet[f"AP{row_idx2}"].set_value(student_info['student_id'])
        
    months_range = [14,16,18,20,22,24,4,6,8,10,12]

    for  counter, month in enumerate(months_range , start =1):
        section_one_row_start , section_one_row_end = int(re.findall(r'\d+',context[month].split('=')[0])[0])-2 , int(re.findall(r'\d+',context[month].split('=')[1])[0]) 
        section_two_row_start , section_two_row_end = int(re.findall(r'\d+',context[month+1].split('=')[0])[0])-2 , int(re.findall(r'\d+',context[month+1].split('=')[1])[0])
        
        if month in [26]:
            print('')
        
        if counter < 7 :
            year = int(year2) 
        elif counter == 7:
            year = int(year1)
            counter+=1        
        else:
            year = int(year1)
            counter+=1
        
        for column in range(8,24):
            day = column-7
            try :
                sheet[section_one_row_start, column-2].set_value( get_day_name_from_date(year , counter , day ) )
                
                if not (day ==1 and counter ==1) :
                    if ("Ø³Ø¨Øª" in get_day_name_from_date(year , counter , day )) or ("Ø¬Ù…Ø¹Ø©" in get_day_name_from_date(year , counter , day )) : 
                        
                        for row in range(section_one_row_start+1 , section_one_row_end ):
                            # FIXME: sheet[row, column].fill = PatternFill(start_color="c0c0c0", fill_type="solid")
                            sheet[row, column-2].set_value('â–’â–’â–’')
            except ValueError:
                pass
            # except AttributeError:
            #     print(section_one_row_start)

        for column in range(24,40):
                day = column-23+16
                try:
                    sheet[section_two_row_start, column].set_value( get_day_name_from_date(year , counter , column-7) )
                    
                    if not (day ==25 and counter ==2) :
                        if ("Ø³Ø¨Øª" in get_day_name_from_date(year , counter , day )) or ("Ø¬Ù…Ø¹Ø©" in get_day_name_from_date(year , counter , day )) : 
                            
                                for row in range(section_two_row_start+1 , section_two_row_end ):
                                    # FIXME: sheet[row, column).fill = PatternFill(start_color="c0c0c0", fill_type="solid")
                                    sheet[row, column].set_value('â–’â–’â–’')
                
                except ValueError:
                    pass
                # except AttributeError:
                #     print(section_two_row_start)         

    doc.saveas(outdir+'one_step_more.ods' )
    
    modeeriah = student_details['school_bridge'].replace('Ù„ÙˆØ§Ø¡ ' , '')
    school_name = student_details['school_name_code'].split(' - ')[1].replace('Ù…Ø¯Ø±Ø³Ø© ', '')
    class_name = student_details['class_name'].split('-')[0].replace('Ø§Ù„ØµÙ' , '')
    sec = student_details['class_name'].split('-')[1]
    teacher = student_details['teacher_incharge_name']
    year1 , year2 = student_details['year_code'].split('-')
    custom_shapes = {
        'modeeriah': f'Ù„ÙˆØ§Ø¡ {modeeriah}',
        'school': school_name,
        'class': class_name,
        'sec': sec,
        'murabee' : teacher,
        'year' : f'{year1}  /  {year2}'
    }

    fill_custom_shape(doc= outdir+'one_step_more.ods', sheet_name='Ø§Ù„ØºÙ„Ø§Ù', custom_shape_values=custom_shapes, outfile= outdir+f'/{class_name}-{sec}.ods')
    
    delete_file(outdir+'one_step_more.ods')

    # outdir = './send_folder'
    filename = f'{class_name}-{sec}.ods'
    command = f'soffice --headless --convert-to pdf:writer_pdf_Export --outdir {outdir} "{outdir}/{filename}"'
    os.system(command)

def get_day_name_from_date(year, month, day):
    """
    Gets the day name in Arabic for a given date.

    Parameters:
    - year (int): The year of the date.
    - month (int): The month of the date.
    - day (int): The day of the date.

    Returns:
    - str: The day name in Arabic.

    Example:
    - day_name = get_day_name_from_date(2024, 1, 9)
      print(day_name)  # Output: 'Ø§Ù„Ø«Ù„Ø§Ø«Ø§Ø¡'
    """    
    # Set the locale to Arabic (Egypt)
    locale.setlocale(locale.LC_ALL, 'ar_EG.utf8')

    # Create a datetime object from the given year, month, and day
    date_object = datetime.date(year, month, day)

    # Get the day name in Arabic
    day_name_arabic = date_object.strftime('%A')

    return day_name_arabic

class InvalidPageRangeError(Exception):
    pass

def print_page_pairs(pair_pages=None,start_page=1 , end_page=None ):
    """
    Prints page pairs based on either a specified number of pairs or a range of pages.

    Parameters:
    - pair_pages (int): The number of page pairs to print. Should be even.
    - start_page (int): The starting page number when range is used.
    - end_page (int): The ending page number when range is used. Should be even.

    Example:
    - print_page_pairs(pair_pages=4)
    - print_page_pairs(start_page=1, end_page=8)
    """    
    if pair_pages is not None:
        if pair_pages % 2 != 0 :
            raise InvalidPageRangeError("Invalid pair pages range: pair_pages should be even")
        pages_length = pair_pages
        counter = pair_pages*2 -2
    else :
        range_ = range(start_page, end_page+1)
        pages_length = int(len(range_)/2)
        pages_list = [i for i in range_]
        counter = len(pages_list)-2
        if len(pages_list) % 2 != 0 :
            raise InvalidPageRangeError("Invalid page range: end_page should be even")
        
    for i in range(start_page,pages_length+1):  
        if i % 2 == 0 :
            print(i ,' ----->' , f'{i + counter+1} - {i}')      
        else :
            print(i ,' ----->', f'{i} - {i + counter+1 }')
        counter -=2

def calculate_age(birth_date, target_date):
    """
    Calculates the age in years, months, and days between two dates.

    Parameters:
    - birth_date (str): The birth date in the format 'YYYY-MM-DD'.
    - target_date (str): The target date for age calculation in the format 'YYYY-MM-DD'.

    Returns:
    - tuple: A tuple containing the age in years, months, and days.

    Example:
    - birth_date = '2007-06-29'
      target_date = '2022-09-01'
      years, months, days = calculate_age(birth_date, target_date)
      print(f"Age on {target_date}: {years} years, {months} months, {days} days")
    """
    birth_date = datetime.datetime.strptime(birth_date, '%Y-%m-%d').date()
    target_date = datetime.datetime.strptime(target_date, '%Y-%m-%d').date()
    age = relativedelta(target_date, birth_date)
    return age.years, age.months, age.days

def fill_Template_With_basic_Student_info(student_details,template='./templet_files/ÙƒØ´Ù Ø§Ù„Ø¨ÙŠØ§Ù†Ø§Øª Ø§Ù„Ø§Ø³Ø§Ø³ÙŠØ© Ù„Ù„Ø·Ù„Ø§Ø¨.xlsx' ,outdir='./send_folder' ):
    """
    Fills an Excel template with basic student information.

    Parameters:
    - student_details (dict): Dictionary containing student details, including class information, school name, and student information.
    - template (str): Path to the Excel template file.
    - outdir (str): Directory where the filled Excel file will be saved.

    Example:
    - student_details = {
        'class_name': 'Ø§Ù„ØµÙ Ø§Ù„Ø«Ø§Ù†ÙŠ Ø¹Ø´Ø± - Ø¹Ù„Ù…ÙŠ',
        'school_name_code': '123456 - Ù…Ø¯Ø±Ø³Ø© Ø§Ù„Ø¹Ù„ÙˆÙ… Ø§Ù„Ø«Ø§Ù†ÙˆÙŠØ©',
        'teacher_incharge_name': 'Teacher Name',
        'principle_name': 'Principal Name',
        'students_info': [
            {'student_id': '123', 'identity_type': 'ID', ...},
            {'student_id': '456', 'identity_type': 'ID', ...},
            ...
        ]
    }
    - fill_Template_With_basic_Student_info(student_details, './templet_files/ÙƒØ´Ù Ø§Ù„Ø¨ÙŠØ§Ù†Ø§Øª Ø§Ù„Ø§Ø³Ø§Ø³ÙŠØ© Ù„Ù„Ø·Ù„Ø§Ø¨.xlsx', './send_folder')
    """
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(template)

    sheet = workbook.active
    # Specify the page ranges and column ranges to enter data into
    page_ranges = [
        (8, 25),
        (39, 56),
        (70, 87),
        (101, 118)
    ]

    counter = 1
    
    class_data = student_details['class_name'].split('-')
    class_name , class_distribution = class_data[1] , class_data[0]
    
    school_name , school_code = student_details['school_name_code'].split(' - ')[1] , [int(digit) for digit in  student_details['school_name_code'].split(' - ')[0] ]
    teacher_incharge_name = student_details['teacher_incharge_name']
    principle_name = student_details['principle_name']
    student_details = student_details['students_info']
                                            
    
    # Iterate over each page range and insert data from the dictionary list
    for start_row, end_row in page_ranges:
        for row_number, dataFrame in zip(range(start_row, end_row+1), student_details):
            if counter > len(student_details):
                break
            sheet.cell(row=row_number, column=1).value = counter
            sheet.cell(row=row_number, column=2).value = dataFrame['student_id']
            sheet.cell(row=row_number, column=3).value = dataFrame['identity_type']
            sheet.cell(row=row_number, column=4).value = dataFrame['first_name']
            sheet.cell(row=row_number, column=5).value = dataFrame['second_name']
            sheet.cell(row=row_number, column=6).value = dataFrame['third_name']
            sheet.cell(row=row_number, column=7).value = dataFrame['last_name']
            sheet.cell(row=row_number, column=8).value = dataFrame['birthPlace_area']
            sheet.cell(row=row_number, column=9).value = dataFrame['birth_date'].replace('-','/')
            sheet.cell(row=row_number, column=10).value = dataFrame['nationality']
            sheet.cell(row=row_number, column=11).value = dataFrame['sex']
            sheet.cell(row=row_number, column=12).value = dataFrame['resident_governorate']
            sheet.cell(row=row_number, column=13).value = dataFrame['resident_district']
            sheet.cell(row=row_number, column=14).value = dataFrame['resident_quarter']
            sheet.cell(row=row_number, column=15).value = ''                                # dataFrame['identity_type']
            sheet.cell(row=row_number, column=16).value = ''
            sheet.cell(row=row_number, column=17).value = counter                                 # dataFrame['student_id']
            sheet.cell(row=row_number, column=18).value = dataFrame['student_id']
            sheet.cell(row=row_number, column=19).value = dataFrame['first_name']+' '+dataFrame['last_name']
            sheet.cell(row=row_number, column=20).value = dataFrame['marital_status']
            sheet.cell(row=row_number, column=21).value = dataFrame['mother_name']
            sheet.cell(row=row_number, column=22).value = dataFrame['study_type']
            sheet.cell(row=row_number, column=23).value = dataFrame['father_education_level']
            sheet.cell(row=row_number, column=24).value = dataFrame['mother_education_level']
            sheet.cell(row=row_number, column=25).value = dataFrame['guardian_name']
            sheet.cell(row=row_number, column=26).value = dataFrame['guardian_student_relationship']
            sheet.cell(row=row_number, column=27).value = dataFrame['guardian_employment']
            sheet.cell(row=row_number, column=28).value = dataFrame['family_size']
            sheet.cell(row=row_number, column=29).value = dataFrame['student_siblings_rank']
            sheet.cell(row=row_number, column=30).value = dataFrame['student_health_status']
            sheet.cell(row=row_number, column=31).value = dataFrame['student_academic_status']
            sheet.cell(row=row_number, column=32).value = dataFrame['external_aid_available']
            sheet.cell(row=row_number, column=33).value = dataFrame['monthly_family_income']
            sheet.cell(row=row_number, column=34).value = dataFrame['religion']
            sheet.cell(row=row_number, column=35).value = dataFrame['govt_card_attribute']
            sheet.cell(row=row_number, column=35).value = dataFrame['guardian_phone_number']
    
    sheet.cell(row=1, column=39).value = school_name
    sheet.cell(row=3, column=39).value = class_distribution
    sheet.cell(row=4, column=39).value = class_name.replace('Ø§Ù„ØµÙ', '')
    sheet.cell(row=5, column=39).value = teacher_incharge_name
    sheet.cell(row=6, column=39).value = principle_name
    
    for column_idx , digit in enumerate(school_code,start=39):
        sheet.cell(row=2, column=column_idx).value = digit

        
        counter += 1
    
    # Save the workbook
    workbook.save(outdir+'/your_file.xlsx')

def get_student_statistic_info(username,password, identity_nos=None, students_openemis_nos=None, student_ids=None, session=None , teacher_full_name=False ,auth=None):
    """
    The function `get_student_statistic_info` retrieves statistical information about students based on
    various parameters.
    
    :param username: The username is a string that represents the username of the user who is trying to
    access the student statistic information. This is typically used for authentication purposes
    :param password: The password parameter is used to authenticate the user and verify their identity
    :param identity_nos: A list of identity numbers of the students for which you want to retrieve
    statistics information. If not provided, statistics information will be retrieved for all students
    :param students_openemis_nos: A list of OpenEMIS numbers of the students for which you want to
    retrieve statistics information
    :param student_ids: A list of student IDs for which you want to retrieve statistics information
    :param session: The session parameter is used to specify the session of the student. It can be a
    specific session or None to get the statistic info for all sessions
    :param teacher_full_name: A boolean value indicating whether to include the teacher's full name in
    the output, defaults to False (optional)
    """
    if auth is None:
        auth = get_auth(username,password)
    final_dict_info = []
    identity_types = get_IdentityTypes(auth, session=session)
    area_data = get_AreaAdministrativeLevels(auth, session=session)['data']
    nationality_data = {i['id']: i['name'] for i in make_request(auth=auth, url='https://emis.moe.gov.jo/openemis-core/restful/v2/User-NationalityNames')['data']}
    curr_period_data = get_curr_period(auth,session=session)
    curr_period = curr_period_data['data'][0]['id']
    inst_id = inst_name(auth,session=session)['data'][0]['Institutions']['id']
    class_data_url = f'https://emis.moe.gov.jo/openemis-core/restful/v2/Institution-InstitutionClasses?_contain=Institutions&_fields=id,name,institution_id,academic_period_id,Institutions.code,Institutions.name,Institutions.area_administrative_id&_finder=classesByInstitutionAndAcademicPeriod[institution_id:{inst_id};academic_period_id:{curr_period}]'
    class_data = make_request(auth=auth , url=class_data_url ,session=session)['data'][0]
    academic_period_id, institution_id , institution_name_code ,modeeriah= class_data['academic_period_id'],class_data['institution_id'],class_data['institution']['code_name'], [i['name'] for i in area_data if i['id'] == class_data['institution']['area_administrative_id']][0]
    class_name ,teacher_incharge_name= '' , ['']*4
    
    # Ø¨ÙŠØ§Ù†Ø§Øª Ø§Ù„Ù…Ø¯Ø±Ø³Ø© Ù„Ù„Ø´Ù‡Ø§Ø¯Ø§Øª Ø§Ù„Ù…Ù„ÙˆÙ†Ø©
    school_data = inst_name(auth,session=session)['data'][0]
    inst_id = school_data['Institutions']['id']
    school_info = make_request(auth=auth , url=f'https://emis.moe.gov.jo/openemis-core/restful/Institution-Institutions.json?_limit=1&id={inst_id}&_contain=InstitutionLands.CustomFieldValues,AreaAdministratives,Areas',session=session)['data'][0]
    
    school_address = school_info['address']
    school_phone_number = school_info['telephone']
    school_national_id = school_info['code']
    school_directorate = ' Ù…Ø¯ÙŠØ±ÙŠØ© Ù„ÙˆØ§Ø¡ '+school_info['area']['name']
    school_bridge = ' Ù„ÙˆØ§Ø¡ '+school_info['area']['name']

    
    staff = get_school_teachers(auth,id=institution_id ,session=session)['staff'] 

    working_teachers = [teacher 
                        for teacher in staff 
                            if teacher['staff_status'] == 1
                        ]
    
    principle_name = [
                        i['name_list']
                        for i in working_teachers 
                        if '- Ù…Ø¯ÙŠØ±' in i['position']
                    ][0]
    
    # Ø§Ø­Ø¶Ø± Ø¨ÙŠØ§Ù†Ø§Øª Ø§Ù„ØµÙ Ø§Ù„ÙŠ Ù…Ø¹ Ø§Ù„Ù…Ø¹Ù„Ù…
    institution_class_id ,class_name = class_data['id'] , class_data['name']
    
    
    if identity_nos is not None:
        for chunk in chunks(identity_nos, 20):
            joined_string = ','.join([f'identity_number:{i}' for i in chunk])
            url='https://emis.moe.gov.jo/openemis-core/restful/Institution-StudentUser?_limit=0&_contain=BirthplaceAreas,CustomFieldValues,Identities&_orWhere='+joined_string
            students_info_data = make_request(auth=auth , url=url,session=session)['data']
            final_dict_info.extend(process_students_info(students_info_data, identity_types, nationality_data , area_data))

    elif students_openemis_nos is not None:
        for chunk in chunks(students_openemis_nos, 20):
            joined_string = ','.join([f'openemis_no:{i}' for i in chunk])
            url='https://emis.moe.gov.jo/openemis-core/restful/Institution-StudentUser?_limit=0&_contain=BirthplaceAreas,CustomFieldValues,Identities&_orWhere='+joined_string
            students_info_data = make_request(auth=auth , url=url,session=session)['data']
            final_dict_info.extend(process_students_info(students_info_data, identity_types, nationality_data , area_data))            

    elif student_ids is not None:
        for chunk in chunks(student_ids, 20):
            joined_string = ','.join([f'id:{i}' for i in chunk])
            url='https://emis.moe.gov.jo/openemis-core/restful/Institution-StudentUser?_limit=0&_contain=BirthplaceAreas,CustomFieldValues,Identities&_orWhere='+joined_string
            students_info_data = make_request(auth=auth , url=url,session=session)['data']
            final_dict_info.extend(process_students_info(students_info_data, identity_types, nationality_data , area_data))

    else :
        
        # # Ø§Ø­Ø¶Ø± Ø§Ø³Ù…Ø§Ø¡ Ø§Ù„Ø·Ù„Ø§Ø¨ ÙÙŠ Ø§Ù„ØµÙ
        url = f"https://emis.moe.gov.jo/openemis-core/restful/v2/Institution.InstitutionSubjectStudents?_fields=student_id&_limit=0&academic_period_id={academic_period_id}&institution_class_id={institution_class_id}&institution_id={institution_id}"
        student_ids= make_request(url,auth,session=session)
        student_ids =list(set([i['student_id'] for i in student_ids['data'] ]))
        for chunk in chunks(student_ids, 20):
            joined_string = ','.join(str(i) for i in [f'id:{i}' for i in chunk])
            url='https://emis.moe.gov.jo/openemis-core/restful/Institution-StudentUser?_limit=0&_contain=BirthplaceAreas,CustomFieldValues,Identities&_orWhere='+joined_string
            students_info_data = make_request(auth=auth , url=url,session=session)['data']
            final_dict_info.extend(process_students_info(students_info_data, identity_types, nationality_data , area_data))

    try : 
        teacher_incharge_name = [
                        i['name_list'] 
                        for i in working_teachers 
                            if i['nat_id'] == str(username) or i['default_nat_id'] == str(username) 
                        ][0]
    except :
        t =make_request(auth=auth , url=f'https://emis.moe.gov.jo/openemis-core/restful/Institution-Institutions.json?_limit=1&id={inst_id}&_contain=Staff.Users,Staff.Positions' )
        
        teacher_incharge_name = [
            [i['user']['first_name'],
                i['user']['middle_name'],
                i['user']['third_name'],
                i['user']['last_name']] for i in t['data'][0]['staff'] 
                                            if i['user']['username'] == str(username)
                                            ][0]
    
    
    sorted_final_dict_info=sorted(final_dict_info, key=lambda x: x['full_name'])
    
    # c['code'] ====> '2022-2023'
    # c['code'].split('-')[0]  ====> '2022'
    # c['code'].split('-')[1]  ====> '2023'

    year_code = curr_period_data['data'][0]['code']

    start_of_the_year_date = curr_period_data['data'][0]['start_date']
    end_of_the_year_date = curr_period_data['data'][0]['end_date']
    return {'students_info':sorted_final_dict_info ,
            'class_name':class_name ,
            'school_name_code':institution_name_code ,
            'modeeriah' : modeeriah,
            'principle_name': principle_name[0]+' '+principle_name[1]+' '+principle_name[3],
            'teacher_incharge_name': teacher_incharge_name[0]+' '+teacher_incharge_name[1]+' '+teacher_incharge_name[3] if not teacher_full_name else ' '.join(teacher_incharge_name),
            'year_code': year_code,
            'school_address' :school_address ,
            'school_phone_number' :school_phone_number ,
            'school_national_id' :school_national_id ,
            'school_directorate' :school_directorate ,
            'school_bridge' :school_bridge ,
            'academic_year_1':year_code.split('-')[0],
            'academic_year_2':year_code.split('-')[1],
            'start_date':start_of_the_year_date,
            'end_date': end_of_the_year_date
            }

def chunks(lst, n):
    """Yield successive n-sized chunks from lst."""
    for i in range(0, len(lst), n):
        yield lst[i:i + n]

def process_students_info(students_info_data, identity_types, nationality_data , area_data):
    """
    The function "process_students_info" takes in four parameters and does some processing on student
    information data.
    
    :param students_info_data: This parameter is a list of dictionaries containing information about
    students. Each dictionary represents a student and contains keys such as "name", "age",
    "identity_type", "identity_number", "nationality", and "area"
    :param identity_types: A list of different types of identities that a student can have, such as
    "student ID", "passport number", etc
    :param nationality_data: The nationality_data parameter is a data structure that contains
    information about different nationalities. It could be a dictionary, list, or any other data
    structure that allows you to store and retrieve information about nationalities
    :param area_data: The area_data parameter is a dictionary that contains information about different
    areas or regions. It could include details such as the area name, population, location, and any
    other relevant information about each area
    """
    dic_list=[]
    options_values_dic ={88: 'Ø§Ø¹Ø²Ø¨',89: 'Ù…ØªØ²ÙˆØ¬',90: 'Ø§Ø±Ù…Ù„',91: 'Ù…Ø·Ù„Ù‚Ø©',124: 'Ù†Ø¸Ø§Ù…ÙŠØ©',125: 'Ù…Ù†Ø²Ù„ÙŠØ©',80: 'Ø§Ù…ÙŠ',81: 'Ø§Ø³Ø§Ø³ÙŠ',82: 'Ø«Ø§Ù†ÙˆÙŠ',83: 'ÙƒÙ„ÙŠØ© Ù…Ø¬ØªÙ…Ø¹',84: 'Ø¨ÙƒØ§Ù„ÙˆØ±ÙŠÙˆØ³',85: 'Ø¯Ù„ÙˆÙ… Ø¹Ø§Ù„ÙŠ'
                        ,86: 'Ù…Ø§Ø¬Ø³ØªÙŠØ±',87: 'Ø¯ÙƒØªÙˆØ±Ø§Ø©',92: 'Ø§Ù…ÙŠ',93: 'Ø§Ø³Ø§Ø³ÙŠ',94: 'Ø«Ø§Ù†ÙˆÙŠ',95: 'ÙƒÙ„ÙŠØ© Ù…Ø¬ØªÙ…Ø¹',96: 'Ø¨ÙƒØ§Ù„ÙˆØ±ÙŠÙˆØ³',97: 'Ø¯Ù„ÙˆÙ… Ø¹Ø§Ù„ÙŠ',98: 'Ù…Ø§Ø¬Ø³ØªÙŠØ±',99: 'Ø¯ÙƒØªÙˆØ±Ø§Ø©',115: 'Ø§Ø¨',116: 'Ø§Ù…'
                        ,117: 'Ù†ÙØ³Ù‡',118: 'Ø¹Ù…-Ø¹Ù…Ù‡',119: 'Ø¬Ø¯-Ø¬Ø¯Ø©',120: 'Ø®Ø§Ù„-Ø®Ø§Ù„Ø©',121: 'Ø§Ø®-Ø§Ø®Øª',136: 'Ø§Ø®Ø±Ù‰',100: 'Ø³Ù„ÙŠÙ…',101: 'ØºÙŠØ± Ø³Ù„ÙŠÙ…',110: 'Ù†Ø§Ø¬Ø­',111: 'Ù…Ø¹ÙŠØ¯',112: 'Ù…ØªØ³Ø±Ø¨'
                        ,122: 'Ù„Ø§ ÙŠÙˆØ¬Ø¯',123: 'ÙŠÙˆØ¬Ø¯',144: 'Ù†Ø¹Ù…',145: 'Ù„Ø§',127: 'Ø§Ù„Ø§Ø³Ù„Ø§Ù…',128: 'Ø§Ù„Ù…Ø³ÙŠØ­ÙŠØ©',113: 'Ù„Ø§ ÙŠØ­Ù…Ù„ Ø¨Ø·Ø§Ù‚Ø©',114: 'Ù„Ø§Ø¬Ø¦',137: 'Ø±ÙˆØ¶Ø© 2 (ØªÙ…Ù‡ÙŠØ¯ÙŠ)',138: 'Ø±ÙˆØ¶Ø© 1 (Ø¨Ø³ØªØ§Ù†)'
                        ,141: 'Ø±ÙˆØ¶Ø© 2 (ØªÙ…Ù‡ÙŠØ¯ÙŠ) Ø±ÙˆØ¶Ø© 1 (Ø¨Ø³ØªØ§Ù†)',142: 'Ù„Ù… ÙŠÙ„ØªØ­Ù‚',158: 'ÙŠØ±Ø³Ù…',159: 'Ø§Ù„Ø®Ø·',160: 'Ø§Ù„ØµÙˆØª Ø§Ù„Ø¬Ù…ÙŠÙ„',161: 'Ø§Ù„Ø¹Ø²Ù',162: 'Ø±ÙŠØ§ØµÙŠØ©',164: 'Ø§Ù„ØªÙ…Ø«ÙŠÙ„',165: 'Ø§Ù„Ø´Ø¹Ø±'
                        ,166: 'Ø§Ù„Ø±ÙˆØ§ÙŠØ©',167: 'Ø§Ø®Ø±Ù‰',168: 'Ø§Ù„ØªØ³Ø±ÙŠØ¹ Ø§Ù„Ø£ÙƒØ§Ø¯ÙŠÙ…ÙŠ',169: 'Ù…Ø¯Ø§Ø±Ø³ Ø§Ù„Ù…Ù„Ùƒ Ø¹Ø¨Ø¯ Ø§Ù„Ù„Ù‡ Ø§Ù„Ø«Ø§Ù†ÙŠ Ù„Ù„ØªÙ…ÙŠØ²',140: 'Ø§Ù„Ù…Ø±Ø§ÙƒØ² Ø§Ù„Ø±ÙŠØ§Ø¯ÙŠØ©',171: 'ØºØ±Ù Ù…ØµØ§Ø¯Ø± Ø§Ù„Ø·Ù„Ø¨Ø© Ø§Ù„Ù…ÙˆÙ‡ÙˆØ¨ÙŠÙ†',172: 'Ø¬Ø§Ø¦Ø²Ø© Ø§Ù†ØªÙ„'
                        ,173: 'Ø¬Ø§Ø¦Ø²Ø© Ø±ÙˆØ¨ÙˆØªÙƒØ³',174: 'Ø¬Ø§Ø¦Ø²Ø© Ø§Ø®Ø±Ù‰',175: 'Ø§Ø®ØªØ±Ø§Ø¹',176: 'Ø§Ø¨ØªÙƒØ§Ø±',177: 'ÙÙƒØ±Ø© Ø§Ø¨Ø¯Ø§Ø¹ÙŠØ©',178: 'Ø§Ø³ØªÙƒØ´Ø§Ù Ù…Ù‚ØµÙˆØ¯', 179:'Ù†Ø¹Ù…' ,180 :'Ù„Ø§' ,1:"Ø°ÙƒØ±" ,2:"Ø§Ù†Ø«Ù‰" }
    variables = {'mother_name': 1,'guardian_employment': 5,'student_siblings_rank': 7,'family_size': 6,'address':'',
        'monthly_family_income': 9,'guardian_name': 11,'father_education_level': 12,'marital_status': 13,
        'mother_education_level': 14,'student_health_status': 15,'student_academic_status': 17,'govt_card_attribute': 18,
        'guardian_student_relationship': 19,'external_aid_available': 20,'study_type': 21,'religion': 22,'did_student_attend_kindergarten': 28,
        'is_family_registered_nationally': 30,'guardian_phone_number': 31,'mother_nationality': 32,'did_student_join_international_program': 37,
        'did_student_attend_kindergarten' : '','intelligence_giftedness' : '','talent_and_giftedness' : '','talent' : '',
        }
    
    for data_item in students_info_data:
        """ 
        ØªÙØ§ØµÙŠÙ„ Ø­Ù‚ÙˆÙ„ Ø§Ù„Ø¨ÙŠÙŠØ§Ù†Ø§Øª Ø§Ù„Ø§Ø­ØµØ§Ø¦ÙŠØ© Ù„Ù„Ø·Ø§Ù„Ø¨ custom_field_values keys
            1 ==> Ø§Ø³Ù… Ø§Ù„Ø£Ù… (Ø§Ù„Ø§Ø³Ù… Ø§Ù„Ø§ÙˆÙ„ ÙˆØ§Ù„Ø¹Ø§Ø¦Ù„Ø©)  variable: 'mother_name'
            5 ==> Ø¹Ù…Ù„ ÙˆÙ„ÙŠ Ø§Ù„Ø£Ù…Ø±    variable: 'guardian_employment'
            6 ==> Ø¹Ø¯Ø¯ Ø£ÙØ±Ø§Ø¯ Ø§Ù„Ø£Ø³Ø±Ø©     variable: 'family_size'
            7 ==> ØªØ±ØªÙŠØ¨ Ø§Ù„Ø·Ø§Ù„Ø¨ Ø¨ÙŠÙ† Ø§Ø®ÙˆØ§Ù†Ù‡  variable: 'student_siblings_rank'
            9 ==> Ø¯Ø®Ù„ Ø§Ù„Ø£Ø³Ø±Ø© Ø§Ù„Ø´Ù‡Ø±ÙŠ    variable: 'monthly_family_income'
            11 ==> Ø§Ø³Ù… ÙˆÙ„ÙŠ Ø§Ù„Ø£Ù…Ø±   variable: 'guardian_name'
            12 ==> Ù…Ø³ØªÙˆÙ‰ ØªØ¹Ù„ÙŠÙ… Ø§Ù„Ø§Ø¨    variable: 'father_education_level'
            13 ==> Ø§Ù„Ø­Ø§Ù„Ø© Ø§Ù„Ø§Ø¬ØªÙ…Ø§Ø¹ÙŠØ©   variable: 'marital_status'
            14 ==> Ù…Ø³ØªÙˆÙ‰ ØªØ¹Ù„ÙŠÙ… Ø§Ù„Ø§Ù…    variable: 'mother_education_level'
            15 ==> Ø§Ù„ÙˆØ¶Ø¹ Ø§Ù„ØµØ­ÙŠ Ù„Ù„Ø·Ø§Ù„Ø¨  variable: 'student_health_status'
            17 ==> Ø§Ù„ÙˆØ¶Ø¹ Ø§Ù„Ø¯Ø±Ø§Ø³ÙŠ Ù„Ù„Ø·Ø§Ù„Ø¨    variable: 'student_academic_status'
            18 ==> ØµÙØ© Ø¨Ø·Ø§Ù‚Ø© Ø§Ù„ØºÙˆØ«     variable: 'govt_card_attribute'
            19 ==> Ø¹Ù„Ø§Ù‚Ø© ÙˆÙ„ÙŠ Ø§Ù„Ø§Ù…Ø± Ø¨Ø§Ù„Ø·Ø§Ù„Ø¨     variable: 'guardian_student_relationship'
            20 ==> Ø§Ù„Ù…Ø³Ø§Ø¹Ø¯Ø§Øª Ø§Ù„Ø®Ø§Ø±Ø¬ÙŠØ© (Ø¥Ù† ÙˆØ¬Ø¯Øª)    variable: 'external_aid_available'
            21 ==> Ù†ÙˆØ¹ Ø§Ù„Ø¯Ø±Ø§Ø³Ø©     variable: 'study_type'
            22 ==> Ø§Ù„Ø¯ÙŠØ§Ù†Ø©     variable: 'religion'
            28 ==> Ù‡Ù„ Ø§Ù„ØªØ­Ù‚ Ø§Ù„Ø·Ø§Ù„Ø¨/Ø§Ù„Ø·Ø§Ù„Ø¨Ø© Ø¨Ø±ÙŠØ§Ø¶ Ø§Ù„Ø§Ø·ÙØ§Ù„ØŸ  variable: 'did_student_attend_kindergarten'
            30 ==> Ù‡Ù„ Ø§Ù„Ø§Ø³Ø±Ø© Ù…Ø³Ø¬Ù„Ø© Ø¨Ø§Ù„Ù…Ø¹ÙˆÙ†Ø© Ø§Ù„ÙˆØ·Ù†ÙŠØ©ØŸ   variable: 'is_family_registered_nationally'
            31 ==> Ù‡Ø§ØªÙ ÙˆÙ„ÙŠ Ø§Ù…Ø± Ø§Ù„Ø·Ø§Ù„Ø¨     variable: 'guardian_phone_number'
            32 ==> Ø¬Ù†Ø³ÙŠØ© Ø§Ù„Ø§Ù…  variable: 'mother_nationality'
            37 ==> Ù‡Ù„ Ø§Ù„ØªØ­Ù‚ Ø§Ù„Ø·Ø§Ù„Ø¨ Ø¨Ø§Ù„Ø¨Ø±Ù†Ø§Ù…Ø¬ Ø§Ù„Ø¯ÙˆÙ„ÙŠ    variable: 'did_student_join_international_program'
                                    Ø§Ù„Ù…ÙˆÙ‡Ø¨Ø©    variable: 'talent'
                                Ø§Ù„ØªÙÙˆÙ‚ Ùˆ Ø§Ù„Ù…ÙˆÙ‡Ø¨Ø©   variable: 'excellence_and_talent'
                                Ø§Ù„ØªÙÙˆÙ‚ Ø§Ù„Ø¹Ù‚Ù„ÙŠ     variable: 'intellectual_excellence'   
        """

        custom_field_values = data_item['custom_field_values'] 
        custom_field_values_dict = {item['student_custom_field_id']: str(item[key]) for item in custom_field_values 
                                                                                        for key in ['text_value', 
                                                                                                    'number_value', 
                                                                                                    'decimal_value', 
                                                                                                    'textarea_value', 
                                                                                                    'date_value', 
                                                                                                    'time_value'] 
                                                                                            if item.get(key) is not None }

        result = {var_name: custom_field_values_dict.get(var_id, '') for var_name, var_id in variables.items()}
        result = {
            key: options_values_dic[int(val)] if val.isdigit() 
                and int(val) in options_values_dic  
                    and key not in ['family_size', 'monthly_family_income', 'student_siblings_rank', 'guardian_phone_number'] 
            else val
            for key, val in result.items()
        }

        area_chain = find_area_chain(data_item['address_area_id'], area_data).split(' - ')
        result['student_id'] = data_item['id']
        result['birthPlace_area'] = '' if data_item['birthplace_area'] is None else data_item['birthplace_area']['name'] 
        result['identity_type'] =  '' if data_item['identity_type_id'] != 825 else identity_types[data_item['identity_type_id']]
        result['identity_number'] = '' if len(data_item['identities']) ==0 else data_item['identities'][0]['number']
        result['full_name'] = '' if data_item['name'] is None else data_item['name'] 
        result['first_name'] = '' if data_item['first_name'] is None else data_item['first_name'] 
        result['second_name'] = '' if data_item['middle_name'] is None else data_item['middle_name'] 
        result['third_name'] = '' if data_item['third_name'] is None else data_item['third_name'] 
        result['last_name'] = '' if data_item['last_name'] is None else data_item['last_name'] 
        result['birth_date'] = '' if data_item['date_of_birth'] is None else data_item['date_of_birth'] 
        result['nationality'] = '' if data_item['nationality_id']  is None else nationality_data[data_item['nationality_id']]  
        result['sex'] = '' if data_item['gender_id']  is None else options_values_dic[data_item['gender_id']]  
        result['resident_governorate'] = area_chain[0] if len(area_chain) == 1 else ''
        result['resident_district'] = area_chain[1] if len(area_chain) == 2 else ''
        result['resident_quarter'] = area_chain[2] if len(area_chain) == 3 else ''    
        result['address'] = '' if data_item['address'] is None else data_item['address'] 
        dic_list.append(result)
    
    return dic_list

def find_parent_info(item_id ,area_data):
    """
    The function find_parent_info takes an item_id and area_data as input and returns information about
    the parent of the item with the given item_id.
    
    :param item_id: The unique identifier of the item for which you want to find the parent information
    :param area_data: The area_data parameter is a dictionary that contains information about different
    areas. Each key in the dictionary represents an area ID, and the corresponding value is a dictionary
    that contains information about that area
    """
    for item in area_data:
        if item['id'] == item_id:
            parent_id = item['parent_id']
            name = item['name']
            if parent_id in [3, 4, 5 ,1]:
                return None, name
            return parent_id, name
    return None, None

def find_area_chain(id,area_data):
    """
    The function "find_area_chain" takes an ID and a dictionary of area data and returns the chain of
    areas associated with that ID.
    
    :param id: The id parameter is a unique identifier for a specific chain
    :param area_data: A list of dictionaries containing information about different areas. Each
    dictionary in the list represents an area and has the following keys: 'id' (unique identifier for
    the area), 'name' (name of the area), and 'chain' (name of the chain the area belongs to)
    """
    names = []

    while id is not None:
        id, name = find_parent_info(id ,area_data)
        if name:
            names.append(name)
    
    names.reverse()  # Reverse the order of names            
    output = ' - '.join(names)
    return output

def get_AreaAdministrativeLevels(auth,session=None):
    """
    The function retrieves the administrative levels for a given area.
    
    :param auth: The auth parameter is used for authentication purposes. It typically contains
    information such as an API key or credentials that are required to access certain resources or
    perform certain actions
    :param session: The `session` parameter is an optional parameter that allows you to pass an existing
    session object to the `get_AreaAdministrativeLevels` function. A session object is typically used to
    persist certain data or settings across multiple requests to a server. If you don't provide a
    session object, the function
    """
    url='https://emis.moe.gov.jo/openemis-core/restful/v2/Area-AreaAdministratives?_limit=0&_contain=AreaAdministrativeLevels&_fields=id,name,parent_id,area_administrative_level_id'
    return make_request(auth=auth , url= url,session=session)

def get_IdentityTypes(auth,session=None):
    """
    The function `get_IdentityTypes` retrieves the identity types using the provided authentication and
    session.
    
    :param auth: The auth parameter is used to authenticate the user. It could be a token,
    username/password combination, or any other form of authentication required by the system you are
    working with
    :param session: The `session` parameter is an optional parameter that represents the user's session.
    It can be used to maintain state or store user-specific information throughout multiple requests.
    """
    url='https://emis.moe.gov.jo/openemis-core/restful/v2/FieldOption-IdentityTypes.json?_limit=0&_fields=id,name'
    return { i['id'] : i['name']  for i in make_request(auth=auth , url=url ,session=session)['data']}

def find_default_teachers_creds(auth ,id=None , nat_school=None ,session=None):
    """
    The function "find_default_teachers_creds" is used to find the default credentials of teachers based
    on the provided parameters.
    
    :param auth: The auth parameter is used for authentication purposes. It could be a token or a
    username/password combination
    :param id: The ID of the school where you want to find the default credentials
    :param nat_school: The parameter "nat_school" is used to specify the national school to search for
    default teachers' credentials
    :param session: The session parameter is used to specify the current session or term for which the
    default teachers' credentials are being searched
    """
    if id == None:
        teachers = get_school_teachers(auth,nat_school=nat_school,session=session)['staff']
    else:
        teachers = get_school_teachers(auth,id=id,session=session)['staff']

    working_teachers = [(teacher['name'],teacher['nat_id']) for teacher in teachers if teacher['staff_status'] == 1]

    found_creds,unfound_creds = [],[]
    for teacher in working_teachers:
        if get_auth(teacher [1] , teacher [1]):
            found_creds.append(teacher[1])
            print('found password for this teacher ' + teacher[0]+' -----> '+teacher[1])
        else: 
            unfound_creds.append(str(teacher[0]+'-'+teacher[1]))
    return {'institution_staff': teachers , 'found_creds': found_creds ,'unfound_creds': unfound_creds}

def five_names_every_class_wrapper(auth , emp_number ,term=1 , session=None):
    """
    This function is a wrapper that takes in authentication, employee number, term, and session as
    parameters.
    
    :param auth: An authentication token or object that verifies the user's identity and permissions
    :param emp_number: The employee number of the person accessing the function
    :param term: The term parameter represents the current term or semester of the class. It is an
    optional parameter with a default value of 1, defaults to 1 (optional)
    :param session: The session parameter is an optional parameter that represents the requests.Session
    that will make fetching data faster, it will default to None
    """
    data = five_names_every_class(auth , emp_number ,session=session)
    term = 'term1' if term == 1 else 'term2'
    long_text = ''

    for subject in data['row_data']:
        if ('Ø¹Ø´Ø±' not in subject['className']) and ('Ø¹Ø§Ø´Ø±' not in subject['className']):
            text =''
            middle_index = len(subject['marks_and_name']) // 2
            first_two = subject['marks_and_name'][:2]
            middle_one = subject['marks_and_name'][middle_index]
            last_two = subject['marks_and_name'][-2:]
            for item_dic in first_two : 
                text +=  item_dic['name'] +'\n'+'\t'+ ' Øª1 ---> ' + str(item_dic[term]['assessment1']) +'\n'+'\t'+ ' Øª2 ---> ' + str(item_dic[term]['assessment2'])+'\n'+'\t'+' Øª3 ---> ' +str(item_dic[term]['assessment3']) +'\n'+'\t'+'Ø§Ù„Ù†Ù‡Ø§Ø¦ÙŠ ---> ' +str(item_dic[term]['assessment4'])+'\n' 
            text += '[ .......... ]'+'\n'        
            text +=  middle_one['name'] +'\n'+'\t'+ ' Øª1 ---> ' + str(middle_one[term]['assessment1']) +'\n'+'\t'+ ' Øª2 ---> ' + str(middle_one[term]['assessment2'])+'\n'+'\t'+' Øª3 ---> ' +str(middle_one[term]['assessment3']) +'\n'+'\t'+'Ø§Ù„Ù†Ù‡Ø§Ø¦ÙŠ ---> ' +str(middle_one[term]['assessment4']) +'\n' 
            text += '[ .......... ]'+'\n'            
            for item_dic in last_two : 
                text +=  item_dic['name'] +'\n'+'\t'+ ' Øª1 ---> ' + str(item_dic[term]['assessment1']) +'\n'+'\t'+ ' Øª2 ---> ' + str(item_dic[term]['assessment2'])+'\n'+'\t'+' Øª3 ---> ' +str(item_dic[term]['assessment3']) +'\n'+'\t'+'Ø§Ù„Ù†Ù‡Ø§Ø¦ÙŠ ---> ' +str(item_dic[term]['assessment4'])+'\n' 
            long_text += '\n'+subject['subject']+'//'+subject['className']+'\n'+text + '-'*70
        
    return long_text

def five_names_every_class(auth, emp_username ,session=None ):
    """
    The function takes in authentication, employee username, and an optional session parameter and
    returns a list of five names for every class.
    
    :param auth: This parameter is likely used for authentication purposes. It could be a token or a
    username/password combination that is used to verify the identity of the user making the request
    :param emp_username: The username of the employee who is accessing the function
    :param session: An optional parameter that represents the current session of the user. It is used to
    maintain the state of the user's interaction with the system
    """
    period_id = get_curr_period(auth,session=session)['data'][0]['id']
    user = user_info(auth , emp_username,session=session)
    userInfo = user['data'][0]
    user_id , user_name = userInfo['id'] , userInfo['first_name']+' '+ userInfo['last_name']+'-' + str(emp_username)
    # years = get_curr_period(auth)
    school_data = inst_name(auth,session=session)['data'][0]
    inst_id = school_data['Institutions']['id']
    # school_name = school_data['Institutions']['name']
    # grades = make_request(auth=auth , url=FILL_OFFICIAL_MARKS_DOC_WRAPPER_OFFLINE_URL_2)
    # school_year = get_curr_period(auth)['data']

    
    # Ù…Ø§ Ø¨Ø¹Ø±Ù ÙƒÙŠÙ Ø³ÙˆÙŠØªÙ‡Ø§ Ù„ÙƒÙ† Ø²Ø¨Ø·Øª 
    classes_id_2 = [lst for lst in get_teacher_classes_v2(auth ,inst_id, user_id, period_id)['data'] if lst]
    assessments = ['assessment1','assessment2','assessment3','assessment4']
    terms = ['term1','term2']
    upload_percentage,modified_classes,classes_id_3 ,classes,mawad,row_data =[],[],[],[],[],[]
    row_d={}

    for class_info in classes_id_2:
        classes_id_3.append([
                                {
                                    'institution_class_id': class_info['institution_class_id'] ,
                                    'sub_name': class_info['institution_subject']['name'],
                                    'class_name': class_info['institution_class']['name'] ,
                                    'subject_id': class_info['institution_subject']['education_subject_id'] ,
                                    'education_grade_id':class_info['institution_subject']['education_grade_id'],
                                    'institution_subject_id': class_info['institution_subject_id']
                                }
                            ])

    for v in range(len(classes_id_2)):
        # id
        # print (classes_id_3[v][0]['institution_class_id'])
        # id = classes_id_3[v][0]['institution_class_id']
        # subject name 
        # print (classes_id_3[v][0]['sub_name'])
        # class name
        # print (classes_id_3[v][0]['class_name'])
        # class_name = classes_id_3[v][0]['class_name']
        # subject id 
        # print (classes_id_3[v][0]['subject_id'])

        mawad.append(classes_id_3[v][0]['sub_name'])
        classes.append(classes_id_3[v][0]['class_name'])
        class_name = classes_id_3[v][0]['class_name'].split('-')[0].replace('Ø§Ù„ØµÙ ' , '')
        # class_char = classes_id_3[v][0]['class_name'].split('-')[1]
        # sub_name = classes_id_3[v][0]['sub_name']    
        
        # institution_subject_id,institution_class_id,institution_id,education_grade_id
        
        students = get_class_students(auth
                                    ,period_id
                                    ,classes_id_3[v][0]['institution_subject_id']
                                    ,classes_id_3[v][0]['institution_class_id']
                                    ,inst_id
                                    ,classes_id_3[v][0]['education_grade_id']
                                    ,session=session)
        # students_names = sorted([i['user']['name'] for i in students['data']])
        # print(students_names)
        students_id_and_names = []
        
        for IdAndName in students['data']:
            students_id_and_names.append({'student_name': IdAndName['user']['name'] , 'student_id':IdAndName['student_id']})

        assessments_json = make_request(auth=auth , url=f'https://emis.moe.gov.jo/openemis-core/restful/Assessment.AssessmentItemResults?academic_period_id={period_id}&education_subject_id='+str(classes_id_3[v][0]['subject_id'])+'&institution_classes_id='+ str(classes_id_3[v][0]['institution_class_id'])+ f'&institution_id={inst_id}&_limit=0&_fields=AssessmentGradingOptions.name,AssessmentGradingOptions.min,AssessmentGradingOptions.max,EducationSubjects.name,EducationSubjects.code,AssessmentPeriods.code,AssessmentPeriods.name,AssessmentPeriods.academic_term,marks,assessment_grading_option_id,student_id,assessment_id,education_subject_id,education_grade_id,assessment_period_id,institution_classes_id&_contain=AssessmentPeriods,AssessmentGradingOptions,EducationSubjects')

        marks_and_name = []

        dic = {'id':'' ,'name': '','term1':{ 'assessment1': '' ,'assessment2': '' , 'assessment3': '' , 'assessment4': ''} ,'term2':{ 'assessment1': '' ,'assessment2': '' , 'assessment3': '' , 'assessment4': ''}}
        for student_data_item in students_id_and_names:   
            for student_assessment_item in assessments_json['data']:
                if student_assessment_item['student_id'] == student_data_item['student_id'] :  
                    # FIXME: ØºÙŠØ± Ø§Ù„Ø´Ø±Ø· Ø§Ø°Ø§ ÙƒØ§Ù† None Ø§Ø³ØªØ¨Ø¯Ù„ Ø§Ù„Ù‚ÙŠÙ…Ø© Ø¨Ù„Ø§ Ø´ÙŠØ¡                    
                    if student_assessment_item["marks"] is not None :
                        dic['id'] = student_data_item['student_id'] 
                        dic['name'] = student_data_item['student_name'] 
                        if student_assessment_item['assessment_period']['name'] == 'Ø§Ù„ØªÙ‚ÙˆÙŠÙ… Ø§Ù„Ø£ÙˆÙ„' and student_assessment_item['assessment_period']['academic_term'] == 'Ø§Ù„ÙØµÙ„ Ø§Ù„Ø£ÙˆÙ„':
                            dic['term1']['assessment1'] = student_assessment_item["marks"] 
                        elif student_assessment_item['assessment_period']['name'] == 'Ø§Ù„ØªÙ‚ÙˆÙŠÙ… Ø§Ù„Ø«Ø§Ù†ÙŠ' and student_assessment_item['assessment_period']['academic_term'] == 'Ø§Ù„ÙØµÙ„ Ø§Ù„Ø£ÙˆÙ„':
                            dic['term1']['assessment2']  = student_assessment_item["marks"]
                        elif student_assessment_item['assessment_period']['name'] == 'Ø§Ù„ØªÙ‚ÙˆÙŠÙ… Ø§Ù„Ø«Ø§Ù„Ø«' and student_assessment_item['assessment_period']['academic_term'] == 'Ø§Ù„ÙØµÙ„ Ø§Ù„Ø£ÙˆÙ„':
                            dic['term1']['assessment3']  = student_assessment_item["marks"]
                        elif student_assessment_item['assessment_period']['name'] == 'Ø§Ù„ØªÙ‚ÙˆÙŠÙ… Ø§Ù„Ø±Ø§Ø¨Ø¹' and student_assessment_item['assessment_period']['academic_term'] == 'Ø§Ù„ÙØµÙ„ Ø§Ù„Ø£ÙˆÙ„':
                            dic['term1']['assessment4']  = student_assessment_item["marks"]
                        elif student_assessment_item['assessment_period']['name'] == 'Ø§Ù„ØªÙ‚ÙˆÙŠÙ… Ø§Ù„Ø£ÙˆÙ„' and student_assessment_item['assessment_period']['academic_term'] == 'Ø§Ù„ÙØµÙ„ Ø§Ù„Ø«Ø§Ù†ÙŠ':
                            dic['term2']['assessment1']  = student_assessment_item["marks"]
                        elif student_assessment_item['assessment_period']['name'] == 'Ø§Ù„ØªÙ‚ÙˆÙŠÙ… Ø§Ù„Ø«Ø§Ù†ÙŠ' and student_assessment_item['assessment_period']['academic_term'] == 'Ø§Ù„ÙØµÙ„ Ø§Ù„Ø«Ø§Ù†ÙŠ':
                            dic['term2']['assessment2']  = student_assessment_item["marks"]
                        elif student_assessment_item['assessment_period']['name'] == 'Ø§Ù„ØªÙ‚ÙˆÙŠÙ… Ø§Ù„Ø«Ø§Ù„Ø«' and student_assessment_item['assessment_period']['academic_term'] == 'Ø§Ù„ÙØµÙ„ Ø§Ù„Ø«Ø§Ù†ÙŠ':
                            dic['term2']['assessment3']  = student_assessment_item["marks"]
                        elif student_assessment_item['assessment_period']['name'] == 'Ø§Ù„ØªÙ‚ÙˆÙŠÙ… Ø§Ù„Ø±Ø§Ø¨Ø¹' and student_assessment_item['assessment_period']['academic_term'] == 'Ø§Ù„ÙØµÙ„ Ø§Ù„Ø«Ø§Ù†ÙŠ':
                            dic['term2']['assessment4']  = student_assessment_item["marks"]
            marks_and_name.append(dic)
            dic = {'id':'' ,'name': '','term1':{ 'assessment1': '' ,'assessment2': '' , 'assessment3': '' , 'assessment4': ''} ,'term2':{ 'assessment1': '' ,'assessment2': '' , 'assessment3': '' , 'assessment4': ''} }


        marks_and_name = [d for d in marks_and_name if d['name'] != '']
        marks_and_name = sorted(marks_and_name, key=lambda x: x['name'])
        
        percent_dict ={'subject': '' , 'className' :'', 'term1' : {'assessment1_percentage': '', 'assessment2_percentage': '', 'assessment3_percentage': '', 'assessment4_percentage': ''} ,
        'term2':{'assessment1_percentage': '', 'assessment2_percentage': '', 'assessment3_percentage': '', 'assessment4_percentage': ''}}
        row_d={}        
        
        if 'Ø¹Ø´Ø±' in class_name or 'Ø¹Ø§Ø´Ø±' in class_name :
            percent_dict ={'subject': '' , 'className' :'', 'term1' : {'assessment1_percentage': 0, 'assessment2_percentage': 0, 'assessment3_percentage': 0, 'assessment4_percentage': 0} ,
                            'term2':{'assessment1_percentage': 0, 'assessment2_percentage': 0, 'assessment3_percentage': 0, 'assessment4_percentage': 0}}
            
            percent_dict['subject']= classes_id_3[v][0]['sub_name']
            percent_dict['className']= classes_id_3[v][0]['class_name']
            row_d['subject'] = classes_id_3[v][0]['sub_name']
            row_d['className'] = classes_id_3[v][0]['class_name']
            # row_d['marks_and_name'] = marks_and_name
        else:
            for term in terms : 
                for assessment in assessments :
                    total_marks ,marks_uploaded= len([i[term][assessment] for i in marks_and_name ]) , len([i[term][assessment] for i in marks_and_name if i[term][assessment] != ''])
                    percentage = int((marks_uploaded / total_marks) * 100)
                    percent_dict[term][assessment+'_percentage']=percentage
                    
            percent_dict['subject']= classes_id_3[v][0]['sub_name']
            percent_dict['className']= classes_id_3[v][0]['class_name']
            row_d['subject'] = classes_id_3[v][0]['sub_name']
            row_d['className'] = classes_id_3[v][0]['class_name']
            row_d['marks_and_name'] = marks_and_name
        
        row_data.append(row_d)
        upload_percentage.append(percent_dict)
    
    return {'teacher': userInfo['name'] ,'upload_percentage' :upload_percentage , 'row_data':row_data}

def convert_official_marks_doc(ods_name='send', outdir='./send_folder' ,ods_num=1,file_path=None, file_content=None , color="#ffffff"):
    """
    The function `convert_official_marks_doc` converts an official marks document to a specified format
    and saves it in a specified directory.
    
    :param ods_name: The name of the ODS file to be converted. If not provided, it will default to
    'send', defaults to send (optional)
    :param outdir: The `outdir` parameter specifies the directory where the converted document will be
    saved. The default value is "./send_folder", which means the converted document will be saved in a
    folder named "send_folder" in the current directory, defaults to ./send_folder (optional)
    :param ods_num: The parameter `ods_num` is used to specify the number of the ODS (Open Document
    Spreadsheet) file. It is an optional parameter with a default value of 1, defaults to 1 (optional)
    :param file_path: The file path of the document that needs to be converted. This parameter is
    optional if the file content is provided
    :param file_content: The content of the file that needs to be converted. This can be provided as a
    string
    :param color: The color parameter is used to specify the background color of the converted document.
    The default value is "#ffffff", which represents white, defaults to #ffffff (optional)
    """
    ods_file = f'{ods_name}{ods_num}.ods'
    
    if file_content is None:
        doc = ezodf.opendoc(file_path)
    else:
        # Save the file content to a temporary file
        with tempfile.NamedTemporaryFile(delete=True) as temp_file:
            temp_file.write(file_content.getvalue())
            temp_file.flush()
            doc = ezodf.opendoc(temp_file.name)
            
            shutil.copy(temp_file.name, outdir+'/final_'+ods_file)
            
    os.system(f'soffice --headless --convert-to pdf:writer_pdf_Export --outdir {outdir}  {outdir}/final_{ods_file} ')
    add_margins(f"{outdir}/final_{ods_name}{ods_num}.pdf", f"{outdir}/output_file.pdf",top_rec=30, bottom_rec=50, left_rec=68, right_rec=120,color_name=color)
    add_margins(f"{outdir}/output_file.pdf", f"{outdir}/Ø³Ø¬Ù„ Ø§Ù„Ø¹Ù„Ø§Ù…Ø§Øª Ø§Ù„Ø±Ø³Ù…ÙŠ.pdf",page=1 , top_rec=60, bottom_rec=80, left_rec=70, right_rec=120,color_name=color)
    split_A3_pages(f"{outdir}/output_file.pdf" , outdir)
    reorder_official_marks_to_A4(f"{outdir}/output.pdf" , f"{outdir}/reordered.pdf")

    add_margins(f"{outdir}/reordered.pdf", f"{outdir}/output_file.pdf",top_rec=60, bottom_rec=50, left_rec=68, right_rec=20,color_name=color)
    add_margins(f"{outdir}/output_file.pdf", f"{outdir}/output_file1.pdf",page=1 , top_rec=100, bottom_rec=80, left_rec=90, right_rec=120,color_name=color)
    add_margins(f"{outdir}/output_file1.pdf", f"{outdir}/output_file2.pdf",page=50 , top_rec=100, bottom_rec=80, left_rec=70, right_rec=60,color_name=color)
    add_margins(f"{outdir}/output_file2.pdf", f"{outdir}/Ø³Ø¬Ù„ Ø§Ù„Ø¹Ù„Ø§Ù…Ø§Øª Ø§Ù„Ø±Ø³Ù…ÙŠ_A4.pdf",page=51 , top_rec=100, bottom_rec=80, left_rec=90, right_rec=120,color_name=color)
    delete_files_except([f"Ø³Ø¬Ù„ Ø§Ù„Ø¹Ù„Ø§Ù…Ø§Øª Ø§Ù„Ø±Ø³Ù…ÙŠ.pdf",f"Ø³Ø¬Ù„ Ø§Ù„Ø¹Ù„Ø§Ù…Ø§Øª Ø§Ù„Ø±Ø³Ù…ÙŠ_A4.pdf"], outdir)

def check_file_if_official_marks_file(file_path=None, file_content=None):
    """
    The function checks if a file is an official marks file.
    
    :param file_path: The file path is the location of the file on the computer's file system. It is a
    string that specifies the directory and file name of the file
    :param file_content: The content of the file that you want to check if it is an official marks file
    """
    if file_content is None:
        doc = ezodf.opendoc(file_path)
    else:
        # Save the file content to a temporary file
        with tempfile.NamedTemporaryFile(delete=True) as temp_file:
            temp_file.write(file_content.getvalue())
            temp_file.flush()
            doc = ezodf.opendoc(temp_file.name)

    exists = 'sheet' in [sheet.name for sheet in doc.sheets]
    return exists

def teachers_marks_upload_percentage_wrapper(auth ,term=1 ,inst_id=None , inst_nat=None , session=None , template='./templet_files/ÙƒØ´Ù Ù†Ø³Ø¨Ø© Ø§Ù„Ø§Ø¯Ø®Ø§Ù„ Ù…Ø¹Ø¯Ù„.xlsx' ,outdir='./send_folder/' ):
    """
    The `teachers_marks_upload_percentage_wrapper` function calculates the percentage of marks uploaded
    by teachers and saves the results in an Excel file.
    
    :param auth: The `auth` parameter is used for authentication purposes. It could be a token or any
    other form of authentication required to access the necessary data
    :param term: The term parameter specifies the term for which the marks upload percentage is being
    calculated. It can be either 1 or 2, representing term 1 or term 2 respectively. The default value
    is 1, defaults to 1 (optional)
    :param inst_id: The `inst_id` parameter is the ID of the institution for which you want to calculate
    the marks upload percentage. If it is not provided, the function will use the ID of the institution
    associated with the authenticated user
    :param inst_nat: The `inst_nat` parameter is used to specify the nationality of the institution. It
    is an optional parameter that can be used to filter the list of teachers based on their nationality
    :param session: The `session` parameter is used to maintain the user's session throughout the
    function. It is typically a session object that stores information about the user's authentication
    and session state
    :param template: The `template` parameter is the file path of the Excel template that will be used
    as a basis for the output file. It should be in the format `./templet_files/ÙƒØ´Ù Ù†Ø³Ø¨Ø© Ø§Ù„Ø§Ø¯Ø®Ø§Ù„
    Ù…Ø¹Ø¯Ù„.xlsx`, defaults to ./templet_files/ÙƒØ´Ù Ù†Ø³Ø¨Ø© Ø§Ù„Ø§Ø¯Ø®Ø§Ù„ Ù…Ø¹Ø¯Ù„.xlsx (optional)
    :param outdir: The `outdir` parameter is the directory where the output file will be saved. It
    specifies the folder path where the generated Excel file will be stored, defaults to ./send_folder/
    (optional)
    """
    
    term = 'term1' if term == 1 else 'term2'
    if inst_id is None and inst_nat is None : 
        inst_id = inst_name(auth ,session=session)['data'][0]['Institutions']['id']
        
    teachers_nats = [teacher['nat_id'] for teacher in get_school_teachers( auth , id=inst_id ,nat_school=inst_nat , session=session)['staff'] if 'Ù…Ø¹Ù„Ù…' in teacher['position']]
    all_teachers_data  = []
    for nat in teachers_nats :
        all_teachers_data.append(teachers_marks_upload_percentage(auth , nat ,session=session))

    # load the existing workbook
    existing_wb = load_workbook(template)
    # Select the worksheet
    existing_ws = existing_wb.active

    
    row_number=11

    for teacher in all_teachers_data : 
        row = teacher['row_data']
        flattened_list = []
        marks_and_name = [r['marks_and_name'] for r in row if "Ø¹Ø´Ø±" not in r['className']]
        
        for group in marks_and_name:
            flattened_list.extend(group)

        total_marks = len([i[term]['assessment1'] for i in flattened_list ]) 
        if total_marks > 0:

            existing_ws.cell(row=row_number, column=1).value = teacher['teacher']
            existing_ws.cell(row=row_number, column=3).value = total_marks
            
            first_marks_uploaded = len([i[term]['assessment1'] for i in flattened_list if i[term]['assessment1'] != '']) 
            existing_ws.cell(row=row_number, column=4).value = first_marks_uploaded
            
            second_marks_uploaded =  len([i[term]['assessment2'] for i in flattened_list if i[term]['assessment2'] != ''])
            existing_ws.cell(row=row_number, column=5).value = second_marks_uploaded
            
            third_marks_uploaded =  len([i[term]['assessment3'] for i in flattened_list if i[term]['assessment3'] != ''])
            existing_ws.cell(row=row_number, column=6).value = third_marks_uploaded
            
            fourth_marks_uploaded =  len([i[term]['assessment4'] for i in flattened_list if i[term]['assessment4'] != ''])
            existing_ws.cell(row=row_number, column=7).value = fourth_marks_uploaded
            
            try :
                percentage = int((first_marks_uploaded + second_marks_uploaded + third_marks_uploaded + fourth_marks_uploaded) / (total_marks * 4) * 100)
            except : 
                percentage = 0
                
            existing_ws.cell(row=row_number, column=2).value = percentage          
            row_number+=1
        else:
            row_number-=1
    existing_wb.save( outdir + f'output.xlsx')
    
    playsound()

def teachers_marks_upload_percentage(auth, emp_username, template='./templet_files/side_marks_note_2.docx' ,outdir='./send_folder/' ,first_page='side_mark_first_page.docx', template_dir='./templet_files/',term=1 ,session=None ):
    """
    The function `teachers_marks_upload_percentage` is used to upload teachers' marks percentage using a
    specified template and directory.
    
    :param auth: The authentication token or credentials required to access the system or API
    :param emp_username: The username of the employee who is uploading the marks
    :param template: The template parameter is the path to the template file that will be used for
    generating the marks upload document. It is set to './templet_files/side_marks_note_2.docx' by
    default, defaults to ./templet_files/side_marks_note_2.docx (optional)
    :param outdir: The `outdir` parameter specifies the directory where the generated files will be
    saved, defaults to ./send_folder/ (optional)
    :param first_page: The parameter `first_page` is the file name of the first page of the document
    that will be generated. It is a Word document file (.docx) and is located in the `template_dir`
    directory. This file will be used as the first page of the final document, defaults to
    side_mark_first_page.docx (optional)
    :param template_dir: The directory where the template files are located. By default, it is set to
    './templet_files/', defaults to ./templet_files/ (optional)
    :param term: The term parameter represents the current term or semester for which the marks are
    being uploaded. It is an integer value, defaults to 1 (optional)
    :param session: The session parameter is used to make the fetching data multiple times faster 
    """
    period_id = get_curr_period(auth,session=session)['data'][0]['id']
    user = user_info(auth , emp_username,session=session)
    userInfo = user['data'][0]
    user_id , user_name = userInfo['id'] , userInfo['first_name']+' '+ userInfo['last_name']+'-' + str(emp_username)
    # years = get_curr_period(auth)
    school_data = inst_name(auth,session=session)['data'][0]
    inst_id = school_data['Institutions']['id']
    # school_name = school_data['Institutions']['name']
    # grades = make_request(auth=auth , url=FILL_OFFICIAL_MARKS_DOC_WRAPPER_OFFLINE_URL_2)
    # school_year = get_curr_period(auth)['data']

    
    # Ù…Ø§ Ø¨Ø¹Ø±Ù ÙƒÙŠÙ Ø³ÙˆÙŠØªÙ‡Ø§ Ù„ÙƒÙ† Ø²Ø¨Ø·Øª 
    classes_id_1 = [[value for key , value in i['InstitutionSubjects'].items() if key == "id"][0] for i in get_teacher_classes1(auth,inst_id,user_id,period_id,session=session)['data']]
    classes_id_2 =[get_teacher_classes2( auth , classes_id_1[i],session=session)['data'] for i in range(len(classes_id_1))]
    assessments = ['assessment1','assessment2','assessment3','assessment4']
    terms = ['term1','term2']
    upload_percentage,modified_classes,classes_id_3 ,classes,mawad,row_data =[],[],[],[],[],[]
    row_d={}

    for class_info in classes_id_2:
        classes_id_3.append([{"institution_class_id": class_info[0]['institution_class_id'] ,"sub_name": class_info[0]['institution_subject']['name'],"class_name": class_info[0]['institution_class']['name'] , 'subject_id': class_info[0]['institution_subject']['education_subject_id']}])

    for v in range(len(classes_id_1)):
        # id
        # print (classes_id_3[v][0]['institution_class_id'])
        # id = classes_id_3[v][0]['institution_class_id']
        # subject name 
        # print (classes_id_3[v][0]['sub_name'])
        # class name
        # print (classes_id_3[v][0]['class_name'])
        # class_name = classes_id_3[v][0]['class_name']
        # subject id 
        # print (classes_id_3[v][0]['subject_id'])

        mawad.append(classes_id_3[v][0]['sub_name'])
        classes.append(classes_id_3[v][0]['class_name'])
        class_name = classes_id_3[v][0]['class_name'].split('-')[0].replace('Ø§Ù„ØµÙ ' , '')
        class_char = classes_id_3[v][0]['class_name'].split('-')[1]
        # sub_name = classes_id_3[v][0]['sub_name']    
        
        students = get_class_students(auth
                                    ,period_id
                                    ,classes_id_1[v]
                                    ,classes_id_3[v][0]['institution_class_id']
                                    ,inst_id
                                    ,session=session)
        students_names = sorted([i['user']['name'] for i in students['data']])
        # print(students_names)
        students_id_and_names = []
        
        for IdAndName in students['data']:
            students_id_and_names.append({'student_name': IdAndName['user']['name'] , 'student_id':IdAndName['student_id']})

        assessments_json = make_request(auth=auth , url=f'https://emis.moe.gov.jo/openemis-core/restful/Assessment.AssessmentItemResults?academic_period_id={period_id}&education_subject_id='+str(classes_id_3[v][0]['subject_id'])+'&institution_classes_id='+ str(classes_id_3[v][0]['institution_class_id'])+ f'&institution_id={inst_id}&_limit=0&_fields=AssessmentGradingOptions.name,AssessmentGradingOptions.min,AssessmentGradingOptions.max,EducationSubjects.name,EducationSubjects.code,AssessmentPeriods.code,AssessmentPeriods.name,AssessmentPeriods.academic_term,marks,assessment_grading_option_id,student_id,assessment_id,education_subject_id,education_grade_id,assessment_period_id,institution_classes_id&_contain=AssessmentPeriods,AssessmentGradingOptions,EducationSubjects',session=session)

        marks_and_name = []

        dic = {'id':'' ,'name': '','term1':{ 'assessment1': '' ,'assessment2': '' , 'assessment3': '' , 'assessment4': ''} ,'term2':{ 'assessment1': '' ,'assessment2': '' , 'assessment3': '' , 'assessment4': ''}}
        for student_data_item in students_id_and_names:   
            for student_assessment_item in assessments_json['data']:
                if student_assessment_item['student_id'] == student_data_item['student_id'] :  
                    # FIXME: ØºÙŠØ± Ø§Ù„Ø´Ø±Ø· Ø§Ø°Ø§ ÙƒØ§Ù† None Ø§Ø³ØªØ¨Ø¯Ù„ Ø§Ù„Ù‚ÙŠÙ…Ø© Ø¨Ù„Ø§ Ø´ÙŠØ¡                    
                    if student_assessment_item["marks"] is not None :
                        dic['id'] = student_data_item['student_id'] 
                        dic['name'] = student_data_item['student_name'] 
                        if student_assessment_item['assessment_period']['name'] == 'Ø§Ù„ØªÙ‚ÙˆÙŠÙ… Ø§Ù„Ø£ÙˆÙ„' and student_assessment_item['assessment_period']['academic_term'] == 'Ø§Ù„ÙØµÙ„ Ø§Ù„Ø£ÙˆÙ„':
                            dic['term1']['assessment1'] = student_assessment_item["marks"] 
                        elif student_assessment_item['assessment_period']['name'] == 'Ø§Ù„ØªÙ‚ÙˆÙŠÙ… Ø§Ù„Ø«Ø§Ù†ÙŠ' and student_assessment_item['assessment_period']['academic_term'] == 'Ø§Ù„ÙØµÙ„ Ø§Ù„Ø£ÙˆÙ„':
                            dic['term1']['assessment2']  = student_assessment_item["marks"]
                        elif student_assessment_item['assessment_period']['name'] == 'Ø§Ù„ØªÙ‚ÙˆÙŠÙ… Ø§Ù„Ø«Ø§Ù„Ø«' and student_assessment_item['assessment_period']['academic_term'] == 'Ø§Ù„ÙØµÙ„ Ø§Ù„Ø£ÙˆÙ„':
                            dic['term1']['assessment3']  = student_assessment_item["marks"]
                        elif student_assessment_item['assessment_period']['name'] == 'Ø§Ù„ØªÙ‚ÙˆÙŠÙ… Ø§Ù„Ø±Ø§Ø¨Ø¹' and student_assessment_item['assessment_period']['academic_term'] == 'Ø§Ù„ÙØµÙ„ Ø§Ù„Ø£ÙˆÙ„':
                            dic['term1']['assessment4']  = student_assessment_item["marks"]
                        elif student_assessment_item['assessment_period']['name'] == 'Ø§Ù„ØªÙ‚ÙˆÙŠÙ… Ø§Ù„Ø£ÙˆÙ„' and student_assessment_item['assessment_period']['academic_term'] == 'Ø§Ù„ÙØµÙ„ Ø§Ù„Ø«Ø§Ù†ÙŠ':
                            dic['term2']['assessment1']  = student_assessment_item["marks"]
                        elif student_assessment_item['assessment_period']['name'] == 'Ø§Ù„ØªÙ‚ÙˆÙŠÙ… Ø§Ù„Ø«Ø§Ù†ÙŠ' and student_assessment_item['assessment_period']['academic_term'] == 'Ø§Ù„ÙØµÙ„ Ø§Ù„Ø«Ø§Ù†ÙŠ':
                            dic['term2']['assessment2']  = student_assessment_item["marks"]
                        elif student_assessment_item['assessment_period']['name'] == 'Ø§Ù„ØªÙ‚ÙˆÙŠÙ… Ø§Ù„Ø«Ø§Ù„Ø«' and student_assessment_item['assessment_period']['academic_term'] == 'Ø§Ù„ÙØµÙ„ Ø§Ù„Ø«Ø§Ù†ÙŠ':
                            dic['term2']['assessment3']  = student_assessment_item["marks"]
                        elif student_assessment_item['assessment_period']['name'] == 'Ø§Ù„ØªÙ‚ÙˆÙŠÙ… Ø§Ù„Ø±Ø§Ø¨Ø¹' and student_assessment_item['assessment_period']['academic_term'] == 'Ø§Ù„ÙØµÙ„ Ø§Ù„Ø«Ø§Ù†ÙŠ':
                            dic['term2']['assessment4']  = student_assessment_item["marks"]
            marks_and_name.append(dic)
            dic = {'id':'' ,'name': '','term1':{ 'assessment1': '' ,'assessment2': '' , 'assessment3': '' , 'assessment4': ''} ,'term2':{ 'assessment1': '' ,'assessment2': '' , 'assessment3': '' , 'assessment4': ''} }


        marks_and_name = [d for d in marks_and_name if d['name'] != '']
        marks_and_name = sorted(marks_and_name, key=lambda x: x['name'])
        
        percent_dict ={'subject': '' , 'className' :'', 'term1' : {'assessment1_percentage': '', 'assessment2_percentage': '', 'assessment3_percentage': '', 'assessment4_percentage': ''} ,
        'term2':{'assessment1_percentage': '', 'assessment2_percentage': '', 'assessment3_percentage': '', 'assessment4_percentage': ''}}
        row_d={}        
        
        if 'Ø¹Ø´Ø±' in class_name :
            percent_dict ={'subject': '' , 'className' :'', 'term1' : {'assessment1_percentage': 0, 'assessment2_percentage': 0, 'assessment3_percentage': 0, 'assessment4_percentage': 0} ,
                            'term2':{'assessment1_percentage': 0, 'assessment2_percentage': 0, 'assessment3_percentage': 0, 'assessment4_percentage': 0}}
            
            percent_dict['subject']= classes_id_3[v][0]['sub_name']
            percent_dict['className']= classes_id_3[v][0]['class_name']
            row_d['subject'] = classes_id_3[v][0]['sub_name']
            row_d['className'] = classes_id_3[v][0]['class_name']
            # row_d['marks_and_name'] = marks_and_name
        else:
            for term in terms : 
                for assessment in assessments :
                    total_marks ,marks_uploaded= len([i[term][assessment] for i in marks_and_name ]) , len([i[term][assessment] for i in marks_and_name if i[term][assessment] != ''])
                    percentage = int((marks_uploaded / total_marks) * 100)
                    percent_dict[term][assessment+'_percentage']=percentage
                    
            percent_dict['subject']= classes_id_3[v][0]['sub_name']
            percent_dict['className']= classes_id_3[v][0]['class_name']
            row_d['subject'] = classes_id_3[v][0]['sub_name']
            row_d['className'] = classes_id_3[v][0]['class_name']
            row_d['marks_and_name'] = marks_and_name
        
        row_data.append(row_d)
        upload_percentage.append(percent_dict)
    
    return {'teacher': userInfo['name'] ,'upload_percentage' :upload_percentage , 'row_data':row_data}

def side_marks_document_with_marks(username=None , password=None ,classes_data=None,template='./templet_files/side_marks_note_2.docx' ,outdir='./send_folder/' ,first_page='side_mark_first_page.docx', template_dir='./templet_files/',term=1 , names_only=False, session=None):
    """Ø¯Ø§Ù„Ø© ØªÙ‚ÙˆÙ… Ø¨Ø§Ù†Ø´Ø§Ø¡ Ø³Ø¬Ù„ Ø¹Ù„Ø§Ù…Ø§Øª Ø¬Ø§Ù†Ø¨ÙŠ ÙˆØªØ¹Ø¨Ø¦Ø© Ø§Ù„Ø¹Ù„Ø§Ù…Ø§Øª 

    Args:
        username (_type_, optional): Ø§Ø³Ù… Ø§Ù„Ù…Ø³ØªØ®Ø¯Ù…. Defaults to None.
        password (_type_, optional): ÙƒÙ„Ù… Ø§Ù„Ø³Ø±. Defaults to None.
        classes_data (_type_, optional): Ø¨ÙŠØ§Ù†Ø§Øª Ø§Ù„ØµÙÙˆÙ Ø§Ø°Ø§ Ø§Ø³ØªØ¹Ù…Ù„Øª Ø§Ù„Ø¯Ø§Ù„Ø© Ø§ÙˆÙÙ„Ø§ÙŠÙ†. Defaults to None.
        template (str, optional): Ø§Ù„Ù†Ù…ÙˆØ°Ø¬ Ø§Ù„Ø°ÙŠ ÙŠØªÙ… Ø§Ø³ØªØ®Ø¯Ø§Ù…Ù‡. Defaults to './templet_files/side_marks_note_2.docx'.
        outdir (str, optional): Ø§Ù„Ù…Ø¬Ù„Ø¯ Ø§Ù„Ø°ÙŠ ÙŠØªÙ… Ø§Ù„Ø­ÙØ¸ ÙÙŠÙ‡ . Defaults to './send_folder/'.
        first_page (str, optional): Ø§Ù„ØµÙØ­Ø© Ø§Ù„Ø§ÙˆÙ„Ù‰ Ù„Ù„Ø³Ø¬Ù„ Ø§Ù„Ø¬Ø§Ù†Ø¨ÙŠ . Defaults to 'side_mark_first_page.docx'.
        template_dir (str, optional): Ù…Ø¬Ù„Ø¯ Ø§Ù„Ù†Ù…Ø§Ø°Ø¬. Defaults to './templet_files/'.
        term (int, optional): Ø§Ù„ÙØµÙ„ Ø§Ù…Ø§ Ø§Ù„Ø§ÙˆÙ„ Ø§Ùˆ Ø§Ù„Ø«Ø§Ù†ÙŠ. Defaults to 1.
        names_only (bool, optional): Ø®ÙŠØ§Ø± Ù„Ø·Ø¨Ø§Ø¹Ø© Ø§Ù„Ø§Ø³Ù…Ø§Ø¡ ÙÙ‚Ø·. Defaults to False.
    """    
    
    classes=[]
    mawad=[]
    modified_classes=[]
    context={}
    
    if username is not None and password is not None : 
        auth = get_auth(username , password)
        period_id = get_curr_period(auth)['data'][0]['id']
        user = user_info(auth , username)
        userInfo = user['data'][0]
        user_id , user_name = userInfo['id'] , userInfo['first_name']+' '+ userInfo['last_name']+'-' + str(username)
        # years = get_curr_period(auth)
        school_data = inst_name(auth)['data'][0]
        inst_id = school_data['Institutions']['id']
        school_name = school_data['Institutions']['name']
        school_name_id = f'{school_name}={inst_id}'
        baldah = make_request(auth=auth , url=FILL_OFFICIAL_MARKS_DOC_WRAPPER_OFFLINE_URL_1.format(inst_id=inst_id))['data'][0]['address'].split('-')[0]
        # grades = make_request(auth=auth , url=FILL_OFFICIAL_MARKS_DOC_WRAPPER_OFFLINE_URL_2)
        modeeriah = inst_area(auth)['data'][0]['Areas']['name']
        school_year = get_curr_period(auth)['data']
        melady1 = str(school_year[0]['start_year'])
        melady2 = str(school_year[0]['end_year'])
        teacher = user['data'][0]['name'].split(' ')[0]+' '+user['data'][0]['name'].split(' ')[-1]
        
        # Ù…Ø§ Ø¨Ø¹Ø±Ù ÙƒÙŠÙ Ø³ÙˆÙŠØªÙ‡Ø§ Ù„ÙƒÙ† Ø²Ø¨Ø·Øª 
        
        assessments = make_request(auth =auth,url=f'https://emis.moe.gov.jo/openemis-core/restful/v2/Assessment-AssessmentPeriods.json?_limit=0' , session=session)
        classes_id_2 =[lst for lst in get_teacher_classes_v2(auth ,inst_id, user_id, period_id)['data'] if lst]
        classes= [i['institution_class']['name'] for i in classes_id_2]
        mawad = [i['institution_subject']['name'] for i in classes_id_2]
        grades_info = get_grade_info(auth)
        assessments_period_data = []
        
        necessary_data_dict = {
                                'userInfo' : userInfo , 
                                'school_name' : school_name , 
                                'modeeriah' : modeeriah , 
                                'melady1' : melady1 , 
                                'melady2' : melady2 , 
                                }

        assessments_period_data = get_marks(auth, inst_id , period_id , classes_id_2 , grades_info, assessments=assessments ,insert_function=insert_to_side_marks_document_with_marks ,necessary_data_dict=necessary_data_dict , template_sheet_or_file = template)
    else:
        student_details = classes_data
        school_name = student_details['custom_shapes']['school']
        # modified_classes =student_details['custom_shapes']['classes']
        teacher = student_details['custom_shapes']['teacher'] 
        melady2 = student_details['custom_shapes']['melady1']
        melady1 = student_details['custom_shapes']['melady2']
        modeeriah = student_details['custom_shapes']['modeeriah']
        
        # Iterate over student data
        for v ,students_data_list in enumerate(student_details['file_data']):
            class_name = students_data_list['class_name'].split('=')[0].replace('Ø§Ù„ØµÙ', '')
            mawad.append(students_data_list['class_name'].split('=')[1])
            classes.append(class_name)

            if 'Ø¹Ø´Ø±' in class_name :
                counter = 0
                for item in students_data_list['students_data'] :
                    context[f'name{counter}'] = item['name']
                    if not names_only :
                        assessments = [
                                    item[f'term{term}']['assessment1'],
                                    item[f'term{term}']['assessment2'],
                                    item[f'term{term}']['assessment3'],
                                    item[f'term{term}']['assessment4']
                                    ]
                        context[f'A1_{counter}'] = item[f'term{term}']['assessment1']
                        context[f'A2_{counter}'] = item[f'term{term}']['assessment2']
                        context[f'A3_{counter}'] = item[f'term{term}']['assessment3']
                        context[f'A4_{counter}'] = item[f'term{term}']['assessment4']
                        SUM = sum(int(assessment) if assessment != '' else 0 for assessment in assessments)                    
                        context[f'S_{counter}'] = SUM if SUM !=0 else ''
                        total = item[f'term{term}']['assessment3']

                        try :                    
                            variables = [random.randint(3, min(total, 5)) for _ in range(3) if total > 0]
                            variables.append(total - sum(variables))       
                            context[f'M1_{counter}'] ,context[f'M2_{counter}'] ,context[f'M3_{counter}'] ,context[f'M4_{counter}'] = variables
                        except : 
                            context[f'M1_{counter}'] ,context[f'M2_{counter}'] ,context[f'M3_{counter}'] ,context[f'M4_{counter}'] =['']*4                        
                    counter+=1 
                    context['teacher'] = teacher                    
            else:
                counter = 0
                for item in students_data_list['students_data'] :
                    context[f'name{counter}'] = item['name']
                    if not names_only :
                        assessments = [
                                    item[f'term{term}']['assessment1'],
                                    item[f'term{term}']['assessment2'],
                                    item[f'term{term}']['assessment3'],
                                    item[f'term{term}']['assessment4']
                                    ]
                        context[f'A1_{counter}'] = item[f'term{term}']['assessment1']
                        context[f'A2_{counter}'] = item[f'term{term}']['assessment2']
                        context[f'A3_{counter}'] = item[f'term{term}']['assessment3']
                        context[f'A4_{counter}'] = item[f'term{term}']['assessment4']
                        SUM = sum(int(assessment) if assessment != '' else 0 for assessment in assessments)                    
                        context[f'S_{counter}'] = SUM if SUM !=0 else ''
                        total = item[f'term{term}']['assessment3']

                        try :
                            variables = [random.randint(3, min(total, 5)) for _ in range(3) if total > 0]
                            variables.append(total - sum(variables))       
                            context[f'M1_{counter}'] ,context[f'M2_{counter}'] ,context[f'M3_{counter}'] ,context[f'M4_{counter}'] = variables
                        except : 
                            context[f'M1_{counter}'] ,context[f'M2_{counter}'] ,context[f'M3_{counter}'] ,context[f'M4_{counter}'] =['']*4                        
                        
                    counter+=1 
                    context['teacher'] = teacher
            context[f'class_name'] = class_name
            context[f'term'] = 'Ø§Ù„Ø£ÙˆÙ„' if term == 1 else 'Ø§Ù„Ø«Ø§Ù†ÙŠ'
            context['school'] = school_name
            context['directory'] = modeeriah
            context['y1'] = melady1
            context['y2'] = melady2
            context['sub'] = students_data_list['class_name'].split('=')[1]
            fill_doc(template , context , outdir+f'send{v}.docx' )
            context.clear()
            generate_pdf(outdir+f'send{v}.docx' , outdir ,v)
            delete_pdf_page(outdir+f'send{v}.pdf', outdir+f'SEND{v}.pdf', 1)
            delete_file(outdir+f'send{v}.pdf')    

    for i in classes: 
        modified_classes.append(get_class_short(i))
    
    # modified_classes = modified_classes if modified_classes else ' ØŒ '.join(modified_classes)
    # modified_classes = sorted(set(modified_classes))
    mawad = [madah.replace('Ø£', 'Ø§').replace('Ø¥', 'Ø§') for madah in mawad]
    mawad = sorted(set(mawad))
    mawad = ' ØŒ '.join(mawad)
    context = {'school':school_name 
            ,'classes' : ' , '.join(modified_classes) 
            ,'subjects' : mawad
            ,'teacher' : teacher if teacher else userInfo['first_name']+' '+ userInfo['middle_name'] +' '+ userInfo['last_name'] 
            ,'y1' : melady2 
            ,'y2' : melady1}
    fill_doc(template_dir+first_page , context , outdir+first_page )
    generate_pdf(outdir+first_page , outdir ,first_page)
    
    input_files = get_pdf_files(outdir)
    # Put ready_side_mark_first_page first on the list.
    input_files.sort(reverse=False)
    input_files.insert(0, input_files.pop(input_files.index(outdir+first_page.replace('docx','pdf'))))
    output_path = "Ø§Ù„Ø³Ø¬Ù„ Ø§Ù„Ø¬Ø§Ù†Ø¨ÙŠ.pdf"
    merge_pdfs(input_files, outdir+output_path)
    [delete_file(i) for i in input_files]

def merge_pdfs(input_files, output_file):
    """Merges a list of PDF files into a single PDF file.

    Args:
    input_files: A list of PDF files to merge.
    output_file: The name of the output PDF file.

    Returns:
    None.
    """

    merger = PdfFileMerger()

    for file in input_files:
        merger.append(file)

    merger.write(output_file)

def get_pdf_files(directory):
    pdf_files = []

    for filename in os.listdir(directory):
        if filename.endswith('.pdf'):
            pdf_files.append(os.path.join(directory, filename))

    return pdf_files

def delete_pdf_page(input_path, output_path, page_number):
    """
    The function deletes a specific page from a PDF file and saves the modified file to a new location.
    
    :param input_path: The path to the input PDF file that you want to delete a page from
    :param output_path: The path where the modified PDF file will be saved
    :param page_number: The page number of the PDF page you want to delete
    """
    with open(input_path, 'rb') as file:
        pdf_reader = PyPDF4.PdfFileReader(file)
        total_pages = len(pdf_reader.pages)

        if page_number < 0 or page_number >= total_pages:
            print("Invalid page number.")
            return

        pdf_writer = PyPDF4.PdfFileWriter()

        for page in range(total_pages):
            if page != page_number:
                pdf_writer.addPage(pdf_reader.pages[page])

        with open(output_path, 'wb') as output_file:
            pdf_writer.write(output_file)

        print("Page deletion completed.")

def create_zip(file_paths, zip_name='Ù…Ù„Ù Ù…Ø¶ØºÙˆØ·' , zip_path='./send_folder/', extension='.rar'):
    """
    The function creates a compressed zip file from a list of file paths, with a specified name, path,
    and extension.
    
    :param file_paths: A list of file paths that you want to include in the zip file
    :param zip_name: The name of the zip file that will be created. The default value is 'Ù…Ù„Ù Ù…Ø¶ØºÙˆØ·',
    defaults to Ù…Ù„Ù Ù…Ø¶ØºÙˆØ· (optional)
    :param zip_path: The `zip_path` parameter specifies the path where the zip file will be created. By
    default, it is set to `./send_folder/`, which means the zip file will be created in a folder named
    "send_folder" in the current directory, defaults to ./send_folder/ (optional)
    :param extension: The extension parameter is used to specify the file extension for the compressed
    file. In this case, the default extension is set to '.rar', which means that the compressed file
    will be in RAR format, defaults to .rar (optional)
    """
    zip_file_path = os.path.join(zip_path, zip_name + extension)
    
    with zipfile.ZipFile(zip_file_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
        for file_path in file_paths:
            zipf.write(file_path , arcname=os.path.basename(file_path))

def Read_E_Side_Note_Marks_ods(file_path=None, file_content=None):
    """
    The function reads the E side note marks from an ods file.
    
    :param file_path: The file path is the location of the .ods file that you want to read. It should be
    a string that specifies the full path to the file, including the file name and extension
    :param file_content: The content of the file to be read. This can be a string containing the file
    content or a file object
    """
    if file_content is None:
        doc = ezodf.opendoc(file_path)
    else:
        # Save the file content to a temporary file
        with tempfile.NamedTemporaryFile(delete=True) as temp_file:
            temp_file.write(file_content.getvalue())
            temp_file.flush()
            doc = ezodf.opendoc(temp_file.name)

    sheets = [sheet for sheet in doc.sheets][:-1]
    info_sheet = [sheet for sheet in doc.sheets][-1]
    read_file_output_lists = []

    for sheet in sheets:
        data = []

        for i, row in enumerate(sheet.rows()):
            if i < 2:
                continue  # Skip the first two rows
            row_data = [cell.value if cell.value is not None else '' for cell in row]
            if row_data[1] != '':
                dic = {
                    'id': int(row_data[1]),
                    'name': row_data[2],
                    'term1': {'assessment1': int(row_data[3]) if not isinstance(row_data[3],str) else '', 'assessment2': int(row_data[4]) if not isinstance(row_data[4],str) else '', 'assessment3': int(row_data[5]) if not isinstance(row_data[5],str) else '', 'assessment4': int(row_data[6]) if not isinstance(row_data[6],str) else ''},
                    'term2': {'assessment1': int(row_data[8]) if not isinstance(row_data[8],str) else '', 'assessment2': int(row_data[9]) if not isinstance(row_data[9],str) else '', 'assessment3': int(row_data[10]) if not isinstance(row_data[10],str) else '', 'assessment4': int(row_data[11]) if not isinstance(row_data[11],str) else ''}
                }
                data.append(dic)
        idsClass=sheet['AM1'].value
        temp_dic = {'class_name':f"{sheet.name}={idsClass}" ,"students_data": data}
        # temp_dic = {'class_name': sheet.name, "students_data": data}
        read_file_output_lists.append(temp_dic)

    modified_classes = []

    classes = [i['class_name'].split('=')[0] for i in read_file_output_lists]
    mawad = [i['class_name'].split('=')[1] for i in read_file_output_lists]
    for i in classes:
        modified_classes.append(get_class_short(i))

    school_id=info_sheet['A1'].value    
    school_name = info_sheet['A2'].value.split('=')[0]
    modeeriah = info_sheet['A3'].value
    hejri1 = info_sheet['A4'].value
    hejri2 = info_sheet['A5'].value
    melady1 = info_sheet['A6'].value
    melady2 = info_sheet['A7'].value
    baldah = info_sheet['A8'].value
    modified_classes = ' ØŒ '.join(modified_classes)
    mawad = sorted(set(mawad))
    mawad = ' ØŒ '.join(mawad)
    teacher = info_sheet['A9'].value
    required_data_mrks_text = info_sheet['A10'].value
    period_id = info_sheet['A11'].value

    custom_shapes = {
        'modeeriah': f'{modeeriah}',
        'hejri1': hejri1,
        'hejri2': hejri2,
        'melady1': melady1,
        'melady2': melady2,
        'baldah': baldah,
        'school': school_name,
        'classes': modified_classes,
        'mawad': mawad,
        'teacher': teacher,
        'modeeriah_20_2': f'{modeeriah}',
        'hejri_20_1': hejri1,
        'hejri_20_2': hejri2,
        'melady_20_1': melady1,
        'melady_20_2': melady2,
        'hejri_20_5': hejri1,
        'hejri_20_6': hejri2,
        'melady_20_7': melady1,
        'melady_20_8': melady2,        
        'baldah_20_2': baldah,
        'school_20_2': school_name,
        'classes_20_2': modified_classes,
        'mawad_20_2': mawad,
        'teacher_20_2': teacher,
        'modeeriah_20_1': f'{modeeriah}',
        'hejri1': hejri1,
        'hejri2': hejri2,
        'melady1': melady1,
        'melady2': melady2,
        'baldah_20_1': baldah,
        'school_20_1': school_name,
        'classes_20_1': modified_classes,
        'mawad_20_1': mawad,
        'teacher_20_1': teacher,
        'period_id': period_id,
        'school_id': school_id
    }

    try:
        required_data_mrks_dic_list = {
                                        int(item.split('-')[0]): 
                                            {
                                                'assessment_grade_id': int(item.split('-')[1].split(',')[0]),
                                                'grade_id': int(item.split(',')[0].split('-')[2]), 
                                                'assessments_period_ids': item.split(',')[1:]
                                            }
                                        for item in required_data_mrks_text.split('\\\\')
                                    }
    except Exception as e:
        required_data_mrks_dic_list = {
                                        0:
                                            {
                                                'assessment_grade_id': 0,
                                                'grade_id': 0, 
                                                'assessments_period_ids': 0
                                            }
                                        }

    read_file_output_dict = {'file_data': read_file_output_lists,
                            'custom_shapes': custom_shapes,
                            'required_data_for_mrks_enter': required_data_mrks_dic_list}

    return read_file_output_dict

def upload_marks(username , password , classess_data ):
    """
    The function is used to upload marks for different classes using a username and password.
    
    :param username: The username is a string that represents the username of the user who is uploading
    the marks. It is used for authentication purposes
    :param password: The password parameter is a string that represents the password for the user's
    account
    :param classess_data: The `classess_data` parameter is a data structure that contains information
    about the marks of students in different classes. It could be a list of dictionaries, where each
    dictionary represents a class and contains the following keys:
    """
    auth = get_auth(username , password)
    period_id = classess_data['custom_shapes']['period_id']
    school_id = classess_data['custom_shapes']['school_id']
    # term1_assessment_codes = ['S1A1', 'S1A2', 'S1A3', 'S1A4']
    # term2_assessment_codes = ['S2A1', 'S2A2', 'S2A3', 'S2A4']
    assessment_codes = ['S1A1', 'S1A2', 'S1A3', 'S1A4' , 'S2A1', 'S2A2', 'S2A3', 'S2A4']
    assessment_code_dic = {'S1A1': {'term' :'term1' , 'assess' : 'assessment1'},
                            'S1A2': {'term' :'term1' , 'assess' : 'assessment2'},
                            'S1A3': {'term' :'term1' , 'assess' : 'assessment3'},
                            'S1A4': {'term' :'term1' , 'assess' : 'assessment4'},
                            'S2A1': {'term' :'term2' , 'assess' : 'assessment1'},
                            'S2A2': {'term' :'term2' , 'assess' : 'assessment2'},
                            'S2A3': {'term' :'term2' , 'assess' : 'assessment3'},
                            'S2A4': {'term' :'term2' , 'assess' : 'assessment4'}}
    
    assessments_periods_data = classess_data['required_data_for_mrks_enter']
    for class_data in classess_data['file_data']:
        class_id = class_data['class_name'].split('=')[2] 
        class_subject = class_data['class_name'].split('=')[3]
        class_name = classess_data['file_data'][1]['class_name'].split('=')[0]
        if 'Ø¹Ø´Ø±' not in class_name : 
            students_marks_ids = class_data['students_data']
            assessment_grade_id = assessments_periods_data[int(class_id)]['assessment_grade_id']
            grade_id = assessments_periods_data[int(class_id)]['grade_id']
            assessment_periods = get_editable_assessments(auth,username,assessment_grade_id,class_subject)
            # assessment_ids = assessments_periods_data[int(class_id)]['assessments_period_ids']
            # s1a1, s1a2, s1a3, s1a4, s2a1, s2a2, s2a3, s2a4 = [assessment_ids[i] if i < len(assessment_ids) else None for i in range(8)]
            for student_info in students_marks_ids:
                for code in assessment_codes:
                    if len([i for i in assessment_periods if code in i['code']]) != 0:
                        assessment_period_id = [i for i in assessment_periods if code in i['code']][0]['AssesId']
                        term = assessment_code_dic[code]['term']
                        assess = assessment_code_dic[code]['assess']
                        term_marks = student_info[term]
                        mark = term_marks.get(assess)
                        if mark != '':
                            enter_mark(
                                    auth,
                                    marks=str("{:.2f}".format(float(mark))),
                                    assessment_grading_option_id=8,
                                    assessment_id=assessment_grade_id,
                                    education_subject_id=class_subject,
                                    education_grade_id=grade_id,
                                    institution_id=school_id,
                                    academic_period_id=period_id,
                                    institution_classes_id=class_id,
                                    student_status_id=1,
                                    student_id=student_info['id'],
                                    assessment_period_id=assessment_period_id
                                    )                   

def scrape_schools(username, password , limit = 10, pages = 10*6 ,sector=11):
    """
    The function scrape_schools takes in a username and password, along with optional parameters for
    limit, pages, and sector, and returns a list of scraped schools.
    
    :param username: The username is the login username for the website or platform you are scraping
    data from. It is used to authenticate and access the data
    :param password: The password parameter is the password for the user's account. It is used to
    authenticate the user and grant access to the scraping functionality
    :param limit: The limit parameter determines the maximum number of schools to scrape data for. It
    specifies the maximum number of schools to retrieve data for, defaults to 10 (optional)
    :param pages: The "pages" parameter determines the number of pages to scrape. Each page typically
    contains multiple schools. The default value is set to 10*6, which means it will scrape 10 pages,
    with each page containing 6 schools
    :param sector: The "sector" parameter refers to the sector code of the schools you want to scrape.
    It is used to filter the schools based on their sector Ø­ÙƒÙˆÙ…ÙŠØ© Ø§Ùˆ Ø®Ø§ØµØ©, defaults to 11 (optional)
    """
    dic_list = []
    for page in range(1,pages):
        auth = get_auth(username,password)
        institutions = make_request(auth=auth , url=f'https://emis.moe.gov.jo/openemis-core/restful/Institution-Institutions.json?_limit={limit}&institution_sector_id={sector}&_page={page}&_fields=name,code,address,institution_sector_id,area_id,area_administrative_id,longitude,latitude')
        if len(institutions['data']) == 0:
            break
        else:
            dic_list.append(institutions['data'])
    return dic_list

def Vacancies (username , password , schools_nats):
    """
    The function Vacancies takes in a username, password, and a list of school names and returns
    what could be open vacancies (Ø´ÙˆØ§ØºØ±) in the school.
    
    :param username: The username parameter is a string that represents the username of the user trying
    to access the vacancies
    :param password: The password parameter is a string that represents the password for the user
    :param schools_nats: The parameter "schools_nats" is likely a list or dictionary that contains
    information about schools and their nationalities. It could be used to store data such as the name
    of the school and the nationality of the students attending that school
    """
    dic_list=[]
    faulty_inst_nat = []
    school_name_code = []
    error = []
    for school_nat in schools_nats:
        try:    
            auth = get_auth(username,password)
            school_name_staff = get_school_teachers(auth,nat_school=school_nat)
            teachers = school_name_staff['staff']
            school_name = school_name_staff['school_code_name']
            school_id = school_name_staff['school_id']
            school_load = get_school_load(auth, school_id)
            teachers_load = get_school_teachers_load(auth , school_id)
            print(school_name)


            working_teachers = [teacher['name'] for teacher in teachers if teacher['staff_status'] == 1]
            sub_teachers = [teacher['name'] for teacher in teachers if teacher['staff_type'] == 197605]
            english_teachers = [name for name in [ i for i in get_teacher_load_with_name(teachers_load , 1)] if name[0] in working_teachers]
            arabic_teachers = [name for name in [ i for i in get_teacher_load_with_name(teachers_load , 2)] if name[0] in working_teachers]
            math_teachers = [name for name in [ i for i in get_teacher_load_with_name(teachers_load , 3)] if name[0] in working_teachers]
            english_teachers_final = [[name[0]+'**',name[1],name[2]] if name[0] in sub_teachers else name for name in english_teachers]
            arabic_teachers_final = [[name[0]+'**',name[1],name[2]] if name[0] in sub_teachers else name for name in arabic_teachers]
            math_teachers_final = [[name[0]+'**',name[1],name[2]] if name[0] in sub_teachers else name for name in math_teachers]

            string = str(school_load['english_school_sum'])+' <== Ù†ØµØ§Ø¨ Ø§Ù„Ø§Ù†Ø¬Ù„ÙŠØ²ÙŠ \n'+str(school_load['arabic_school_sum'])+' <== Ù†ØµØ§Ø¨ Ø§Ù„Ø¹Ø±Ø¨ÙŠ \n'+str(school_load['math_school_sum'])+' <== Ù†ØµØ§Ø¨ Ø§Ù„Ø±ÙŠØ§Ø¶ÙŠØ§Øª \n'
            classes = ' ,\n'.join(str(i).replace('Ø§Ù„ØµÙ', '') for i in school_load['classes'])

            long_string = '--------------Ù…Ø¹Ù„Ù…ÙŠÙ† Ø§Ù„Ø§Ù†Ø¬Ù„ÙŠØ²ÙŠ--------------\n'
            for item in english_teachers_final:
                long_string += item[0]+' =======>> '+ str(item[1]) + ' =======>> ' +  ' , '.join(str(i).replace('Ø§Ù„ØµÙ', '') for i in item[2])+'\n'
            long_string += '--------------Ù…Ø¹Ù„Ù…ÙŠÙ† Ø§Ù„Ø¹Ø±Ø¨ÙŠ--------------\n'
            for item in arabic_teachers_final:
                long_string += item[0]+' =======>> '+ str(item[1]) + ' =======>> ' +  ' , '.join(str(i).replace('Ø§Ù„ØµÙ', '') for i in item[2])+'\n'
            long_string += '--------------Ù…Ø¹Ù„Ù…ÙŠÙ† Ø§Ù„Ø±ÙŠØ§Ø¶ÙŠØ§Øª--------------\n'
            for item in math_teachers_final:
                long_string += item[0]+' =======>> '+ str(item[1]) + ' =======>> ' +  ' , '.join(str(i).replace('Ø§Ù„ØµÙ', '') for i in item[2])+'\n'

            dic = { 'school_name' :school_name , 'school_load' : string , 'teachers' : long_string , 'classes': classes }
            dic_list.append(dic)
        
        except Exception as e : 
            faulty_inst_nat.append(school_nat)
            error.append(e)
            print ('----------------------->' ,school_name)
            try:
                school_name_code.append(school_name_staff['school_code_name'])
            except:
                pass
    return (dic_list ,faulty_inst_nat)

def get_school_load(auth , inst_id ,academic_period_id=13):
    """
    The function "get_school_load" retrieves the school load for a specific institution and academic
    period.
    
    :param auth: The auth parameter is used for authentication purposes. It could be a token or any
    other form of authentication that allows the user to access the necessary resources
    :param inst_id: The inst_id parameter is the ID of the institution or school for which you want to
    retrieve the school load
    :param academic_period_id: The academic period ID is an optional parameter that represents the
    specific academic period for which you want to retrieve the school load. If not provided, it
    defaults to 13, defaults to 13 (optional)
    """
    student_classess = make_request(auth=auth, url=f'https://emis.moe.gov.jo/openemis-core/restful/v2/Institution-InstitutionClassStudents.json?institution_id={inst_id}&academic_period_id={academic_period_id}&_limit=0&_contain=Users.Genders')['data']
    institution_class_ids = list(set([i['institution_class_id'] for i in student_classess]))
    joined_string = ','.join(str(i) for i in [f'institution_class_id:{i}' for i in institution_class_ids])
    classes_data = make_request(auth=auth,url='https://emis.moe.gov.jo/openemis-core/restful/Institution.InstitutionClassSubjects?status=1&_contain=InstitutionSubjects,InstitutionClasses&_limit=0&_orWhere='+joined_string)['data']
    class_list = []
    for i in classes_data:
        class_list.append({'class_id': i['institution_class_id'] , 'class_name': i['institution_class']['name'] })
        class_dict = {i['class_id']: i['class_name'] for i in class_list if i['class_id'] != ''}
        classes = [key for value,key in class_dict.items()]

    arabic_class_sum = 0
    english_class_sum = 0
    math_class_sum = 0

    for school_class in classes:
        if 'Ø§ÙˆÙ„' in school_class or 'Ø«Ø§Ù†ÙŠ' in school_class or 'Ø«Ø§Ù„Ø«'in school_class or  'Ø±Ø§Ø¨Ø¹' in school_class or 'Ø¹Ø´Ø±' in school_class:
            if 'Ø¯Ø¨ÙŠ' in school_class:
                math_class_sum+=3
            else:
                math_class_sum+=5
            english_class_sum += 4
        else: 
            english_class_sum+= 5
            math_class_sum += 5

    for school_class in classes:
        if 'Ø³Ø§Ø¨Ø¹' in school_class or 'Ø«Ø§Ù…Ù†' in school_class or 'ØªØ§Ø³Ø¹' in school_class or  'Ø¹Ø§Ø´Ø±' in school_class :
            arabic_class_sum += 6
        else:
            if 'Ø¹Ø´Ø±' in school_class:
                if 'Ø¯Ø¨ÙŠ' in school_class:
                    if  'Ø­Ø§Ø¯ÙŠ' in school_class:
                        arabic_class_sum+=5
                    elif 'Ø«Ø§Ù†ÙŠ' in school_class:
                        arabic_class_sum+=4
                else:
                    arabic_class_sum+=4
            else:
                arabic_class_sum+=7
    return {'english_school_sum' : english_class_sum , 'arabic_school_sum' : arabic_class_sum , 'math_school_sum' :  math_class_sum , 'classes' :  classes}

def get_school_teachers(auth ,id=None , nat_school=None ,session=None ,row=False):
    """
    The function "get_school_teachers" retrieves information about teachers from a school.
    
    :param auth: The auth parameter is used for authentication purposes. It could be a token or a
    username/password combination that allows the user to access the necessary resources
    :param id: The id parameter is used to specify the unique identifier of a specific school. If
    provided, the function will return information about that specific school
    :param nat_school: The parameter "nat_school" is used to specify the national school to which the
    teachers belong. It is an optional parameter and can be set to a specific value to filter the
    teachers based on their national school
    :param session: The session parameter is used to make retreiving data faster if it used again and 
    it is optional
    :param row: The "row" parameter is a boolean value that determines whether the function should
    return the result as row or not, defaults to False
    (optional)
    """
    if id == None:
        teachers =make_request(auth=auth, url=f'https://emis.moe.gov.jo/openemis-core/restful/Institution-Institutions.json?_limit=1&_orWhere=code:{nat_school}&_contain=Staff.Users,Staff.Positions',session=session)
    else:
        teachers =make_request(auth=auth, url=f'https://emis.moe.gov.jo/openemis-core/restful/Institution-Institutions.json?_limit=1&id={id}&_contain=Staff.Users,Staff.Positions',session=session)
    dic_list=[]
    for teacher in teachers['data'][0]['staff'] : 
        if teacher['staff_status_id'] == 1:
        # print(counter ,'-' , each['staff_name'])
            # dic_list.append(teacher['staff_name'])
            dic_list.append({'staffId':teacher['staff_id'],'name':teacher['name'],'name_list':[teacher['user']['first_name'], teacher['user']['middle_name'],teacher['user']['third_name'],teacher['user']['last_name']],'position':teacher['position']['name'],'birthDate':teacher['user']['date_of_birth'], 'nat_id':teacher['user']['identity_number'],'default_nat_id':teacher['user']['default_identity_type'],'staff_type':teacher['staff_type_id'] , 'staff_status': teacher['staff_status_id']})
    if row :
        return teachers
    else:
        return {'school_code_name' : teachers['data'][0]['code_name'], 'staff' : dic_list , 'school_id':teachers['data'][0]['id']}

def get_school_teachers_load(auth , inst_id , academic_period_id=13):
    """
    This function retrieves the load of teachers in a specific school for a given academic period.
    
    :param auth: The auth parameter is used for authentication purposes. It could be a token or any
    other form of authentication that allows the user to access the necessary resources
    :param inst_id: The inst_id parameter is the ID of the institution or school for which you want to
    retrieve the teachers' load
    :param academic_period_id: The academic period ID is an optional parameter that represents the
    specific academic period for which you want to retrieve the teachers' load. If not provided, it
    defaults to 13, defaults to 13 (optional)
    """
    school_load = make_request(auth=auth,url=f'https://emis.moe.gov.jo/openemis-core/restful/v2/Institution-InstitutionSubjectStaff.json?institution_id={inst_id}&_contain=Users,InstitutionSubjects&academic_period_id={academic_period_id}&_limit=0')['data']
    
    institution_class_ids = list(set([i['institution_subject_id'] for i in school_load]))
    class_list = []

    for i in range(0, len(institution_class_ids), 20):
        start = i
        end = i+19 if i+19 < len(institution_class_ids) else i+(len(institution_class_ids)-i)-1
        class_ids = [i for i in  institution_class_ids[start:end]]
        joined_string = ','.join(str(i) for i in [f'institution_subject_id:{i}' for i in class_ids])
        classes_data = make_request(auth=auth,url='https://emis.moe.gov.jo/openemis-core/restful/Institution.InstitutionClassSubjects?status=1&_contain=InstitutionSubjects,InstitutionClasses&_limit=0&_orWhere='+joined_string)['data']
        for i in classes_data:
            class_list.append({'class_id': i['institution_subject_id'] , 'class_name': i['institution_class']['name'] })
            class_dict = {i['class_id']: i['class_name'] for i in class_list if i['class_id'] != ''}
            classes = [key for value,key in class_dict.items()]
            
    grade_data = get_grade_info(auth)
    grade_list = []
    for i in grade_data:
        grade_list.append({'grade_id': i['education_grade_id'] , 'grade_name': re.sub('.*Ù„Ù„ØµÙ','Ø§Ù„ØµÙ', i['name']) })
        grade_dict = {i['grade_id']: i['grade_name'] for i in grade_list if i['grade_name'] != ''}
    grade_dict

    school_load_dictionary = []
    for load in school_load:
        try:
            school_load_dictionary.append({'name':load['user']['name'],'subject':load['institution_subject']['name'],'grade':class_dict[load['institution_subject_id']]})
        except:
            try:
                school_load_dictionary.append({'name':load['user']['name'],'subject':load['institution_subject']['name'],'grade':grade_dict[load['institution_subject']['education_grade_id']]})
            except:
                school_load_dictionary.append({'name':load['user']['name'],'subject':load['institution_subject']['name'],'grade':load['institution_subject_id']})
    return school_load_dictionary   

def count_teachers_grades(teachers_load):
    """
    The function counts the number of grades for each teacher.
    
    :param teachers_load: A dictionary where the keys are the names of the teachers and the values are
    lists of grades for each teacher
    """
    english_teachers = [item for item in teachers_load if 'Ù†Ø¬Ù„ÙŠØ²ÙŠ' in item['subject']]
    arabic_teachers = [item for item in teachers_load if 'Ø¹Ø±Ø¨ÙŠ' in item['subject']]
    math_teachers = [item for item in teachers_load if 'Ø±ÙŠØ§Ø¶ÙŠØ§Øª' in item['subject']]

    unique_english_teachers = set(item['name'] for item in english_teachers )
    unique_arabic_teachers = set(item['name'] for item in arabic_teachers)
    unique_math_teachers = set(item['name'] for item in math_teachers)
    
    loads = {'english': [], 'arabic': [], 'math': []}
    teachers = {'english': english_teachers, 'arabic': arabic_teachers, 'math': math_teachers}
    unique_teachers = {'english': unique_english_teachers, 'arabic': unique_arabic_teachers, 'math': unique_math_teachers}

    for subject in loads.keys():
        for u_name in list(unique_teachers[subject]):
            load = [teacher for teacher in teachers[subject] if teacher['name'] == u_name]
            loads[subject].append(load)
    return loads

def get_teacher_load_with_name(teachers_load , subject):
    """
    This function takes a dictionary of teachers' loads and a subject as input and returns the load of
    the teacher who teaches that subject.
    
    :param teachers_load: A dictionary where the keys are teacher names and the values are lists of
    subjects they teach and the number of classes they have for each subject
    :param subject: The subject parameter is a string that represents the name of a subject
    """
    if subject == 1 :
        subject = 'english'
        subject_sum = 'english_class_sum'
    elif subject == 2 :
        subject = 'arabic'
        subject_sum = 'arabic_class_sum'
    elif subject == 3 :
        subject = 'math'
        subject_sum = 'math_class_sum'
    teachers = []
    for group in count_teachers_grades(teachers_load)[subject]:
        grades = [grade['grade'] for grade in group ] 
        teachers.append((group[0]['name'],count_teacher_load(grades)[subject_sum],count_teacher_load(grades)['classes']))
    return teachers

def count_teacher_load(classes):
    """
    The function counts the number of classes taught by each teacher.
    
    :param classes: A list of dictionaries representing different classes. Each dictionary should have
    the following keys:
    """
    
    arabic_class_sum = 0
    english_class_sum = 0
    math_class_sum = 0

    for school_class in classes:
        if 'Ø§ÙˆÙ„' in school_class or 'Ø«Ø§Ù†ÙŠ' in school_class or 'Ø«Ø§Ù„Ø«'in school_class or  'Ø±Ø§Ø¨Ø¹' in school_class or 'Ø¹Ø´Ø±' in school_class:
            if 'Ø¯Ø¨ÙŠ' in school_class:
                math_class_sum+=3
            else:
                math_class_sum+=5
            english_class_sum += 4
        else: 
            english_class_sum+= 5
            math_class_sum += 5

    for school_class in classes:
        if 'Ø³Ø§Ø¨Ø¹' in school_class or 'Ø«Ø§Ù…Ù†' in school_class or 'ØªØ§Ø³Ø¹' in school_class or  'Ø¹Ø§Ø´Ø±' in school_class :
            arabic_class_sum += 6
        else:
            if 'Ø¹Ø´Ø±' in school_class:
                if 'Ø¯Ø¨ÙŠ' in school_class and 'Ø®ØµØµ' in school_class:
                    if  'Ø­Ø§Ø¯ÙŠ' in school_class:
                        arabic_class_sum+=5
                    elif 'Ø«Ø§Ù†ÙŠ' in school_class:
                        arabic_class_sum+=4
                else:
                    arabic_class_sum+=4
            else:
                arabic_class_sum+=7
    return {'english_class_sum' : english_class_sum , 'arabic_class_sum' : arabic_class_sum , 'math_class_sum' :  math_class_sum , 'classes' :  classes}

def create_tables_wrapper(username , password ,term2=False , just_teacher_class=True , curr_year = None,student_status_ids=[1]): 
    """
    The function creates tables in using the provided username and password. It is wrapper and that 
    make my code more consise 
    
    :param username: The username parameter is used to specify the username for the database connection
    :param password: The password parameter is used to specify the password for the database connection
    :param term2: A boolean parameter that determines whether to include tables in the second term. If set
    to True. If set to False, only the first term marks will be included in tables that will be created,
    defaults to False (optional)
    """
    session = requests.Session()
    auth = get_auth(username, password)
    student_info_marks = get_students_info_subjectsMarks( username , password ,session = session ,just_teacher_class=just_teacher_class,curr_year=curr_year,student_status_ids=student_status_ids)
    dic_list4 = student_info_marks
    #TODO: Ø§Ø­Ø°Ù Ù‡Ø°Ø§ Ø§Ù„Ù…ØªØºÙŠØ± Ø§Ù„Ø¹Ø§Ù„Ù…ÙŠ Ùˆ Ø§Ù„Ø°ÙŠ ØªÙ… ÙˆØ¶Ø¹Ù‡ Ù„ØºØ§ÙŠØ§Øª Ø§Ù„ØªØ­Ù‚Ù‚ Ù…Ù† Ø§Ù„Ø§Ø®Ø·Ø§Ø¡ 
    global grouped_list
    grouped_list = group_students(dic_list4 )
    
    
    add_subject_sum_dictionary(grouped_list)
    add_averages_to_group_list(grouped_list ,skip_art_sport=False)
    
    # save_dictionary_to_json_file(dictionary={'grouped_list':grouped_list})
    create_tables(auth , grouped_list ,term2=term2 )

def create_certs_wrapper(username , password , student_identity_number = None ,term2=False,skip_art_sport=True,just_teacher_class=True,session=None , curr_year = None):
    """
    The function create_certs_wrapper is a Python function that takes in parameters username, password,
    term2 (with a default value of False), and session (with a default value of None).
    
    :param username: The username parameter is a string that represents the username of the user for
    whom the certificates are being created
    :param password: The password parameter is used to specify the password for the user
    :param term2: A boolean value indicating whether the user wants to create certificates for the
    second term or not. The default value is False, defaults to False (optional)
    :param session: The `session` parameter is an optional parameter that allows you to pass an existing
    session object. This can be useful if you want to reuse an existing session for making multiple
    requests
    """
    student_info_marks = get_students_info_subjectsMarks( username , password ,student_identity_number = student_identity_number , just_teacher_class=just_teacher_class,session=session, curr_year = curr_year)
    dic_list4 = student_info_marks
    listo=get_school_absent_days_with_studentID_and_countOfAbsentDays_and_classID_and_className(get_auth(username,password))
    absent_list_with_names=get_absens_studentName_and_countOfDays_and_studentID(dic_list4,listo)
    grouped_list = group_students(dic_list4 )
    
    add_subject_sum_dictionary(grouped_list)
    add_averages_to_group_list(grouped_list ,skip_art_sport=skip_art_sport)
    
    create_certs(grouped_list,absent_list_with_names , term2=term2)

def create_tables(auth , grouped_list ,term2=False ,template='./templet_files/tamplete_table.xlsx'  , outdir='./send_folder/'):
    """
    The function `create_tables` creates tables(Ø¬Ø¯Ø§ÙˆÙ„) based on a grouped list and saves them as Excel files in
    a specified output directory.
    
    :param auth: The auth parameter is used for authentication purposes, such as providing credentials
    or tokens to access certain resources or APIs. It is not clear what specific authentication method
    or library is being used in this code, so you would need to refer to the documentation or code
    implementation to understand how to provide the appropriate authentication
    :param grouped_list: The `grouped_list` parameter is a list of lists. Each inner list represents a
    group of data that will be displayed in a separate table in the output file. Each element in the
    inner list represents a row in the table
    :param term2: The `term2` parameter is a boolean value that determines whether or not to include a
    second term in the table. If `term2` is set to `True`, the table will include a second term. If
    `term2` is set to `False` (default), the table will, defaults to False (optional)
    :param template: The `template` parameter is the path to the template file that will be used to
    create the tables. It should be a .xlsx file, defaults to ./templet_files/tamplete_table.xlsx
    (optional)
    :param outdir: The `outdir` parameter specifies the directory where the output files will be saved,
    defaults to ./send_folder/ (optional)
    """
    # auth = get_auth(username , password)
    institution_area_data = inst_area(auth)
    institution_data = inst_name(auth)
    curr_year_code = get_curr_period(auth)['data'][0]['code']

    for group in grouped_list:
        
        if 'Ø¹Ø´Ø±' not in group[0]['student_grade_name']:
            template_file = openpyxl.load_workbook(template)
            marks_sheet = template_file.worksheets[2]
        
            for row_number, dataFrame in enumerate(sort_dictionary_list_based_on(group ,dictionary_key='student__full_name',simple=False,reverse=False ,sort_names_only=True), start=4):
                islam_subject = [value for key ,value in dataFrame['subject_sums'].items() if 'Ø³Ù„Ø§Ù…ÙŠØ©' in key if len(dataFrame['subjects_assessments_info'])] # Ø§Ù„ØªØ±Ø¨ÙŠØ© Ø§Ù„Ø§Ø³Ù„Ø§Ù…ÙŠØ©
                arabic_subject = [value for key ,value in dataFrame['subject_sums'].items() if 'Ø¹Ø±Ø¨ÙŠ' in key if len(dataFrame['subjects_assessments_info'])] # Ø§Ù„Ù„ØºØ© Ø§Ù„Ø¹Ø±Ø¨ÙŠØ©
                english_subject = [value for key ,value in dataFrame['subject_sums'].items() if 'Ø¬Ù„ÙŠØ²ÙŠØ©' in key if len(dataFrame['subjects_assessments_info'])]   # Ø§Ù„Ù„ØºØ© Ø§Ù„Ø§Ù†Ø¬Ù„ÙŠØ²ÙŠØ© 
                math_subject = [value for key ,value in dataFrame['subject_sums'].items() if 'Ø±ÙŠØ§Ø¶ÙŠØ§Øª' in key if len(dataFrame['subjects_assessments_info'])] # Ø§Ù„Ø±ÙŠØ§Ø¶ÙŠØ§Øª 
                social_subjects = [value for key ,value in dataFrame['subject_sums'].items() if 'Ø§Ø¬ØªÙ…Ø§Ø¹ÙŠØ© Ùˆ Ø§Ù„ÙˆØ·Ù†ÙŠØ©' in key if len(dataFrame['subjects_assessments_info'])]   # Ø§Ù„ØªØ±Ø¨ÙŠØ© Ø§Ù„Ø§Ø¬ØªÙ…Ø§Ø¹ÙŠØ© Ùˆ Ø§Ù„ÙˆØ·Ù†ÙŠØ© 
                science_subjects = [value for key ,value in dataFrame['subject_sums'].items() if 'Ø§Ù„Ø¹Ù„ÙˆÙ…' in key if len(dataFrame['subjects_assessments_info'])]  # Ø§Ù„Ø¹Ù„ÙˆÙ…
                art_subject = [value for key ,value in dataFrame['subject_sums'].items() if 'Ø§Ù„ÙÙ†ÙŠØ© ÙˆØ§Ù„Ù…ÙˆØ³' in key if len(dataFrame['subjects_assessments_info'])]    # Ø§Ù„ØªØ±Ø¨ÙŠØ© Ø§Ù„ÙÙ†ÙŠØ© ÙˆØ§Ù„Ù…ÙˆØ³ÙŠÙ‚ÙŠØ©
                sport_subject = [value for key ,value in dataFrame['subject_sums'].items() if 'Ø±ÙŠØ§Ø¶ÙŠØ©' in key if len(dataFrame['subjects_assessments_info'])] # Ø§Ù„ØªØ±Ø¨ÙŠØ© Ø§Ù„Ø±ÙŠØ§Ø¶ÙŠØ©
                vocational_subject = [value for key ,value in dataFrame['subject_sums'].items() if 'Ù…Ù‡Ù†ÙŠØ©' in key if len(dataFrame['subjects_assessments_info'])] # Ø§Ù„ØªØ±Ø¨ÙŠØ© Ø§Ù„Ù…Ù‡Ù†ÙŠØ© 
                computer_subject = [value for key ,value in dataFrame['subject_sums'].items() if 'Ø­Ø§Ø³ÙˆØ¨' in key if len(dataFrame['subjects_assessments_info'])]   # Ø§Ù„Ø­Ø§Ø³ÙˆØ¨
                financial_subject = [value for key ,value in dataFrame['subject_sums'].items() if 'Ù…Ø§Ù„ÙŠØ©' in key if len(dataFrame['subjects_assessments_info'])]  # Ø§Ù„Ø«Ù‚Ø§ÙØ© Ø§Ù„Ù…Ø§Ù„ÙŠØ©
                franch_subject = [value for key ,value in dataFrame['subject_sums'].items() if 'ÙØ±Ù†Ø³ÙŠØ©' in key if len(dataFrame['subjects_assessments_info'])]    # Ø§Ù„Ù„ØºØ© Ø§Ù„ÙØ±Ù†Ø³ÙŠØ© 
                christian_subject = [value for key ,value in dataFrame['subject_sums'].items() if 'Ø§Ù„Ø¯ÙŠÙ† Ø§Ù„Ù…Ø³ÙŠØ­ÙŠ' in key if len(dataFrame['subjects_assessments_info'])]  # Ø§Ù„Ø¯ÙŠÙ† Ø§Ù„Ù…Ø³ÙŠØ­ÙŠ

                marks_sheet.cell(row=row_number, column=1).value = row_number-3
                marks_sheet.cell(row=row_number, column=2).value = dataFrame['student__full_name']
                marks_sheet.cell(row=row_number, column=3).value = dataFrame['student_nat']
                marks_sheet.cell(row=row_number, column=4).value = dataFrame['student_birth_place']
                marks_sheet.cell(row=row_number, column=5).value = dataFrame['student_birth_date'].split('/')[0]
                marks_sheet.cell(row=row_number, column=6).value = dataFrame['student_birth_date'].split('/')[1]
                marks_sheet.cell(row=row_number, column=7).value = dataFrame['student_birth_date'].split('/')[2]
                marks_sheet.cell(row=row_number, column=8).value = islam_subject[0][0] if islam_subject and len(islam_subject[0]) > 0 else ''
                marks_sheet.cell(row=row_number, column=9).value = islam_subject[0][1] if term2 and islam_subject and len(islam_subject[0]) > 0 else ''
                marks_sheet.cell(row=row_number, column=10).value = islam_subject[0][0]+islam_subject[0][1] if term2 and islam_subject and len(islam_subject[0]) > 0 else ''
                marks_sheet.cell(row=row_number, column=11).value = arabic_subject[0][0] if arabic_subject and len(arabic_subject[0]) > 0 else ''
                marks_sheet.cell(row=row_number, column=12).value = arabic_subject[0][1] if term2 and arabic_subject and len(arabic_subject[0]) > 0 else ''
                marks_sheet.cell(row=row_number, column=13).value = arabic_subject[0][0]+arabic_subject[0][1] if term2 and arabic_subject and len(arabic_subject[0]) > 0 else ''
                marks_sheet.cell(row=row_number, column=14).value = english_subject[0][0] if english_subject and len(english_subject[0]) > 0 else ''
                marks_sheet.cell(row=row_number, column=15).value = english_subject[0][1] if term2 and english_subject and len(english_subject[0]) > 0 else ''
                marks_sheet.cell(row=row_number, column=16).value = english_subject[0][0]+english_subject[0][1] if term2 and english_subject and len(english_subject[0]) > 0 else ''
                marks_sheet.cell(row=row_number, column=17).value = math_subject[0][0] if math_subject and len(math_subject[0]) > 0 else ''
                marks_sheet.cell(row=row_number, column=18).value = math_subject[0][1] if term2 and math_subject and len(math_subject[0]) > 0 else ''
                marks_sheet.cell(row=row_number, column=19).value = math_subject[0][0]+math_subject[0][1] if term2 and math_subject and len(math_subject[0]) > 0 else ''
                marks_sheet.cell(row=row_number, column=20).value = social_subjects[0][0] if social_subjects and len(social_subjects[0]) > 0 else ''
                marks_sheet.cell(row=row_number, column=21).value = social_subjects[0][1] if term2 and social_subjects and len(social_subjects[0]) > 0 else ''
                marks_sheet.cell(row=row_number, column=22).value = social_subjects[0][0]+social_subjects[0][1] if term2 and social_subjects and len(social_subjects[0]) > 0 else ''
                marks_sheet.cell(row=row_number, column=23).value = science_subjects[0][0] if science_subjects and len(science_subjects[0]) > 0 else ''
                marks_sheet.cell(row=row_number, column=24).value = science_subjects[0][1] if term2 and science_subjects and len(science_subjects[0]) > 0 else ''
                marks_sheet.cell(row=row_number, column=25).value = science_subjects[0][0]+science_subjects[0][1] if term2 and science_subjects and len(science_subjects[0]) > 0 else ''
                marks_sheet.cell(row=row_number, column=26).value = art_subject[0][0] if art_subject and len(art_subject[0]) > 0 else ''
                marks_sheet.cell(row=row_number, column=27).value = art_subject[0][1] if term2 and art_subject and len(art_subject[0]) > 0 else ''
                marks_sheet.cell(row=row_number, column=28).value = art_subject[0][0]+art_subject[0][1] if term2 and art_subject and len(art_subject[0]) > 0 else ''
                marks_sheet.cell(row=row_number, column=29).value = sport_subject[0][0] if sport_subject and len(sport_subject[0]) > 0 else ''
                marks_sheet.cell(row=row_number, column=30).value = sport_subject[0][1] if term2 and sport_subject and len(sport_subject[0]) > 0 else ''
                marks_sheet.cell(row=row_number, column=31).value = sport_subject[0][0]+sport_subject[0][1] if term2 and sport_subject and len(sport_subject[0]) > 0 else ''
                marks_sheet.cell(row=row_number, column=32).value = financial_subject[0][0] if financial_subject and len(financial_subject[0]) > 0 else ''
                marks_sheet.cell(row=row_number, column=33).value = financial_subject[0][1] if term2 and financial_subject and len(financial_subject[0]) > 0 else ''
                marks_sheet.cell(row=row_number, column=34).value = financial_subject[0][0]+financial_subject[0][1] if term2 and financial_subject and len(financial_subject[0]) > 0 else ''
                marks_sheet.cell(row=row_number, column=35).value = vocational_subject[0][0] if vocational_subject and len(vocational_subject[0]) > 0 else ''
                marks_sheet.cell(row=row_number, column=36).value = vocational_subject[0][1] if term2 and vocational_subject and len(vocational_subject[0]) > 0 else ''
                marks_sheet.cell(row=row_number, column=37).value = vocational_subject[0][0]+vocational_subject[0][1] if term2 and vocational_subject and len(vocational_subject[0]) > 0 else ''
                marks_sheet.cell(row=row_number, column=38).value = computer_subject[0][0] if computer_subject and len(computer_subject[0]) > 0 else ''
                marks_sheet.cell(row=row_number, column=39).value = computer_subject[0][1] if term2 and computer_subject and len(computer_subject[0]) > 0 else ''
                marks_sheet.cell(row=row_number, column=40).value = computer_subject[0][0]+computer_subject[0][1] if term2 and computer_subject and len(computer_subject[0]) > 0 else ''
                marks_sheet.cell(row=row_number, column=41).value = franch_subject[0][0] if franch_subject and len(franch_subject[0]) > 0 else ''
                marks_sheet.cell(row=row_number, column=42).value = franch_subject[0][1] if term2 and franch_subject and len(franch_subject[0]) > 0 else ''
                marks_sheet.cell(row=row_number, column=43).value = franch_subject[0][0]+franch_subject[0][1] if term2 and franch_subject and len(franch_subject[0]) > 0 else ''
                # marks_sheet.cell(row=row_number, column=44).value = dataFrame[0][] if = and len(=[0]) > 0 else ''
                # marks_sheet.cell(row=row_number, column=45).value = dataFrame[0][] if = and len(=[0]) > 0 else ''
                # marks_sheet.cell(row=row_number, column=46).value = dataFrame[0][] if = and len(=[0]) > 0 else ''
                marks_sheet.cell(row=row_number, column=47).value = christian_subject[0][0] if christian_subject and len(christian_subject[0]) > 0 else ''
                marks_sheet.cell(row=row_number, column=48).value = christian_subject[0][1] if term2 and christian_subject and len(christian_subject[0]) > 0 else ''
                marks_sheet.cell(row=row_number, column=49).value = christian_subject[0][0]+christian_subject[0][1] if term2 and christian_subject and len(christian_subject[0]) > 0 else ''

            if 'Ø§Ù„Ø«Ø§Ù…Ù†' in group[0]['student_grade_name']:
                marks_sheet['a1'] = 1800
                marks_sheet['a3'] =f'Ø¬Ø¯ÙˆÙ„ Ø§Ù„Ø¹Ù„Ø§Ù…Ø§Øª Ø§Ù„Ù…Ø¯Ø±Ø³ÙŠÙ‡ Ù„Ù„ØµÙ Ø§Ù„Ø«Ø§Ù…Ù† Ø§Ù„Ø£Ø³Ø§Ø³ÙŠ Ù„Ù„Ø¹Ø§Ù… Ø§Ù„Ø¯Ø±Ø§Ø³ÙŠ ( {curr_year_code} )'
                # Ø§Ø³Ù„Ø§Ù…ÙŠØ©
                # h/i/j
                marks_sheet['h3'],marks_sheet['i3'],marks_sheet['j3'] = [200]*3
                # Ø¹Ø±Ø¨ÙŠØ© 
                # k/l/m
                marks_sheet['k3'],marks_sheet['l3'],marks_sheet['m3'] = [300]*3
                # Ø§Ù†Ø¬Ù„ÙŠØ²ÙŠØ© 
                # n/o/p
                marks_sheet['n3'],marks_sheet['o3'],marks_sheet['p3'] = [200]*3
                # Ø±ÙŠØ§Ø¶ÙŠØ§Øª
                # q/r/s
                marks_sheet['q3'],marks_sheet['r3'],marks_sheet['s3'] = [200]*3
                # Ø§Ø¬ØªÙ…Ø§Ø¹ÙŠØ§Øª 
                # t/u/v
                marks_sheet['t3'],marks_sheet['u3'],marks_sheet['v3'] = [200]*3
                # Ø¹Ù„ÙˆÙ…
                # w/x/y
                marks_sheet['w3'],marks_sheet['x3'],marks_sheet['y3'] = [200]*3
            elif 'Ø§Ù„ØªØ§Ø³Ø¹' in group[0]['student_grade_name']:
                marks_sheet['a1'] = 2000
                marks_sheet['a3'] =f'Ø¬Ø¯ÙˆÙ„ Ø§Ù„Ø¹Ù„Ø§Ù…Ø§Øª Ø§Ù„Ù…Ø¯Ø±Ø³ÙŠÙ‡ Ù„Ù„ØµÙ Ø§Ù„ØªØ§Ø³Ø¹  Ø§Ù„Ø£Ø³Ø§Ø³ÙŠ Ù„Ù„Ø¹Ø§Ù… Ø§Ù„Ø¯Ø±Ø§Ø³ÙŠ ( {curr_year_code} )'
                # Ø§Ø³Ù„Ø§Ù…ÙŠØ©
                # h/i/j
                marks_sheet['h3'],marks_sheet['i3'],marks_sheet['j3'] = [200]*3
                # Ø¹Ø±Ø¨ÙŠØ© 
                # k/l/m
                marks_sheet['k3'],marks_sheet['l3'],marks_sheet['m3'] = [300]*3
                # Ø§Ù†Ø¬Ù„ÙŠØ²ÙŠØ© 
                # n/o/p
                marks_sheet['n3'],marks_sheet['o3'],marks_sheet['p3'] = [200]*3
                # Ø±ÙŠØ§Ø¶ÙŠØ§Øª
                # q/r/s
                marks_sheet['q3'],marks_sheet['r3'],marks_sheet['s3'] = [200]*3
                # Ø§Ø¬ØªÙ…Ø§Ø¹ÙŠØ§Øª 
                # t/u/v
                marks_sheet['t3'],marks_sheet['u3'],marks_sheet['v3'] = [200]*3
                # Ø¹Ù„ÙˆÙ…
                # w/x/y
                marks_sheet['w3'],marks_sheet['x3'],marks_sheet['y3'] = [400]*3
            elif 'Ø§Ù„Ø¹Ø§Ø´Ø±' in group[0]['student_grade_name']:
                marks_sheet['a1'] = 2000
                marks_sheet['a3'] =f'Ø¬Ø¯ÙˆÙ„ Ø§Ù„Ø¹Ù„Ø§Ù…Ø§Øª Ø§Ù„Ù…Ø¯Ø±Ø³ÙŠÙ‡ Ù„Ù„ØµÙ Ø§Ù„Ø¹Ø§Ø´Ø± Ø§Ù„Ø£Ø³Ø§Ø³ÙŠ Ù„Ù„Ø¹Ø§Ù… Ø§Ù„Ø¯Ø±Ø§Ø³ÙŠ ( {curr_year_code} )'
                # Ø§Ø³Ù„Ø§Ù…ÙŠØ©
                # h/i/j
                marks_sheet['h3'],marks_sheet['i3'],marks_sheet['j3'] = [200]*3
                # Ø¹Ø±Ø¨ÙŠØ© 
                # k/l/m
                marks_sheet['k3'],marks_sheet['l3'],marks_sheet['m3'] = [300]*3
                # Ø§Ù†Ø¬Ù„ÙŠØ²ÙŠØ© 
                # n/o/p
                marks_sheet['n3'],marks_sheet['o3'],marks_sheet['p3'] = [200]*3
                # Ø±ÙŠØ§Ø¶ÙŠØ§Øª
                # q/r/s
                marks_sheet['q3'],marks_sheet['r3'],marks_sheet['s3'] = [200]*3
                # Ø§Ø¬ØªÙ…Ø§Ø¹ÙŠØ§Øª 
                # t/u/v
                marks_sheet['t3'],marks_sheet['u3'],marks_sheet['v3'] = [200]*3
                # Ø¹Ù„ÙˆÙ…
                # w/x/y
                marks_sheet['w3'],marks_sheet['x3'],marks_sheet['y3'] = [400]*3        
            else:
                marks_sheet['a3'] = f'Ø¬Ø¯ÙˆÙ„ Ø§Ù„Ø¹Ù„Ø§Ù…Ø§Øª Ø§Ù„Ø¯Ø±Ø§Ø³ÙŠØ© Ù„Ù„ØµÙÙˆÙ Ù…Ù† Ø§Ù„Ø£ÙˆÙ„ Ø§Ù„Ù‰ Ø§Ù„Ø³Ø§Ø¨Ø¹ Ø§Ù„Ø£Ø³Ø§Ø³ÙŠ ( {curr_year_code} )'
                marks_sheet['h3'],marks_sheet['i3'],marks_sheet['j3'] = [100]*3
                # Ø¹Ø±Ø¨ÙŠØ© 
                # k/l/m
                marks_sheet['k3'],marks_sheet['l3'],marks_sheet['m3'] = [100]*3
                # Ø§Ù†Ø¬Ù„ÙŠØ²ÙŠØ© 
                # n/o/p
                marks_sheet['n3'],marks_sheet['o3'],marks_sheet['p3'] = [100]*3
                # Ø±ÙŠØ§Ø¶ÙŠØ§Øª
                # q/r/s
                marks_sheet['q3'],marks_sheet['r3'],marks_sheet['s3'] = [100]*3
                # Ø§Ø¬ØªÙ…Ø§Ø¹ÙŠØ§Øª 
                # t/u/v
                marks_sheet['t3'],marks_sheet['u3'],marks_sheet['v3'] = [100]*3
                # Ø¹Ù„ÙˆÙ…
                # w/x/y
                marks_sheet['w3'],marks_sheet['x3'],marks_sheet['y3'] = [100]*3    
                if 'Ø³Ø§Ø¨Ø¹' in group[0]['student_grade_name']:
                    marks_sheet['a1'] = 1100
                elif 'Ø³Ø§Ø¯Ø³' in group[0]['student_grade_name']:
                    marks_sheet['a1'] = 900
                else:
                    marks_sheet['a1'] = 800
                
            marks_sheet['b3'] = institution_area_data['data'][0]['Areas']['name']
            marks_sheet['c3'] = ''
            marks_sheet['d3'] = institution_data['data'][0]['Institutions']['code_name']
            marks_sheet['e3'] = institution_area_data['data'][0]['AreaAdministratives']['name']
            marks_sheet['f3'] = group[0]['student_grade_name']
            marks_sheet['g3'] = group[0]['student_class_name_letter']
            
            template_file.save(outdir+' Ø¬Ø¯ÙˆÙ„ '+group[0]['student_class_name_letter']+'.xlsx')

def create_certs(grouped_list ,absent_list_with_names, term2=False ,template='./templet_files/a4_gray_cert.xlsx' ,image='./templet_files/Pasted image.png' , outdir='./send_folder/'):
    """
    The function `create_certs` creates certificates using a template file and an image, and saves them
    in an output directory.
    
    :param grouped_list: The grouped_list parameter is a list of lists. Each inner list represents a
    group of individuals who will receive a certificate together. Each inner list should contain the
    names of the individuals in that group. For example, if there are three groups, the grouped_list
    parameter could look like this:
    :param term2: The term2 parameter is a boolean value that indicates whether the certificate is for
    the second term or not. If term2 is set to True, it means the certificate is for the second term. If
    term2 is set to False, it means the certificate is for the first term, defaults to False (optional)
    :param template: The path to the Excel template file that will be used to create the certificates.
    The default value is './templet_files/a4_gray_cert.xlsx', defaults to
    ./templet_files/a4_gray_cert.xlsx (optional)
    :param image: The `image` parameter is the path to the image file that will be inserted into the
    certificate template, defaults to ./templet_files/Pasted image.png (optional)
    :param outdir: The `outdir` parameter specifies the directory where the generated certificates will
    be saved, defaults to ./send_folder/ (optional)
    """
    gray_cert_cell_positions_context = {
    'student_name': 'E7',
    'class_section': 'B11',
    'nationality': 'H9',
    'national_id': 'B9',
    'birthplace_date': 'H7',
    'religion': '',
    'student_address': '',
    'school_name': 'G11',
    'school_address': '',
    'directorate': 'B13',
    'brigade': 'I13',
    'supervising_authority': '',
    'school_phone_number': 'C18',
    'school_national_id': 'C19',
    'academic_year': '',
    'academic_year_1': '',
    'academic_year_2': '',
    'class': '',
    'islamic_education': 'c18',
    'arabic_language': 'c19',
    'english_language': 'c20',
    'mathematics': 'c21',
    'social_studies': 'c22',
    'science': 'c23',
    'visual_arts': 'c24',
    'physical_education': 'c25',
    'vocational_education': 'c26',
    'computer': 'c27',
    'financial_culture': 'c28',
    'french_language': 'c29',
    'christian_religion': 'C30',
    'result': '',
    'average': 'C32',
    'school_days': 'G35',
    'class_teacher_name': 'J35',
    'principal_name': 'I36',
    'semester_1_student_absent': '',
    'semester_2_student_absent': '',
    'student_name': '',
    }
    for group in grouped_list:
        if 'Ø¹Ø´Ø±' not in group[0]['student_grade_name']:
            template_file = load_workbook(template)
            sheet1 = template_file.worksheets[0]
            
            names_averages =  sort_dictionary_list_based_on(group)

            group = sort_dictionary_list_based_on(group ,simple=False)

            for row_number, dataFrame in enumerate(names_averages, start=5):
                sheet1.cell(row=row_number, column=2).value = dataFrame[1]
                sheet1.cell(row=row_number, column=4).value = dataFrame[0]
                
            counter = 1
            try:
                for group_item in group:
                    
                    sheet2 = template_file.copy_worksheet(template_file.worksheets[1])
                    sheet2.title = str(counter)
                    counter += 1
                    sheet2.sheet_view.rightToLeft = True    
                    sheet2.sheet_view.rightToLeft = True   

                    img = openpyxl.drawing.image.Image(image)
                    img.anchor = 'e2'
                    sheet2.add_image(img)

                    # group_item = grouped_list[0][0]
                    # print(sheet2)
                    sheet2['b7'] = group_item['student__full_name']
                    # Ù…ÙƒØ§Ù† Ùˆ ØªØ§Ø±ÙŠØ® Ø§Ù„ÙˆÙ„Ø§Ø¯Ø©
                    sheet2['h7']= str(group_item['student_birth_place']) + ' ' + str(group_item['student_birth_date'])
                    #Ø§Ù„Ø±Ù‚Ù… Ø§Ù„ÙˆØ·Ù†ÙŠ
                    sheet2['b9']= group_item['student_nat_id']
                    #Ø§Ù„Ø¬Ù†Ø³ÙŠØ©
                    sheet2['h9']= group_item['student_nat']
                    #Ø§Ù„ØµÙ Ùˆ Ø§Ù„Ø´Ø¹Ø¨Ø© 
                    sheet2['b11']= group_item['student_class_name_letter']
                    #Ø§Ù„Ù…Ø¯Ø±Ø³Ø© Ùˆ Ø±Ù‚Ù…Ù‡Ø§ Ø§Ù„ÙˆØ·Ù†ÙŠ
                    sheet2['g11']= group_item['student_school_name']
                    #Ø§Ù„Ù…Ù†Ø·Ù‚Ø© Ø§Ù„ØªØ¹Ù„ÙŠÙ…ÙŠØ© 
                    sheet2['b13']= group_item['student_edu_place']  
                    #Ø§Ù„Ø¨Ù„Ø¯Ø© 
                    sheet2['f13']= ''
                    #Ø§Ù„Ù„ÙˆØ§Ø¡
                    sheet2['i13']= group_item['student_directory']
                    # put the subjects cells inder here 
                    i ='c,d,e,g,j,f'.split(',')
                    r = range(18,32)

                    value_item = 0

                    # Ø§Ù„ØªØ±Ø¨ÙŠØ© Ø§Ù„Ø§Ø³Ù„Ø§Ù…ÙŠØ©
                    islam_subject = [value for key ,value in group_item['subject_sums'].items() if 'Ø³Ù„Ø§Ù…ÙŠØ©' in key]
                    if 'Ø«Ø§Ù…Ù†' in group_item['student_grade_name'] or 'ØªØ§Ø³Ø¹' in group_item['student_grade_name'] or 'Ø¹Ø§Ø´Ø±' in group_item['student_grade_name']:
                        sheet2['C18'] = 200
                        islam_subject_maxMark = 200
                    else:
                        sheet2['C18'] = 100
                        islam_subject_maxMark = 100
                    sheet2['D18'] = islam_subject[0][value_item] if islam_subject and len(islam_subject[0]) != 0 else ''
                    sheet2['E18'] = islam_subject[0][1] if term2 and islam_subject and len(islam_subject[0]) != 0 else ''
                    sheet2['F18'] = (islam_subject[0][value_item] + islam_subject[0][1])/2 if term2 and islam_subject and len(islam_subject[0]) != 0 else ''
                    if term2:
                        sheet2['G18'] = convert_avarage_to_words((islam_subject[0][value_item] + islam_subject[0][1])/2) if islam_subject else ''
                        sheet2['J18'] = score_in_words(((islam_subject[0][value_item] + islam_subject[0][1])/2),max_mark=islam_subject_maxMark) if islam_subject else ''
                    else:
                        sheet2['G18'] = convert_avarage_to_words(islam_subject[0][value_item]) if islam_subject else ''
                        sheet2['J18'] = score_in_words(islam_subject[0][value_item],max_mark=islam_subject_maxMark) if islam_subject else ''

                    # Ø§Ù„Ù„ØºØ© Ø§Ù„Ø¹Ø±Ø¨ÙŠØ©
                    arabic_subject = [value for key ,value in group_item['subject_sums'].items() if 'Ø¹Ø±Ø¨ÙŠ' in key]
                    if 'Ø«Ø§Ù…Ù†' in group_item['student_grade_name'] or 'ØªØ§Ø³Ø¹' in group_item['student_grade_name'] or 'Ø¹Ø§Ø´Ø±' in group_item['student_grade_name']:
                        sheet2['C19'] = 300
                        arabic_subject_maxMark = 300
                    else:
                        sheet2['C19'] = 100
                        arabic_subject_maxMark = 100
                    sheet2['D19'] = arabic_subject[0][value_item] if arabic_subject and len(arabic_subject[0]) != 0 else ''
                    sheet2['E19'] = arabic_subject[0][1] if term2 and arabic_subject and len(arabic_subject[0]) != 0 else ''
                    sheet2['F19'] = (arabic_subject[0][value_item] + arabic_subject[0][1])/2 if term2 and arabic_subject and len(arabic_subject[0]) != 0 else ''
                    if term2:
                        sheet2['G19'] = convert_avarage_to_words((arabic_subject[0][value_item] + arabic_subject[0][1])/2) if arabic_subject else ''
                        sheet2['J19'] = score_in_words(((arabic_subject[0][value_item] + arabic_subject[0][1])/2),max_mark=arabic_subject_maxMark) if arabic_subject else ''
                    else:
                        sheet2['G19'] = convert_avarage_to_words(arabic_subject[0][value_item]) if arabic_subject else ''
                        sheet2['J19'] = score_in_words(arabic_subject[0][value_item],max_mark=arabic_subject_maxMark) if arabic_subject else ''

                    # Ø§Ù„Ù„ØºØ© Ø§Ù„Ø§Ù†Ø¬Ù„ÙŠØ²ÙŠØ© 
                    english_subject = [value for key ,value in group_item['subject_sums'].items() if 'Ø¬Ù„ÙŠØ²ÙŠØ©' in key]
                    if 'Ø«Ø§Ù…Ù†' in group_item['student_grade_name'] or 'ØªØ§Ø³Ø¹' in group_item['student_grade_name'] or 'Ø¹Ø§Ø´Ø±' in group_item['student_grade_name']:
                        sheet2['C20'] = 200
                        english_subject_maxMark = 200
                    else:
                        sheet2['C20'] = 100
                        english_subject_maxMark = 100
                    sheet2['D20'] = english_subject[0][value_item] if english_subject and len(english_subject[0]) != 0 else ''
                    sheet2['E20'] = english_subject[0][1] if term2 and english_subject and len(english_subject[0]) != 0 else ''
                    sheet2['F20'] = (english_subject[0][value_item] + english_subject[0][1])/2 if term2 and english_subject and len(english_subject[0]) != 0 else ''
                    if term2:
                        sheet2['G20'] = convert_avarage_to_words((english_subject[0][value_item] + english_subject[0][1])/2) if english_subject else ''
                        sheet2['J20'] = score_in_words(((english_subject[0][value_item] + english_subject[0][1])/2),max_mark=english_subject_maxMark) if english_subject else ''
                    else:
                        sheet2['G20'] = convert_avarage_to_words(english_subject[0][value_item]) if english_subject else ''
                        sheet2['J20'] = score_in_words(english_subject[0][value_item],max_mark=english_subject_maxMark) if english_subject else ''

                    # Ø§Ù„Ø±ÙŠØ§Ø¶ÙŠØ§Øª 
                    math_subject = [value for key ,value in group_item['subject_sums'].items() if 'Ø±ÙŠØ§Ø¶ÙŠØ§Øª' in key]
                    if 'Ø«Ø§Ù…Ù†' in group_item['student_grade_name'] or 'ØªØ§Ø³Ø¹' in group_item['student_grade_name'] or 'Ø¹Ø§Ø´Ø±' in group_item['student_grade_name']:
                        sheet2['C21'] = 200
                        math_subject_maxMark = 200
                    else:
                        sheet2['C21'] = 100
                        math_subject_maxMark = 100
                    sheet2['D21'] = math_subject[0][value_item] if math_subject and len(math_subject[0]) != 0 else ''
                    sheet2['E21'] = math_subject[0][1] if term2 and math_subject and len(math_subject[0]) != 0 else ''
                    sheet2['F21'] = (math_subject[0][value_item] + math_subject[0][1])/2 if term2 and math_subject and len(math_subject[0]) != 0 else ''
                    if term2:
                        sheet2['G21'] = convert_avarage_to_words((math_subject[0][value_item] + math_subject[0][1])/2) if math_subject else ''
                        sheet2['J21'] = score_in_words(((math_subject[0][value_item] + math_subject[0][1])/2),max_mark=math_subject_maxMark) if math_subject else ''
                    else:
                        sheet2['G21'] = convert_avarage_to_words(math_subject[0][value_item]) if math_subject else ''
                        sheet2['J21'] = score_in_words(math_subject[0][value_item],max_mark=math_subject_maxMark) if math_subject else ''

                    # Ø§Ù„ØªØ±Ø¨ÙŠØ© Ø§Ù„Ø§Ø¬ØªÙ…Ø§Ø¹ÙŠØ© Ùˆ Ø§Ù„ÙˆØ·Ù†ÙŠØ© 
                    social_subjects = [value for key ,value in group_item['subject_sums'].items() if 'Ø§Ø¬ØªÙ…Ø§Ø¹ÙŠØ© Ùˆ Ø§Ù„ÙˆØ·Ù†ÙŠØ©' in key]
                    if 'Ø«Ø§Ù…Ù†' in group_item['student_grade_name'] or 'ØªØ§Ø³Ø¹' in group_item['student_grade_name'] or 'Ø¹Ø§Ø´Ø±' in group_item['student_grade_name']:
                        sheet2['C22'] = 200
                        social_subjects_maxMark = 200
                        sheet2['D22'] = int(social_subjects[0][value_item]*(2/3)) if social_subjects and len(social_subjects[0]) != 0 else ''
                    elif 'Ø³Ø§Ø¯Ø³' in group_item['student_grade_name'] or 'Ø³Ø§Ø¨Ø¹' in group_item['student_grade_name']:
                        sheet2['D22'] = int(social_subjects[0][value_item]/3) if social_subjects and len(social_subjects[0]) != 0 else ''
                        sheet2['C22'] = 100
                        social_subjects_maxMark = 100                
                    else:
                        sheet2['D22'] = social_subjects[0][value_item]
                        sheet2['C22'] = 100
                        social_subjects_maxMark = 100
                        
                    # sheet2['D22'] = social_subjects[0][value_item] if social_subjects and len(social_subjects[0]) != 0 else ''
                    sheet2['E22'] = social_subjects[0][1] if term2 and social_subjects and len(social_subjects[0]) != 0 else ''
                    sheet2['F22'] = (social_subjects[0][value_item] + social_subjects[0][1])/2 if term2 and social_subjects and len(social_subjects[0]) != 0 else ''
                    if term2:
                        sheet2['G22'] = convert_avarage_to_words((social_subjects[0][value_item] + social_subjects[0][1])/2) if social_subjects else ''
                        sheet2['J22'] = score_in_words(int(((social_subjects[0][value_item] + social_subjects[0][1])/2)*(2/3)),max_mark=social_subjects_maxMark) if social_subjects else ''
                    else:
                        sheet2['G22'] = convert_avarage_to_words(social_subjects[0][value_item]) if social_subjects else ''
                        sheet2['J22'] = score_in_words(int(social_subjects[0][value_item]*(2/3)),max_mark=social_subjects_maxMark) if social_subjects else ''

                    # Ø§Ù„Ø¹Ù„ÙˆÙ…
                    science_subjects = [value for key ,value in group_item['subject_sums'].items() if 'Ø¹Ù„ÙˆÙ…' in key]
                    if 'Ø«Ø§Ù…Ù†' in group_item['student_grade_name'] :
                        sheet2['C23'] = 200
                        science_subjects_maxMark = 200
                    elif 'ØªØ§Ø³Ø¹' in group_item['student_grade_name'] or 'Ø¹Ø§Ø´Ø±' in group_item['student_grade_name']:
                        sheet2['C23'] = 400
                        science_subjects_maxMark = 400
                    else:
                        sheet2['C23'] = 100
                        science_subjects_maxMark = 100
                    sheet2['D23'] = science_subjects[0][value_item] if science_subjects and len(science_subjects[0]) != 0 else ''
                    sheet2['E23'] = science_subjects[0][1] if term2 and science_subjects and len(science_subjects[0]) != 0 else ''
                    sheet2['F23'] = (science_subjects[0][value_item] + science_subjects[0][1])/2 if term2 and science_subjects and len(science_subjects[0]) != 0 else ''
                    if term2:
                        sheet2['G23'] = convert_avarage_to_words((science_subjects[0][value_item] + science_subjects[0][1])/2) if science_subjects else ''
                        sheet2['J23'] = score_in_words( ((science_subjects[0][value_item] + science_subjects[0][1])/2),max_mark=science_subjects_maxMark) if  science_subjects else ''
                    else:
                        sheet2['G23'] = convert_avarage_to_words(science_subjects[0][value_item]) if science_subjects else ''
                        sheet2['J23'] = score_in_words( science_subjects[0][value_item],max_mark=science_subjects_maxMark) if  science_subjects else ''

                    # Ø§Ù„ØªØ±Ø¨ÙŠØ© Ø§Ù„ÙÙ†ÙŠØ© ÙˆØ§Ù„Ù…ÙˆØ³ÙŠÙ‚ÙŠØ©
                    art_subject = [value for key ,value in group_item['subject_sums'].items() if 'Ø§Ù„ÙÙ†ÙŠØ© ÙˆØ§Ù„Ù…ÙˆØ³' in key]
                    sheet2['C24'] = 100 if art_subject and len(art_subject[0]) != 0 else ''
                    sheet2['D24'] = art_subject[0][value_item] if art_subject and len(art_subject[0]) != 0 else ''
                    sheet2['E24'] = art_subject[0][1] if term2 and art_subject and len(art_subject[0]) != 0 else ''
                    sheet2['F24'] = (art_subject[0][value_item] + art_subject[0][1])/2 if term2 and art_subject and len(art_subject[0]) != 0 else ''
                    if term2:
                        sheet2['G24'] = convert_avarage_to_words((art_subject[0][value_item] + art_subject[0][1])/2) if art_subject else ''
                        sheet2['J24'] = score_in_words(((art_subject[0][value_item] + art_subject[0][1])/2) ) if art_subject else ''
                    else:
                        sheet2['G24'] = convert_avarage_to_words(art_subject[0][value_item]) if art_subject else ''
                        sheet2['J24'] = score_in_words(art_subject[0][value_item] ) if art_subject else ''

                    # Ø§Ù„ØªØ±Ø¨ÙŠØ© Ø§Ù„Ø±ÙŠØ§Ø¶ÙŠØ©
                    sport_subject = [value for key ,value in group_item['subject_sums'].items() if 'Ø±ÙŠØ§Ø¶ÙŠØ©' in key]
                    sheet2['C25'] = 100 if sport_subject and len(sport_subject[0]) != 0 else ''
                    sheet2['D25'] = sport_subject[0][value_item] if sport_subject and len(sport_subject[0]) != 0 else ''
                    sheet2['E25'] = sport_subject[0][1] if term2 and sport_subject and len(sport_subject[0]) != 0 else ''
                    sheet2['F25'] = (sport_subject[0][value_item] + sport_subject[0][1])/2 if term2 and sport_subject and len(sport_subject[0]) != 0 else ''
                    if term2:
                        sheet2['G25'] = convert_avarage_to_words((sport_subject[0][value_item] + sport_subject[0][1])/2) if sport_subject else ''
                        sheet2['J25'] = score_in_words(((sport_subject[0][value_item] + sport_subject[0][1])/2) ) if sport_subject else ''
                    else:
                        sheet2['G25'] = convert_avarage_to_words(sport_subject[0][value_item]) if sport_subject else ''
                        sheet2['J25'] = score_in_words(sport_subject[0][value_item] ) if sport_subject else ''

                    # Ø§Ù„ØªØ±Ø¨ÙŠØ© Ø§Ù„Ù…Ù‡Ù†ÙŠØ© 
                    vocational_subject = [value for key ,value in group_item['subject_sums'].items() if 'Ù…Ù‡Ù†ÙŠØ©' in key]
                    sheet2['C26'] = 100 if vocational_subject and len(vocational_subject[0]) != 0 else ''
                    sheet2['D26'] = vocational_subject[0][value_item] if vocational_subject and len(vocational_subject[0]) != 0 else ''
                    sheet2['E26'] = vocational_subject[0][1] if term2 and vocational_subject and len(vocational_subject[0]) != 0 else ''
                    sheet2['F26'] = (vocational_subject[0][value_item] + vocational_subject[0][1])/2 if term2 and vocational_subject and len(vocational_subject[0]) != 0 else ''
                    if term2:
                        sheet2['G26'] = convert_avarage_to_words((vocational_subject[0][value_item] + vocational_subject[0][1])/2) if vocational_subject else ''
                        sheet2['J26'] = score_in_words(((vocational_subject[0][value_item] + vocational_subject[0][1])/2) ) if vocational_subject else ''
                    else:
                        sheet2['G26'] = convert_avarage_to_words(vocational_subject[0][value_item]) if vocational_subject else ''
                        sheet2['J26'] = score_in_words(vocational_subject[0][value_item] ) if vocational_subject else ''

                    # Ø§Ù„Ø­Ø§Ø³ÙˆØ¨
                    computer_subject = [value for key ,value in group_item['subject_sums'].items() if 'Ø­Ø§Ø³ÙˆØ¨' in key]
                    sheet2['C27'] = 100 if computer_subject and len(computer_subject[0]) != 0 else ''
                    sheet2['D27'] = computer_subject[0][value_item] if computer_subject and len(computer_subject[0]) != 0 else ''
                    sheet2['E27'] = computer_subject[0][1] if term2 and computer_subject and len(computer_subject[0]) != 0 else ''
                    sheet2['F27'] = (computer_subject[0][value_item] + computer_subject[0][1])/2 if term2 and computer_subject and len(computer_subject[0]) != 0 else ''
                    if term2:
                        sheet2['G27'] = convert_avarage_to_words((computer_subject[0][value_item] + computer_subject[0][1])/2) if computer_subject else ''
                        sheet2['J27'] = score_in_words(((computer_subject[0][value_item] + computer_subject[0][1])/2) ) if computer_subject else ''
                    else:
                        sheet2['G27'] = convert_avarage_to_words(computer_subject[0][value_item]) if computer_subject else ''
                        sheet2['J27'] = score_in_words(computer_subject[0][value_item] ) if computer_subject else ''

                    # Ø§Ù„Ø«Ù‚Ø§ÙØ© Ø§Ù„Ù…Ø§Ù„ÙŠØ©
                    financial_subject = [value for key ,value in group_item['subject_sums'].items() if 'Ù…Ø§Ù„ÙŠØ©' in key]
                    sheet2['C28'] = 100 if financial_subject and len(financial_subject[0]) != 0 else ''
                    sheet2['D28'] = financial_subject[0][value_item] if financial_subject and len(financial_subject[0]) != 0 else ''
                    sheet2['E28'] = financial_subject[0][1] if term2 and financial_subject and len(financial_subject[0]) != 0 else ''
                    sheet2['F28'] = (financial_subject[0][value_item] + financial_subject[0][1])/2 if term2 and financial_subject and len(financial_subject[0]) != 0 else ''
                    if term2:
                        sheet2['G28'] = convert_avarage_to_words((financial_subject[0][value_item] + financial_subject[0][1])/2) if financial_subject else ''
                        sheet2['J28'] = score_in_words(((financial_subject[0][value_item] + financial_subject[0][1])/2) ) if financial_subject else ''
                    else:
                        sheet2['G28'] = convert_avarage_to_words(financial_subject[0][value_item]) if financial_subject else ''
                        sheet2['J28'] = score_in_words(financial_subject[0][value_item] ) if financial_subject else ''

                    # Ø§Ù„Ù„ØºØ© Ø§Ù„ÙØ±Ù†Ø³ÙŠØ© 
                    franch_subject = [value for key ,value in group_item['subject_sums'].items() if 'ÙØ±Ù†Ø³ÙŠØ©' in key]
                    sheet2['C29'] = 100 if franch_subject and len(franch_subject[0]) != 0 else ''
                    sheet2['D29'] = franch_subject[0][value_item] if franch_subject and len(franch_subject[0]) != 0 else ''
                    sheet2['E29'] = franch_subject[0][1] if term2 and franch_subject and len(franch_subject[0]) != 0 else ''
                    sheet2['F29'] = (franch_subject[0][value_item] + franch_subject[0][1])/2 if term2 and franch_subject and len(franch_subject[0]) != 0 else ''
                    if term2:
                        sheet2['G29'] = convert_avarage_to_words((franch_subject[0][value_item] + franch_subject[0][1])/2) if franch_subject else ''
                        sheet2['J29'] = score_in_words(((franch_subject[0][value_item] + franch_subject[0][1])/2) ) if franch_subject else ''
                    else:
                        sheet2['G29'] = convert_avarage_to_words(franch_subject[0][value_item]) if franch_subject else ''
                        sheet2['J29'] = score_in_words(franch_subject[0][value_item] ) if franch_subject else ''

                    # Ø§Ù„Ø¯ÙŠÙ† Ø§Ù„Ù…Ø³ÙŠØ­ÙŠ
                    christian_subject = [value for key ,value in group_item['subject_sums'].items() if 'Ø§Ù„Ø¯ÙŠÙ† Ø§Ù„Ù…Ø³ÙŠØ­ÙŠ' in key]
                    sheet2['C30'] = 100 if christian_subject and len(christian_subject[0]) != 0 else ''
                    sheet2['D30'] = christian_subject[0][value_item] if christian_subject and len(christian_subject[0]) != 0 else ''
                    sheet2['E30'] = christian_subject[0][1] if term2 and christian_subject and len(christian_subject[0]) != 0 else ''
                    sheet2['F30'] = (christian_subject[0][value_item] + christian_subject[0][1])/2 if term2 and christian_subject and len(christian_subject[0]) != 0 else ''
                    if term2:
                        sheet2['G30'] = convert_avarage_to_words((christian_subject[0][value_item] + christian_subject[0][1])/2) if christian_subject else ''
                        sheet2['J30'] = score_in_words(((christian_subject[0][value_item] + christian_subject[0][1])/2) ) if christian_subject else ''
                    else:
                        sheet2['G30'] = convert_avarage_to_words(christian_subject[0][value_item]) if christian_subject else ''
                        sheet2['J30'] = score_in_words(christian_subject[0][value_item] ) if christian_subject else ''

                counter = 0
                for subject_name  ,S1_S2 in group_item['subject_sums'].items():
                    average = (S1_S2[0]+S1_S2[1])/2
                    print( subject_name, S1_S2)
                    if 'Ø³Ù„Ø§Ù…ÙŠ' in subject_name and average < islam_subject_maxMark / 2 : 
                        counter+=1
                    elif "Ø¹Ø±Ø¨ÙŠ"  in subject_name and average < arabic_subject_maxMark / 2 : 
                        counter+=1
                    elif "Ù†Ø¬Ù„ÙŠØ²ÙŠ"  in subject_name and average < english_subject_maxMark / 2 : 
                        counter+=1
                    elif "Ø±ÙŠØ§Ø¶ÙŠØ§Øª"  in subject_name and average < math_subject_maxMark / 2 : 
                        counter+=1
                    elif "Ø¬ØªÙ…Ø§Ø¹ÙŠØ©"  in subject_name and average < social_subjects_maxMark / 2 : 
                        counter+=1
                    elif "Ø¹Ù„ÙˆÙ…"  in subject_name and average < science_subjects_maxMark / 2 : 
                        counter+=1
                    elif  average < 50: 
                        counter+=1
                    # Ø·Ø±ÙŠÙ‚Ø© Ø·Ø¨Ø§Ø¹Ø© Ø§Ù„Ø±Ù‚Ù… ØµØ­ÙŠØ­ Ø§Ø°Ø§ ÙƒØ§Ù† Ø¨Ø¯ÙˆÙ† Ø§Ø¹Ø´Ø§Ø± 
                    # print(subject_name , int((S1_S2[0]+S1_S2[1])/2) if str((S1_S2[0]+S1_S2[1])/2).split('.')[1] == '0' else (S1_S2[0]+S1_S2[1])/2 )
                
                # print(counter)
                if counter > 4 : 
                    print('ÙŠØ¨Ù‚Ù‰ ÙÙŠ ØµÙÙ‡')
                    result = 2
                elif counter == 0 :
                    print("Ù†Ø§Ø¬Ø­")
                    result = 0
                else :     
                    print('Ù…Ù‚ØµØ±')
                    result = 1
                
                result_value = ['Ù†Ø§Ø¬Ø­' , 'Ù…Ù‚ØµØ±', "ÙŠØ¨Ù‚Ù‰ ÙÙŠ ØµÙÙ‡"]
                if term2 :
                    # Ø§Ù„Ù…Ø¹Ø¯Ù„ Ø§Ù„Ù…Ø¦ÙˆÙŠ Ø¨Ø§Ù„Ø±Ù‚Ø§Ù… 
                    sheet2['c32']= group_item['t1+t2+year_avarage'][2]
                    #Ø¨Ø§Ù„Ø­Ø±ÙˆÙ
                    sheet2['e32']= convert_avarage_to_words(group_item['t1+t2+year_avarage'][2]) if group_item else ''
                    #ØªØ±ØªÙŠØ¨ Ø§Ù„Ø·Ø§Ù„Ø¨ Ø¹Ù„Ù‰ Ø§Ù„ØµÙ 
                    sheet2['j32']= counter-1

                    #Ø§Ù„Ù†ØªÙŠØ¬Ø© 
                    sheet2['b33']= result_value[result]
                else:
                    
                    #Ø§Ù„Ù…Ø¹Ø¯Ù„ Ø§Ù„Ù…Ø¦ÙˆÙŠ Ø¨Ø§Ù„Ø±Ù‚Ø§Ù… 
                    sheet2[c['average']]= group_item['t1+t2+year_avarage'][term]
                    #Ø¨Ø§Ù„Ø­Ø±ÙˆÙ
                    sheet2['e32']= convert_avarage_to_words(group_item['t1+t2+year_avarage'][0]) if group_item else ''
                    #ØªØ±ØªÙŠØ¨ Ø§Ù„Ø·Ø§Ù„Ø¨ Ø¹Ù„Ù‰ Ø§Ù„ØµÙ 
                    sheet2['j32']= counter-1
                    #Ø§Ù„Ù†ØªÙŠØ¬Ø© 
                    if result == 2 : # Ø§Ø°Ø§ ÙƒØ§Ù† Ù…ÙƒÙ…Ù„ ÙÙŠ ØµÙÙ‡ Ø§Ù„ÙØµÙ„ Ø§Ù„Ø§ÙˆÙ„ Ø®Ù„ÙŠÙ‡Ø§ Ø§Ù„Ù‡ Ø¨Ø³ Ø±Ø§Ø³Ø¨ Ù„Ø§Ù†Ù‡ Ø¨Ø¬ÙˆØ² Ø§Ù„ÙØµÙ„ Ø§Ù„Ø«Ø§Ù†ÙŠ ÙŠØªØ­Ø³Ù† 
                        sheet2['b33']= 'Ù…Ù‚ØµØ±'
                    else:    
                        sheet2['b33']= "Ù†Ø§Ø¬Ø­"
                
                    # Ø¹Ø¯Ø¯ Ø§ÙŠØ§Ù… ØºÙŠØ§Ø¨ Ø§Ù„Ø·Ø§Ù„Ø¨ 
                    student_absent_count = [i for i in absent_list_with_names if i[0] == group_item['student_id'] ]
                    sheet2['c35']= student_absent_count[0][1] if len(student_absent_count) else 0
                    
                    #Ø¹Ø¯Ø¯ Ø§ÙŠØ§Ù… Ø§Ù„Ø¯ÙˆØ§Ù… Ø§Ù„Ø±Ø³Ù…ÙŠ Ø§Ù„ÙƒØ§Ù…Ù„ 
                    sheet2['g35']= ''
                    #Ø§Ø³Ù… Ùˆ ØªÙˆÙ‚ÙŠØ¹ Ù…Ø±Ø¨ÙŠ Ø§Ù„ØµÙ 
                    sheet2['j35']= ''
                    #Ø§Ù„ØªØ§Ø±ÙŠØ®
                    sheet2['b36']= ''
                    #Ø§Ø³Ù… Ùˆ ØªÙˆÙ‚ÙŠØ¹ Ù…Ø¯ÙŠØ± Ø§Ù„Ù…Ø¯Ø±Ø³Ø©
                    sheet2['i36']= ''
            except:
                pass
            template_file.remove(template_file['sheet'])
            template_file.save(outdir+group[0]['student_class_name_letter']+'.xlsx')

def create_coloured_certs_excel(grouped_list ,absent_list_with_names, term2=False ,template='./templet_files/Ù†Ù…ÙˆØ°Ø¬ Ø´Ù‡Ø§Ø¯Ø§Øª Ù…Ù„ÙˆÙ†Ø©.xlsx' , outdir='./send_folder/'):
    """
    The function creates coloured certificates in Excel format based on a grouped list of data.
    
    :param grouped_list: The `grouped_list` parameter is a list of dictionaries where each dictionary
    represents a group of data for which a certificate needs to be created. Each dictionary should
    contain the necessary information for creating a certificate, such as the recipient's name, date,
    and any other relevant details
    :param term2: A boolean value indicating whether the certificates are for the second term or not. If
    set to True, it means the certificates are for the second term. If set to False, it means the
    certificates are for the first term, defaults to False (optional)
    :param template: The `template` parameter is the path to the template file that contains the design
    and layout of the coloured certificates. It should be an Excel file (.xlsx) format, defaults to
    ./templet_files/Ù†Ù…ÙˆØ°Ø¬ Ø´Ù‡Ø§Ø¯Ø§Øª Ù…Ù„ÙˆÙ†Ø©.xlsx (optional)
    :param outdir: The `outdir` parameter specifies the directory where the generated certificates will
    be saved, defaults to ./send_folder/ (optional)
    """
    
    colored_cert_cells_position_context = {
                                        'student_name': 'E11',
                                        'student_name2': 'B44',
                                        'class_section': 'E12',
                                        'class_name':'B45',
                                        'nationality': 'D13',
                                        'national_id': 'E14',
                                        'birthplace_date': 'E15',
                                        'religion': 'E16',
                                        'student_address': 'E17',
                                        'school_name': 'E18',
                                        'school_address': 'E19',
                                        'directorate': 'D20',
                                        'school_bridge': 'D21',
                                        'supervising_authority': 'E22',
                                        'school_phone_number': 'E23',
                                        'academic_year_1': 'F42',
                                        'academic_year_2': 'G42',
                                        'class': 'B45',
                                        'islamic_education': 'E50',
                                        'arabic_language': 'E51',
                                        'english_language': 'E52',
                                        'mathematics': 'E53',
                                        'social_studies': 'E54',
                                        'science': 'E55',
                                        'visual_arts': 'E56',
                                        'physical_education': 'E57',
                                        'vocational_education': 'E58',
                                        'computer': 'E59',
                                        'financial_culture': 'E60',
                                        'french_language': 'E61',
                                        'christian_religion': 'E62',
                                        'average': 'B67',
                                        'school_days': 'I64',
                                        'class_teacher_name': 'B69',
                                        'principal_name': 'L69',
                                        'semester_1_student_absent': 'H66',
                                        'semester_2_student_absent': 'L66',
                                        }
    
    c = colored_cert_cells_position_context
    result_cell_positions = ['B64','B65','B66']
    
    statistic_data = grouped_list['students_info']
    assessments_data_groups = grouped_list['assessments_data']
    term = 1 if term2 else 0
    
    
    for group in assessments_data_groups:
        if not any("Ø¹Ø´Ø±" in i['student_grade_name'] for i in group) : 
            template_file = load_workbook(template)
            sheet1 = template_file.worksheets[0]
            
            names_averages =  sort_dictionary_list_based_on(group,item_in_list=term)

            group = sort_dictionary_list_based_on(group ,simple=False,item_in_list=term)

            for row_number, dataFrame in enumerate(names_averages, start=5):
                sheet1.cell(row=row_number, column=2).value = dataFrame[1]
                sheet1.cell(row=row_number, column=4).value = dataFrame[0]
                
            counter = 1
            for group_item in group:

                sheet2 = template_file.copy_worksheet(template_file.worksheets[1])
                sheet2.title = str(counter)
                counter += 1
                sheet2.sheet_view.rightToLeft = True    
                sheet2.sheet_view.rightToLeft = True   
                
                wanted_id = group_item['student_id']
                student_statistic_info = [i for i in statistic_data if i['student_id'] == wanted_id ][0]

                # img = openpyxl.drawing.image.Image(image)
                # img.anchor = 'e2'
                # sheet2.add_image(img)

                # print(sheet2)
                sheet2[c['student_name']] = group_item['student__full_name']
                # Ù…ÙƒØ§Ù† Ùˆ ØªØ§Ø±ÙŠØ® Ø§Ù„ÙˆÙ„Ø§Ø¯Ø©
                sheet2[c['birthplace_date']]= str(group_item['student_birth_place']) + ' ' + str(group_item['student_birth_date'])
                #Ø§Ù„Ø±Ù‚Ù… Ø§Ù„ÙˆØ·Ù†ÙŠ
                sheet2[c['national_id']]= group_item['student_nat_id']
                #Ø§Ù„Ø¬Ù†Ø³ÙŠØ©
                sheet2[c['nationality']]= group_item['student_nat']
                #Ø§Ù„ØµÙ Ùˆ Ø§Ù„Ø´Ø¹Ø¨Ø© 
                sheet2[c['class_section']]= group_item['student_class_name_letter']
                # ØµÙ Ø§Ù„Ø·Ø§Ù„Ø¨
                sheet2[c['class_name']]= group_item['student_grade_name'].replace('Ø§Ù„ØµÙ' ,'')
                #Ø§Ù„Ù…Ø¯Ø±Ø³Ø© Ùˆ Ø±Ù‚Ù…Ù‡Ø§ Ø§Ù„ÙˆØ·Ù†ÙŠ
                sheet2[c['school_name']]= group_item['student_school_name']
                #Ø§Ù„Ù…Ù†Ø·Ù‚Ø© Ø§Ù„ØªØ¹Ù„ÙŠÙ…ÙŠØ©  Ø§Ùˆ Ø§Ù„Ø³Ù„Ø·Ø© Ø§Ù„Ù…Ø´Ø±ÙØ©
                sheet2[c['supervising_authority']]= 'ÙˆØ²Ø§Ø±Ø© Ø§Ù„ØªØ±Ø¨ÙŠØ© Ùˆ Ø§Ù„ØªØ¹Ù„ÙŠÙ…'
                # Ù…Ø¯ÙŠØ±ÙŠØ© Ø§Ù„Ù…Ø¯Ø±Ø³Ø© 
                sheet2[c['directorate']]= grouped_list['school_directorate']
                # Ù„ÙˆØ§Ø¡ Ø§Ù„Ù…Ø¯Ø±Ø³Ø©
                sheet2[c['school_bridge']]= grouped_list['school_bridge']
                # Ø§Ù„Ø¯ÙŠØ§Ù†Ø©
                sheet2[c['religion']]= student_statistic_info['religion']
                # Ø¹Ù†ÙˆØ§Ù† Ø§Ù„Ø·Ø§Ù„Ø¨
                sheet2[c['student_address']]= student_statistic_info['address']
                # Ø¹Ù†ÙˆØ§Ù† Ø§Ù„Ù…Ø¯Ø±Ø³Ø©
                sheet2[c['school_address']]= grouped_list['school_address']
                # Ø§Ù„Ø¹Ø§Ù… Ø§Ù„Ø¯Ø±Ø§Ø³ÙŠ Ø§Ù„Ø§ÙˆÙ„ 
                sheet2[c['academic_year_1']]= grouped_list['academic_year_1']
                # Ø§Ù„Ø¹Ø§Ù… Ø§Ù„Ø¯Ø±Ø§Ø³ÙŠ Ø§Ù„Ø«Ø§Ù†ÙŠ 
                sheet2[c['academic_year_2']]= grouped_list['academic_year_2']
                # Ø§Ù„Ø§Ø³Ù… Ø¹Ù„Ù‰ Ø§Ù„ÙˆØ¬Ù‡ Ø§Ù„Ø«Ø§Ù†ÙŠ
                sheet2[c['student_name2']] = group_item['student__full_name']
                # ØºÙŠØ§Ø¨ Ø§Ù„ÙØµÙ„ Ø§Ù„Ø§ÙˆÙ„ 
                sheet2[c['semester_1_student_absent']]= ''
                # ØºÙŠØ§Ø¨ Ø§Ù„ÙØµÙ„ Ø§Ù„Ø«Ø§Ù†ÙŠ
                sheet2[c['semester_2_student_absent']]= ''
                
                # put the subjects cells inder here 
                i ='c,d,e,g,j,f'.split(',')
                r = range(18,32)


                # Ø§Ù„ØªØ±Ø¨ÙŠØ© Ø§Ù„Ø§Ø³Ù„Ø§Ù…ÙŠØ©
                islam_subject = [value for key ,value in group_item['subject_sums'].items() if 'Ø³Ù„Ø§Ù…ÙŠØ©' in key]
                if 'Ø«Ø§Ù…Ù†' in group_item['student_grade_name'] or 'ØªØ§Ø³Ø¹' in group_item['student_grade_name'] or 'Ø¹Ø§Ø´Ø±' in group_item['student_grade_name']:
                    islam_subject_maxMark = 200
                    sheet2[c['islamic_education']] = islam_subject_maxMark /2
                    sheet2['F'+c['islamic_education'][1:]] = islam_subject_maxMark
                else:
                    islam_subject_maxMark = 100
                    sheet2[c['islamic_education']] = islam_subject_maxMark /2
                    sheet2['F'+c['islamic_education'][1:]] = islam_subject_maxMark
                sheet2['H'+c['islamic_education'][1:]] = islam_subject[0][0] if islam_subject and len(islam_subject[0]) != 0 else ''
                sheet2['I'+c['islamic_education'][1:]] = islam_subject[0][1] if term2 and islam_subject and len(islam_subject[0]) != 0 else ''
                sheet2['K'+c['islamic_education'][1:]] = (islam_subject[0][0] + islam_subject[0][1])/2 if term2 and islam_subject and len(islam_subject[0]) != 0 else ''
                # if term2:
                #     sheet2['G'+c['islamic_education']] = convert_avarage_to_words((islam_subject[0][0] + islam_subject[0][1])/2) if islam_subject else ''
                #     sheet2['J'+c['islamic_education']] = score_in_words(((islam_subject[0][0] + islam_subject[0][1])/2),max_mark=maxMark) if islam_subject else ''
                # else:
                #     sheet2['G'+c['islamic_education']] = convert_avarage_to_words(islam_subject[0][0]) if islam_subject else ''
                #     sheet2['J'+c['islamic_education']] = score_in_words(islam_subject[0][0],max_mark=maxMark) if islam_subject else ''

                # Ø§Ù„Ù„ØºØ© Ø§Ù„Ø¹Ø±Ø¨ÙŠØ©
                arabic_subject = [value for key ,value in group_item['subject_sums'].items() if 'Ø¹Ø±Ø¨ÙŠ' in key]
                if 'Ø«Ø§Ù…Ù†' in group_item['student_grade_name'] or 'ØªØ§Ø³Ø¹' in group_item['student_grade_name'] or 'Ø¹Ø§Ø´Ø±' in group_item['student_grade_name']:
                    arabic_subject_maxMark = 300
                    sheet2[c['arabic_language']] = arabic_subject_maxMark / 2
                    sheet2['F'+c['arabic_language'][1:]] = arabic_subject_maxMark
                else:
                    arabic_subject_maxMark = 100
                    sheet2[c['arabic_language']] = arabic_subject_maxMark / 2
                    sheet2['F'+c['arabic_language'][1:]] = arabic_subject_maxMark
                sheet2['H' + c['arabic_language'][1:]] = arabic_subject[0][0] if arabic_subject and len(arabic_subject[0]) != 0 else ''
                sheet2['I' + c['arabic_language'][1:]] = arabic_subject[0][1] if term2 and arabic_subject and len(arabic_subject[0]) != 0 else ''
                sheet2['K' + c['arabic_language'][1:]] = (arabic_subject[0][0] + arabic_subject[0][1])/2 if term2 and arabic_subject and len(arabic_subject[0]) != 0 else ''
                # if term2:
                #     sheet2['J19'] = score_in_words(((arabic_subject[0][0] + arabic_subject[0][1])/2),max_mark=maxMark) if arabic_subject else ''
                #     sheet2['F'+] = maxMark / 2
                #     sheet2[] = convert_avarage_to_words((arabic_subject[0][0] + arabic_subject[0][1])/2) if arabic_subject else maxMark
                # else:
                #     sheet2['J19'] = score_in_words(arabic_subject[0][0],max_mark=maxMark) if arabic_subject else ''
                #     sheet2['F'+] = maxMark / 2
                #     sheet2[] = convert_avarage_to_words(arabic_subject[0][0]) if arabic_subject else maxMark

                # Ø§Ù„Ù„ØºØ© Ø§Ù„Ø§Ù†Ø¬Ù„ÙŠØ²ÙŠØ© 
                english_subject = [value for key ,value in group_item['subject_sums'].items() if 'Ø¬Ù„ÙŠØ²ÙŠØ©' in key]
                if 'Ø«Ø§Ù…Ù†' in group_item['student_grade_name'] or 'ØªØ§Ø³Ø¹' in group_item['student_grade_name'] or 'Ø¹Ø§Ø´Ø±' in group_item['student_grade_name']:
                    english_subject_maxMark = 200
                    sheet2[c['english_language']] = english_subject_maxMark / 2
                    sheet2['F'+c['english_language'][1:]] = english_subject_maxMark
                else:
                    english_subject_maxMark = 100
                    sheet2[c['english_language']] = english_subject_maxMark / 2
                    sheet2['F'+c['english_language'][1:]] = english_subject_maxMark
                sheet2['H'+c['english_language'][1:]] = english_subject[0][0] if english_subject and len(english_subject[0]) != 0 else ''
                sheet2['I'+c['english_language'][1:]] = english_subject[0][1] if term2 and english_subject and len(english_subject[0]) != 0 else ''
                sheet2['K'+c['english_language'][1:]] = (english_subject[0][0] + english_subject[0][1])/2 if term2 and english_subject and len(english_subject[0]) != 0 else ''
                # if term2:
                #     sheet2['J20'] = score_in_words(((english_subject[0][0] + english_subject[0][1])/2),max_mark=maxMark) if english_subject else ''
                #     sheet2['F'+c[''][1:]] = maxMark / 2
                #     sheet2[] = convert_avarage_to_words((english_subject[0][0] + english_subject[0][1])/2) if english_subject else maxMark
                # else:
                #     sheet2['J20'] = score_in_words(english_subject[0][0],max_mark=maxMark) if english_subject else ''
                #     sheet2['F'+c[''][1:]] = maxMark / 2
                #     sheet2[] = convert_avarage_to_words(english_subject[0][0]) if english_subject else maxMark

                # Ø§Ù„Ø±ÙŠØ§Ø¶ÙŠØ§Øª 
                math_subject = [value for key ,value in group_item['subject_sums'].items() if 'Ø±ÙŠØ§Ø¶ÙŠØ§Øª' in key]
                if 'Ø«Ø§Ù…Ù†' in group_item['student_grade_name'] or 'ØªØ§Ø³Ø¹' in group_item['student_grade_name'] or 'Ø¹Ø§Ø´Ø±' in group_item['student_grade_name']:
                    math_subject_maxMark = 200
                    sheet2[c['mathematics']] = math_subject_maxMark / 2
                    sheet2['F'+c['mathematics'][1:]] = math_subject_maxMark
                else:
                    math_subject_maxMark = 100
                    sheet2[c['mathematics']] = math_subject_maxMark / 2
                    sheet2['F'+c['mathematics'][1:]] = math_subject_maxMark
                sheet2['H'+c['mathematics'][1:]] = math_subject[0][0] if math_subject and len(math_subject[0]) != 0 else ''
                sheet2['I'+c['mathematics'][1:]] = math_subject[0][1] if term2 and math_subject and len(math_subject[0]) != 0 else ''
                sheet2['K'+c['mathematics'][1:]] = (math_subject[0][0] + math_subject[0][1])/2 if term2 and math_subject and len(math_subject[0]) != 0 else ''
                # if term2:
                #     sheet2['J21'] = score_in_words(((math_subject[0][0] + math_subject[0][1])/2),max_mark=maxMark) if math_subject else ''
                #     sheet2['F'+c[''][1:]] = maxMark / 2
                #     sheet2[] = convert_avarage_to_words((math_subject[0][0] + math_subject[0][1])/2) if math_subject else maxMark
                # else:
                #     sheet2['J21'] = score_in_words(math_subject[0][0],max_mark=maxMark) if math_subject else ''
                #     sheet2['F'+c[''][1:]] = maxMark / 2
                #     sheet2[] = convert_avarage_to_words(math_subject[0][0]) if math_subject else maxMark

                # Ø§Ù„ØªØ±Ø¨ÙŠØ© Ø§Ù„Ø§Ø¬ØªÙ…Ø§Ø¹ÙŠØ© Ùˆ Ø§Ù„ÙˆØ·Ù†ÙŠØ© 
                social_subjects = [value for key ,value in group_item['subject_sums'].items() if 'Ø§Ø¬ØªÙ…Ø§Ø¹ÙŠØ© Ùˆ Ø§Ù„ÙˆØ·Ù†ÙŠØ©' in key]
                if 'Ø«Ø§Ù…Ù†' in group_item['student_grade_name'] or 'ØªØ§Ø³Ø¹' in group_item['student_grade_name'] or 'Ø¹Ø§Ø´Ø±' in group_item['student_grade_name']:
                    social_subjects_maxMark = 200
                    sheet2[c['social_studies']] = social_subjects_maxMark / 2
                    sheet2['F'+c['social_studies'][1:]] = social_subjects_maxMark
                    sheet2['H'+c['social_studies'][1:]] = int(social_subjects[0][0]*(2/3)) if social_subjects and len(social_subjects[0]) != 0 else ''
                    sheet2['I'+c['social_studies'][1:]] = int(social_subjects[0][1]*(2/3)) if term2 and social_subjects and len(social_subjects[0]) != 0 else ''
                elif 'Ø³Ø§Ø¯Ø³' in group_item['student_grade_name'] or 'Ø³Ø§Ø¨Ø¹' in group_item['student_grade_name']:
                    social_subjects_maxMark = 100                
                    sheet2[c['social_studies']] = social_subjects_maxMark / 2
                    sheet2['F'+c['social_studies'][1:]] = social_subjects_maxMark
                    sheet2['H'+c['social_studies'][1:]] = int(social_subjects[0][0]/3) if social_subjects and len(social_subjects[0]) != 0 else ''
                    sheet2['I'+c['social_studies'][1:]] = int(social_subjects[0][1]*(2/3)) if term2 and social_subjects and len(social_subjects[0]) != 0 else ''
                else:
                    social_subjects_maxMark = 100
                    sheet2[c['social_studies']] = social_subjects_maxMark / 2
                    sheet2['F'+c['social_studies'][1:]] = social_subjects_maxMark
                    sheet2['H'+c['social_studies'][1:]] = int(social_subjects[0][0]) if social_subjects and len(social_subjects[0]) != 0 else ''
                    sheet2['I'+c['social_studies'][1:]] = int(social_subjects[0][1]*(2/3)) if term2 and social_subjects and len(social_subjects[0]) != 0 else ''
                    
                sheet2['K'+c['social_studies'][1:]] = (social_subjects[0][0] + social_subjects[0][1])/2 if term2 and social_subjects and len(social_subjects[0]) != 0 else ''
                # if term2:
                #     sheet2['J22'] = score_in_words(int(((social_subjects[0][0] + social_subjects[0][1])/2)*(2/3)),max_mark=maxMark) if social_subjects else ''
                #     sheet2['F'+c[''][1:]] = maxMark / 2
                #     sheet2[] = convert_avarage_to_words((social_subjects[0][0] + social_subjects[0][1])/2) if social_subjects else maxMark
                # else:
                #     sheet2['J22'] = score_in_words(int(social_subjects[0][0]*(2/3)),max_mark=maxMark) if social_subjects else ''
                #     sheet2['F'+c[''][1:]] = maxMark / 2
                #     sheet2[] = convert_avarage_to_words(social_subjects[0][0]) if social_subjects else maxMark

                # Ø§Ù„Ø¹Ù„ÙˆÙ…
                science_subjects = [value for key ,value in group_item['subject_sums'].items() if 'Ø§Ù„Ø¹Ù„ÙˆÙ…' in key]
                if 'Ø«Ø§Ù…Ù†' in group_item['student_grade_name'] :
                    science_subjects_maxMark = 200
                    sheet2[c['science']] = science_subjects_maxMark / 2
                    sheet2['F'+c['science'][1:]] = science_subjects_maxMark
                elif 'ØªØ§Ø³Ø¹' in group_item['student_grade_name'] or 'Ø¹Ø§Ø´Ø±' in group_item['student_grade_name']:
                    science_subjects_maxMark = 400
                    sheet2[c['science']] = science_subjects_maxMark / 2
                    sheet2['F'+c['science'][1:]] = science_subjects_maxMark
                else:
                    science_subjects_maxMark = 100
                    sheet2[c['science']] = science_subjects_maxMark / 2
                    sheet2['F'+c['science'][1:]] = science_subjects_maxMark
                sheet2['H'+c['science'][1:]] = science_subjects[0][0] if science_subjects and len(science_subjects[0]) != 0 else ''
                sheet2['I'+c['science'][1:]] = science_subjects[0][1] if term2 and science_subjects and len(science_subjects[0]) != 0 else ''
                sheet2['K'+c['science'][1:]] = (science_subjects[0][0] + science_subjects[0][1])/2 if term2 and science_subjects and len(science_subjects[0]) != 0 else ''
                # if term2:
                #     sheet2['J23'] = score_in_words( ((science_subjects[0][0] + science_subjects[0][1])/2),max_mark=maxMark) if  science_subjects else ''
                #     sheet2['F'+c[''][1:]] = maxMark / 2
                #     sheet2[] = convert_avarage_to_words((science_subjects[0][0] + science_subjects[0][1])/2) if science_subjects else maxMark
                # else:
                #     sheet2['J23'] = score_in_words( science_subjects[0][0],max_mark=maxMark) if  science_subjects else ''
                #     sheet2['F'+c[''][1:]] = maxMark / 2
                #     sheet2[] = convert_avarage_to_words(science_subjects[0][0]) if science_subjects else maxMark

                # Ø§Ù„ØªØ±Ø¨ÙŠØ© Ø§Ù„ÙÙ†ÙŠØ© ÙˆØ§Ù„Ù…ÙˆØ³ÙŠÙ‚ÙŠØ©
                art_subject = [value for key ,value in group_item['subject_sums'].items() if 'Ø§Ù„ÙÙ†ÙŠØ© ÙˆØ§Ù„Ù…ÙˆØ³' in key]
                sheet2[c['visual_arts']] = 50 if art_subject and len(art_subject[0]) != 0 else ''
                sheet2['F'+c['visual_arts'][1:]] = 100 if art_subject and len(art_subject[0]) != 0 else ''
                sheet2['H'+c['visual_arts'][1:]] = art_subject[0][0] if art_subject and len(art_subject[0]) != 0 else ''
                sheet2['I'+c['visual_arts'][1:]] = art_subject[0][1] if term2 and art_subject and len(art_subject[0]) != 0 else ''
                sheet2['K'+c['visual_arts'][1:]] = (art_subject[0][0] + art_subject[0][1])/2 if term2 and art_subject and len(art_subject[0]) != 0 else ''
                # if term2:
                #     sheet2['G24'] = convert_avarage_to_words((art_subject[0][0] + art_subject[0][1])/2) if art_subject else ''
                #     sheet2['J24'] = score_in_words(((art_subject[0][0] + art_subject[0][1])/2) ) if art_subject else ''
                # else:
                #     sheet2['G24'] = convert_avarage_to_words(art_subject[0][0]) if art_subject else ''
                #     sheet2['J24'] = score_in_words(art_subject[0][0] ) if art_subject else ''

                # Ø§Ù„ØªØ±Ø¨ÙŠØ© Ø§Ù„Ø±ÙŠØ§Ø¶ÙŠØ©
                sport_subject = [value for key ,value in group_item['subject_sums'].items() if 'Ø±ÙŠØ§Ø¶ÙŠØ©' in key]
                sheet2[c['physical_education']] = 50 if sport_subject and len(sport_subject[0]) != 0 else ''
                sheet2['F'+c['physical_education'][1:]] = 100 if sport_subject and len(sport_subject[0]) != 0 else ''
                sheet2['H'+c['physical_education'][1:]] = sport_subject[0][0] if sport_subject and len(sport_subject[0]) != 0 else ''
                sheet2['I'+c['physical_education'][1:]] = sport_subject[0][1] if term2 and sport_subject and len(sport_subject[0]) != 0 else ''
                sheet2['K'+c['physical_education'][1:]] = (sport_subject[0][0] + sport_subject[0][1])/2 if term2 and sport_subject and len(sport_subject[0]) != 0 else ''
                # if term2:
                #     sheet2['G25'] = convert_avarage_to_words((sport_subject[0][0] + sport_subject[0][1])/2) if sport_subject else ''
                #     sheet2['J25'] = score_in_words(((sport_subject[0][0] + sport_subject[0][1])/2) ) if sport_subject else ''
                # else:
                #     sheet2['G25'] = convert_avarage_to_words(sport_subject[0][0]) if sport_subject else ''
                #     sheet2['J25'] = score_in_words(sport_subject[0][0] ) if sport_subject else ''

                # Ø§Ù„ØªØ±Ø¨ÙŠØ© Ø§Ù„Ù…Ù‡Ù†ÙŠØ© 
                vocational_subject = [value for key ,value in group_item['subject_sums'].items() if 'Ù…Ù‡Ù†ÙŠØ©' in key]
                sheet2[c['vocational_education']] = 50 if vocational_subject and len(vocational_subject[0]) != 0 else ''
                sheet2['F'+c['vocational_education'][1:]] = 100 if vocational_subject and len(vocational_subject[0]) != 0 else ''
                sheet2['H'+c['vocational_education'][1:]] = vocational_subject[0][0] if vocational_subject and len(vocational_subject[0]) != 0 else ''
                sheet2['I'+c['vocational_education'][1:]] = vocational_subject[0][1] if term2 and vocational_subject and len(vocational_subject[0]) != 0 else ''
                sheet2['K'+c['vocational_education'][1:]] = (vocational_subject[0][0] + vocational_subject[0][1])/2 if term2 and vocational_subject and len(vocational_subject[0]) != 0 else ''
                # if term2:
                #     sheet2['G26'] = convert_avarage_to_words((vocational_subject[0][0] + vocational_subject[0][1])/2) if vocational_subject else ''
                #     sheet2['J26'] = score_in_words(((vocational_subject[0][0] + vocational_subject[0][1])/2) ) if vocational_subject else ''
                # else:
                #     sheet2['G26'] = convert_avarage_to_words(vocational_subject[0][0]) if vocational_subject else ''
                #     sheet2['J26'] = score_in_words(vocational_subject[0][0] ) if vocational_subject else ''

                # Ø§Ù„Ø­Ø§Ø³ÙˆØ¨
                computer_subject = [value for key ,value in group_item['subject_sums'].items() if 'Ø­Ø§Ø³ÙˆØ¨' in key]
                sheet2[c['computer']] = 50 if computer_subject and len(computer_subject[0]) != 0 else ''
                sheet2['F'+c['computer'][1:]] = 100 if computer_subject and len(computer_subject[0]) != 0 else ''
                sheet2['H'+c['computer'][1:]] = computer_subject[0][0] if computer_subject and len(computer_subject[0]) != 0 else ''
                sheet2['I'+c['computer'][1:]] = computer_subject[0][1] if term2 and computer_subject and len(computer_subject[0]) != 0 else ''
                sheet2['K'+c['computer'][1:]] = (computer_subject[0][0] + computer_subject[0][1])/2 if term2 and computer_subject and len(computer_subject[0]) != 0 else ''
                # if term2:
                #     sheet2['G27'] = convert_avarage_to_words((computer_subject[0][0] + computer_subject[0][1])/2) if computer_subject else ''
                #     sheet2['J27'] = score_in_words(((computer_subject[0][0] + computer_subject[0][1])/2) ) if computer_subject else ''
                # else:
                #     sheet2['G27'] = convert_avarage_to_words(computer_subject[0][0]) if computer_subject else ''
                #     sheet2['J27'] = score_in_words(computer_subject[0][0] ) if computer_subject else ''

                # Ø§Ù„Ø«Ù‚Ø§ÙØ© Ø§Ù„Ù…Ø§Ù„ÙŠØ©
                financial_subject = [value for key ,value in group_item['subject_sums'].items() if 'Ù…Ø§Ù„ÙŠØ©' in key]
                sheet2[c['financial_culture']] = 50 if financial_subject and len(financial_subject[0]) != 0 else ''
                sheet2['F'+c['financial_culture'][1:]] = 100 if financial_subject and len(financial_subject[0]) != 0 else ''
                sheet2['H'+c['financial_culture'][1:]] = financial_subject[0][0] if financial_subject and len(financial_subject[0]) != 0 else ''
                sheet2['I'+c['financial_culture'][1:]] = financial_subject[0][1] if term2 and financial_subject and len(financial_subject[0]) != 0 else ''
                sheet2['K'+c['financial_culture'][1:]] = (financial_subject[0][0] + financial_subject[0][1])/2 if term2 and financial_subject and len(financial_subject[0]) != 0 else ''
                # if term2:
                #     sheet2['G28'] = convert_avarage_to_words((financial_subject[0][0] + financial_subject[0][1])/2) if financial_subject else ''
                #     sheet2['J28'] = score_in_words(((financial_subject[0][0] + financial_subject[0][1])/2) ) if financial_subject else ''
                # else:
                #     sheet2['G28'] = convert_avarage_to_words(financial_subject[0][0]) if financial_subject else ''
                #     sheet2['J28'] = score_in_words(financial_subject[0][0] ) if financial_subject else ''

                # Ø§Ù„Ù„ØºØ© Ø§Ù„ÙØ±Ù†Ø³ÙŠØ© 
                franch_subject = [value for key ,value in group_item['subject_sums'].items() if 'ÙØ±Ù†Ø³ÙŠØ©' in key]
                sheet2[c['french_language']] = 50 if franch_subject and len(franch_subject[0]) != 0 else ''
                sheet2['F'+c['french_language'][1:]] = 100 if franch_subject and len(franch_subject[0]) != 0 else ''
                sheet2['H'+c['french_language'][1:]] = franch_subject[0][0] if franch_subject and len(franch_subject[0]) != 0 else ''
                sheet2['I'+c['french_language'][1:]] = franch_subject[0][1] if term2 and franch_subject and len(franch_subject[0]) != 0 else ''
                sheet2['K'+c['french_language'][1:]] = (franch_subject[0][0] + franch_subject[0][1])/2 if term2 and franch_subject and len(franch_subject[0]) != 0 else ''
                # if term2:
                #     sheet2['G29'] = convert_avarage_to_words((franch_subject[0][0] + franch_subject[0][1])/2) if franch_subject else ''
                #     sheet2['J29'] = score_in_words(((franch_subject[0][0] + franch_subject[0][1])/2) ) if franch_subject else ''
                # else:
                #     sheet2['G29'] = convert_avarage_to_words(franch_subject[0][0]) if franch_subject else ''
                #     sheet2['J29'] = score_in_words(franch_subject[0][0] ) if franch_subject else ''

                # Ø§Ù„Ø¯ÙŠÙ† Ø§Ù„Ù…Ø³ÙŠØ­ÙŠ
                christian_subject = [value for key ,value in group_item['subject_sums'].items() if 'Ø§Ù„Ø¯ÙŠÙ† Ø§Ù„Ù…Ø³ÙŠØ­ÙŠ' in key]
                sheet2[c['christian_religion']] = 50 if christian_subject and len(christian_subject[0]) != 0 else ''
                sheet2['F'+c['christian_religion'][1:]] = 100 if christian_subject and len(christian_subject[0]) != 0 else ''
                sheet2['H'+c['christian_religion'][1:]] = christian_subject[0][0] if christian_subject and len(christian_subject[0]) != 0 else ''
                sheet2['I'+c['christian_religion'][1:]] = christian_subject[0][1] if term2 and christian_subject and len(christian_subject[0]) != 0 else ''
                sheet2['K'+c['christian_religion'][1:]] = (christian_subject[0][0] + christian_subject[0][1])/2 if term2 and christian_subject and len(christian_subject[0]) != 0 else ''
                # if term2:
                #     sheet2['G30'] = convert_avarage_to_words((christian_subject[0][0] + christian_subject[0][1])/2) if christian_subject else ''
                #     sheet2['J30'] = score_in_words(((christian_subject[0][0] + christian_subject[0][1])/2) ) if christian_subject else ''
                # else:
                #     sheet2['G30'] = convert_avarage_to_words(christian_subject[0][0]) if christian_subject else ''
                #     sheet2['J30'] = score_in_words(christian_subject[0][0] ) if christian_subject else ''

                counter = 0
                for subject_name  ,S1_S2 in group_item['subject_sums'].items():
                    average = (S1_S2[0]+S1_S2[1])/2
                    print( subject_name, S1_S2)
                    if 'Ø³Ù„Ø§Ù…ÙŠ' in subject_name and average < islam_subject_maxMark / 2 : 
                        counter+=1
                    elif "Ø¹Ø±Ø¨ÙŠ"  in subject_name and average < arabic_subject_maxMark / 2 : 
                        counter+=1
                    elif "Ù†Ø¬Ù„ÙŠØ²ÙŠ"  in subject_name and average < english_subject_maxMark / 2 : 
                        counter+=1
                    elif "Ø±ÙŠØ§Ø¶ÙŠØ§Øª"  in subject_name and average < math_subject_maxMark / 2 : 
                        counter+=1
                    elif "Ø¬ØªÙ…Ø§Ø¹ÙŠØ©"  in subject_name and average < social_subjects_maxMark / 2 : 
                        counter+=1
                    elif "Ø¹Ù„ÙˆÙ…"  in subject_name and average < science_subjects_maxMark / 2 : 
                        counter+=1
                    elif  average < 50: 
                        counter+=1
                    # Ø·Ø±ÙŠÙ‚Ø© Ø·Ø¨Ø§Ø¹Ø© Ø§Ù„Ø±Ù‚Ù… ØµØ­ÙŠØ­ Ø§Ø°Ø§ ÙƒØ§Ù† Ø¨Ø¯ÙˆÙ† Ø§Ø¹Ø´Ø§Ø± 
                    # print(subject_name , int((S1_S2[0]+S1_S2[1])/2) if str((S1_S2[0]+S1_S2[1])/2).split('.')[1] == '0' else (S1_S2[0]+S1_S2[1])/2 )
                
                # print(counter)
                if counter > 4 : 
                    print('ÙŠØ¨Ù‚Ù‰ ÙÙŠ ØµÙÙ‡')
                    result = 2
                elif counter == 0 :
                    print("Ù†Ø§Ø¬Ø­")
                    result = 0
                else :     
                    print('Ù…ÙƒÙ…Ù„')
                    result = 1
                
                if term2 :
                    # Ø§Ù„Ù…Ø¹Ø¯Ù„ Ø§Ù„Ù…Ø¦ÙˆÙŠ Ø¨Ø§Ù„Ø±Ù‚Ø§Ù… 
                    sheet2[c['average']]= group_item['t1+t2+year_avarage'][2]
                    # #Ø¨Ø§Ù„Ø­Ø±ÙˆÙ
                    # sheet2['e32']= convert_avarage_to_words(group_item['t1+t2+year_avarage'][2]) if group_item else ''
                    # #ØªØ±ØªÙŠØ¨ Ø§Ù„Ø·Ø§Ù„Ø¨ Ø¹Ù„Ù‰ Ø§Ù„ØµÙ 
                    # sheet2['j32']= counter-1

                    #Ø§Ù„Ù†ØªÙŠØ¬Ø© 
                    sheet2[result_cell_positions[result]]= 'âœ“'
                else:
                    
                    #Ø§Ù„Ù…Ø¹Ø¯Ù„ Ø§Ù„Ù…Ø¦ÙˆÙŠ Ø¨Ø§Ù„Ø±Ù‚Ø§Ù… 
                    sheet2[c['average']]= group_item['t1+t2+year_avarage'][term]
                    # #Ø¨Ø§Ù„Ø­Ø±ÙˆÙ
                    # sheet2['e32']= convert_avarage_to_words(group_item['t1+t2+year_avarage'][0]) if group_item else ''
                    # #ØªØ±ØªÙŠØ¨ Ø§Ù„Ø·Ø§Ù„Ø¨ Ø¹Ù„Ù‰ Ø§Ù„ØµÙ 
                    # sheet2['j32']= counter-1
                    # #Ø§Ù„Ù†ØªÙŠØ¬Ø© 
                    if result == 2 : # Ø§Ø°Ø§ ÙƒØ§Ù† Ù…ÙƒÙ…Ù„ ÙÙŠ ØµÙÙ‡ Ø§Ù„ÙØµÙ„ Ø§Ù„Ø§ÙˆÙ„ Ø®Ù„ÙŠÙ‡Ø§ Ø§Ù„Ù‡ Ø¨Ø³ Ø±Ø§Ø³Ø¨ Ù„Ø§Ù†Ù‡ Ø¨Ø¬ÙˆØ² Ø§Ù„ÙØµÙ„ Ø§Ù„Ø«Ø§Ù†ÙŠ ÙŠØªØ­Ø³Ù† 
                        sheet2[result_cell_positions[1]]= 'âœ“'
                    else:    
                        sheet2[result_cell_positions[result]]= 'âœ“'
                
                #Ø¹Ø¯Ø¯ Ø§ÙŠØ§Ù… ØºÙŠØ§Ø¨ Ø§Ù„Ø·Ø§Ù„Ø¨ 
                student_absent_count = [i for i in absent_list_with_names if i[0] == group_item['student_id'] ]
                sheet2['c35']= student_absent_count[0][1] if len(student_absent_count) else 0
                # #Ø¹Ø¯Ø¯ Ø§ÙŠØ§Ù… Ø§Ù„Ø¯ÙˆØ§Ù… Ø§Ù„Ø±Ø³Ù…ÙŠ Ø§Ù„ÙƒØ§Ù…Ù„ 
                sheet2[c['school_days']]= ''
                #Ø§Ø³Ù… Ùˆ ØªÙˆÙ‚ÙŠØ¹ Ù…Ø±Ø¨ÙŠ Ø§Ù„ØµÙ 
                sheet2[c['class_teacher_name']]= grouped_list['teacher_incharge_name']
                # #Ø§Ù„ØªØ§Ø±ÙŠØ®
                # sheet2['b36']= ''
                #Ø§Ø³Ù… Ùˆ ØªÙˆÙ‚ÙŠØ¹ Ù…Ø¯ÙŠØ± Ø§Ù„Ù…Ø¯Ø±Ø³Ø©
                sheet2[c['principal_name']]= grouped_list['principle_name'] 
            template_file.remove(template_file['Sheet1'])
            template_file.save(outdir+group[0]['student_class_name_letter']+'.xlsx')

def create_coloured_certs_ods(grouped_list ,absent_list_with_names, term2=False ,template='./templet_files/Ø­Ø´Ùˆ  Ø´Ù‡Ø§Ø¯Ø§Øª Ø§Ù„ÙƒØ±ØªÙˆÙ†.ods' , outdir='./send_folder/'):
    """
    The function `create_coloured_certs_ods` creates coloured certificates in ODS format based on a
    grouped list of data.
    
    :param grouped_list: A list of dictionaries where each dictionary represents a group of certificates
    to be created. Each dictionary should have the following keys:
    :param term2: A boolean value indicating whether the certificates are for the second term or not. If
    set to True, it means the certificates are for the second term. If set to False, it means the
    certificates are for the first term, defaults to False (optional)
    :param template: The `template` parameter is the path to the template file that will be used to
    create the coloured certificates. It should be a path to an OpenDocument Spreadsheet (ODS) file,
    defaults to ./templet_files/Ø­Ø´Ùˆ  Ø´Ù‡Ø§Ø¯Ø§Øª Ø§Ù„ÙƒØ±ØªÙˆÙ†.ods (optional)
    :param outdir: The `outdir` parameter specifies the directory where the generated coloured
    certificates will be saved, defaults to ./send_folder/ (optional)
    """
    
    colored_cert_cells_position_context = {
                                        'student_name': 'E11',
                                        'student_name2': 'B44',
                                        'class_section': 'E12',
                                        'class_name':'B45',
                                        'nationality': 'D13',
                                        'national_id': 'E14',
                                        'birthplace_date': 'E15',
                                        'religion': 'D16',
                                        'student_address': 'E17',
                                        'school_name': 'E18',
                                        'school_address': 'D19',
                                        'directorate': 'D20',
                                        'school_bridge': 'D21',
                                        'supervising_authority': 'E22',
                                        'school_phone_number': 'E23',
                                        'school_natioanl_id': 'E24',
                                        'academic_year_1': 'F42',
                                        'academic_year_2': 'G42',
                                        'class': 'B45',
                                        'islamic_education': 'E50',
                                        'arabic_language': 'E51',
                                        'english_language': 'E52',
                                        'mathematics': 'E53',
                                        'social_studies': 'E54',
                                        'science': 'E55',
                                        'visual_arts': 'E56',
                                        'physical_education': 'E57',
                                        'vocational_education': 'E58',
                                        'computer': 'E59',
                                        'financial_culture': 'E60',
                                        'french_language': 'E61',
                                        'christian_religion': 'E62',
                                        'average': 'B67',
                                        'school_days': 'I64',
                                        'class_teacher_name': 'B69',
                                        'principal_name': 'L69',
                                        'semester_1_student_absent': 'H66',
                                        'semester_2_student_absent': 'L66',
                                        }
    
    c = colored_cert_cells_position_context
    result_cell_positions = ['B64','B65','B66']
    
    statistic_data = grouped_list['students_info']
    assessments_data_groups = grouped_list['assessments_data']
    term = 1 if term2 else 0
    
    
    for group in assessments_data_groups:
        if not any("Ø¹Ø´Ø±" in i['student_grade_name'] for i in group) : 
            template_file = ezodf.opendoc(template)
            sheet1 = template_file.sheets[0]
            filler_sheet = template_file.sheets[1]

            sheet2 =filler_sheet
            
            names_averages =  sort_dictionary_list_based_on(group,item_in_list=term)

            group = sort_dictionary_list_based_on(group ,simple=False,item_in_list=term)

            for row_number, dataFrame in enumerate(names_averages, start=4):
                sheet1[row_number, 1].set_value( dataFrame[1])
                sheet1[row_number, 3].set_value( dataFrame[0])
                
            for sheet_number , group_item in enumerate(group , start=2):

                # sheet2 = template_file.copy_worksheet(template_file.worksheets[1])
                # sheet2.title = str(counter)
                # counter += 1
                # sheet2.sheet_view.rightToLeft = True    
                # sheet2.sheet_view.rightToLeft = True   
                
                wanted_id = group_item['student_id']
                student_statistic_info = [i for i in statistic_data if i['student_id'] == wanted_id ][0]

                # img = openpyxl.drawing.image.Image(image)
                # img.anchor = 'e2'
                # sheet2.add_image(img)

                # print(sheet2)
                sheet2[c['student_name']].set_value( group_item['student__full_name'])
                # Ù…ÙƒØ§Ù† Ùˆ ØªØ§Ø±ÙŠØ® Ø§Ù„ÙˆÙ„Ø§Ø¯Ø©
                sheet2[c['birthplace_date']].set_value( str(group_item['student_birth_place']) + ' ' + str(group_item['student_birth_date']))
                # Ø§Ù„Ø¯ÙŠØ§Ù†Ø©
                sheet2[c['religion']].set_value( student_statistic_info['religion'])
                # Ø¹Ù†ÙˆØ§Ù† Ø§Ù„Ø·Ø§Ù„Ø¨
                sheet2[c['student_address']].set_value( student_statistic_info['address'])
                #Ø§Ù„Ø±Ù‚Ù… Ø§Ù„ÙˆØ·Ù†ÙŠ
                sheet2[c['national_id']].set_value( group_item['student_nat_id'])
                #Ø§Ù„Ø¬Ù†Ø³ÙŠØ©
                sheet2[c['nationality']].set_value( group_item['student_nat'])
                #Ø§Ù„ØµÙ Ùˆ Ø§Ù„Ø´Ø¹Ø¨Ø© 
                sheet2[c['class_section']].set_value( group_item['student_class_name_letter'])
                # ØµÙ Ø§Ù„Ø·Ø§Ù„Ø¨
                sheet2[c['class_name']].set_value( c)
                # Ø§Ø³Ù… Ø§Ù„Ù…Ø¯Ø±Ø³Ø©
                sheet2[c['school_name']].set_value( group_item['student_school_name'])
                # Ø¹Ù†ÙˆØ§Ù† Ø§Ù„Ù…Ø¯Ø±Ø³Ø©
                sheet2[c['school_address']].set_value( grouped_list['school_address'])
                # Ù…Ø¯ÙŠØ±ÙŠØ© Ø§Ù„Ù…Ø¯Ø±Ø³Ø© 
                sheet2[c['directorate']].set_value( grouped_list['school_directorate'])
                # Ù„ÙˆØ§Ø¡ Ø§Ù„Ù…Ø¯Ø±Ø³Ø©
                sheet2[c['school_bridge']].set_value( grouped_list['school_bridge'])
                #Ø§Ù„Ù…Ù†Ø·Ù‚Ø© Ø§Ù„ØªØ¹Ù„ÙŠÙ…ÙŠØ©  Ø§Ùˆ Ø§Ù„Ø³Ù„Ø·Ø© Ø§Ù„Ù…Ø´Ø±ÙØ©
                sheet2[c['supervising_authority']].set_value( 'ÙˆØ²Ø§Ø±Ø© Ø§Ù„ØªØ±Ø¨ÙŠØ© Ùˆ Ø§Ù„ØªØ¹Ù„ÙŠÙ…')
                # Ù‡Ø§ØªÙ Ø§Ù„Ù…Ø¯Ø±Ø³Ø©
                sheet2[c['school_phone_number']].set_value( grouped_list['school_phone_number'])
                # Ø±Ù‚Ù… Ø§Ù„Ù…Ø¯Ø±Ø³Ø© Ø§Ù„ÙˆØ·Ù†ÙŠ
                sheet2[c['school_natioanl_id']].set_value( grouped_list['school_national_id'])
                # Ø§Ù„Ø¹Ø§Ù… Ø§Ù„Ø¯Ø±Ø§Ø³ÙŠ Ø§Ù„Ø§ÙˆÙ„ 
                sheet2[c['academic_year_1']].set_value( grouped_list['academic_year_1'])
                # Ø§Ù„Ø¹Ø§Ù… Ø§Ù„Ø¯Ø±Ø§Ø³ÙŠ Ø§Ù„Ø«Ø§Ù†ÙŠ 
                sheet2[c['academic_year_2']].set_value( grouped_list['academic_year_2'])
                # Ø§Ù„Ø§Ø³Ù… Ø¹Ù„Ù‰ Ø§Ù„ÙˆØ¬Ù‡ Ø§Ù„Ø«Ø§Ù†ÙŠ
                sheet2[c['student_name2']].set_value( group_item['student__full_name'])
                # ØºÙŠØ§Ø¨ Ø§Ù„ÙØµÙ„ Ø§Ù„Ø§ÙˆÙ„ 
                sheet2[c['semester_1_student_absent']].set_value( '')
                # ØºÙŠØ§Ø¨ Ø§Ù„ÙØµÙ„ Ø§Ù„Ø«Ø§Ù†ÙŠ
                sheet2[c['semester_2_student_absent']].set_value( '')
                
                # put the subjects cells inder here 
                i ='c,d,e,g,j,f'.split(',')
                r = range(18,32)


                # Ø§Ù„ØªØ±Ø¨ÙŠØ© Ø§Ù„Ø§Ø³Ù„Ø§Ù…ÙŠØ©
                islam_subject = [value for key ,value in group_item['subject_sums'].items() if 'Ø³Ù„Ø§Ù…ÙŠØ©' in key]
                if 'Ø«Ø§Ù…Ù†' in group_item['student_grade_name'] or 'ØªØ§Ø³Ø¹' in group_item['student_grade_name'] or 'Ø¹Ø§Ø´Ø±' in group_item['student_grade_name']:
                    islam_subject_maxMark = 200
                    sheet2[c['islamic_education']].set_value( islam_subject_maxMark /2)
                    sheet2['F'+c['islamic_education'][1:]].set_value( islam_subject_maxMark)
                else:
                    islam_subject_maxMark = 100
                    sheet2[c['islamic_education']].set_value( islam_subject_maxMark /2)
                    sheet2['F'+c['islamic_education'][1:]].set_value( islam_subject_maxMark)
                sheet2['H'+c['islamic_education'][1:]].set_value( islam_subject[0][0] if islam_subject and len(islam_subject[0]) != 0 else '')
                sheet2['I'+c['islamic_education'][1:]].set_value( islam_subject[0][1] if term2 and islam_subject and len(islam_subject[0]) != 0 else '')
                sheet2['K'+c['islamic_education'][1:]].set_value( (islam_subject[0][0] + islam_subject[0][1])/2 if term2 and islam_subject and len(islam_subject[0]) != 0 else '')
                # if term2:
                #     sheet2['G'+c['islamic_education']].set_value( convert_avarage_to_words((islam_subject[0][0] + islam_subject[0][1])/2) if islam_subject else '')
                #     sheet2['J'+c['islamic_education']].set_value( score_in_words(((islam_subject[0][0] + islam_subject[0][1])/2),max_mark=maxMark) if islam_subject else '')
                # else:
                #     sheet2['G'+c['islamic_education']].set_value( convert_avarage_to_words(islam_subject[0][0]) if islam_subject else '')
                #     sheet2['J'+c['islamic_education']].set_value( score_in_words(islam_subject[0][0],max_mark=maxMark) if islam_subject else '')

                # Ø§Ù„Ù„ØºØ© Ø§Ù„Ø¹Ø±Ø¨ÙŠØ©
                arabic_subject = [value for key ,value in group_item['subject_sums'].items() if 'Ø¹Ø±Ø¨ÙŠ' in key]
                if 'Ø«Ø§Ù…Ù†' in group_item['student_grade_name'] or 'ØªØ§Ø³Ø¹' in group_item['student_grade_name'] or 'Ø¹Ø§Ø´Ø±' in group_item['student_grade_name']:
                    arabic_subject_maxMark = 300
                    sheet2[c['arabic_language']].set_value( arabic_subject_maxMark / 2)
                    sheet2['F'+c['arabic_language'][1:]].set_value( arabic_subject_maxMark)
                else:
                    arabic_subject_maxMark = 100
                    sheet2[c['arabic_language']].set_value( arabic_subject_maxMark / 2)
                    sheet2['F'+c['arabic_language'][1:]].set_value( arabic_subject_maxMark)
                sheet2['H' + c['arabic_language'][1:]].set_value( arabic_subject[0][0] if arabic_subject and len(arabic_subject[0]) != 0 else '')
                sheet2['I' + c['arabic_language'][1:]].set_value( arabic_subject[0][1] if term2 and arabic_subject and len(arabic_subject[0]) != 0 else '')
                sheet2['K' + c['arabic_language'][1:]].set_value( (arabic_subject[0][0] + arabic_subject[0][1])/2 if term2 and arabic_subject and len(arabic_subject[0]) != 0 else '')
                # if term2:
                #     sheet2['J19'].set_value( score_in_words(((arabic_subject[0][0] + arabic_subject[0][1])/2),max_mark=maxMark) if arabic_subject else '')
                #     sheet2['F'+].set_value( maxMark / 2)
                #     sheet2[].set_value( convert_avarage_to_words((arabic_subject[0][0] + arabic_subject[0][1])/2) if arabic_subject else maxMark)
                # else:
                #     sheet2['J19'].set_value( score_in_words(arabic_subject[0][0],max_mark=maxMark) if arabic_subject else '')
                #     sheet2['F'+].set_value( maxMark / 2)
                #     sheet2[].set_value( convert_avarage_to_words(arabic_subject[0][0]) if arabic_subject else maxMark)

                # Ø§Ù„Ù„ØºØ© Ø§Ù„Ø§Ù†Ø¬Ù„ÙŠØ²ÙŠØ© 
                english_subject = [value for key ,value in group_item['subject_sums'].items() if 'Ø¬Ù„ÙŠØ²ÙŠØ©' in key]
                if 'Ø«Ø§Ù…Ù†' in group_item['student_grade_name'] or 'ØªØ§Ø³Ø¹' in group_item['student_grade_name'] or 'Ø¹Ø§Ø´Ø±' in group_item['student_grade_name']:
                    english_subject_maxMark = 200
                    sheet2[c['english_language']].set_value( english_subject_maxMark / 2)
                    sheet2['F'+c['english_language'][1:]].set_value( english_subject_maxMark)
                else:
                    english_subject_maxMark = 100
                    sheet2[c['english_language']].set_value( english_subject_maxMark / 2)
                    sheet2['F'+c['english_language'][1:]].set_value( english_subject_maxMark)
                sheet2['H'+c['english_language'][1:]].set_value( english_subject[0][0] if english_subject and len(english_subject[0]) != 0 else '')
                sheet2['I'+c['english_language'][1:]].set_value( english_subject[0][1] if term2 and english_subject and len(english_subject[0]) != 0 else '')
                sheet2['K'+c['english_language'][1:]].set_value( (english_subject[0][0] + english_subject[0][1])/2 if term2 and english_subject and len(english_subject[0]) != 0 else '')
                # if term2:
                #     sheet2['J20'].set_value( score_in_words(((english_subject[0][0] + english_subject[0][1])/2),max_mark=maxMark) if english_subject else '')
                #     sheet2['F'+c[''][1:]].set_value( maxMark / 2)
                #     sheet2[].set_value( convert_avarage_to_words((english_subject[0][0] + english_subject[0][1])/2) if english_subject else maxMark)
                # else:
                #     sheet2['J20'].set_value( score_in_words(english_subject[0][0],max_mark=maxMark) if english_subject else '')
                #     sheet2['F'+c[''][1:]].set_value( maxMark / 2)
                #     sheet2[].set_value( convert_avarage_to_words(english_subject[0][0]) if english_subject else maxMark)

                # Ø§Ù„Ø±ÙŠØ§Ø¶ÙŠØ§Øª 
                math_subject = [value for key ,value in group_item['subject_sums'].items() if 'Ø±ÙŠØ§Ø¶ÙŠØ§Øª' in key]
                if 'Ø«Ø§Ù…Ù†' in group_item['student_grade_name'] or 'ØªØ§Ø³Ø¹' in group_item['student_grade_name'] or 'Ø¹Ø§Ø´Ø±' in group_item['student_grade_name']:
                    math_subject_maxMark = 200
                    sheet2[c['mathematics']].set_value( math_subject_maxMark / 2)
                    sheet2['F'+c['mathematics'][1:]].set_value( math_subject_maxMark)
                else:
                    math_subject_maxMark = 100
                    sheet2[c['mathematics']].set_value( math_subject_maxMark / 2)
                    sheet2['F'+c['mathematics'][1:]].set_value( math_subject_maxMark)
                sheet2['H'+c['mathematics'][1:]].set_value( math_subject[0][0] if math_subject and len(math_subject[0]) != 0 else '')
                sheet2['I'+c['mathematics'][1:]].set_value( math_subject[0][1] if term2 and math_subject and len(math_subject[0]) != 0 else '')
                sheet2['K'+c['mathematics'][1:]].set_value( (math_subject[0][0] + math_subject[0][1])/2 if term2 and math_subject and len(math_subject[0]) != 0 else '')
                # if term2:
                #     sheet2['J21'].set_value( score_in_words(((math_subject[0][0] + math_subject[0][1])/2),max_mark=maxMark) if math_subject else '')
                #     sheet2['F'+c[''][1:]].set_value( maxMark / 2)
                #     sheet2[].set_value( convert_avarage_to_words((math_subject[0][0] + math_subject[0][1])/2) if math_subject else maxMark)
                # else:
                #     sheet2['J21'].set_value( score_in_words(math_subject[0][0],max_mark=maxMark) if math_subject else '')
                #     sheet2['F'+c[''][1:]].set_value( maxMark / 2)
                #     sheet2[].set_value( convert_avarage_to_words(math_subject[0][0]) if math_subject else maxMark)

                # Ø§Ù„ØªØ±Ø¨ÙŠØ© Ø§Ù„Ø§Ø¬ØªÙ…Ø§Ø¹ÙŠØ© Ùˆ Ø§Ù„ÙˆØ·Ù†ÙŠØ© 
                social_subjects = [value for key ,value in group_item['subject_sums'].items() if 'Ø§Ø¬ØªÙ…Ø§Ø¹ÙŠØ© Ùˆ Ø§Ù„ÙˆØ·Ù†ÙŠØ©' in key]
                if 'Ø«Ø§Ù…Ù†' in group_item['student_grade_name'] or 'ØªØ§Ø³Ø¹' in group_item['student_grade_name'] or 'Ø¹Ø§Ø´Ø±' in group_item['student_grade_name']:
                    social_subjects_maxMark = 200
                    sheet2[c['social_studies']].set_value( social_subjects_maxMark / 2)
                    sheet2['F'+c['social_studies'][1:]].set_value( social_subjects_maxMark)
                    sheet2['H'+c['social_studies'][1:]].set_value( int(social_subjects[0][0]*(2/3)) if social_subjects and len(social_subjects[0]) != 0 else '')
                    sheet2['I'+c['social_studies'][1:]].set_value( int(social_subjects[0][1]*(2/3)) if term2 and social_subjects and len(social_subjects[0]) != 0 else '')
                    sheet2['K'+c['social_studies'][1:]].set_value( round((((social_subjects[0][0] + social_subjects[0][1])/2)*(2/3)),1) if term2 and social_subjects and len(social_subjects[0]) != 0 else '')
                elif 'Ø³Ø§Ø¯Ø³' in group_item['student_grade_name'] or 'Ø³Ø§Ø¨Ø¹' in group_item['student_grade_name']:
                    social_subjects_maxMark = 100                
                    sheet2[c['social_studies']].set_value( social_subjects_maxMark / 2)
                    sheet2['F'+c['social_studies'][1:]].set_value( social_subjects_maxMark)
                    sheet2['H'+c['social_studies'][1:]].set_value( int(social_subjects[0][0]/3) if social_subjects and len(social_subjects[0]) != 0 else '')
                    sheet2['I'+c['social_studies'][1:]].set_value( int(social_subjects[0][1]/3)) if term2 and social_subjects and len(social_subjects[0]) != 0 else ''
                    sheet2['K'+c['social_studies'][1:]].set_value( round(((social_subjects[0][0] + social_subjects[0][1])/2)/3 , 1) if term2 and social_subjects and len(social_subjects[0]) != 0 else '')
                else:
                    social_subjects_maxMark = 100
                    sheet2[c['social_studies']].set_value( social_subjects_maxMark / 2)
                    sheet2['F'+c['social_studies'][1:]].set_value( social_subjects_maxMark)
                    sheet2['H'+c['social_studies'][1:]].set_value( int(social_subjects[0][0]) if social_subjects and len(social_subjects[0]) != 0 else '')
                    sheet2['I'+c['social_studies'][1:]].set_value( int(social_subjects[0][1]) if term2 and social_subjects and len(social_subjects[0]) != 0 else '')
                    sheet2['K'+c['social_studies'][1:]].set_value( (social_subjects[0][0] + social_subjects[0][1])/2 if term2 and social_subjects and len(social_subjects[0]) != 0 else '')
                    
                # if term2:
                #     sheet2['J22'].set_value( score_in_words(int(((social_subjects[0][0] + social_subjects[0][1])/2)*(2/3)),max_mark=maxMark) if social_subjects else '')
                #     sheet2['F'+c[''][1:]].set_value( maxMark / 2)
                #     sheet2[].set_value( convert_avarage_to_words((social_subjects[0][0] + social_subjects[0][1])/2) if social_subjects else maxMark)
                # else:
                #     sheet2['J22'].set_value( score_in_words(int(social_subjects[0][0]*(2/3)),max_mark=maxMark) if social_subjects else '')
                #     sheet2['F'+c[''][1:]].set_value( maxMark / 2)
                #     sheet2[].set_value( convert_avarage_to_words(social_subjects[0][0]) if social_subjects else maxMark)

                # Ø§Ù„Ø¹Ù„ÙˆÙ…
                science_subjects = [value for key ,value in group_item['subject_sums'].items() if 'Ø§Ù„Ø¹Ù„ÙˆÙ…' in key]
                if 'Ø«Ø§Ù…Ù†' in group_item['student_grade_name'] :
                    science_subjects_maxMark = 200
                    sheet2[c['science']].set_value( science_subjects_maxMark / 2)
                    sheet2['F'+c['science'][1:]].set_value( science_subjects_maxMark)
                elif 'ØªØ§Ø³Ø¹' in group_item['student_grade_name'] or 'Ø¹Ø§Ø´Ø±' in group_item['student_grade_name']:
                    science_subjects_maxMark = 400
                    sheet2[c['science']].set_value( science_subjects_maxMark / 2)
                    sheet2['F'+c['science'][1:]].set_value( science_subjects_maxMark)
                else:
                    science_subjects_maxMark = 100
                    sheet2[c['science']].set_value( science_subjects_maxMark / 2)
                    sheet2['F'+c['science'][1:]].set_value( science_subjects_maxMark)
                sheet2['H'+c['science'][1:]].set_value( science_subjects[0][0] if science_subjects and len(science_subjects[0]) != 0 else '')
                sheet2['I'+c['science'][1:]].set_value( science_subjects[0][1] if term2 and science_subjects and len(science_subjects[0]) != 0 else '')
                sheet2['K'+c['science'][1:]].set_value( (science_subjects[0][0] + science_subjects[0][1])/2 if term2 and science_subjects and len(science_subjects[0]) != 0 else '')
                # if term2:
                #     sheet2['J23'].set_value( score_in_words( ((science_subjects[0][0] + science_subjects[0][1])/2),max_mark=maxMark) if  science_subjects else '')
                #     sheet2['F'+c[''][1:]].set_value( maxMark / 2)
                #     sheet2[].set_value( convert_avarage_to_words((science_subjects[0][0] + science_subjects[0][1])/2) if science_subjects else maxMark)
                # else:
                #     sheet2['J23'].set_value( score_in_words( science_subjects[0][0],max_mark=maxMark) if  science_subjects else '')
                #     sheet2['F'+c[''][1:]].set_value( maxMark / 2)
                #     sheet2[].set_value( convert_avarage_to_words(science_subjects[0][0]) if science_subjects else maxMark)

                # Ø§Ù„ØªØ±Ø¨ÙŠØ© Ø§Ù„ÙÙ†ÙŠØ© ÙˆØ§Ù„Ù…ÙˆØ³ÙŠÙ‚ÙŠØ©
                art_subject = [value for key ,value in group_item['subject_sums'].items() if 'Ø§Ù„ÙÙ†ÙŠØ© ÙˆØ§Ù„Ù…ÙˆØ³' in key]
                sheet2[c['visual_arts']].set_value( 50 if art_subject and len(art_subject[0]) != 0 else '')
                sheet2['F'+c['visual_arts'][1:]].set_value( 100 if art_subject and len(art_subject[0]) != 0 else '')
                sheet2['H'+c['visual_arts'][1:]].set_value( art_subject[0][0] if art_subject and len(art_subject[0]) != 0 else '')
                sheet2['I'+c['visual_arts'][1:]].set_value( art_subject[0][1] if term2 and art_subject and len(art_subject[0]) != 0 else '')
                sheet2['K'+c['visual_arts'][1:]].set_value( (art_subject[0][0] + art_subject[0][1])/2 if term2 and art_subject and len(art_subject[0]) != 0 else '')
                # if term2:
                #     sheet2['G24'].set_value( convert_avarage_to_words((art_subject[0][0] + art_subject[0][1])/2) if art_subject else '')
                #     sheet2['J24'].set_value( score_in_words(((art_subject[0][0] + art_subject[0][1])/2) ) if art_subject else '')
                # else:
                #     sheet2['G24'].set_value( convert_avarage_to_words(art_subject[0][0]) if art_subject else '')
                #     sheet2['J24'].set_value( score_in_words(art_subject[0][0] ) if art_subject else '')

                # Ø§Ù„ØªØ±Ø¨ÙŠØ© Ø§Ù„Ø±ÙŠØ§Ø¶ÙŠØ©
                sport_subject = [value for key ,value in group_item['subject_sums'].items() if 'Ø±ÙŠØ§Ø¶ÙŠØ©' in key]
                sheet2[c['physical_education']].set_value( 50 if sport_subject and len(sport_subject[0]) != 0 else '')
                sheet2['F'+c['physical_education'][1:]].set_value( 100 if sport_subject and len(sport_subject[0]) != 0 else '')
                sheet2['H'+c['physical_education'][1:]].set_value( sport_subject[0][0] if sport_subject and len(sport_subject[0]) != 0 else '')
                sheet2['I'+c['physical_education'][1:]].set_value( sport_subject[0][1] if term2 and sport_subject and len(sport_subject[0]) != 0 else '')
                sheet2['K'+c['physical_education'][1:]].set_value( (sport_subject[0][0] + sport_subject[0][1])/2 if term2 and sport_subject and len(sport_subject[0]) != 0 else '')
                # if term2:
                #     sheet2['G25'].set_value( convert_avarage_to_words((sport_subject[0][0] + sport_subject[0][1])/2) if sport_subject else '')
                #     sheet2['J25'].set_value( score_in_words(((sport_subject[0][0] + sport_subject[0][1])/2) ) if sport_subject else '')
                # else:
                #     sheet2['G25'].set_value( convert_avarage_to_words(sport_subject[0][0]) if sport_subject else '')
                #     sheet2['J25'].set_value( score_in_words(sport_subject[0][0] ) if sport_subject else '')

                # Ø§Ù„ØªØ±Ø¨ÙŠØ© Ø§Ù„Ù…Ù‡Ù†ÙŠØ© 
                vocational_subject = [value for key ,value in group_item['subject_sums'].items() if 'Ù…Ù‡Ù†ÙŠØ©' in key]
                sheet2[c['vocational_education']].set_value( 50 if vocational_subject and len(vocational_subject[0]) != 0 else '')
                sheet2['F'+c['vocational_education'][1:]].set_value( 100 if vocational_subject and len(vocational_subject[0]) != 0 else '')
                sheet2['H'+c['vocational_education'][1:]].set_value( vocational_subject[0][0] if vocational_subject and len(vocational_subject[0]) != 0 else '')
                sheet2['I'+c['vocational_education'][1:]].set_value( vocational_subject[0][1] if term2 and vocational_subject and len(vocational_subject[0]) != 0 else '')
                sheet2['K'+c['vocational_education'][1:]].set_value( (vocational_subject[0][0] + vocational_subject[0][1])/2 if term2 and vocational_subject and len(vocational_subject[0]) != 0 else '')
                # if term2:
                #     sheet2['G26'].set_value( convert_avarage_to_words((vocational_subject[0][0] + vocational_subject[0][1])/2) if vocational_subject else '')
                #     sheet2['J26'].set_value( score_in_words(((vocational_subject[0][0] + vocational_subject[0][1])/2) ) if vocational_subject else '')
                # else:
                #     sheet2['G26'].set_value( convert_avarage_to_words(vocational_subject[0][0]) if vocational_subject else '')
                #     sheet2['J26'].set_value( score_in_words(vocational_subject[0][0] ) if vocational_subject else '')

                # Ø§Ù„Ø­Ø§Ø³ÙˆØ¨
                computer_subject = [value for key ,value in group_item['subject_sums'].items() if 'Ø­Ø§Ø³ÙˆØ¨' in key]
                sheet2[c['computer']].set_value( 50 if computer_subject and len(computer_subject[0]) != 0 else '')
                sheet2['F'+c['computer'][1:]].set_value( 100 if computer_subject and len(computer_subject[0]) != 0 else '')
                sheet2['H'+c['computer'][1:]].set_value( computer_subject[0][0] if computer_subject and len(computer_subject[0]) != 0 else '')
                sheet2['I'+c['computer'][1:]].set_value( computer_subject[0][1] if term2 and computer_subject and len(computer_subject[0]) != 0 else '')
                sheet2['K'+c['computer'][1:]].set_value( (computer_subject[0][0] + computer_subject[0][1])/2 if term2 and computer_subject and len(computer_subject[0]) != 0 else '')
                # if term2:
                #     sheet2['G27'].set_value( convert_avarage_to_words((computer_subject[0][0] + computer_subject[0][1])/2) if computer_subject else '')
                #     sheet2['J27'].set_value( score_in_words(((computer_subject[0][0] + computer_subject[0][1])/2) ) if computer_subject else '')
                # else:
                #     sheet2['G27'].set_value( convert_avarage_to_words(computer_subject[0][0]) if computer_subject else '')
                #     sheet2['J27'].set_value( score_in_words(computer_subject[0][0] ) if computer_subject else '')

                # Ø§Ù„Ø«Ù‚Ø§ÙØ© Ø§Ù„Ù…Ø§Ù„ÙŠØ©
                financial_subject = [value for key ,value in group_item['subject_sums'].items() if 'Ù…Ø§Ù„ÙŠØ©' in key]
                sheet2[c['financial_culture']].set_value( 50 if financial_subject and len(financial_subject[0]) != 0 else '')
                sheet2['F'+c['financial_culture'][1:]].set_value( 100 if financial_subject and len(financial_subject[0]) != 0 else '')
                sheet2['H'+c['financial_culture'][1:]].set_value( financial_subject[0][0] if financial_subject and len(financial_subject[0]) != 0 else '')
                sheet2['I'+c['financial_culture'][1:]].set_value( financial_subject[0][1] if term2 and financial_subject and len(financial_subject[0]) != 0 else '')
                sheet2['K'+c['financial_culture'][1:]].set_value( (financial_subject[0][0] + financial_subject[0][1])/2 if term2 and financial_subject and len(financial_subject[0]) != 0 else '')
                # if term2:
                #     sheet2['G28'].set_value( convert_avarage_to_words((financial_subject[0][0] + financial_subject[0][1])/2) if financial_subject else '')
                #     sheet2['J28'].set_value( score_in_words(((financial_subject[0][0] + financial_subject[0][1])/2) ) if financial_subject else '')
                # else:
                #     sheet2['G28'].set_value( convert_avarage_to_words(financial_subject[0][0]) if financial_subject else '')
                #     sheet2['J28'].set_value( score_in_words(financial_subject[0][0] ) if financial_subject else '')

                # Ø§Ù„Ù„ØºØ© Ø§Ù„ÙØ±Ù†Ø³ÙŠØ© 
                franch_subject = [value for key ,value in group_item['subject_sums'].items() if 'ÙØ±Ù†Ø³ÙŠØ©' in key]
                sheet2[c['french_language']].set_value( 50 if franch_subject and len(franch_subject[0]) != 0 else '')
                sheet2['F'+c['french_language'][1:]].set_value( 100 if franch_subject and len(franch_subject[0]) != 0 else '')
                sheet2['H'+c['french_language'][1:]].set_value( franch_subject[0][0] if franch_subject and len(franch_subject[0]) != 0 else '')
                sheet2['I'+c['french_language'][1:]].set_value( franch_subject[0][1] if term2 and franch_subject and len(franch_subject[0]) != 0 else '')
                sheet2['K'+c['french_language'][1:]].set_value( (franch_subject[0][0] + franch_subject[0][1])/2 if term2 and franch_subject and len(franch_subject[0]) != 0 else '')
                # if term2:
                #     sheet2['G29'].set_value( convert_avarage_to_words((franch_subject[0][0] + franch_subject[0][1])/2) if franch_subject else '')
                #     sheet2['J29'].set_value( score_in_words(((franch_subject[0][0] + franch_subject[0][1])/2) ) if franch_subject else '')
                # else:
                #     sheet2['G29'].set_value( convert_avarage_to_words(franch_subject[0][0]) if franch_subject else '')
                #     sheet2['J29'].set_value( score_in_words(franch_subject[0][0] ) if franch_subject else '')

                # Ø§Ù„Ø¯ÙŠÙ† Ø§Ù„Ù…Ø³ÙŠØ­ÙŠ
                christian_subject = [value for key ,value in group_item['subject_sums'].items() if 'Ø§Ù„Ø¯ÙŠÙ† Ø§Ù„Ù…Ø³ÙŠØ­ÙŠ' in key]
                sheet2[c['christian_religion']].set_value( 50 if christian_subject and len(christian_subject[0]) != 0 else '')
                sheet2['F'+c['christian_religion'][1:]].set_value( 100 if christian_subject and len(christian_subject[0]) != 0 else '')
                sheet2['H'+c['christian_religion'][1:]].set_value( christian_subject[0][0] if christian_subject and len(christian_subject[0]) != 0 else '')
                sheet2['I'+c['christian_religion'][1:]].set_value( christian_subject[0][1] if term2 and christian_subject and len(christian_subject[0]) != 0 else '')
                sheet2['K'+c['christian_religion'][1:]].set_value( (christian_subject[0][0] + christian_subject[0][1])/2 if term2 and christian_subject and len(christian_subject[0]) != 0 else '')
                # if term2:
                #     sheet2['G30'].set_value( convert_avarage_to_words((christian_subject[0][0] + christian_subject[0][1])/2) if christian_subject else '')
                #     sheet2['J30'].set_value( score_in_words(((christian_subject[0][0] + christian_subject[0][1])/2) ) if christian_subject else '')
                # else:
                #     sheet2['G30'].set_value( convert_avarage_to_words(christian_subject[0][0]) if christian_subject else '')
                #     sheet2['J30'].set_value( score_in_words(christian_subject[0][0] ) if christian_subject else '')

                counter = 0
                for subject_name  ,S1_S2 in group_item['subject_sums'].items():
                    average = (S1_S2[0]+S1_S2[1])/2
                    print( subject_name, S1_S2)
                    if 'Ø³Ù„Ø§Ù…ÙŠ' in subject_name and average < islam_subject_maxMark / 2 : 
                        counter+=1
                    elif "Ø¹Ø±Ø¨ÙŠ"  in subject_name and average < arabic_subject_maxMark / 2 : 
                        counter+=1
                    elif "Ù†Ø¬Ù„ÙŠØ²ÙŠ"  in subject_name and average < english_subject_maxMark / 2 : 
                        counter+=1
                    elif "Ø±ÙŠØ§Ø¶ÙŠØ§Øª"  in subject_name and average < math_subject_maxMark / 2 : 
                        counter+=1
                    elif "Ø¬ØªÙ…Ø§Ø¹ÙŠØ©"  in subject_name and average < social_subjects_maxMark / 2 : 
                        counter+=1
                    elif "Ø¹Ù„ÙˆÙ…"  in subject_name and average < science_subjects_maxMark / 2 : 
                        counter+=1
                    elif  average < 50: 
                        counter+=1
                    # Ø·Ø±ÙŠÙ‚Ø© Ø·Ø¨Ø§Ø¹Ø© Ø§Ù„Ø±Ù‚Ù… ØµØ­ÙŠØ­ Ø§Ø°Ø§ ÙƒØ§Ù† Ø¨Ø¯ÙˆÙ† Ø§Ø¹Ø´Ø§Ø± 
                    # print(subject_name , int((S1_S2[0]+S1_S2[1])/2) if str((S1_S2[0]+S1_S2[1])/2).split('.')[1] == '0' else (S1_S2[0]+S1_S2[1])/2 )
                    
                # print(counter)
                if counter > 4 : 
                    print('ÙŠØ¨Ù‚Ù‰ ÙÙŠ ØµÙÙ‡')
                    result = 2
                elif counter == 0 :
                    print("Ù†Ø§Ø¬Ø­")
                    result = 0
                else :     
                    print('Ù…ÙƒÙ…Ù„')
                    result = 1
                
                if term2 :
                    # Ø§Ù„Ù…Ø¹Ø¯Ù„ Ø§Ù„Ù…Ø¦ÙˆÙŠ Ø¨Ø§Ù„Ø±Ù‚Ø§Ù… 
                    sheet2[c['average']].set_value( group_item['t1+t2+year_avarage'][2])
                    # #Ø¨Ø§Ù„Ø­Ø±ÙˆÙ
                    # sheet2['e32'].set_value( convert_avarage_to_words(group_item['t1+t2+year_avarage'][2]) if group_item else '')
                    # #ØªØ±ØªÙŠØ¨ Ø§Ù„Ø·Ø§Ù„Ø¨ Ø¹Ù„Ù‰ Ø§Ù„ØµÙ 
                    # sheet2['j32'].set_value( counter-1)

                    #Ø§Ù„Ù†ØªÙŠØ¬Ø© 
                    sheet2[result_cell_positions[result]].set_value( 'âœ“')
                else:
                    
                    #Ø§Ù„Ù…Ø¹Ø¯Ù„ Ø§Ù„Ù…Ø¦ÙˆÙŠ Ø¨Ø§Ù„Ø±Ù‚Ø§Ù… 
                    sheet2[c['average']].set_value( group_item['t1+t2+year_avarage'][term])
                    # #Ø¨Ø§Ù„Ø­Ø±ÙˆÙ
                    # sheet2['e32'].set_value( convert_avarage_to_words(group_item['t1+t2+year_avarage'][0]) if group_item else '')
                    # #ØªØ±ØªÙŠØ¨ Ø§Ù„Ø·Ø§Ù„Ø¨ Ø¹Ù„Ù‰ Ø§Ù„ØµÙ 
                    # sheet2['j32'].set_value( counter-1)
                    # #Ø§Ù„Ù†ØªÙŠØ¬Ø© 
                    if result == 2 : # Ø§Ø°Ø§ ÙƒØ§Ù† Ù…ÙƒÙ…Ù„ ÙÙŠ ØµÙÙ‡ Ø§Ù„ÙØµÙ„ Ø§Ù„Ø§ÙˆÙ„ Ø®Ù„ÙŠÙ‡Ø§ Ø§Ù„Ù‡ Ø¨Ø³ Ø±Ø§Ø³Ø¨ Ù„Ø§Ù†Ù‡ Ø¨Ø¬ÙˆØ² Ø§Ù„ÙØµÙ„ Ø§Ù„Ø«Ø§Ù†ÙŠ ÙŠØªØ­Ø³Ù† 
                        sheet2[result_cell_positions[1]].set_value( 'âœ“')
                    else:    
                        sheet2[result_cell_positions[result]].set_value( 'âœ“')
                
                
                # #Ø¹Ø¯Ø¯ Ø§ÙŠØ§Ù… Ø§Ù„Ø¯ÙˆØ§Ù… Ø§Ù„Ø±Ø³Ù…ÙŠ Ø§Ù„ÙƒØ§Ù…Ù„ 
                sheet2[c['school_days']].set_value( '')
                                
                # #Ø¹Ø¯Ø¯ Ø§ÙŠØ§Ù… ØºÙŠØ§Ø¨ Ø§Ù„Ø·Ø§Ù„Ø¨ 
                student_absent_count = [i for i in absent_list_with_names if i[0] == group_item['student_id'] ]
                sheet2['I66'].set_value(student_absent_count if len(student_absent_count) else 0)
                
                if grouped_list['teacher_incharge_name'] != '  ':   # Ø§Ø°Ø§ Ù„Ù… ÙŠÙƒÙ† Ø§Ø³Ù… Ø§Ù„Ù…Ø¹Ù„Ù… ÙØ§Ø±Øº Ø¹Ø¨ÙŠ Ø§Ø³Ù… Ø§Ù„Ù…Ø¹Ù„Ù…
                    #Ø§Ø³Ù… Ùˆ ØªÙˆÙ‚ÙŠØ¹ Ù…Ø±Ø¨ÙŠ Ø§Ù„ØµÙ 
                    sheet2[c['class_teacher_name']].set_value( grouped_list['teacher_incharge_name'])
                # #Ø§Ù„ØªØ§Ø±ÙŠØ®
                # sheet2['b36'].set_value( '')
                #Ø§Ø³Ù… Ùˆ ØªÙˆÙ‚ÙŠØ¹ Ù…Ø¯ÙŠØ± Ø§Ù„Ù…Ø¯Ø±Ø³Ø©
                
                sheet2[c['principal_name']].set_value( grouped_list['principle_name'] )
                sheet2 = filler_sheet.copy(newname=str(sheet_number))
                template_file.sheets += sheet2  
            del template_file.sheets[-1]
            # template_file.remove(template_file['Sheet1'])
            template_file.saveas(outdir+group[0]['student_class_name_letter']+'.ods')

def sort_dictionary_list_based_on(dictionary_list, dictionary_key='t1+t2+year_avarage', item_in_list=0, reverse=True, simple=True ,sort_names_only=False):
    """
    This function sorts a list of dictionaries based on a specified key and returns the sorted list.
    
    :param dictionary_list: A list of dictionaries that you want to sort
    :param dictionary_key: The key in the dictionaries within the list that will be used for sorting. By
    default, it is set to 't1+t2+year_avarage', defaults to t1+t2+year_avarage (optional)
    :param item_in_list: The index of the item in each dictionary within the list that you want to sort
    by. For example, if each dictionary in the list has a 'name' key and you want to sort based on the
    value of 'name', you would set item_in_list=0, defaults to 0 (optional)
    :param reverse: The "reverse" parameter determines whether the sorting should be in ascending or
    descending order. If set to True, the list will be sorted in descending order. If set to False, the
    list will be sorted in ascending order, defaults to True (optional)
    :param simple: The "simple" parameter is a boolean value that determines whether the sorting should
    be done in a simple or complex manner. If set to True, the sorting will be done based on the
    dictionary key specified in the "dictionary_key" parameter. If set to False, the sorting will be
    done based on, defaults to True (optional)
    """
    if simple:
        return [
                    (
                        0 if len(i.get('subjects_assessments_info', [])) == 0 else i.get(dictionary_key, [''])[item_in_list],
                        i.get('student__full_name', '')
                    )
                    for i in sorted(
                        dictionary_list,
                        key=lambda x: 0 if len(x.get('subjects_assessments_info', [])) == 0 else x.get(dictionary_key, [''])[item_in_list],
                        reverse=reverse
                    )
                ]
    elif sort_names_only:
        return sorted(dictionary_list, key=lambda x: x['student__full_name'])
    else:
        return sorted(
                dictionary_list,
                key=lambda x: 0 if len(x.get('subjects_assessments_info', [])) == 0 else x.get(dictionary_key, [''])[item_in_list],
                reverse=reverse
            )

def convert_avarage_to_words(digit):
    """
    The function takes a digit as input and converts it into words.
    
    :param digit: The digit parameter represents the average value that you want to convert into words
    """
    number_fraction = str(digit).split('.')
    if '.' in str(digit):
        number_in_words = re.sub("Ø±ÙŠØ§Ù„.*", "",num2words(int(number_fraction[0]),lang='ar', to='currency'))
        fraction = int(number_fraction[1])
        if fraction == 1:
            fraction_in_words = 'Ø¹Ø´Ø±'
        elif fraction == 2:
            fraction_in_words = 'Ø¹Ø´Ø±ÙŠÙ†'
        else:
            fraction_in_words = str(num2words(fraction,lang='ar', to='year')) + ' Ø§Ø¹Ø´Ø§Ø±'
        lst = ["Ù…Ø§Ø¦Ø©", "Ù…Ø¦ØªØ§Ù†", "Ø«Ù„Ø§Ø«Ù…Ø§Ø¦Ø©", "Ø£Ø±Ø¨Ø¹Ù…Ø§Ø¦Ø©", "Ø®Ù…Ø³Ù…Ø§Ø¦Ø©", "Ø³ØªÙ…Ø§Ø¦Ø©", "Ø³Ø¨Ø¹Ù…Ø§Ø¦Ø©", "Ø«Ù…Ø§Ù†Ù…Ø§Ø¦Ø©", "ØªØ³Ø¹Ù…Ø§Ø¦Ø©"]
        for item in lst:
            if item in number_in_words:
                word = number_in_words.replace(item, item + ' ').split()
                if len(word) > 1:
                    number_in_words = number_in_words.replace(item, item + ' Ùˆ')
                else:
                    number_in_words = number_in_words.replace(item, item )
        if fraction == 0:
            return number_in_words.replace(' Ùˆ', ' Ùˆ ')
        else:
            return (number_in_words + ' Ùˆ '+ fraction_in_words )
    else:
        number_in_words = re.sub("Ø±ÙŠØ§Ù„.*", "",num2words(int(number_fraction[0]),lang='ar', to='currency'))
        lst = ["Ù…Ø§Ø¦Ø©", "Ù…Ø¦ØªØ§Ù†", "Ø«Ù„Ø§Ø«Ù…Ø§Ø¦Ø©", "Ø£Ø±Ø¨Ø¹Ù…Ø§Ø¦Ø©", "Ø®Ù…Ø³Ù…Ø§Ø¦Ø©", "Ø³ØªÙ…Ø§Ø¦Ø©", "Ø³Ø¨Ø¹Ù…Ø§Ø¦Ø©", "Ø«Ù…Ø§Ù†Ù…Ø§Ø¦Ø©", "ØªØ³Ø¹Ù…Ø§Ø¦Ø©"]
        for item in lst:
            if item in number_in_words:
                word = number_in_words.replace(item, item + ' ').split()
                if len(word) > 1:
                    number_in_words = number_in_words.replace(item, item + ' Ùˆ ')
                else:
                    number_in_words = number_in_words.replace(item, item )
        number_in_words = number_in_words.replace('Ùˆ  Ùˆ', 'Ùˆ')
        return re.sub(r" Ùˆ $", "", number_in_words).replace('Ùˆ  Ùˆ', 'Ùˆ')

def score_in_words(digit, max_mark=100):
    """
    The function "score_in_words" converts a numerical score into words (Ù…Ù…ØªØ§Ø² Ø§Ùˆ Ø¬ÙŠØ¯ Ø§Ùˆ Ø¬ÙŠØ¯ Ø¬Ø¯Ø§ Ø§Ùˆ Ø¶Ø¹ÙŠÙ Ø§Ùˆ Ù…Ù‚ØµØ±).
    
    :param digit: The digit parameter represents the numerical score that you want to convert into words
    :param max_mark: The maximum possible score that can be achieved. It is set to 100 by default,
    defaults to 100 (optional)
    """
    excellent_threshold = 0.9
    very_good_threshold = 0.8
    good_threshold = 0.7
    average_threshold = 0.6
    pass_threshold = 0.5
    
    if digit >= excellent_threshold * max_mark:
        return 'Ù…Ù…ØªØ§Ø²'
    elif digit >= very_good_threshold * max_mark:
        return 'Ø¬ÙŠØ¯ Ø¬Ø¯Ø§'
    elif digit >= good_threshold * max_mark:
        return 'Ø¬ÙŠØ¯'
    elif digit >= average_threshold * max_mark:
        return 'Ù…ØªÙˆØ³Ø·'
    elif digit >= pass_threshold * max_mark:
        return 'Ù…Ù‚Ø¨ÙˆÙ„'
    else:
        return 'Ù…Ù‚ØµØ±'

def add_averages_to_group_list(grouped_list , skip_art_sport=True):
    """
    The function adds the average values to a grouped list, with an option to skip certain subjects.
    
    :param grouped_list: A list of lists, where each inner list represents a group of items. Each inner
    list should contain the items that belong to that group
    :param skip_art_sport: A boolean parameter that determines whether to skip adding averages for
    the subjects art and sport , defaults to True (optional)
    """
    for group in grouped_list:
        for item in group:
            term_1_avarage ,term_2_avarage , year_avarage = [0]*3
            try :
                if 'Ø®Ø§Ù…Ø³' in  item['student_grade_name']:
                    for key, value in item['subject_sums'].items():
                        # if 'Ø±Ø¨ÙŠØ© Ø§Ù„Ø§Ø¬ØªÙ…Ø§Ø¹ÙŠØ© Ùˆ Ø§Ù„ÙˆØ·Ù†ÙŠØ©' in key or 'Ø¬ØªÙ…Ø§Ø¹ÙŠ' in key  or 'ÙˆØ·Ù†ÙŠ' in key :
                        #     # print(key ,round(value[0]*2/3),1)
                        #     term_1_avarage +=round(value[0]/3,1)
                        #     term_2_avarage +=round(value[1]/3,1)
                        #     # year_avarage += round((term_1_avarage + term_2_avarage)/2,1)
                        if ('Ø§Ù„ØªØ±Ø¨ÙŠØ© Ø§Ù„ÙÙ†ÙŠØ© ÙˆØ§Ù„Ù…ÙˆØ³ÙŠÙ‚ÙŠØ©' in key or 'Ø§Ù„ØªØ±Ø¨ÙŠØ© Ø§Ù„Ø±ÙŠØ§Ø¶ÙŠØ©' in key) and skip_art_sport :
                            pass
                        else:
                            # print(key , value[0])
                            term_1_avarage += value[0]
                            term_2_avarage += value[1]
                            # year_avarage += round((term_1_avarage + term_2_avarage)/2,1)
                    term_1_avarage ,term_2_avarage ,year_avarage =round((term_1_avarage / 900)* 100,1) , round((term_2_avarage / 900)* 100,1) , round((((term_1_avarage+term_2_avarage)/2) / 900)* 100,1)
                    item['t1+t2+year_avarage'] = [term_1_avarage ,term_2_avarage ,year_avarage ]
                
                elif 'Ø³Ø§Ø¯Ø³' in  item['student_grade_name']:
                    for key, value in item['subject_sums'].items():
                        if 'Ø±Ø¨ÙŠØ© Ø§Ù„Ø§Ø¬ØªÙ…Ø§Ø¹ÙŠØ© Ùˆ Ø§Ù„ÙˆØ·Ù†ÙŠØ©' in key or 'Ø¬ØªÙ…Ø§Ø¹ÙŠ' in key  or 'ÙˆØ·Ù†ÙŠ' in key :
                            # print(key ,round(value[0]*2/3),1)
                            term_1_avarage +=round(value[0]/3,1)
                            term_2_avarage +=round(value[1]/3,1)
                            # year_avarage += round((term_1_avarage + term_2_avarage)/2,1)
                        elif ('Ø§Ù„ØªØ±Ø¨ÙŠØ© Ø§Ù„ÙÙ†ÙŠØ© ÙˆØ§Ù„Ù…ÙˆØ³ÙŠÙ‚ÙŠØ©' in key or 'Ø§Ù„ØªØ±Ø¨ÙŠØ© Ø§Ù„Ø±ÙŠØ§Ø¶ÙŠØ©' in key) and skip_art_sport :
                            pass
                        else:
                            # print(key , value[0])
                            term_1_avarage += value[0]
                            term_2_avarage += value[1]
                            # year_avarage += round((term_1_avarage + term_2_avarage)/2,1)
                    term_1_avarage ,term_2_avarage ,year_avarage =round((term_1_avarage / 900)* 100,1) , round((term_2_avarage / 900)* 100,1) , round((((term_1_avarage+term_2_avarage)/2) / 900)* 100,1)
                    item['t1+t2+year_avarage'] = [term_1_avarage ,term_2_avarage ,year_avarage ]

                elif 'Ø³Ø§Ø¨Ø¹' in  item['student_grade_name']:
                    for key, value in item['subject_sums'].items():
                        # if 'Ø±Ø¨ÙŠØ© Ø§Ù„Ø§Ø¬ØªÙ…Ø§Ø¹ÙŠØ© Ùˆ Ø§Ù„ÙˆØ·Ù†ÙŠØ©' in key or 'Ø¬ØªÙ…Ø§Ø¹ÙŠ' in key  or 'ÙˆØ·Ù†ÙŠ' in key :
                        #     # print(key ,round(value[0]*2/3),1)
                        #     term_1_avarage +=round(value[0]/3,1)
                        #     term_2_avarage +=round(value[1]/3,1)
                        #     # year_avarage += round((term_1_avarage + term_2_avarage)/2,1)
                        if ('Ø§Ù„ØªØ±Ø¨ÙŠØ© Ø§Ù„ÙÙ†ÙŠØ© ÙˆØ§Ù„Ù…ÙˆØ³ÙŠÙ‚ÙŠØ©' in key or 'Ø§Ù„ØªØ±Ø¨ÙŠØ© Ø§Ù„Ø±ÙŠØ§Ø¶ÙŠØ©' in key) and skip_art_sport :
                            pass                        
                        else:
                            # print(key , value[0])
                            term_1_avarage += value[0]
                            term_2_avarage += value[1]
                            # year_avarage += round((term_1_avarage + term_2_avarage)/2,1)
                    term_1_avarage ,term_2_avarage ,year_avarage =round((term_1_avarage / 1100)* 100,1) , round((term_2_avarage / 1100)* 100,1) , round((((term_1_avarage+term_2_avarage)/2) / 1100)* 100,1)
                    item['t1+t2+year_avarage'] = [term_1_avarage ,term_2_avarage ,year_avarage ]

                elif 'Ø«Ø§Ù…Ù†' in  item['student_grade_name']:
                    for key, value in item['subject_sums'].items():
                        if 'Ø±Ø¨ÙŠØ© Ø§Ù„Ø§Ø¬ØªÙ…Ø§Ø¹ÙŠØ© Ùˆ Ø§Ù„ÙˆØ·Ù†ÙŠØ©' in key or 'Ø¬ØªÙ…Ø§Ø¹ÙŠ' in key  or 'ÙˆØ·Ù†ÙŠ' in key :
                            # print(key ,round(value[0]*2/3),1)
                            term_1_avarage += round(value[0]*2/3,1)
                            term_2_avarage += round(value[1]*2/3,1)
                            # year_avarage += round((term_1_avarage + term_2_avarage)/2,1)
                        elif ('Ø§Ù„ØªØ±Ø¨ÙŠØ© Ø§Ù„ÙÙ†ÙŠØ© ÙˆØ§Ù„Ù…ÙˆØ³ÙŠÙ‚ÙŠØ©' in key or 'Ø§Ù„ØªØ±Ø¨ÙŠØ© Ø§Ù„Ø±ÙŠØ§Ø¶ÙŠØ©' in key) and skip_art_sport :
                            pass                        
                        else:
                            # print(key , value[0])
                            term_1_avarage += value[0]
                            term_2_avarage += value[1]
                            # year_avarage += round((term_1_avarage + term_2_avarage)/2,1)
                    term_1_avarage ,term_2_avarage ,year_avarage =round((term_1_avarage / 1800)* 100,1) , round((term_2_avarage / 1800)* 100,1) , round((((term_1_avarage+term_2_avarage)/2) / 1800)* 100,1)
                    item['t1+t2+year_avarage'] = [term_1_avarage ,term_2_avarage ,year_avarage ]

                elif 'ØªØ§Ø³Ø¹' in  item['student_grade_name']:
                    for key, value in item['subject_sums'].items():
                        if 'Ø±Ø¨ÙŠØ© Ø§Ù„Ø§Ø¬ØªÙ…Ø§Ø¹ÙŠØ© Ùˆ Ø§Ù„ÙˆØ·Ù†ÙŠØ©' in key :
                            # print(key ,round(value[0]*2/3),1)
                            term_1_avarage +=round(value[0]*2/3,1)
                            term_2_avarage +=round(value[1]*2/3,1)
                            # year_avarage += round((term_1_avarage + term_2_avarage)/2,1)
                        elif ('Ø§Ù„ØªØ±Ø¨ÙŠØ© Ø§Ù„ÙÙ†ÙŠØ© ÙˆØ§Ù„Ù…ÙˆØ³ÙŠÙ‚ÙŠØ©' in key or 'Ø§Ù„ØªØ±Ø¨ÙŠØ© Ø§Ù„Ø±ÙŠØ§Ø¶ÙŠØ©' in key) and skip_art_sport :
                            pass                        
                        else:
                            # print(key , value[0])
                            term_1_avarage += value[0]
                            term_2_avarage += value[1]
                            # year_avarage += round((term_1_avarage + term_2_avarage)/2,1)
                    term_1_avarage ,term_2_avarage ,year_avarage =round((term_1_avarage / 2000)* 100,1) , round((term_2_avarage / 2000)* 100,1) , round((((term_1_avarage+term_2_avarage)/2) / 2000)* 100,1)
                    item['t1+t2+year_avarage'] = [term_1_avarage ,term_2_avarage ,year_avarage ]

                elif 'Ø¹Ø§Ø´Ø±' in  item['student_grade_name']:
                    for key, value in item['subject_sums'].items():
                        if 'Ø±Ø¨ÙŠØ© Ø§Ù„Ø§Ø¬ØªÙ…Ø§Ø¹ÙŠØ© Ùˆ Ø§Ù„ÙˆØ·Ù†ÙŠØ©' in key :
                            # print(key ,round(value[0]*2/3),1)
                            term_1_avarage +=round(value[0]*2/3,1)
                            term_2_avarage +=round(value[1]*2/3,1)
                            # year_avarage += round((term_1_avarage + term_2_avarage)/2,1)
                        elif ('Ø§Ù„ØªØ±Ø¨ÙŠØ© Ø§Ù„ÙÙ†ÙŠØ© ÙˆØ§Ù„Ù…ÙˆØ³ÙŠÙ‚ÙŠØ©' in key or 'Ø§Ù„ØªØ±Ø¨ÙŠØ© Ø§Ù„Ø±ÙŠØ§Ø¶ÙŠØ©' in key) and skip_art_sport :
                            pass
                        else:
                            # print(key , value[0])
                            term_1_avarage += value[0]
                            term_2_avarage += value[1]
                            # year_avarage += round((term_1_avarage + term_2_avarage)/2,1)
                    term_1_avarage ,term_2_avarage ,year_avarage =round((term_1_avarage / 2000)* 100,1) , round((term_2_avarage / 2000)* 100,1) , round((((term_1_avarage+term_2_avarage)/2) / 2000)* 100,1)
                    item['t1+t2+year_avarage'] = [term_1_avarage ,term_2_avarage ,year_avarage ]

                else:
                    if 'Ø¹Ø´Ø±' not in item['student_grade_name']:
                        for key, value in item['subject_sums'].items():
                            if 'Ø±Ø¨ÙŠØ© Ø§Ù„Ø§Ø¬ØªÙ…Ø§Ø¹ÙŠØ© Ùˆ Ø§Ù„ÙˆØ·Ù†ÙŠØ©' in key :
                                # print(key ,round(value[0]*2/3),1)
                                term_1_avarage +=round(value[0]*2/3,1)
                                term_2_avarage +=round(value[1]*2/3,1)
                                # year_avarage += round((term_1_avarage + term_2_avarage)/2,1)
                            elif ('Ø§Ù„ØªØ±Ø¨ÙŠØ© Ø§Ù„ÙÙ†ÙŠØ© ÙˆØ§Ù„Ù…ÙˆØ³ÙŠÙ‚ÙŠØ©' in key or 'Ø§Ù„ØªØ±Ø¨ÙŠØ© Ø§Ù„Ø±ÙŠØ§Ø¶ÙŠØ©' in key) and skip_art_sport :
                                pass                        
                            else:
                                # print(key , value[0])
                                term_1_avarage += value[0]
                                term_2_avarage += value[1]
                                # year_avarage += round((term_1_avarage + term_2_avarage)/2,1)
                        term_1_avarage ,term_2_avarage ,year_avarage =round((term_1_avarage / 800)* 100,1) , round((term_2_avarage / 800)* 100,1) , round((((term_1_avarage+term_2_avarage)/2) / 800)* 100,1)
                        item['t1+t2+year_avarage'] = [term_1_avarage ,term_2_avarage ,year_avarage ]
            except:
                pass

def add_subject_sum_dictionary (grouped_dict_list , based_on_subject_name=True , based_on_subject_id = False):
    """
    The function takes a list of dictionaries and returns a new dictionary with the sum of values for
    each subject.
    
    :param grouped_dict_list: A list of dictionaries where each dictionary represents a group of
    subjects and their corresponding sums
    """
    subject_sums = {}
    for group in grouped_dict_list:
        for items in group:
            if len(items['subjects_assessments_info']) > 0 :
                science_sum ,social_sum ,subject_sum ,science_sum_t2 ,social_sum_t2 ,subject_sum_t2 =  [0] * 6
                for i in items['subjects_assessments_info'][0]:
                    if "Ø¹Ù„ÙˆÙ… Ø§Ù„Ø£Ø±Ø¶" in i['subject_name'] or 'Ø§Ù„ÙƒÙŠÙ…ÙŠØ§Ø¡' in i['subject_name'] or 'Ø§Ù„Ø­ÙŠØ§ØªÙŠØ©' in i['subject_name'] or 'Ø§Ù„ÙÙŠØ²ÙŠØ§Ø¡' in i['subject_name'] or 'Ø§Ù„Ø¹Ù„ÙˆÙ…' in i['subject_name']:
                        # compute sum for science subjects
                        science_sum +=  sum(int(i['term1'][key]) for key in i['term1'] if re.compile(r'^assessment\d+$').match(key) and 'max_mark' not in key and i['term1'][key])
                        science_sum_t2 +=  sum(int(i['term2'][key]) for key in i['term2'] if re.compile(r'^assessment\d+$').match(key) and 'max_mark' not in key and i['term2'][key])
                    elif 'ØªØ±Ø¨ÙŠØ© Ø§Ù„ÙˆØ·Ù†ÙŠØ©' in i['subject_name'] or 'Ø¬ØºØ±Ø§ÙÙŠØ§' in i['subject_name'] or 'ØªØ§Ø±ÙŠØ®' in i['subject_name'] or 'ØªØ±Ø¨ÙŠØ© Ø§Ù„Ø§Ø¬ØªÙ…Ø§Ø¹ÙŠ' in i['subject_name'] or 'Ø¬ØªÙ…Ø§Ø¹ÙŠ' in i['subject_name'] or 'ÙˆØ·Ù†ÙŠ' in i['subject_name']:
                        # compute sum for social subjects
                        social_sum +=  sum(int(i['term1'][key]) for key in i['term1'] if re.compile(r'^assessment\d+$').match(key) and 'max_mark' not in key and i['term1'][key])
                        social_sum_t2 +=  sum(int(i['term2'][key]) for key in i['term2'] if re.compile(r'^assessment\d+$').match(key) and 'max_mark' not in key and i['term2'][key])
                    else:
                        # compute sum for other subjects
                        subject_sum = sum(int(i['term1'][key]) for key in i['term1'] if re.compile(r'^assessment\d+$').match(key) and 'max_mark' not in key and i['term1'][key])
                        subject_sum_t2 = sum(int(i['term2'][key]) for key in i['term2'] if re.compile(r'^assessment\d+$').match(key) and 'max_mark' not in key and i['term2'][key])
                        # update dictionary with other subject sum
                        subject_sums[i['subject_name']] = [subject_sum,subject_sum_t2]
                    # Ø§Ø¶Ù Ù‚ÙŠÙ… Ø§Ù„Ø¹Ù„ÙˆÙ… Ùˆ Ø§Ù„Ø§Ø¬ØªÙ…Ø§Ø¹ÙŠØ§Øª Ù…Ù‡Ù…Ø§ ÙƒØ§Ù†Øª Ù‚ÙŠÙ… Ø§Ù„Ù…ÙˆØ§Ø¯
                    subject_sums['Ø§Ù„Ø¹Ù„ÙˆÙ…'] = [science_sum ,science_sum_t2]
                    subject_sums['Ø§Ù„ØªØ±Ø¨ÙŠØ© Ø§Ù„Ø§Ø¬ØªÙ…Ø§Ø¹ÙŠØ© Ùˆ Ø§Ù„ÙˆØ·Ù†ÙŠØ©'] = [social_sum,social_sum_t2]

                    # update dictionary with social subject sum
                # print (items['student__full_name'],items['student_class_name_letter'],subject_sums)
                items['subject_sums'] = subject_sums
                subject_sums={}

def playsound(debug =False):
    """
    The function playsound plays a sound if the debug parameter is set to True.
    
    :param debug: A boolean parameter that determines whether or not to print debug information. If set
    to True, debug information will be printed. If set to False, debug information will not be printed,
    defaults to False (optional)
    """
    # Execute the shell command to play a sine wave sound with frequency 440Hz for 2 seconds
    subprocess.run(['play', '-n', 'synth', '2', 'sin', '440'])
    if debug :
        pdb.set_trace()

def group_students(dic_list4 , i = None):
    """
    The function "group_students" takes a list of dictionaries and an optional index as input and
    returns a new list of dictionaries grouped by the value of the specified index.
    
    :param dic_list4: A list of dictionaries where each dictionary represents a student and their
    information
    :param i: The parameter "i" is an optional parameter that represents the index of the student group
    to be returned. If no index is provided, the function will return all student groups
    """
    # sort the list based on the 'class_name' key
    sorted_list = sorted(dic_list4, key=lambda x: x['student_class_name_letter'])

    # group the sorted list by the 'class_name' key
    grouped_list = []
    for key, group in itertools.groupby(sorted_list, key=lambda x: x['student_class_name_letter']):
        group_list = list(group)
        if all(x.get('student_class_name_letter') for x in group_list):
            grouped_list.append(group_list)
    if i :
        for i in grouped_list:
            print(len(i),i[0]['student_class_name_letter'])
        return 0
    else : 
        return grouped_list

def get_students_info_subjectsMarks(username , password , student_identity_number = None , empty_marks = False , just_teacher_class=True , session=None , curr_year=None ,student_status_ids=[1] , auth=None):
    """
    Ø¯Ø§Ù„Ø© Ù„Ø§Ø³ØªØ®Ø±Ø§Ø¬ Ù…Ø¹Ù„ÙˆÙ…Ø§Øª Ùˆ Ø¹Ù„Ø§Ù…Ø§Øª Ø§Ù„Ø·Ù„Ø§Ø¨ Ù„Ø§Ø³ØªØ®Ø¯Ø§Ù…Ù‡Ø§ Ù„Ø§Ø­Ù‚Ø§ ÙÙŠ Ø§Ù†Ø´Ø§Ø¡ Ø§Ù„Ø¬Ø¯Ø§ÙˆÙ„ Ùˆ Ø§Ù„Ø¹Ù„Ø§Ù…Ø§Øª
    """
    if auth is None:
        auth=get_auth(username,password)
    dic_list=[]
    target_student_marks=[]
    inst_data = inst_name(session=session,auth=auth)['data'][0]
    school_name = inst_data['Institutions']['name']
    inst_id = inst_data['Institutions']['id']
    edu_directory = inst_area(session=session,auth=auth)['data'][0]['Areas']['name']
    if curr_year is None :
        curr_year = get_curr_period(auth,session)['data'][0]['id']

    # get the assessment periods dictionary 
    assessment_periods  = { 'data':get_assessment_periods_list(auth)}
    
    # class_subject_teacher_mapping = get_class_subject_teacher_mapping_dictionary( class_data_with_subjects_dictionary , subject_mapping_for_teachers , teacher_with_subject_mapping)
    subjects_list = get_subjects_dictionary_list_from_the_site(auth ,session)
    
    subjects_assessments_info=[]

    if not student_identity_number : 
        class_data_dic , students_with_data_dic = get_school_classes_and_students_with_classes(auth ,inst_id , curr_year ,session=session)
        
        if just_teacher_class:
            teacher_class_id = get_teacher_homeroom_class(auth, inst_id , curr_year)['data'][0]['id']
            students = get_teacher_class_students(auth , teacher_class_id , inst_id , curr_year , just_id_and_name_and_empty_marks=False,session=session,student_status_ids =student_status_ids)['data']
            class_data_dic = {teacher_class_id : class_data_dic[teacher_class_id]}
        else : 
            students = get_school_students_ids(session=session,auth=auth,curr_year=curr_year,student_status_ids =student_status_ids)
        
        for i in students:
            dic_list.append(
                        {
                            'student_id':i['student_id'],
                            'student__full_name':i['user']['name'],
                            'student_nat':i['user']['nationality_id'],
                            'student_birth_place':i['user']['birthplace_area_id'] if i['user']['birthplace_area_id'] is not None and i['user']['birthplace_area_id'] != 'None' else '' ,
                            'student_birth_date' : i['user']['date_of_birth'] ,
                            'student_nat_id': '' if i['user']['identity_number'] is None else i['user']['identity_number'],
                            'student_grade_id':i['education_grade_id'],
                            'student_grade_name' : i['education_grade_id'] ,
                            'student_class_name_letter': '' if not isinstance(i['institution_class_id'], int) else i['institution_class_id'],
                            'student_edu_place' : edu_directory ,
                            'student_directory':edu_directory,
                            'student_school_name':school_name,
                            'subjects_assessments_info':[] ,
                        }
                        )

        # add subjects to the class dictionary variable which is class_data_dic
        class_data_with_subjects_dictionary = add_subjects_to_class_data_dic(auth,inst_id , curr_year ,class_data_dic,session=session)
        global open_emis_core_marks
        open_emis_core_marks = get_school_marks_version_2(auth,inst_id , curr_year ,class_data_dic)

        # get the teachers or staff data (what the subjects they teach and the class names)
        SubjectStaff_data = make_request(auth=auth , url=f'https://emis.moe.gov.jo/openemis-core/restful/v2/Institution-InstitutionSubjectStaff.json?institution_id={inst_id}&academic_period_id={curr_year}&_contain=Users,InstitutionSubjects&_limit=0',session=session)['data']

        # map the followings 
        # teachers load  
        # subjects for each teacher  
        # the teacher with subjects
        staff_load_mapping = {
                            x['staff_id'] : {
                                'name': x['user']['name'],
                                'teacher_subjects':
                                    [
                                        {
                                            'subject_class_id' :i['institution_subject']['id'] ,
                                            'subject_name' :i['institution_subject']['name'] ,
                                            'subject_grade_id' :i['institution_subject']['education_grade_id'],
                                            'subject_id' :i['institution_subject']['education_subject_id'] ,
                                        
                                        } for i in SubjectStaff_data if x['staff_id'] == i['staff_id']
                                    ]
                                }
                            for x in SubjectStaff_data
                                if x['end_date'] is None
                            }
        subject_mapping_for_teachers = {
                                        i['id'] : { 
                                                'name': i['name'] , 
                                                'class_id': class_id ,
                                                'class_name' : class_data_dic[class_id]['name'] ,
                                                'education_subject_id': i['education_subject_id']
                                                }    
                                        for class_id in class_data_with_subjects_dictionary 
                                        for i in class_data_with_subjects_dictionary[class_id]['subjects']
                                        }
        teacher_with_subject_mapping = {
                                            i['subject_class_id'] : { 
                                                    'teacher_name': staff_load_mapping[teacher_id]['name'] , 
                                                    'education_subject_name': i['subject_name'],
                                                    'education_subject_id': i['subject_id']
                                                    }    
                                            for teacher_id in staff_load_mapping 
                                            for i in staff_load_mapping[teacher_id]['teacher_subjects']
                                        }

    else:
        for i in make_request(auth=auth, url=f'https://emis.moe.gov.jo/openemis-core/restful/v2/Institution-InstitutionClassStudents.json?_limit=5&_finder=Users.address_area_id,Users.birthplace_area_id,Users.gender_id,Users.date_of_birth,Users.date_of_death,Users.nationality_id,Users.identity_number,Users.external_reference,Users.status&identity_number={student_identity_number}&academic_period_id={curr_year}&_contain=Users',session=session)['data']:
            dic_list.append(
                            {
                                'student_id':i['student_id'],
                                'student__full_name':i['user']['name'],
                                'student_nat':i['user']['nationality_id'],
                                'student_birth_place':i['user']['birthplace_area_id'] if i['user']['birthplace_area_id'] is not None and i['user']['birthplace_area_id'] != 'None' else '' ,
                                'student_birth_date' : i['user']['date_of_birth'] ,
                                'student_nat_id': '' if i['user']['identity_number'] is None else i['user']['identity_number'],
                                'student_grade_id':i['education_grade_id'],
                                'student_grade_name' : i['education_grade_id'] ,
                                'student_class_name_letter': '' if not isinstance(i['institution_class_id'], int) else i['institution_class_id'],
                                'student_edu_place' : edu_directory ,
                                'student_directory':edu_directory,
                                'student_school_name':school_name,
                                'subjects_assessments_info':[] ,
                            }
                            )
        open_emis_core_marks = make_request(auth=auth , url=f'https://emis.moe.gov.jo/openemis-core/restful/Assessment.AssessmentItemResults?_fields=created_user_id,AssessmentGradingOptions.name,AssessmentGradingOptions.min,AssessmentGradingOptions.max,EducationSubjects.name,EducationSubjects.code,AssessmentPeriods.code,AssessmentPeriods.name,AssessmentPeriods.academic_term,marks,assessment_grading_option_id,student_id,assessment_id,education_subject_id,education_grade_id,assessment_period_id,institution_classes_id&academic_period_id={curr_year}&_contain=AssessmentPeriods,AssessmentGradingOptions,EducationSubjects&student_id={dic_list[0]["student_id"]}&_limit=1000')['data'] # 2001419515
        
    
    # Ù…Ø¹Ù†Ù‰ Ù‡Ø°Ù‡ Ø§Ù„Ø¬Ù…Ù„Ø© Ø§Ù„ØªÙƒØ±Ø§Ø±ÙŠØ© Ù‡Ùˆ Ù„ÙƒÙ„ Ø·Ø§Ù„Ø¨ Ù…Ù† Ø§Ù„Ø·Ù„Ø§Ø¨ Ø§Ù„Ù…ÙˆØ¬ÙˆØ¯ÙŠÙ† ÙÙŠ Ø§Ù„Ù‚Ø§Ù…ÙˆØ³
    for student_data in dic_list:
        student_id= student_data['student_id']
        subject_dict = {'subject_name':'','subject_number':'',
                        'term1':{ 'assessment1': '','max_mark_assessment1':'' ,'assessment2': '','max_mark_assessment2':'' , 'assessment3': '','max_mark_assessment3':'' , 'assessment4': '','max_mark_assessment4':''} ,
                        'term2':{ 'assessment1': '','max_mark_assessment1':'' ,'assessment2': '','max_mark_assessment2':'' , 'assessment3': '','max_mark_assessment3':'' , 'assessment4': '','max_mark_assessment4':''}}
        # Ø§Ù„Ø¹Ù„Ø§Ù…Ø§Øª Ø§Ù„ØªÙŠ Ø§Ø³ØªØ®Ø±Ø¬ØªÙ‡Ø§ Ù…Ù† Ø§Ù„Ø±Ø§Ø¨Ø·
        # Ø§Ø³ØªØ·ÙŠØ¹ Ø§Ù„Ø§Ø³ØªØºÙ†Ø§Ø¡ Ø¹Ù†Ù‡Ø§ Ø¨Ø§Ø³ØªØ®Ø¯Ø§Ù… Ø¯Ø§Ù„Ø© 
        target_student_marks = [ mark for mark in open_emis_core_marks if mark['student_id'] == student_id ]
        
        #TODO: Ø§Ø­Ø°Ù Ù‡Ø°Ø§ Ø§Ù„Ù…ØªØºÙŠØ± Ø§Ù„Ø¹Ø§Ù„Ù…ÙŠ Ù„Ø§Ù†Ù‡
        global target_student_subjects
        
        # Ø±ØªØ¨ Ø§Ù„Ù…ÙˆØ§Ø¯ Ø­Ø³Ø¨ Ø±Ù‚Ù… Ø§Ù„Ù…Ø§Ø¯Ø© Ùˆ Ø§Ø­Ø°Ù Ø§Ù„Ù…ØªÙƒØ±Ø± 
        target_student_subjects = list(set(d['education_subject_id'] for d in target_student_marks))
        for subject in target_student_subjects:
            assessments_list = [assessment for assessment in target_student_marks if subject == assessment['education_subject_id']]
            
            subject_data = [i for i in subjects_list['data'] if i['assessment_period_id'] == int(assessments_list[0]['assessment_period_id']) and i['education_subject_id'] == subject][0]
            
            subject_dict['subject_name'] = subject_data['education_subject']['name']
            subject_dict['subject_number']= subject
            
            # Ø§Ø°Ø§ ØªÙ… Ø§Ø¹Ø·Ø§Ø¡ Ø±Ù‚Ù… Ø§Ù„Ø·Ø§Ù„Ø¨ Ø§Ù„ÙˆØ·Ù†ÙŠ
            if student_identity_number is not None:
                mark_key = 'marks'
            else:
                mark_key = 'mark'
            
            values = offline_sort_assessement_period_ids_v2( assessments_list , assessment_periods)
            subject_dict['assessments_periods_ides'] = [int(x) for x in [i['assessment_period_id'] for i in values ] if x is not None]
            subject_dict['term1']['assessment1'] = float(values[0][mark_key]) if values[0][mark_key] is not None and not empty_marks else ''
            subject_dict['term1']['assessment2'] = float(values[1][mark_key]) if values[1][mark_key] is not None and not empty_marks else ''
            subject_dict['term1']['assessment3'] = float(values[2][mark_key]) if values[2][mark_key] is not None and not empty_marks else ''
            subject_dict['term1']['assessment4'] = float(values[3][mark_key]) if values[3][mark_key] is not None and not empty_marks else ''
            subject_dict['term2']['assessment1'] = float(values[4][mark_key]) if values[4][mark_key] is not None and not empty_marks else ''
            subject_dict['term2']['assessment2'] = float(values[5][mark_key]) if values[5][mark_key] is not None and not empty_marks else ''
            subject_dict['term2']['assessment3'] = float(values[6][mark_key]) if values[6][mark_key] is not None and not empty_marks else ''
            subject_dict['term2']['assessment4'] = float(values[7][mark_key]) if values[7][mark_key] is not None and not empty_marks else ''
            
            # Ù„Ø§ Ø§Ø­ØªØ§Ø¬ Ø§Ù„Ø§Ù† Ù„Ù„Ø¹Ù„Ø§Ù…Ø© Ø§Ù„ÙƒØ¨Ø±Ù‰ 
            # sub_dic['term1']['max_mark_assessment1'] = [assessments['assessment_grading_option']['max'] for assessments in dictionaries if 'S1A1' in assessment_periods_dictionary[int(assessments['assessment_period_id'])]['code']][0] if [assessments['assessment_grading_option']['max'] for assessments in dictionaries if 'S1A1' in assessment_periods_dictionary[int(assessments['assessment_period_id'])]['code']] else ''
            # sub_dic['term1']['max_mark_assessment2'] = [assessments['assessment_grading_option']['max'] for assessments in dictionaries if 'S1A2' in assessment_periods_dictionary[int(assessments['assessment_period_id'])]['code']][0] if [assessments['assessment_grading_option']['max'] for assessments in dictionaries if 'S1A2' in assessment_periods_dictionary[int(assessments['assessment_period_id'])]['code']] else ''
            # sub_dic['term1']['max_mark_assessment3'] = [assessments['assessment_grading_option']['max'] for assessments in dictionaries if 'S1A3' in assessment_periods_dictionary[int(assessments['assessment_period_id'])]['code']][0] if [assessments['assessment_grading_option']['max'] for assessments in dictionaries if 'S1A3' in assessment_periods_dictionary[int(assessments['assessment_period_id'])]['code']] else ''
            # sub_dic['term1']['max_mark_assessment4'] = [assessments['assessment_grading_option']['max'] for assessments in dictionaries if 'S1A4' in assessment_periods_dictionary[int(assessments['assessment_period_id'])]['code']][0] if [assessments['assessment_grading_option']['max'] for assessments in dictionaries if 'S1A4' in assessment_periods_dictionary[int(assessments['assessment_period_id'])]['code']] else ''
            # sub_dic['term2']['max_mark_assessment1'] = [assessments['assessment_grading_option']['max'] for assessments in dictionaries if 'S2A1' in assessment_periods_dictionary[int(assessments['assessment_period_id'])]['code']][0] if [assessments['assessment_grading_option']['max'] for assessments in dictionaries if 'S2A1' in assessment_periods_dictionary[int(assessments['assessment_period_id'])]['code']] else ''
            # sub_dic['term2']['max_mark_assessment2'] = [assessments['assessment_grading_option']['max'] for assessments in dictionaries if 'S2A2' in assessment_periods_dictionary[int(assessments['assessment_period_id'])]['code']][0] if [assessments['assessment_grading_option']['max'] for assessments in dictionaries if 'S2A2' in assessment_periods_dictionary[int(assessments['assessment_period_id'])]['code']] else ''
            # sub_dic['term2']['max_mark_assessment3'] = [assessments['assessment_grading_option']['max'] for assessments in dictionaries if 'S2A3' in assessment_periods_dictionary[int(assessments['assessment_period_id'])]['code']][0] if [assessments['assessment_grading_option']['max'] for assessments in dictionaries if 'S2A3' in assessment_periods_dictionary[int(assessments['assessment_period_id'])]['code']] else ''
            # sub_dic['term2']['max_mark_assessment4'] = [assessments['assessment_grading_option']['max'] for assessments in dictionaries if 'S2A4' in assessment_periods_dictionary[int(assessments['assessment_period_id'])]['code']][0] if [assessments['assessment_grading_option']['max'] for assessments in dictionaries if 'S2A4' in assessment_periods_dictionary[int(assessments['assessment_period_id'])]['code']] else ''
            subjects_assessments_info.append(subject_dict)   
            subject_dict = {'subject_name':'','subject_number':'','term1':{ 'assessment1': '','max_mark_assessment1':'' ,'assessment2': '','max_mark_assessment2':'' , 'assessment3': '','max_mark_assessment3':'' , 'assessment4': '','max_mark_assessment4':''} ,'term2':{ 'assessment1': '','max_mark_assessment1':'' ,'assessment2': '','max_mark_assessment2':'' , 'assessment3': '','max_mark_assessment3':'' , 'assessment4': '','max_mark_assessment4':''}}
            # [dic for dic in dic_list if dic['student_id']==3439303][0]['subjects_assessments_info']
        target_index = next((i for i, dic in enumerate(dic_list) if dic['student_id'] == student_id != 0 ), None)
        if target_index is not None and len(target_student_subjects) != 0:
            dic_list[target_index]['subjects_assessments_info'].append(subjects_assessments_info)
            # dic_list[target_index]['student_class_name_letter'] = dictionaries[0]['institution_classes_id']
            # print(dic_list[target_index])
            subjects_assessments_info=[]
            target_student_marks = []
    
    class_name_letter = list(set([i['student_class_name_letter'] for i in dic_list if i['student_class_name_letter'] != '' ]))
    joined_string = ','.join(str(i) for i in [f'institution_class_id:{i}' for i in class_name_letter])
    classes_data = make_request(session=session,auth=auth,url='https://emis.moe.gov.jo/openemis-core/restful/Institution.InstitutionClassSubjects?status=1&_contain=InstitutionSubjects,InstitutionClasses&_limit=0&_orWhere='+joined_string)['data']
    class_list = []
    for i in classes_data:
        class_list.append({'class_id': i['institution_class_id'] , 'class_name': i['institution_class']['name'] })
        class_dict = {i['class_id']: i['class_name'] for i in class_list if i['class_id'] != ''}
    for i in dic_list:
        class_id = i['student_class_name_letter']
        if class_id != '':
            i['student_class_name_letter'] = class_dict.get(class_id, class_id)
    grade_id = list(set([i['student_grade_id'] for i in dic_list if i['student_grade_id'] != '' ]))
    grade_data = get_grade_info(auth=auth,session=session,period_id=curr_year)
    grade_list = []
    for i in grade_data:
        grade_list.append({'grade_id': i['education_grade_id'] , 'grade_name': re.sub('.*Ù„Ù„ØµÙ','Ø§Ù„ØµÙ', i['name']) })
        grade_dict = {i['grade_id']: i['grade_name'] for i in grade_list if i['grade_name'] != ''}
    for i in dic_list:
        grade_id = i['student_grade_id']
        if grade_id != '':
            i['student_grade_name'] = grade_dict.get(grade_id, grade_id)
    
    nat_id = list(set([i['student_grade_id'] for i in dic_list if i['student_grade_id'] != '' ]))
    birth_place_id = list(set([i['student_grade_id'] for i in dic_list if i['student_grade_id'] != '' ]))
    birth_place_data = make_request(session=session,auth=auth , url='https://emis.moe.gov.jo/openemis-core/restful/v2/Area-AreaAdministratives?_limit=0&_contain=AreaAdministrativeLevels')['data']
    nationality_data = make_request(session=session,auth=auth , url='https://emis.moe.gov.jo/openemis-core/restful/v2/User-NationalityNames')['data']
    nationality_data_list = []
    birth_place_list = []
    for i in nationality_data:
        nationality_data_list.append({'nat_id': i['id'] , 'nat_name': i['name'] })
        nattionality_dict = {i['nat_id']: i['nat_name'] for i in nationality_data_list if i['nat_name'] != ''}
    for i in dic_list:
        nat_id = i['student_nat']
        if nat_id != '':
            i['student_nat'] = nattionality_dict.get(nat_id, nat_id)

    for i in birth_place_data:
        birth_place_list.append({'birth_place_id': i['id'] , 'birth_place_name': i['name'] })
        birth_place_dict = {i['birth_place_id']: i['birth_place_name'] for i in birth_place_list if i['birth_place_name'] != ''}
    for i in dic_list:
        birth_place_id = i['student_birth_place']
        if birth_place_id != '':
            i['student_birth_place'] = birth_place_dict.get(birth_place_id, birth_place_id)

    return dic_list

def get_school_students_ids(auth, inst_id=None ,curr_year=None,student_status_ids=[1],session=None ):
    """
    Retrieves the IDs of students enrolled in a school for the current academic year.

    Parameters:
        auth (dict): Authentication information.
        inst_id (int, optional): Institution ID. Defaults to None.
        curr_year (int, optional): Current academic year ID. Defaults to None.
        session (requests.sessions.Session, optional): Requests session. Defaults to None.

    Returns:
        list: List of student IDs.
    """    
    if inst_id is None:
        inst_id = inst_name(auth,session=session)['data'][0]['Institutions']['id']
    if curr_year is None:
        curr_year = get_curr_period(auth,session=session)['data'][0]['id']
    students = [
                i['student_id'] 
                for i in make_request(session=session ,auth=auth,url=f'https://emis.moe.gov.jo/openemis-core/restful/v2/Institution.Students?_limit=0&_finder=Users.address_area_id,Users.birthplace_area_id,Users.gender_id,Users.date_of_birth,Users.date_of_death,Users.nationality_id,Users.identity_number,Users.external_reference,Users.status&institution_id={inst_id}&academic_period_id={curr_year}&_contain=Users')['data']
                    
                    if i['student_status_id'] in student_status_ids
                ]
    InstitutionClassStudents = [
                                i 
                                for i in make_request(auth=auth, url=f'https://emis.moe.gov.jo/openemis-core/restful/v2/Institution-InstitutionClassStudents.json?_limit=0&_finder=Users.address_area_id,Users.birthplace_area_id,Users.gender_id,Users.date_of_birth,Users.date_of_death,Users.nationality_id,Users.identity_number,Users.external_reference,Users.status&institution_id={inst_id}&academic_period_id={curr_year}&_contain=Users',session=session)['data'] 
                                    if i['student_status_id'] in student_status_ids and i['student_id'] in students
                                ]
    return [
            i
            for i in InstitutionClassStudents
            
            ]

def fill_official_marks_a3_two_face_doc2_offline_version(students_data_lists, ods_file ):
    """
    This function fills the official marks for A3 Two Face Doc2 in the offline version.
    
    :param students_data_lists: A list of lists containing the data of each student. Each inner list
    should contain the following information in order:
    :param ods_file: The ods_file parameter is the file path or name of the OpenDocument Spreadsheet
    (ODS) file that contains the official marks for the students
    """
    
    context = {'46': 'A6:A30', '4': 'A39:A63', '3': 'L6:L30', '45': 'L39:L63', '44': 'A71:A95', '6': 'A103:A127', '5': 'L71:L95', '43': 'L103:L127', '42': 'A135:A159', '8': 'A167:A191', '7': 'L135:L159', '41': 'L167:L191', '40': 'A199:A223', '10': 'A231:A255', '9': 'L199:L223', '39': 'L231:L255', '38': 'A263:A287', '12': 'A295:A319', '11': 'L263:L287', '37': 'L295:L319', '36': 'A327:A351', '14': 'A359:A383', '13': 'L327:L351', '35': 'L359:L383', '34': 'A391:A415', '16': 'A423:A447', '15': 'L391:L415', '33': 'L423:L447', '32': 'A455:A479', '18': 'A487:A511', '17': 'L455:L479', '31': 'L487:L511', '30': 'A519:A543', '20': 'A551:A575', '19': 'L519:L543', '29': 'L551:L575', '28': 'A583:A607', '22': 'A615:A639', '21': 'L583:L607', '27': 'L615:L639', '26': 'A647:A671', '24': 'A679:A703', '23': 'L647:L671', '25': 'L679:L703'}
    
    page = 4
    name_counter = 1
    name_counter = 1
    
    # Open the ODS file and load the sheet you want to fill
    doc = ezodf.opendoc(ods_file)
    
    sheet_name = 'sheet'
    sheet = doc.sheets[sheet_name]


    for students_data_list in students_data_lists:
        
#         ['Ø§Ù„ØµÙ Ø§Ù„Ø³Ø§Ø¨Ø¹', 'Ø£', 'Ø§Ù„Ù„ØºØ© Ø§Ù„Ø§Ù†Ø¬Ù„ÙŠØ²ÙŠØ©', '786118']
        
        class_data = students_data_list['class_name'].split('=')
        # mawad.append(class_data[2])
        # classes.append('-'.join([class_data[0],class_data[1]]))
        class_name = class_data[0].replace('Ø§Ù„ØµÙ ' , '').split('-')[0]
        class_char = class_data[0].split('-')[1]
        sub_name = class_data[1]
        
        sheet[f"D{int(context[str(page)].split(':')[0][1:])-5 }"].set_value(f' Ø§Ù„ØµÙ: {class_name}')
        sheet[f"I{int(context[str(page)].split(':')[0][1:])-5 }"].set_value(f'Ø§Ù„Ø´Ø¹Ø¨Ø© (   {class_char}    )')    
        sheet[f"O{int(context[str(page+1)].split(':')[0][1:])-5}"].set_value(sub_name)

        for counter,student_info in enumerate(students_data_list['students_data'], start=1):
            if counter >= 26:
                page += 2
                counter = 1
                
                sheet[f"D{int(context[str(page)].split(':')[0][1:])-5}"].set_value(f' Ø§Ù„ØµÙ: {class_name}')
                sheet[f"I{int(context[str(page)].split(':')[0][1:])-5}"].set_value(f'Ø§Ù„Ø´Ø¹Ø¨Ø© (   {class_char}    )')  
                sheet[f"O{int(context[str(page+1)].split(':')[0][1:])-5}"].set_value(sub_name)
                #    Ø§Ù„Ù…Ø§Ø¯Ø© Ø§Ù„Ø¯Ø±Ø§Ø³ÙŠØ©     
                
                # {'id': 3824166, 'name': 'Ù†ÙˆØ±Ø§Ù„Ø¯ÙŠÙ† Ù…Ø­Ù…ÙˆØ¯ Ø±Ø§Ø¶ÙŠ Ø§Ù„Ø¯ØºÙŠÙ…Ø§Øª', 'term1': {'assessment1': 9, 'assessment2': 10, 'assessment3': 11, 'assessment4': 20}}
                
                for student_info in students_data_list['students_data'][25:] :
                    row_idx = counter + int(context[str(page)].split(':')[0][1:]) - 1  # compute the row index based on the counter
                    sheet[f"A{row_idx}"].set_value(name_counter)
                    sheet[f"B{row_idx}"].set_value(student_info['name'])
                    if 'term1' in student_info and 'assessment1' in student_info['term1'] and 'assessment2' in student_info['term1'] and 'assessment3' in student_info['term1'] and 'assessment4' in student_info['term1']:
                        sheet[f"D{row_idx}"].set_value(student_info['term1']['assessment1']) 
                        sheet[f"E{row_idx}"].set_value(student_info['term1']['assessment2']) 
                        sheet[f"F{row_idx}"].set_value(student_info['term1']['assessment3'])
                        sheet[f"G{row_idx}"].set_value(student_info['term1']['assessment4'])
                    if 'term2' in student_info:
                        row_idx2 = counter + int(context[str(page+1)].split(':')[0][1:]) - 1  # compute the row index based on the counter 
                        sheet[f"L{row_idx2}"].set_value(student_info['term2']['assessment1']) 
                        sheet[f"M{row_idx2}"].set_value(student_info['term2']['assessment2']) 
                        sheet[f"N{row_idx2}"].set_value(student_info['term2']['assessment3'])
                        sheet[f"O{row_idx2}"].set_value(student_info['term2']['assessment4'])                       
                    counter += 1
                    name_counter += 1              
                break                    
            row_idx = counter + int(context[str(page)].split(':')[0][1:]) - 1  # compute the row index based on the counter
            sheet[f"A{row_idx}"].set_value(name_counter)
            sheet[f"B{row_idx}"].set_value(student_info['name']) 
            if 'term1' in student_info and 'assessment1' in student_info['term1'] and 'assessment2' in student_info['term1'] and 'assessment3' in student_info['term1'] and 'assessment4' in student_info['term1']:
                sheet[f"D{row_idx}"].set_value(student_info['term1']['assessment1']) 
                sheet[f"E{row_idx}"].set_value(student_info['term1']['assessment2']) 
                sheet[f"F{row_idx}"].set_value(student_info['term1']['assessment3'])
                sheet[f"G{row_idx}"].set_value(student_info['term1']['assessment4'])
            if 'term2' in student_info:
                row_idx2 = counter + int(context[str(page+1)].split(':')[0][1:]) - 1  # compute the row index based on the counter 
                sheet[f"L{row_idx2}"].set_value(student_info['term2']['assessment1']) 
                sheet[f"M{row_idx2}"].set_value(student_info['term2']['assessment2']) 
                sheet[f"N{row_idx2}"].set_value(student_info['term2']['assessment3'])
                sheet[f"O{row_idx2}"].set_value(student_info['term2']['assessment4'])                
            name_counter += 1 
        name_counter = 1
        page += 2

    # FIXME: make the customshapes crop _20_ to the rest of the key in the custom_shapes
    # Iterate through the cells of the sheet and fill in the values you want
    doc.save()
    
    return 0
    # return custom_shapes 

def Read_E_Side_Note_Marks_xlsx(file_path=None , file_content=None):
    """
    The function reads the content of an Excel file that contains E side note marks.
    
    :param file_path: The file path is the location of the Excel file that you want to read. It should
    be a string that specifies the full path to the file, including the file name and extension
    :param file_content: The `file_content` parameter is used to pass the content of the Excel file as a
    string. This can be useful if you already have the content of the file stored in a variable and want
    to pass it directly to the function without reading it from a file
    :returns: the read content as dictionary
    """
    if file_content is None:
        # Load the workbook
        wb = load_workbook(file_path)
    else:
        wb = load_workbook(filename=file_content)
        
    sheets = wb.sheetnames[:-1]
    # sheet = wb[wb.sheetnames[0]]
    info_sheet = wb[wb.sheetnames[-1]]
    read_file_output_lists = []

    for sheet in sheets :
        rows = []
        data = []
        # Loop over the rows in each sheet
        for row in wb[sheet].iter_rows(min_row=3, values_only=True):
            row = [cell if cell is not None else '' for cell in row]
            # Append the row data to the list
            rows.append(list(row))    

        for row in rows:
            if row[1] != '' or row[2] != '': 
                dic = {
                    'id': row[1], 
                    'name':  row[2],
                    'term1': {'assessment1':  row[3], 'assessment2':row[4], 'assessment3': row[5], 'assessment4': row[6]},
                    'term2': {'assessment1': row[8], 'assessment2': row[9], 'assessment3': row[10], 'assessment4': row[11]}
                        }
                data.append(dic)
        idsClass=wb[sheet]['AM1'].value
        temp_dic = {'class_name':f"{sheet}={idsClass}" ,"students_data": data}
        read_file_output_lists.append(temp_dic)
    
    modified_classes = []

    classes = [i['class_name'].split('=')[0] for i in read_file_output_lists]
    mawad = [i['class_name'].split('=')[1] for i in read_file_output_lists]
    for i in classes: 
        modified_classes.append(get_class_short(i))
        
    school_id=info_sheet['A1'].value    
    school_name = info_sheet['A2'].value.split('=')[0]
    modeeriah = info_sheet['A3'].value
    hejri1 = info_sheet['A4'].value
    hejri2 = info_sheet['A5'].value
    melady1 = info_sheet['A6'].value
    melady2 = info_sheet['A7'].value
    baldah = info_sheet['A8'].value
    modified_classes = ' ØŒ '.join(modified_classes)
    mawad = sorted(set(mawad))
    mawad = ' ØŒ '.join(mawad)
    teacher = info_sheet['A9'].value
    required_data_mrks_text = info_sheet['A10'].value
    period_id = info_sheet['A11'].value
    custom_shapes = {
    'modeeriah': f'{modeeriah}',
    'hejri1': hejri1,
    'hejri2': hejri2,
    'melady1': melady1,
    'melady2': melady2,
    'baldah': baldah,
    'school': school_name,
    'classes': modified_classes,
    'mawad': mawad,
    'teacher' : teacher,
    'modeeriah_20_2': f'{modeeriah}',
    'hejri_20_1': hejri1,
    'hejri_20_2': hejri2,
    'melady_20_1': melady1,
    'melady_20_2': melady2,
    'baldah_20_2': baldah,
    'school_20_2': school_name,
    'classes_20_2': modified_classes,
    'mawad_20_2': mawad,
    'teacher_20_2': teacher ,
    'modeeriah_20_1': f'{modeeriah}',
    'hejri1': hejri1,
    'hejri2': hejri2,
    'melady1': melady1,
    'melady2': melady2,
    'baldah_20_1': baldah,
    'school_20_1': school_name,
    'classes_20_1': modified_classes,
    'mawad_20_1': mawad,
    'teacher_20_1': teacher,
    'period_id': period_id,
    'school_id' : school_id 
    }
    
    try:
        required_data_mrks_dic_list = {
                                        int(item.split('-')[0]): 
                                            {
                                                'assessment_grade_id': int(item.split('-')[1].split(',')[0]),
                                                'grade_id': int(item.split(',')[0].split('-')[2]), 
                                                'assessments_period_ids': item.split(',')[1:]
                                            }
                                        for item in required_data_mrks_text.split('\\\\')
                                    }
    except Exception as e:
        required_data_mrks_dic_list = {
                                        0:
                                            {
                                                'assessment_grade_id': 0,
                                                'grade_id': 0, 
                                                'assessments_period_ids': 0
                                            }
                                        }

    read_file_output_dict = {'file_data': read_file_output_lists ,
                            'custom_shapes' : custom_shapes ,
                            'required_data_for_mrks_enter' : required_data_mrks_dic_list }
    
    return read_file_output_dict

def enter_marks_arbitrary_controlled_version(username , password , required_data_list ,AssessId=None, assess_period_data=None ,range1='' , range2=''):
    """
    This function allows a user to enter marks for a specific assessment, with optional range
    restrictions. and if the function is provided without range1 or range2 then it will empty the 
    marks.
    
    :param username: The username is a string that represents the username of the user trying to access
    the system. It is used for authentication purposes
    :param password: The password parameter is used to authenticate the user and ensure that only
    authorized users can access and enter marks
    :param required_data_list: A list of required data for entering marks. This could include student
    names, IDs, or any other relevant information
    :param AssessId: The AssessId parameter is used to identify the assessment for which the marks are
    being entered. It could be a unique identifier or a name that helps identify the assessment
    :param range1: range1 is the lower limit of the range of marks that can be entered. If a value is
    provided for range1, it means that the marks entered must be greater than or equal to range1. If no
    value is provided, there is no lower limit for the marks and it will but empty mark.
    :param range2: The parameter "range2" is an optional parameter that specifies the upper range limit
    for the marks. If not provided, the upper range limit will be considered as the maximum possible
    value
    """
    auth = get_auth(username , password)
    period_id = get_curr_period(auth)['data'][0]['id']
    inst_id = inst_name(auth)['data'][0]['Institutions']['id']
    fuzz_postdata_list ,grade_period_ids= [] , []
    
    for item in required_data_list : 
        if  assess_period_data : 
            grade_period_ids = [i for i in assess_period_data if i.get('gradeId') == item['assessment_id']]
            
        for AssessPeriod in grade_period_ids :
            for Student_id in item['students_ids']:
                fuzz_postdata = {
                                    'marks': str("{:.2f}".format(float(random.randint(range1, range2)))) if range1 !='' and range2 !=''  else None,
                                    'assessment_id': item['assessment_id'],
                                    'education_subject_id': item['education_subject_id'],
                                    'education_grade_id': item['education_grade_id'],
                                    'institution_classes_id': item['institution_classes_id'],
                                    'student_id': Student_id,
                                    'assessment_period_id': AssessPeriod['AssesId'] if not AssessId else AssessId,
                                    'action_type': 'default'
                                }
                fuzz_postdata_list.append(json.dumps(fuzz_postdata).replace('{','').replace('}',''))
        
    body_postdata = json.dumps({
            'assessment_grading_option_id': 8,
            'institution_id': inst_id,
            'academic_period_id': period_id,
            'student_status_id': 1,
            'action_type': 'default'}).replace('}',', FUZZ }')

    headers = [("User-Agent" , "python-requests/2.28.1"),("Accept-Encoding" , "gzip, deflate"),("Accept" , "*/*"),("Connection" , "close"),("Authorization" , f"{auth}"),("ControllerAction" , "Results"),("Content-Type" , "application/json")]
    
    url = ENTER_MARK_URL
    
    unsuccessful_requests = wfuzz_function(url , fuzz_postdata_list,headers,body_postdata , proxies = [("127.0.0.1","8080","HTTP")])
    
    while len(unsuccessful_requests) != 0:
        unsuccessful_requests = wfuzz_function(url , unsuccessful_requests,headers,body_postdata , proxies = [("127.0.0.1","8080","HTTP")])

    print("All requests were successful!")

def assessments_commands_text(lst):
    """
    The function takes a list as input and returns a string to print it in the telegram bot as massage for the available tests .
    
    :param lst: The parameter "lst" is a list of commands
    """
    S1 = [i for i in lst if i.get('SEname') !='Ø§Ù„ÙØµÙ„ Ø§Ù„Ø«Ø§Ù†ÙŠ']
    S2 = [i for i in lst if i.get('SEname') =='Ø§Ù„ÙØµÙ„ Ø§Ù„Ø«Ø§Ù†ÙŠ']    
    text = ""

    if S1:
        text = 'Ø§Ù„ÙØµÙ„ Ø§Ù„Ø§ÙˆÙ„\n'
        for item in S1:
            text += '/' + item['code'] + ' Ø§Ù„ØµÙ Ø§Ù„' + num2words(int(re.match('G\d{1,}', item['code']).group()[1:]), lang='ar', to='ordinal') + ' ' + item['AssesName'] + ' ' + ' Ø¹Ù„Ø§Ù…Ø© Ø§Ù„Ù†Ø¬Ø§Ø­ ' + str(item['pass_mark']) + ' Ùˆ Ø§Ù„Ø¹Ù„Ø§Ù…Ø© Ø§Ù„Ù‚ØµÙˆÙ‰ ' + str(item['max_mark']) + '\n'

    if S2:
        text += 'Ø§Ù„ÙØµÙ„ Ø§Ù„Ø«Ø§Ù†ÙŠ\n'
        for item in S2:
            text += '/' + item['code'] + ' Ø§Ù„ØµÙ Ø§Ù„' + num2words(int(re.match('G\d{1,}', item['code']).group()[1:]), lang='ar', to='ordinal') + ' ' + item['AssesName'] + ' ' + ' Ø¹Ù„Ø§Ù…Ø© Ø§Ù„Ù†Ø¬Ø§Ø­ ' + str(item['pass_mark']) + ' Ùˆ Ø§Ù„Ø¹Ù„Ø§Ù…Ø© Ø§Ù„Ù‚ØµÙˆÙ‰ ' + str(item['max_mark']) + '\n'

    if not S1 and not S2:
        # change this to send message for user that there is no assessement to fill now
        print("Both S1 and S2 lists are empty.")
    else:
        text += '/All_asses ØªØ¹Ø¨Ø¦Ø© ÙƒÙ„ Ø§Ù„Ø§Ù…ØªØ­Ø§Ù†Ø§Øª Ø§Ù„Ù…ØªÙˆÙØ±Ø© ØªÙ„Ù‚Ø§Ø¦ÙŠØ§'
        return text

def get_editable_assessments( auth , username ,assessment_grade_id=None , class_subject=None,session=None):
    """
    This function returns a list of editable assessments or tests on emis.moe.gov.jo/openemis-core site based on the provided parameters.
    
    :param auth: An authentication object that contains information about the user's credentials and
    permissions
    :param username: The username parameter is used to specify the username of the user for whom we want
    to retrieve the editable assessments
    :param assessment_grade_id: The ID of the grade level for which you want to retrieve editable
    assessments. If not provided, assessments for all grade levels will be retrieved
    :param class_subject: The subject of the class for which you want to get editable assessments
    :param session: The session parameter is used to specify the academic session for which the
    assessments are being retrieved. It could be a specific year or term, for example, "2021-2022" or
    "Term 1"
    """
    if assessment_grade_id is None or class_subject is None:
        required_data_list = get_required_data_to_enter_marks(auth=auth ,username=username,session=session)
        ass_data = [[y['assessment_id'],y['education_subject_id']] for y in required_data_list ]
        ass_data = [item for sublist in [get_all_assessments_periods_data2(auth, i[0],i[1],session=session) for i in ass_data] for item in sublist if item.get('editable')==True]
        # unique_lst = [dict(t) for t in {tuple(sorted(d.items())) for d in lst}]
        unique_dict_list = [dict(t) for t in {tuple(sorted(d.items())) for d in ass_data}]
        sorted_dict = sorted(unique_dict_list , key=lambda x: x['code'])
        return sorted_dict
    else:
        ass_data = [item for sublist in [get_all_assessments_periods_data2(auth, assessment_grade_id ,class_subject ,session=session)] for item in sublist if item.get('editable')==True]
        # unique_lst = [dict(t) for t in {tuple(sorted(d.items())) for d in lst}]
        unique_dict_list = [dict(t) for t in {tuple(sorted(d.items())) for d in ass_data}]
        sorted_dict = sorted(unique_dict_list , key=lambda x: x['code'])
        return sorted_dict    

def assessments_periods_min_max_mark(auth , assessment_id , education_subject_id ,session=None):
    from setting import ASSESSMENTS_PERIODS_MIN_MAX_MARK_URL
    """
         Ø§Ø³ØªØ¹Ù„Ø§Ù… Ø¹Ù† Ø§Ù„Ù‚ÙŠÙ…Ø© Ø§Ù„Ù‚ØµÙˆÙ‰ Ùˆ Ø§Ù„Ø¯Ù†ÙŠØ§ Ù„ÙƒÙ„ Ø§Ù„ØªÙ‚ÙˆÙŠÙ…Ø§Øª  
        Ø¹ÙˆØ§Ù…Ù„ Ø§Ù„Ø¯Ø§Ù„Ø© ØªØ¹Ø±ÙŠÙÙŠ Ø§Ù„Ø³Ù†Ø© Ø§Ù„Ø¯Ø±Ø§Ø³ÙŠØ© Ùˆ Ø§Ù„ØªÙˆÙƒÙ†
        ØªØ¹ÙˆØ¯ Ø¨Ù…Ø¹Ù„ÙˆÙ…Ø§Øª Ø¹Ù† ØªÙ‚ÙŠÙ…Ø§Øª Ø§Ù„ØµÙÙˆÙ ÙÙŠ Ø§Ù„Ø³Ù†Ø© Ø§Ù„Ø¯Ø±Ø§Ø³ÙŠØ©  
    """
    url = ASSESSMENTS_PERIODS_MIN_MAX_MARK_URL.format(assessment_id=assessment_id,education_subject_id=education_subject_id)
    return make_request(url,auth,session=session)

def get_all_assessments_periods_data2(auth , assessment_id ,education_subject_id,session=None):
    """
         Ø§Ø³ØªØ¹Ù„Ø§Ù… Ø¹Ù† ØªØ¹Ø±ÙŠÙØ§Øª Ø§Ù„ØªÙ‚ÙˆÙŠÙ…Ø§Øª ÙÙŠ Ø§Ù„Ø³Ù†Ø© Ø§Ù„Ø¯Ø±Ø§Ø³ÙŠØ© Ùˆ Ø§Ù…ÙƒØ§Ù†ÙŠØ© ØªØ­Ø±ÙŠØ± Ø§Ù„ØªÙ‚ÙˆÙŠÙ… Ùˆ  Ø§Ù„Ø¹Ù„Ø§Ù…Ø© Ø§Ù„Ù‚ØµÙˆÙ‰ Ùˆ Ø§Ù„Ø¯Ù†ÙŠØ§
        Ø¹ÙˆØ§Ù…Ù„ Ø§Ù„Ø¯Ø§Ù„Ø© ØªØ¹Ø±ÙŠÙÙŠ Ø§Ù„Ø³Ù†Ø© Ø§Ù„Ø¯Ø±Ø§Ø³ÙŠØ© Ùˆ Ø§Ù„ØªÙˆÙƒÙ†
        ØªØ¹ÙˆØ¯ ØªØ¹Ø±ÙŠÙØ§Øª Ø§Ù„ØªÙ‚ÙˆÙŠÙ…Ø§Øª ÙÙŠ Ø§Ù„Ø³Ù†Ø© Ø§Ù„Ø¯Ø±Ø§Ø³ÙŠØ© Ùˆ Ø§Ù…ÙƒØ§Ù†ÙŠØ© ØªØ­Ø±ÙŠØ± Ø§Ù„ØªÙ‚ÙˆÙŠÙ… Ùˆ  Ø§Ù„Ø¹Ù„Ø§Ù…Ø© Ø§Ù„Ù‚ØµÙˆÙ‰ Ùˆ Ø§Ù„Ø¯Ù†ÙŠØ§  
    """
    terms = get_AcademicTerms(auth=auth , assessment_id=assessment_id,session=session)['data']
    season_assessments = []
    dic =  {'SEname': '', 'AssesName': '' ,'AssesId': '' , 'pass_mark': '' , 'max_mark' : '' , 'editable' : '' , 'code':'' , 'gradeId':''}
    min_max=[]
    for i in assessments_periods_min_max_mark(auth , assessment_id, education_subject_id,session=session)['data']:
        min_max.append({'id': i['assessment_period_id'] , 'pass_mark':i['assessment_grading_type']['pass_mark'] , 'max_mark' : i['assessment_grading_type']['max'] } )
    for term in terms:
        for asses in get_assessments_periods(auth, term['name'], assessment_id=assessment_id,session=session)['data']:
            dic = {
                    # Key: 'SEname', Value: Academic term from the 'asses' dictionary
                    'SEname': asses["academic_term"],
                    # Key: 'AssesName', Value: Name from the 'asses' dictionary
                    'AssesName': asses["name"],
                    # Key: 'AssesId', Value: ID from the 'asses' dictionary
                    'AssesId': asses["id"],
                    # Key: 'pass_mark', Value: Pass mark from the 'min_max' list where the 'id' matches the 'asses' dictionary's ID
                    'pass_mark': [dictionary['pass_mark'] for dictionary in min_max if dictionary.get('id') == asses["id"]][0],
                    # Key: 'max_mark', Value: Max mark from the 'min_max' list where the 'id' matches the 'asses' dictionary's ID
                    'max_mark': [dictionary['max_mark'] for dictionary in min_max if dictionary.get('id') == asses["id"]][0],
                    # Key: 'editable', Value: Editable flag from the 'asses' dictionary
                    'editable': asses['editable'],
                    # Key: 'code', Value: Code from the 'asses' dictionary
                    'code': asses['code'],
                    # Key: 'gradeId', Value: Assessment ID from the 'asses' dictionary
                    'gradeId': asses['assessment_id']
                }
            season_assessments.append(dic)
    return season_assessments

def enter_marks_arbitrary(username , password , assessment_period_id ,range1 ,range2):
    """
    This function allows to enter marks for an assessment period within a specified range arbitrary.
    
    :param username: The username parameter is a string that represents the username of the user who
    wants to enter marks
    :param password: The password parameter is used to authenticate the user and ensure that only
    authorized users can access and enter marks
    :param assessment_period_id: The assessment period ID is a unique identifier for a specific
    assessment period. It is used to identify and retrieve information related to a particular
    assessment period
    :param range1: The starting range of marks for the assessment period
    :param range2: The range2 parameter is used to specify the upper limit of the range of marks that
    can be entered for a particular assessment period
    """
    auth = get_auth(username , password)
    period_id = get_curr_period(auth)['data'][0]['id']
    inst_id = inst_name(auth)['data'][0]['Institutions']['id']
    
    required_data_list = get_required_data_to_enter_marks(auth , username)
    for item in required_data_list : 
        for Student_id in item['students_ids']:
            enter_mark(auth 
                ,marks= random.randint(range1, range2)
                ,assessment_grading_option_id= 8
                ,assessment_id= item['assessment_id']
                ,education_subject_id= item['education_subject_id']
                ,education_grade_id= item['education_grade_id']
                ,institution_id= inst_id
                ,academic_period_id= period_id
                ,institution_classes_id= item['institution_classes_id']
                ,student_status_id= 1
                ,student_id= Student_id
                ,assessment_period_id= assessment_period_id)

def get_class_students_ids(auth,academic_period_id,institution_subject_id,institution_class_id,institution_id,session=None):
    from setting import GET_CLASS_STUDENTS_IDS_URL
    """
    Ø§Ø³ØªØ¯Ø¹Ø§Ø¡ Ù…Ø¹Ù„ÙˆÙ…Ø§Øª Ø¹Ù† Ø§Ù„Ø·Ù„Ø§Ø¨ ÙÙŠ Ø§Ù„ØµÙ
    Ø¹ÙˆØ§Ù…Ù„ Ø§Ù„Ø¯Ø§Ù„Ø© Ù‡ÙŠ Ø§Ù„Ø±Ø§Ø¨Ø· Ùˆ Ø§Ù„ØªÙˆÙƒÙ† Ùˆ ØªØ¹Ø±ÙŠÙÙŠ Ø§Ù„ÙØªØ±Ø© Ø§Ù„Ø§ÙƒØ§Ø¯ÙŠÙ…ÙŠØ© Ùˆ ØªØ¹Ø±ÙŠÙÙŠ Ù…Ø§Ø¯Ø© Ø§Ù„Ù…Ø¤Ø³Ø³Ø© Ùˆ ØªØ¹Ø±ÙŠÙÙŠ ØµÙ Ø§Ù„Ù…Ø¤Ø³Ø³Ø© Ùˆ ØªØ¹Ø±ÙŠÙÙŠ Ø§Ù„Ù…Ø¤Ø³Ø³Ø©
    ØªØ¹ÙˆØ¯ Ø¨Ù…Ø¹Ù„ÙˆÙ…Ø§Øª ØªÙØµÙŠÙ„ÙŠØ© Ø¹Ù† ÙƒÙ„ Ø·Ø§Ù„Ø¨ ÙÙŠ Ø§Ù„ØµÙ Ø¨Ù…Ø§ ÙÙŠ Ø°Ù„Ùƒ Ø§Ø³Ù…Ù‡ Ø§Ù„Ø±Ø¨Ø§Ø¹ÙŠ Ùˆ Ø§Ù„ØªØ¹Ø±ÙŠÙÙŠ Ùˆ Ù…ÙƒØ§Ù† Ø³ÙƒÙ†Ù‡
    """
    url = GET_CLASS_STUDENTS_IDS_URL.format(academic_period_id=academic_period_id,institution_subject_id=institution_subject_id,institution_class_id=institution_class_id,institution_id=institution_id)
    student_ids = [student['student_id'] for student in make_request(url,auth,session=session)['data']]
    return student_ids

def get_required_data_to_enter_marks(auth ,username,session=None):
    """
    This function is used to get the required data to enter marks for a specific class or students or the students.
    
    :param auth: This parameter is used to authenticate the user. It could be a token, a
    username/password combination, or any other form of authentication required to access the data
    :param username: The username parameter is a string that represents the username of the user who
    wants to enter marks
    :param session: The session parameter is an optional parameter that requests.Session() incase function used again to make it faster
    """
    period_id = get_curr_period(auth,session=session)['data'][0]['id']
    inst_id = inst_name(auth,session=session)['data'][0]['Institutions']['id']
    user_id = user_info(auth,username,session=session)['data'][0]['id']
    years = get_curr_period(auth,session=session)
    # Ù…Ø§ Ø¨Ø¹Ø±Ù ÙƒÙŠÙ Ø³ÙˆÙŠØªÙ‡Ø§ Ù„ÙƒÙ† Ø²Ø¨Ø·Øª 
    classes_id_1 = [[value for key , value in i['institution_subject'].items() if key == "id"][0] for i in get_teacher_classes1(auth,inst_id,user_id,period_id)['data']]
    required_data_to_enter_marks = []
    
    for class_id in classes_id_1 : 
        try:
            class_info = get_teacher_classes2( auth , class_id,session=session)['data']
            dic = {'assessment_id':'','education_subject_id':'' ,'education_grade_id':'','institution_classes_id':'','students_ids':[] }
            dic['assessment_id'] = get_assessment_id_from_grade_id(auth , class_info[0]['institution_subject']['education_grade_id'],session=session)
            dic['education_subject_id'] = class_info[0]['institution_subject']['education_subject_id']
            dic['education_grade_id'] = class_info[0]['institution_subject']['education_grade_id']
            dic['institution_classes_id'] = class_info[0]['institution_class_id']
            dic['class_name'] = class_info[0]['institution_class']['name']
            dic['students_ids'] = get_class_students_ids(auth,period_id,class_info[0]['institution_subject_id'],class_info[0]['institution_class_id'],inst_id,session=session)

            required_data_to_enter_marks.append(dic)
        except IndexError:
            pass
    
    return required_data_to_enter_marks

def get_grade_info(auth,period_id=None,session=None):
    from setting import GET_GRADE_ID_FROM_ASSESSMENT_ID_URL    
    """
    The function "get_grade_info" takes an authentication token as input and returns information about a
    student's grades.
    :param auth: The auth parameter is used to authenticate the user and ensure that they have the
    necessary permissions to access the grade information
    :param session: The "session" parameter is used for requests.Session() incase function used again to make it faster
    :return: a sorted list of dictionaries containing assessment data.
    """
    if period_id is None :
        period_id = get_curr_period(auth)['data'][0]['id']
    my_list = make_request(session=session ,auth=auth , url=f'{GET_GRADE_ID_FROM_ASSESSMENT_ID_URL}&academic_period_id={period_id}')['data']
    return my_list

def get_grade_name_from_grade_id(auth , grade_id):
    """
    The function `get_grade_name_from_grade_id` takes in an authentication token and a grade ID and
    returns the name of the grade.
    
    :param auth: An authentication token or object that is used to authenticate the user making the
    request
    :param grade_id: The grade_id parameter is the unique identifier for a specific grade
    """
    from setting import GET_GRADE_ID_FROM_ASSESSMENT_ID_URL
    my_list = make_request(auth=auth , url=f'{GET_GRADE_ID_FROM_ASSESSMENT_ID_URL}')['data']

    return [d['name'] for d in my_list if d.get('education_grade_id') == grade_id][0].replace('Ø§Ù„ÙØªØ±Ø§Øª Ø§Ù„ØªÙ‚ÙˆÙŠÙ…ÙŠØ© Ù„','Ø§')

def get_assessment_id_from_grade_id(auth , grade_id,period_id=None,session=None):
    from setting import GET_GRADE_ID_FROM_ASSESSMENT_ID_URL
    """
    This function retrieves the assessment ID associated with a given grade ID.
    
    :param auth: The auth parameter is used for authentication purposes. It could be a token or a
    username/password combination, depending on the authentication method being used
    :param grade_id: The grade ID is a unique identifier for a specific grade in a system. It is used to
    track and manage grades for assessments or assignments
    :param session: The session parameter is an optional parameter that represents the current session
    or connection to the database. It is used to execute the SQL query to retrieve the assessment ID
    from the grade ID. If a session is not provided, a new session will be created
    """
    if period_id is None :
        period_id = get_curr_period(auth)['data'][0]['id']
    my_list = make_request(auth=auth , url=f'{GET_GRADE_ID_FROM_ASSESSMENT_ID_URL}&academic_period_id={period_id}',session=session)['data']

    return [d['id'] for d in my_list if d.get('education_grade_id') == grade_id][0]

def create_e_side_marks_doc(username , password ,template='./templet_files/e_side_marks.xlsx' ,outdir='./send_folder' ,student_status_ids = [1], period_id = None , empty_marks = False , session=None):
    """
    The function `create_e_side_marks_doc` creates a document with e-side marks using a specified
    template and saves it in a specified output directory.
    
    :param username: The username is the username of the user who wants to create the document. It is
    used for authentication purposes
    :param password: The password is a string that represents the password for the user's account
    :param template: The template parameter is the path to the Excel file that will be used as a
    template for creating the document. It should be in the format './templet_files/e_side_marks.xlsx',
    defaults to ./templet_files/e_side_marks.xlsx (optional)
    :param outdir: The `outdir` parameter specifies the directory where the generated document will be
    saved, defaults to ./send_folder (optional)
    :param session: The `session` parameter is an optional parameter that can be used to pass an
    existing session object. This can be useful if you want to reuse an existing session for
    authentication or other purposes. If no session object is provided, a new session will be created
    """
    auth = get_auth(username , password )
    if period_id is None :
        period_id = get_curr_period(auth,session=session)['data'][0]['id']
    user = user_info(auth , username,session=session)
    userInfo = user['data'][0]
    user_id , user_name = userInfo['id'] ,f"{userInfo['first_name']} {userInfo['middle_name']} {userInfo['last_name']} - {str(username)}"  
    # years = get_curr_period(auth)
    school_data = inst_name(auth,session=session)['data'][0]
    inst_id = school_data['Institutions']['id']
    school_name = school_data['Institutions']['name']
    school_name_id = f'{school_name}={inst_id}'
    school_id=inst_id

    baldah = make_request(auth=auth , url=f'https://emis.moe.gov.jo/openemis-core/restful/Institution-Institutions.json?_limit=1&id={inst_id}&_contain=InstitutionLands.CustomFieldValues',session=session)['data'][0]['address'].split('-')[0]
    # grades = make_request(auth=auth , url='https://emis.moe.gov.jo/openemis-core/restful/Education.EducationGrades?_limit=0')
    school_place_data= make_request(auth=auth , url=f'https://emis.moe.gov.jo/openemis-core/restful/Institution-Institutions.json?_limit=1&id={inst_id}&_contain=InstitutionLands.CustomFieldValues', session=session)['data'][0]
    indcator_of_private_techers_sector=school_place_data['institution_sector_id']
    if indcator_of_private_techers_sector == 12 : 
        area_data = get_AreaAdministrativeLevels(auth, session=session)['data']
        area_chain_list = find_area_chain(school_place_data['area_administrative_id'], area_data).split(' - ')
        modeeriah_v2=area_chain_list[1]
        modeeriah=f'Ø§Ù„ØªØ¹Ù„ÙŠÙ… Ø§Ù„Ø®Ø§Øµ / {modeeriah_v2}'
    else:
        modeeriah = inst_area(auth , session=session)['data'][0]['Areas']['name']
        modeeriah=f'{modeeriah}'
    school_year = get_curr_period(auth,session=session)['data']
    hejri1 = str(hijri_converter.convert.Gregorian(school_year[0]['start_year'], 1, 1).to_hijri().year)
    hejri2 =  str(hijri_converter.convert.Gregorian(school_year[0]['end_year'], 1, 1).to_hijri().year)
    melady1 = str(school_year[0]['start_year'])
    melady2 = str(school_year[0]['end_year'])
    teacher = user['data'][0]['name'].split(' ')[0]+' '+user['data'][0]['name'].split(' ')[-1]
    
    assessment_periods = make_request(auth =auth,url=f'https://emis.moe.gov.jo/openemis-core/restful/v2/Assessment-AssessmentPeriods.json?_limit=0' , session=session)
    # Ù…Ø§ Ø¨Ø¹Ø±Ù ÙƒÙŠÙ Ø³ÙˆÙŠØªÙ‡Ø§ Ù„ÙƒÙ† Ø²Ø¨Ø·Øª 
    classes_id_2 =[lst for lst in get_teacher_classes_v2(auth ,inst_id, user_id, period_id)['data'] if lst]
    assessments_period_data = []
    grades_info = get_grade_info(auth,period_id,session=session)
    
    # load the existing workbook
    existing_wb = load_workbook(template)

    teacher_load_marks_data = get_marks_v2(auth , inst_id, period_id, classes_id_2 , grades_info , assessment_periods,session,student_status_ids=student_status_ids, empty_marks=empty_marks)
    
    # assessments_period_data = get_marks(auth, inst_id , period_id , classes_id_2 , grades_info, assessments=assessment_periods ,insert_function = insert_to_e_side_marks_doc ,template_sheet_or_file=existing_wb)
    insert_to_e_side_marks_doc(teacher_load_marks_data , template_sheet_or_file=existing_wb)
    
    classes_institution_assessment_education_ides = [f"{i['institution_class_id']}-{i['assessment_id']}-{i['education_grade_id']}" for i in teacher_load_marks_data]
    assessments_periods_lists = [i['students_data'][0]['assessments_periods_ides'] for i in teacher_load_marks_data if i['students_data']]
    assessments_periods_lists_strings = [','.join(map(str,i)) for i in assessments_periods_lists]
    joined_class_data_and_assessment_periods_string = [ ','.join(class_data_and_assessment_periods_string) for  _, class_data_and_assessment_periods_string in enumerate(zip(classes_institution_assessment_education_ides, assessments_periods_lists_strings)) if len(class_data_and_assessment_periods_string[1])]
    
    assessments_period_data_text ='\\\\'.join(joined_class_data_and_assessment_periods_string)
    
    existing_wb.remove(existing_wb['Sheet1'])

    # Create a new sheet
    new_sheet = existing_wb.create_sheet("info_sheet")
    new_sheet.sheet_view.rightToLeft = True    
    # existing_ws.sheet_view.rightToLeft = True  
    
    # Access the new sheet by name
    info_sheet = existing_wb["info_sheet"]

    # Write data to the new sheet
    info_sheet["A1"] = school_id
    info_sheet["A2"] = school_name_id
    info_sheet["A3"] = modeeriah
    info_sheet["A4"] = hejri1
    info_sheet["A5"] = hejri2
    info_sheet["A6"] = melady1
    info_sheet["A7"] = melady2
    info_sheet["A8"] = baldah
    info_sheet["A9"] = teacher
    info_sheet["A10"] = assessments_period_data_text
    info_sheet["A11"] = str(period_id)

    # save the modified workbook
    existing_wb.save(f'{outdir}/{user_name}.xlsx')

def split_A3_pages(input_file, outdir):
    """
    The function `split_A3_pages` takes an A3 PDF file as input, splits each page into two A4 pages, and
    saves the resulting pages to a new PDF file.
    
    :param input_file: The input_file parameter is the path to the A3 PDF file that you want to split
    into A4 pages
    :param outdir: The `outdir` parameter is the directory where the output PDF file will be saved. It
    specifies the path to the directory where the `output.pdf` file will be created
    """
    # Open the A3 PDF file in read-binary mode
    with open(input_file, 'rb') as pdf_file:
        # Create a PDF reader object
        pdf_reader = PyPDF4.PdfFileReader(pdf_file)

        # Create a new PDF writer object
        pdf_writer = PyPDF4.PdfFileWriter()

        # Iterate through each page of the PDF file
        for page_num in range(pdf_reader.getNumPages()):
            # Get the current page of the PDF file
            page = pdf_reader.getPage(page_num)

            # Create a new page with A4 size by splitting the A3 page into two A4 pages
            x1, y1, x2, y2 = page.mediaBox.lowerLeft + page.mediaBox.upperRight
            x_mid = (x1 + x2) / 2
            a4_size = (100, -25)
            page1 = PyPDF4.pdf.PageObject.createBlankPage(None, a4_size[0], a4_size[1])
            page1.mergeTranslatedPage(page, -x1, -y1)
            page1.mediaBox.upperRight = (x2 - x_mid + 50, y2 - a4_size[1])
            page2 = PyPDF4.pdf.PageObject.createBlankPage(None, a4_size[0], a4_size[1])
            page2.mergeTranslatedPage(page, -x_mid, -y1)
            page2.mediaBox.upperRight = (x2 - x_mid , y2 - a4_size[1])

            # Add the two A4 pages to the PDF writer object
            pdf_writer.addPage(page1)
            pdf_writer.addPage(page2)

        # Save the new A4 pages to a new PDF file
        with open(f'{outdir}/output.pdf', 'wb') as output_file:
            pdf_writer.write(output_file)

def reorder_official_marks_to_A4(input_file, out_file):
    """
    The function `reorder_official_marks_to_A4` takes an input PDF file, reorders its pages according to
    a predefined list, and saves the reordered PDF to an output file.
    
    :param input_file: The `input_file` parameter is the path to the PDF file that you want to reorder
    the pages of. It should be a string representing the file path, including the file name and
    extension
    :param out_file: The `out_file` parameter is the name of the output file where the reordered PDF
    document will be saved. It should be a string representing the file name or the file path. For
    example, if you want to save the reordered PDF as "reordered.pdf" in the current directory, you can
    """
    # Load the PDF document
    with open(input_file, 'rb') as file:
        pdf_reader = PyPDF2.PdfFileReader(file)

        # List of page locations in the new order
        new_order_list = ["1=1","52=2","51=3","2=4","3=5","50=6","49=7","3=8","4=9","46=10","45=11","4=12","5=13","44=14","43=15","6=16","7=17","42=18","41=19","8=20","9=21","40=22","39=23","10=24","11=25","38=26","37=27","12=28","13=29","36=30","35=31","14=32","15=33","34=34","33=35","16=36","17=37","32=38","31=39","18=40","19=41","30=42","29=43","20=44","21=45","28=46","27=47","22=48","23=49","26=50","25=51","24=52"]

        # Create a dictionary from the new order list
        new_order_dict = {}
        for item in new_order_list:
            location, page_number = item.split('=')
            new_order_dict[int(page_number)] = int(location)

        # Sort the dictionary by values
        sorted_dict = dict(sorted(new_order_dict.items(), key=lambda x: x[1]))

    #     print(len(sorted_dict))
        # Create a new PDF document object
        pdf_writer = PyPDF2.PdfFileWriter()

        # Iterate over the sorted dictionary's keys and add the corresponding page to the new PDF document
        for page_number in sorted_dict.keys():
            pdf_writer.addPage(pdf_reader.getPage(page_number - 1))

    #     Save the new PDF document with the reordered pages
        with open(out_file, 'wb') as file:
            pdf_writer.write(file)

def delete_files_except(filenames, dir_path):
    """
    Deletes every ODS or PDF file in the specified directory except for the files with the specified names.
    """
    for file in os.listdir(dir_path):
        if file not in filenames and (file.endswith(".ods") or file.endswith(".pdf") or file.endswith(".bak") or file.endswith(".docx")or file.endswith(".xlsx") ):
            os.remove(os.path.join(dir_path, file))

def fill_official_marks_doc_wrapper_offline(lst, ods_name='send', outdir='./send_folder' ,ods_num=1 , do_not_delete_send_folder=False , templet_file = './templet_files/official_marks_doc_a3_two_face_white_cover.ods', color="#ffffff"):
    """
    The function `fill_official_marks_doc_wrapper_offline` fills an official marks document template
    with data, adds custom shapes, converts it to PDF, adds margins, splits A3 pages, reorders pages,
    and deletes unnecessary files.
    
    :param lst: The `lst` parameter is a dictionary that contains two keys: 'file_data' and
    'custom_shapes'
    :param ods_name: The `ods_name` parameter is the name of the output ODS file. By default, it is set
    to 'send', defaults to send (optional)
    :param outdir: The `outdir` parameter specifies the directory where the output files will be saved,
    defaults to ./send_folder (optional)
    :param ods_num: The parameter `ods_num` is used to specify the number of the ODS file. It is set to
    1 by default, defaults to 1 (optional)
    :param do_not_delete_send_folder: The parameter `do_not_delete_send_folder` is a boolean flag that
    determines whether or not to delete all files in the `outdir` folder except for the generated PDF
    and ODS files. If `do_not_delete_send_folder` is set to `True`, the files will not be deleted. If,
    defaults to False (optional)
    :param templet_file: The `templet_file` parameter is the path to the template file that will be used
    to create the official marks document. It should be an OpenDocument Spreadsheet (ODS) file, defaults
    to ./templet_files/official_marks_doc_a3_two_face_white_cover.ods (optional)
    :param color: The "color" parameter is used to specify the color of the margins that will be added
    to the PDF files. It should be a hexadecimal color code, such as "#ffffff" for white or "#000000"
    for black, defaults to #ffffff (optional)
    """
    ods_file = f'{ods_name}{ods_num}.ods'
    copy_ods_file(templet_file , f'{outdir}/{ods_file}')
    fill_official_marks_a3_two_face_doc2_offline_version(students_data_lists=lst['file_data'], ods_file=f'{outdir}/{ods_file}')
    custom_shapes = lst['custom_shapes']
    
    fill_custom_shape(doc= f'{outdir}/{ods_file}' ,sheet_name= 'Ø§Ù„ØºÙ„Ø§Ù Ø§Ù„Ø¯Ø§Ø®Ù„ÙŠ' , custom_shape_values= custom_shapes , outfile=f'{outdir}/modified.ods')
    fill_custom_shape(doc=f'{outdir}/modified.ods', sheet_name='Ø§Ù„ØºÙ„Ø§Ù Ø§Ù„Ø§Ø²Ø±Ù‚', custom_shape_values=custom_shapes, outfile=f"{outdir}/{custom_shapes['teacher']}.ods")
    os.system(f'soffice --headless --convert-to pdf:writer_pdf_Export --outdir {outdir} {outdir}/\"{custom_shapes["teacher"]}.ods\" ')
    add_margins(f"{outdir}/{custom_shapes['teacher']}.pdf", f"{outdir}/output_file.pdf",top_rec=30, bottom_rec=50, left_rec=68, right_rec=120, color_name=color)
    add_margins(f"{outdir}/output_file.pdf", f"{outdir}/{custom_shapes['teacher']}.pdf",page=1 , top_rec=60, bottom_rec=80, left_rec=70, right_rec=120, color_name=color)
    split_A3_pages(f"{outdir}/output_file.pdf" , outdir)
    reorder_official_marks_to_A4(f"{outdir}/output.pdf" , f"{outdir}/reordered.pdf")

    add_margins(f"{outdir}/reordered.pdf", f"{outdir}/output_file.pdf",top_rec=60, bottom_rec=50, left_rec=68, right_rec=20, color_name=color)
    add_margins(f"{outdir}/output_file.pdf", f"{outdir}/output_file1.pdf",page=1 , top_rec=100, bottom_rec=80, left_rec=90, right_rec=120, color_name=color)
    add_margins(f"{outdir}/output_file1.pdf", f"{outdir}/output_file2.pdf",page=50 , top_rec=100, bottom_rec=80, left_rec=70, right_rec=60, color_name=color)
    add_margins(f"{outdir}/output_file2.pdf", f"{outdir}/{custom_shapes['teacher']}_A4.pdf",page=51 , top_rec=100, bottom_rec=80, left_rec=90, right_rec=120, color_name=color)  
    
    if not do_not_delete_send_folder :
        delete_files_except([f"{custom_shapes['teacher']}.pdf",f"{custom_shapes['teacher']}_A4.pdf",f'{custom_shapes["teacher"]}.ods'], outdir)

def fill_official_marks_doc_wrapper(username , password , ods_name='send', outdir='./send_folder' ,ods_num=1 , templet_file = './templet_files/official_marks_doc_a3_two_face_white_cover.ods', color="#ffffff"):
    """
    The function `fill_official_marks_doc_wrapper` takes in various parameters, including a username and
    password, and performs a series of operations to fill an official marks document template and
    generate a final PDF output.
    
    :param username: The username is a string that represents the username for authentication purposes.
    It is used in the `fill_official_marks_a3_two_face_doc2` function
    :param password: The `password` parameter is used to provide the password for accessing the document
    or file. It is required to authenticate the user and grant access to the document
    :param ods_name: The `ods_name` parameter is the name of the ODS file that will be generated. By
    default, it is set to 'send', defaults to send (optional)
    :param outdir: The `outdir` parameter specifies the directory where the output files will be saved,
    defaults to ./send_folder (optional)
    :param ods_num: The parameter `ods_num` is used to specify the number of the ODS file. It is set to
    a default value of 1, but you can change it to any desired number when calling the
    `fill_official_marks_doc_wrapper` function, defaults to 1 (optional)
    :param templet_file: The `templet_file` parameter is the path to the template file that will be used
    to create the official marks document. It should be an OpenDocument Spreadsheet (ODS) file, defaults
    to ./templet_files/official_marks_doc_a3_two_face_white_cover.ods (optional)
    :param color: The "color" parameter is used to specify the color of the margins that will be added
    to the PDF files. It is a hexadecimal color code, such as "#ffffff" for white or "#000000" for
    black, defaults to #ffffff (optional)
    """

    ods_file = f'{ods_name}{ods_num}.ods'
    copy_ods_file(templet_file , f'{outdir}/{ods_file}')
    
    custom_shapes = fill_official_marks_a3_two_face_doc2(username= username, password= password , ods_file=f'{outdir}/{ods_file}')
    fill_custom_shape(doc= f'{outdir}/{ods_file}' ,sheet_name= 'Ø§Ù„ØºÙ„Ø§Ù Ø§Ù„Ø¯Ø§Ø®Ù„ÙŠ' , custom_shape_values= custom_shapes , outfile=f'{outdir}/modified.ods')
    fill_custom_shape(doc=f'{outdir}/modified.ods', sheet_name='Ø§Ù„ØºÙ„Ø§Ù Ø§Ù„Ø§Ø²Ø±Ù‚', custom_shape_values=custom_shapes, outfile=f'{outdir}/final_'+ods_file)
    os.system(f'soffice --headless --convert-to pdf:writer_pdf_Export --outdir {outdir} {outdir}/final_{ods_file} ')
    add_margins(f"{outdir}/final_{ods_name}{ods_num}.pdf", f"{outdir}/output_file.pdf",top_rec=30, bottom_rec=50, left_rec=68, right_rec=120, color_name=color)
    add_margins(f"{outdir}/output_file.pdf", f"{outdir}/{custom_shapes['teacher']}.pdf",page=1 , top_rec=60, bottom_rec=80, left_rec=70, right_rec=120, color_name=color)
    split_A3_pages(f"{outdir}/output_file.pdf" , outdir)
    reorder_official_marks_to_A4(f"{outdir}/output.pdf" , f"{outdir}/reordered.pdf")

    add_margins(f"{outdir}/reordered.pdf", f"{outdir}/output_file.pdf",top_rec=60, bottom_rec=50, left_rec=68, right_rec=20, color_name=color)
    add_margins(f"{outdir}/output_file.pdf", f"{outdir}/output_file1.pdf",page=1 , top_rec=100, bottom_rec=80, left_rec=90, right_rec=120, color_name=color)
    add_margins(f"{outdir}/output_file1.pdf", f"{outdir}/output_file2.pdf",page=50 , top_rec=100, bottom_rec=80, left_rec=70, right_rec=60, color_name=color)    
    add_margins(f"{outdir}/output_file2.pdf", f"{outdir}/{custom_shapes['teacher']}_A4.pdf",page=51 , top_rec=100, bottom_rec=80, left_rec=90, right_rec=120, color_name=color)  
    delete_files_except([f"{custom_shapes['teacher']}.pdf",f"{custom_shapes['teacher']}_A4.pdf",f'final_{ods_file}'], outdir)

def delete_file(file_path):
    """Delete a file"""
    os.remove(file_path)

def copy_ods_file(source_file_path, destination_folder):
    """Copy an ODS file to a destination folder"""
    shutil.copy(source_file_path, destination_folder)

def fill_official_marks_a3_two_face_doc2(username, password , ods_file ,session=None):
    """
    doc is the copy that you want to send 
    """
    context = {'46': 'A6:A30', '4': 'A39:A63', '3': 'L6:L30', '45': 'L39:L63', '44': 'A71:A95', '6': 'A103:A127', '5': 'L71:L95', '43': 'L103:L127', '42': 'A135:A159', '8': 'A167:A191', '7': 'L135:L159', '41': 'L167:L191', '40': 'A199:A223', '10': 'A231:A255', '9': 'L199:L223', '39': 'L231:L255', '38': 'A263:A287', '12': 'A295:A319', '11': 'L263:L287', '37': 'L295:L319', '36': 'A327:A351', '14': 'A359:A383', '13': 'L327:L351', '35': 'L359:L383', '34': 'A391:A415', '16': 'A423:A447', '15': 'L391:L415', '33': 'L423:L447', '32': 'A455:A479', '18': 'A487:A511', '17': 'L455:L479', '31': 'L487:L511', '30': 'A519:A543', '20': 'A551:A575', '19': 'L519:L543', '29': 'L551:L575', '28': 'A583:A607', '22': 'A615:A639', '21': 'L583:L607', '27': 'L615:L639', '26': 'A647:A671', '24': 'A679:A703', '23': 'L647:L671', '25': 'L679:L703'}
    
    page = 4
    name_counter = 1
    name_counter = 1
    auth = get_auth(username , password)
    period_id = get_curr_period(auth)['data'][0]['id']
    inst_id = inst_name(auth)['data'][0]['Institutions']['id']
    user_id = user_info(auth , username)['data'][0]['id']
    # Ù…Ø§ Ø¨Ø¹Ø±Ù ÙƒÙŠÙ Ø³ÙˆÙŠØªÙ‡Ø§ Ù„ÙƒÙ† Ø²Ø¨Ø·Øª 
    classes_id_1 = sorted([[value for key , value in i['institution_subject'].items() if key == "id"][0] for i in get_teacher_classes1(auth,inst_id,user_id,period_id,session=session)['data']])
    classes_id_2 =[get_teacher_classes2( auth , classes_id_1[i])['data'] for i in range(len(classes_id_1))]
    classes_id_2 =[lst for lst in classes_id_2 if lst]
    classes_id_3 = []
    
    user = user_info(auth , username)
    school_name = inst_name(auth)['data'][0]['Institutions']['name']
    baldah = make_request(auth=auth , url=FILL_OFFICIAL_MARKS_DOC_WRAPPER_OFFLINE_URL_1.format(inst_id=inst_id))['data'][0]['address'].split('-')[0]
    grades= make_request(auth=auth , url=FILL_OFFICIAL_MARKS_DOC_WRAPPER_OFFLINE_URL_2)
    modeeriah = inst_area(auth)['data'][0]['Areas']['name']
    school_year = get_curr_period(auth)['data']
    hejri1 = str(hijri_converter.convert.Gregorian(school_year[0]['start_year'], 1, 1).to_hijri().year)
    hejri2 =  str(hijri_converter.convert.Gregorian(school_year[0]['end_year'], 1, 1).to_hijri().year)
    melady1 = str(school_year[0]['start_year'])
    melady2 = str(school_year[0]['end_year'])
    teacher = user['data'][0]['name'].split(' ')[0]+' '+user['data'][0]['name'].split(' ')[-1]
    
    classes=[]
    mawad=[]
    modified_classes=[]
    
    # Open the ODS file and load the sheet you want to fill
    doc = ezodf.opendoc(ods_file) 
    
    sheet_name = 'sheet'
    sheet = doc.sheets[sheet_name]

    for class_info in classes_id_2:
        classes_id_3.append([
                                {
                                    'institution_class_id': class_info[0]['institution_class_id'] ,
                                    'sub_name': class_info[0]['institution_subject']['name'],
                                    'class_name': class_info[0]['institution_class']['name'] ,
                                    'subject_id': class_info[0]['institution_subject']['education_subject_id'] ,
                                    'education_grade_id':class_info[0]['institution_subject']['education_grade_id']
                                }
                            ])

    for v in range(len(classes_id_3)):
        # id
        print (classes_id_3[v][0]['institution_class_id'])
        # subject name 
        print (classes_id_3[v][0]['sub_name'])
        # class name
        print (classes_id_3[v][0]['class_name'])
        mawad.append(classes_id_3[v][0]['sub_name'])
        classes.append(classes_id_3[v][0]['class_name'])
        if '-' not in classes_id_3[v][0]['class_name']:
            class_name = ' '.join(class_name.split(' ')[:-1])
            class_char = classes_id_3[v][0]['class_name'][-1]
        else:
            class_name = classes_id_3[v][0]['class_name'].split('-')[0].replace('Ø§Ù„ØµÙ ' , '')
            class_char = classes_id_3[v][0]['class_name'].split('-')[1]
        sub_name = classes_id_3[v][0]['sub_name']    
        students = get_class_students(auth
                                    ,period_id
                                    ,classes_id_1[v]
                                    ,classes_id_3[v][0]['institution_class_id']
                                    ,inst_id
                                    ,classes_id_3[v][0]['education_grade_id'])
        # students_and_marks
        all1 = get_students_marks(auth
                                    ,period_id
                                    ,classes_id_3[v][0]['subject_id']
                                    ,classes_id_3[v][0]['institution_class_id']
                                    ,inst_id)   
        students_names = []
        for IdAndName in students['data']:
            students_names.append({'student_name': IdAndName['user']['name'] , 'student_id':IdAndName['student_id']})
    
        marks_and_name = []
        mark_data =  {'Sid':'' ,'Sname': '','S1':{ 'ass1': '' ,'ass2': '' , 'ass3': '' , 'ass4': ''} ,'S2':{ 'ass1': '' ,'ass2': '' , 'ass3': '' , 'ass4': ''} }
        term_mapping = {
            "Ø§Ù„ÙØµÙ„ Ø§Ù„Ø£ÙˆÙ„": "term1",
            "Ø§Ù„ÙØµÙ„ Ø§Ù„Ø«Ø§Ù†ÙŠ": "term2"
            # add more mappings here
        }

        assessment_mapping = {
            "Ø§Ù„ØªÙ‚ÙˆÙŠÙ… Ø§Ù„Ø£ÙˆÙ„": "assessment1",
            "Ø§Ù„ØªÙ‚ÙˆÙŠÙ… Ø§Ù„Ø«Ø§Ù†ÙŠ": "assessment2",
            "Ø§Ù„ØªÙ‚ÙˆÙŠÙ… Ø§Ù„Ø«Ø§Ù„Ø«": "assessment3",
            "Ø§Ù„ØªÙ‚ÙˆÙŠÙ… Ø§Ù„Ø±Ø§Ø¨Ø¹": "assessment4",
            # add more mappings here
        }

        students_marks = []
        students_info= students_names
        name_and_marks = []
        all_marks= all1

        for student_data in students_info:
            student_marks = {
                'id': int(student_data['student_id']), 
                'name': student_data['student_name'],
                'term1': {'assessment1': '', 'assessment2': '', 'assessment3': '', 'assessment4': ''},
                'term2': {'assessment1': '', 'assessment2': '', 'assessment3': '', 'assessment4': ''}
            }
            for mark_data in all_marks['data']:
                if mark_data['student_id'] == student_data['student_id']:
                    if mark_data['marks'] is not None:
                        term_key = term_mapping.get(mark_data['assessment_period']['academic_term'])
                        if term_key is not None:
                            assessment_key = assessment_mapping.get(mark_data['assessment_period']['name'])
                            if assessment_key is not None:
                                student_marks[term_key][assessment_key] = mark_data['marks']
            students_marks.append(student_marks)

        sorted_students_names_and_marks = sorted(students_marks, key=lambda x: x['name'])
        
        sheet[f"D{int(context[str(page)].split(':')[0][1:])-5 }"].set_value(f' Ø§Ù„ØµÙ: {class_name}')
        sheet[f"I{int(context[str(page)].split(':')[0][1:])-5 }"].set_value(f'Ø§Ù„Ø´Ø¹Ø¨Ø© (   {class_char}    )')    
        sheet[f"O{int(context[str(page+1)].split(':')[0][1:])-5}"].set_value(sub_name)
        
        for counter,student_info in enumerate(sorted_students_names_and_marks, start=1):
            if counter >= 26:
                page += 2
                counter = 1
                
                sheet[f"D{int(context[str(page)].split(':')[0][1:])-5}"].set_value(f' Ø§Ù„ØµÙ: {class_name}')
                sheet[f"I{int(context[str(page)].split(':')[0][1:])-5}"].set_value(f'Ø§Ù„Ø´Ø¹Ø¨Ø© (   {class_char}    )')  
                sheet[f"O{int(context[str(page+1)].split(':')[0][1:])-5}"].set_value(sub_name)
                #    Ø§Ù„Ù…Ø§Ø¯Ø© Ø§Ù„Ø¯Ø±Ø§Ø³ÙŠØ©     
                
                # {'id': 3824166, 'name': 'Ù†ÙˆØ±Ø§Ù„Ø¯ÙŠÙ† Ù…Ø­Ù…ÙˆØ¯ Ø±Ø§Ø¶ÙŠ Ø§Ù„Ø¯ØºÙŠÙ…Ø§Øª', 'term1': {'assessment1': 9, 'assessment2': 10, 'assessment3': 11, 'assessment4': 20}}
                
                for student_info in sorted_students_names_and_marks[25:] :
                    row_idx = counter + int(context[str(page)].split(':')[0][1:]) - 1  # compute the row index based on the counter
                    sheet[f"A{row_idx}"].set_value(name_counter)
                    sheet[f"B{row_idx}"].set_value(student_info['name'])
                    if 'term1' in student_info and 'assessment1' in student_info['term1'] and 'assessment2' in student_info['term1'] and 'assessment3' in student_info['term1'] and 'assessment4' in student_info['term1']:
                        sheet[f"D{row_idx}"].set_value(student_info['term1']['assessment1']) 
                        sheet[f"E{row_idx}"].set_value(student_info['term1']['assessment2']) 
                        sheet[f"F{row_idx}"].set_value(student_info['term1']['assessment3'])
                        sheet[f"G{row_idx}"].set_value(student_info['term1']['assessment4'])
                    if 'term2' in student_info:
                        row_idx2 = counter + int(context[str(page+1)].split(':')[0][1:]) - 1  # compute the row index based on the counter 
                        sheet[f"L{row_idx2}"].set_value(student_info['term2']['assessment1']) 
                        sheet[f"M{row_idx2}"].set_value(student_info['term2']['assessment2']) 
                        sheet[f"N{row_idx2}"].set_value(student_info['term2']['assessment3'])
                        sheet[f"O{row_idx2}"].set_value(student_info['term2']['assessment4'])                       
                    counter += 1
                    name_counter += 1              
                break                    
            row_idx = counter + int(context[str(page)].split(':')[0][1:]) - 1  # compute the row index based on the counter
            sheet[f"A{row_idx}"].set_value(name_counter)
            sheet[f"B{row_idx}"].set_value(student_info['name']) 
            if 'term1' in student_info and 'assessment1' in student_info['term1'] and 'assessment2' in student_info['term1'] and 'assessment3' in student_info['term1'] and 'assessment4' in student_info['term1']:
                sheet[f"D{row_idx}"].set_value(student_info['term1']['assessment1']) 
                sheet[f"E{row_idx}"].set_value(student_info['term1']['assessment2']) 
                sheet[f"F{row_idx}"].set_value(student_info['term1']['assessment3'])
                sheet[f"G{row_idx}"].set_value(student_info['term1']['assessment4'])
            if 'term2' in student_info:
                row_idx2 = counter + int(context[str(page+1)].split(':')[0][1:]) - 1  # compute the row index based on the counter 
                sheet[f"L{row_idx2}"].set_value(student_info['term2']['assessment1']) 
                sheet[f"M{row_idx2}"].set_value(student_info['term2']['assessment2']) 
                sheet[f"N{row_idx2}"].set_value(student_info['term2']['assessment3'])
                sheet[f"O{row_idx2}"].set_value(student_info['term2']['assessment4'])                
            name_counter += 1 
        name_counter = 1
        page += 2

    for i in classes: 
        if '-' not in i:
            i = ' '.join(i.split(' ')[0:-1])+'-'+i.split(' ')[-1]
        modified_classes.append(get_class_short(i))
        
    modified_classes = ' ØŒ '.join(modified_classes)
    mawad = sorted(set(mawad))
    mawad = ' ØŒ '.join(mawad)

    custom_shapes = {
        'modeeriah': f'Ù„ÙˆØ§Ø¡ {modeeriah}',
        'hejri1': hejri1,
        'hejri2': hejri2,
        'melady1': melady1,
        'melady2': melady2,
        'baldah': baldah,
        'school': school_name,
        'classes': modified_classes,
        'mawad': mawad,
        'teacher' : teacher,
        'modeeriah_20_2': f'Ù„ÙˆØ§Ø¡ {modeeriah}',
        'hejri_20_1': hejri1,
        'hejri_20_2': hejri2,
        'melady_20_1': melady1,
        'melady_20_2': melady2,
        'baldah_20_2': baldah,
        'school_20_2': school_name,
        'classes_20_2': modified_classes,
        'mawad_20_2': mawad,
        'teacher_20_2': teacher ,
        'modeeriah_20_1': f'Ù„ÙˆØ§Ø¡ {modeeriah}',
        'hejri1': hejri1,
        'hejri2': hejri2,
        'melady1': melady1,
        'melady2': melady2,
        'baldah_20_1': baldah,
        'school_20_1': school_name,
        'classes_20_1': modified_classes,
        'mawad_20_1': mawad,
        'teacher_20_1': teacher
    }
    # FIXME: make the customshapes crop _20_ to the rest of the key in the custom_shapes
    # Iterate through the cells of the sheet and fill in the values you want
    doc.save()
            
    return custom_shapes 

def get_class_short(string):
    """
    The `mawad_representations` function takes a string representing a class name and returns a modified
    version of the string with abbreviated representations for certain class names.
    
    :param string: The `string` parameter is a string that represents a class name or grade level. It
    should be in the format "Class Name - Grade Level". For example, "Ø§Ù„ØµÙ Ø§Ù„Ø«Ø§Ù†ÙŠ - 2" represents the
    second grade class
    :return: a modified version of the input string. The modified string includes a representation of
    the school grade or level based on the provided mappings in the dictionary 'y'. The modified string
    is in the format 'grade - class_num'.
    """
    y = {'Ø±ÙˆØ¶Ø© - 1': 'Ø±1', 'Ø±ÙˆØ¶Ø© - 2': 'Ø±2', 'Ø§Ù„ØµÙ Ø§Ù„Ø£ÙˆÙ„': '1', 'Ø§Ù„ØµÙ Ø§Ù„Ø«Ø§Ù†ÙŠ': '2', 'Ø§Ù„ØµÙ Ø§Ù„Ø«Ø§Ù„Ø«': '3', 'Ø§Ù„ØµÙ Ø§Ù„Ø³Ø§Ø¨Ø¹': '7', 'Ø§Ù„ØµÙ Ø§Ù„Ø«Ø§Ù…Ù†': '8', 'Ø§Ù„ØµÙ Ø§Ù„ØªØ§Ø³Ø¹': '9', 'Ø§Ù„ØµÙ Ø§Ù„Ø±Ø§Ø¨Ø¹': '4', 'Ø§Ù„ØµÙ Ø§Ù„Ø®Ø§Ù…Ø³': '5', 'Ø§Ù„ØµÙ Ø§Ù„Ø³Ø§Ø¯Ø³': '6', 'Ø§Ù„ØµÙ Ø§Ù„Ø¹Ø§Ø´Ø±': '10', 'Ø§Ù„ØµÙ Ø§Ù„Ø­Ø§Ø¯ÙŠ Ø¹Ø´Ø± Ø§Ù„Ø¹Ù„Ù…ÙŠ': '11', 'Ø§Ù„ØµÙ Ø§Ù„Ø«Ø§Ù†ÙŠ Ø¹Ø´Ø± Ø§Ù„Ø¹Ù„Ù…ÙŠ': '12 Ø¹Ù„Ù…ÙŠ', 'Ø§Ù„ØµÙ Ø§Ù„Ø­Ø§Ø¯ÙŠ Ø¹Ø´Ø± Ø§Ù„Ø£Ø¯Ø¨ÙŠ': '11 Ø§Ø¯Ø¨ÙŠ', 'Ø§Ù„ØµÙ Ø§Ù„Ø«Ø§Ù†ÙŠ Ø¹Ø´Ø± Ø§Ù„Ø£Ø¯Ø¨ÙŠ': '12 Ø§Ø¯Ø¨ÙŠ', 'Ø§Ù„ØµÙ Ø§Ù„Ø­Ø§Ø¯ÙŠ Ø¹Ø´Ø± Ø§Ù„Ø´Ø±Ø¹ÙŠ': '11 Ø´Ø±ØºÙŠ', 'Ø§Ù„ØµÙ Ø§Ù„Ø«Ø§Ù†ÙŠ Ø¹Ø´Ø± Ø§Ù„Ø´Ø±Ø¹ÙŠ': '12 Ø´Ø±Ø¹ÙŠ', 'Ø§Ù„ØµÙ Ø§Ù„Ø­Ø§Ø¯ÙŠ Ø¹Ø´Ø± Ø§Ù„ØµØ­ÙŠ': '11 ØµØ­ÙŠ', 'Ø§Ù„ØµÙ Ø§Ù„Ø«Ø§Ù†ÙŠ Ø¹Ø´Ø± Ø§Ù„ØµØ­ÙŠ': '12 ØµØ­ÙŠ', 'Ø§Ù„ØµÙ Ø§Ù„Ø­Ø§Ø¯ÙŠ Ø¹Ø´Ø± - Ø¥Ø¯Ø§Ø±Ø© Ù…Ø¹Ù„ÙˆÙ…Ø§ØªÙŠØ©': '11 Ø§Ø¯Ø§Ø±Ø©', 'Ø§Ù„ØµÙ Ø§Ù„Ø«Ø§Ù†ÙŠ Ø¹Ø´Ø± - Ø¥Ø¯Ø§Ø±Ø© Ù…Ø¹Ù„ÙˆÙ…Ø§ØªÙŠØ©': '12 Ø§Ø¯Ø§Ø±Ø©', 'Ø§Ù„ØµÙ Ø§Ù„Ø­Ø§Ø¯ÙŠ Ø¹Ø´Ø± - Ø§Ù‚ØªØµØ§Ø¯ Ù…Ù†Ø²Ù„ÙŠ': '11 Ø§Ù‚ØªØµØ§Ø¯', 'Ø§Ù„ØµÙ Ø§Ù„Ø«Ø§Ù†ÙŠ Ø¹Ø´Ø± - Ø§Ù‚ØªØµØ§Ø¯ Ù…Ù†Ø²Ù„ÙŠ': '12 Ø§Ù‚ØªØµØ§Ø¯', 'Ø§Ù„ØµÙ Ø§Ù„Ø­Ø§Ø¯ÙŠ Ø¹Ø´Ø±- ÙÙ†Ø¯Ù‚ÙŠ': '11 ÙÙ†Ø¯Ù‚ÙŠ', 'Ø§Ù„ØµÙ Ø§Ù„Ø«Ø§Ù†ÙŠ Ø¹Ø´Ø± - ÙÙ†Ø¯Ù‚ÙŠ': '12 ÙÙ†Ø¯Ù‚ÙŠ', 'Ø§Ù„ØµÙ Ø§Ù„Ø­Ø§Ø¯ÙŠ Ø¹Ø´Ø± - ØµÙ†Ø§Ø¹ÙŠ': '11 ØµÙ†Ø§Ø¹ÙŠ', 'Ø§Ù„ØµÙ Ø§Ù„Ø«Ø§Ù†ÙŠ Ø¹Ø´Ø± - ØµÙ†Ø§Ø¹ÙŠ': '12 ØµÙ†Ø§Ø¹ÙŠ', 'Ø§Ù„ØµÙ Ø§Ù„Ø­Ø§Ø¯ÙŠ Ø¹Ø´Ø± - Ø²Ø±Ø§Ø¹ÙŠ': '11 Ø²Ø±Ø§Ø¹ÙŠ', 'Ø§Ù„ØµÙ Ø§Ù„Ø«Ø§Ù†ÙŠ Ø¹Ø´Ø± - Ø²Ø±Ø§Ø¹ÙŠ': '12 Ø²Ø±Ø§Ø¹ÙŠ', 'Ø§Ù„ØµÙ Ø§Ù„Ø¹Ø§Ø´Ø± Ø§Ù„Ø£ÙƒØ§ÙŠÙ…ÙŠ': '10 Ø§ÙƒØ§Ø¯ÙŠÙ…ÙŠ'}

    search_str ,class_num = string.split('-')[0] ,string.split('-')[1]

    for key, value in y.items():
        search_key = search_str
        if search_key in key:
            replacement = value
            search_str = search_str.replace(search_key, replacement)

    return f'{search_str} - {class_num}'

def get_students_marks(auth,period_id,sub_id,instit_class_id,instit_id):
    """
    Ø¯Ø§Ù„Ø© Ù„Ø§Ø³ØªØ¯Ø¹Ø§Ø¡ Ø¹Ù„Ø§Ù…Ø§Øª Ø§Ù„Ø·Ù„Ø§Ø¨ Ùˆ Ø§Ø³Ù…Ø§Ø¦Ù‡Ù… 
    Ùˆ Ø¹ÙˆØ§Ù…Ù„Ù‡Ø§ Ø§Ù„ØªÙˆÙƒÙ† Ø±Ù‚Ù… Ø§Ù„Ø³Ù†Ø© Ø§Ù„ØªØ¹Ø±ÙŠÙÙŠ ÙˆØ±Ù‚Ù… Ø§Ù„Ù…Ø§Ø¯Ø© Ø§Ù„ØªØ¹Ø±ÙŠÙÙŠ Ùˆ Ø±Ù‚Ù… Ø§Ù„Ù…Ø¤Ø³Ø³Ø© Ùˆ  Ø±Ù‚Ù… Ø§Ù„ØµÙ Ø§Ù„ØªØ¹Ø±ÙŠÙÙŠ
    Ùˆ ØªØ¹ÙˆØ¯ Ø¨Ø§Ø³Ù…Ø§Ø¡ Ø§Ù„Ø·Ø§Ù„Ø¨ Ùˆ Ø¹Ù„Ø§Ù…Ø§ØªÙ‡Ù…
    """
    url = GET_STUDENTS_MARKS_URL.format(period_id=period_id,sub_id=sub_id,instit_class_id=instit_class_id,instit_id=instit_id)
    return make_request(url,auth)

def get_assessments_periods(auth ,term, assessment_id,session=None):
    """
         Ø§Ø³ØªØ¹Ù„Ø§Ù… Ø¹Ù† ØªØ¹Ø±ÙŠÙØ§Øª Ø§Ù„ØªÙ‚ÙˆÙŠÙ…Ø§Øª ÙÙŠ Ø§Ù„ÙØµÙ„ Ø§Ù„Ø¯Ø±Ø§Ø³ÙŠ 
        Ø¹ÙˆØ§Ù…Ù„ Ø§Ù„Ø¯Ø§Ù„Ø© ØªØ¹Ø±ÙŠÙÙŠ Ø§Ù„Ø³Ù†Ø© Ø§Ù„Ø¯Ø±Ø§Ø³ÙŠØ© Ùˆ Ø§Ù„ØªÙˆÙƒÙ†
        ØªØ¹ÙˆØ¯ Ø¨Ù…Ø¹Ù„ÙˆÙ…Ø§Øª Ø¹Ù† ØªÙ‚ÙŠÙ…Ø§Øª Ø§Ù„ØµÙÙˆÙ ÙÙŠ Ø§Ù„Ø³Ù†Ø© Ø§Ù„Ø¯Ø±Ø§Ø³ÙŠØ©  
    """
    url = GET_ASSESSMENTS_PERIODS_URL_1.format(term=term,assessment_id=assessment_id)
    return make_request(url=url,auth=auth,session=session)

def get_all_assessments_periods(auth , assessment_id):
    """
         Ø§Ø³ØªØ¹Ù„Ø§Ù… Ø¹Ù† ØªØ¹Ø±ÙŠÙØ§Øª Ø§Ù„ØªÙ‚ÙˆÙŠÙ…Ø§Øª ÙÙŠ Ø§Ù„Ø³Ù†Ø© Ø§Ù„Ø¯Ø±Ø§Ø³ÙŠØ© 
        Ø¹ÙˆØ§Ù…Ù„ Ø§Ù„Ø¯Ø§Ù„Ø© ØªØ¹Ø±ÙŠÙÙŠ Ø§Ù„Ø³Ù†Ø© Ø§Ù„Ø¯Ø±Ø§Ø³ÙŠØ© Ùˆ Ø§Ù„ØªÙˆÙƒÙ†
        ØªØ¹ÙˆØ¯ Ø¨Ù…Ø¹Ù„ÙˆÙ…Ø§Øª Ø¹Ù† ÙƒÙ„ ØªÙ‚ÙŠÙ…Ø§Øª Ø§Ù„ØµÙÙˆÙ ÙÙŠ Ø§Ù„Ø³Ù†Ø© Ø§Ù„Ø¯Ø±Ø§Ø³ÙŠØ©  
    """
    terms = get_AcademicTerms(auth=auth , assessment_id=assessment_id)['data']
    season_assessments = []
    dic =  {'SEname': '', 'AssesName': '' ,'AssesId': '' }
    for term in terms:
        for asses in get_assessments_periods(auth, term['name'], assessment_id=assessment_id)['data']:
            dic = {'SEname': asses["academic_term"], 'AssesName': asses["name"], 'AssesId': asses["id"]}
            season_assessments.append(dic)
    return season_assessments

def get_assessments_id( auth ,education_grade_id ):
    """
         Ø§Ø³ØªØ¹Ù„Ø§Ù… Ø¹Ù† ØªØ¹Ø±ÙŠÙÙŠ Ø§Ù„ØµÙ Ø§Ù„Ø¯Ø±Ø§Ø³ÙŠ 
          Ø¹ÙˆØ§Ù…Ù„ Ø§Ù„Ø¯Ø§Ù„Ø© ØªØ¹Ø±ÙŠÙÙŠ Ø§Ù„Ù…Ø±Ø­Ù„Ø© Ø§Ù„Ø¯Ø±Ø§Ø³ÙŠØ© Ùˆ Ø§Ù„ØªÙˆÙƒÙ†
        ØªØ¹ÙˆØ¯ Ø¨Ù…Ø¹Ù„ÙˆÙ…Ø§Øª Ø¹Ù† ØªÙ‚ÙŠÙ…Ø§Øª Ø§Ù„ØµÙÙˆÙ ÙÙŠ Ø§Ù„Ø³Ù†Ø© Ø§Ù„Ø¯Ø±Ø§Ø³ÙŠØ©  
    """
    assessments = get_assessments(auth)
    for assessment in assessments['data'] : 
        if assessment['education_grade_id'] == education_grade_id :
            return assessment['id']

def get_AcademicTerms(auth,assessment_id,session=None):
    """
    Ø¯Ø§Ù„Ø© Ù„Ø§Ø³ØªØ¯Ø¹Ø§Ø¡ Ø§Ø³Ù… Ø§Ù„ÙØµÙ„ 
    Ùˆ Ø¹ÙˆØ§Ù…Ù„Ù‡Ø§ Ø§Ù„ØªÙˆÙƒÙ† Ùˆ Ø±Ù‚Ù… ØªÙ‚ÙŠÙ… Ø§Ù„ØµÙ 
    Ùˆ ØªØ¹ÙˆØ¯ Ø¨Ø§Ø³Ù…Ø§Ø¡ Ø§Ù„ÙØµÙˆÙ„ Ø¹Ù„Ù‰ Ø´ÙƒÙ„ Ø¬ÙŠØ³Ù†
    """
    url = GET_ACADEMIC_TERMS_URL.format(assessment_id=assessment_id)   
    return make_request(url,auth,session=session)        

def draw_rect_top(page, page_width, fill_color , width=50):
    """
    Ø±Ø³Ù… Ù…Ø³ØªØ·ÙŠÙ„ Ø¹Ù„Ù‰ Ø§Ù„Ø¬Ø²Ø¡ Ø§Ù„Ø¹Ù„ÙˆÙŠ Ù…Ù† Ø§Ù„ØµÙØ­Ø©
    """
    rect_top = fitz.Rect(0, 0, page_width, width)
    page.draw_rect(rect_top, color=fill_color, fill=fill_color)

def draw_rect_bottom(page, page_width, page_height, fill_color, width=50):
    """
    Ø±Ø³Ù… Ù…Ø³ØªØ·ÙŠÙ„ Ø¹Ù„Ù‰ Ø§Ù„Ø¬Ø²Ø¡ Ø§Ù„Ø³ÙÙ„ÙŠ Ù…Ù† Ø§Ù„ØµÙØ­Ø©
    """
    rect_bottom = fitz.Rect(0, page_height - width, page_width, page_height)
    page.draw_rect(rect_bottom, color=fill_color, fill=fill_color)

def draw_rect_left(page, page_height, fill_color, width=50):
    """
    Ø±Ø³Ù… Ù…Ø³ØªØ·ÙŠÙ„ Ø¹Ù„Ù‰ Ø§Ù„Ø¬Ø²Ø¡ Ø§Ù„Ø£ÙŠØ³Ø± Ù…Ù† Ø§Ù„ØµÙØ­Ø©
    """
    rect_left = fitz.Rect(0, 0, width, page_height)
    page.draw_rect(rect_left, color=fill_color, fill=fill_color)

def draw_rect_right(page, page_width, page_height, fill_color, width=50):
    """
    Ø±Ø³Ù… Ù…Ø³ØªØ·ÙŠÙ„ Ø¹Ù„Ù‰ Ø§Ù„Ø¬Ø²Ø¡ Ø§Ù„Ø£ÙŠÙ…Ù† Ù…Ù† Ø§Ù„ØµÙØ­Ø©
    """
    rect_right = fitz.Rect(page_width - width, 0, page_width, page_height)
    page.draw_rect(rect_right, color=fill_color, fill=fill_color)

def add_margins(input_pdf1, output_pdf ,color_name="#8cd6e6",top_rec=50,bottom_rec=50,left_rec=50,right_rec=50 ,page=0):
    """
    Ø¥Ø¶Ø§ÙØ© Ù‡ÙˆØ§Ù…Ø´ Ø¨Ø§Ù„Ù„ÙˆÙ† 8cd6e6 Ø¥Ù„Ù‰ Ø¬Ù…ÙŠØ¹ Ø§Ù„Ø¬ÙˆØ§Ù†Ø¨ Ù…Ù† Ø§Ù„ØµÙØ­Ø©
    """
    
    """
    example of how to add colored margin for the first and scond page
    add_margins("existing_file.pdf", "output_file.pdf",top_rec=27, bottom_rec=20, left_rec=90, right_rec=120)
    add_margins("output_file.pdf", "output_file2.pdf",page=1 , top_rec=60, bottom_rec=25, left_rec=90, right_rec=120)
    """
    # Open the PDF file
    input_pdf = fitz.open(input_pdf1)
    
    # Get the first page
    page = input_pdf[page]
    # Get the page dimensions
    page_width = page.rect.width
    page_height = page.rect.height

    # Convert the color from hex to RGB
    color_name = color_name
    color = webcolors.hex_to_rgb(color_name)
    color = tuple(c / 255 for c in color)  # Convert to RGB values between 0 and 1

    # Set the color
    fill_color = color  # Color code in RGB format

    # Draw rectangles on all four sides of the page
    draw_rect_top(page, page_width, fill_color , top_rec)
    draw_rect_bottom(page, page_width, page_height, fill_color, bottom_rec)
    draw_rect_left(page, page_height, fill_color , left_rec)
    draw_rect_right(page, page_width, page_height, fill_color, right_rec)

    # Save the modified PDF file
    input_pdf.save(output_pdf)

def mawad(string):
    """
    The `mawad` function takes a string input representing a class name and returns a modified version
    of the string with abbreviated class names.
    
    :param string: The `string` parameter is a string that represents a class name or grade level. It is
    in the format "Class Name - Grade Level". For example, "Ø§Ù„ØµÙ Ø§Ù„Ø«Ø§Ù†ÙŠ Ø¹Ø´Ø± Ø§Ù„Ø¹Ù„Ù…ÙŠ - 12 Ø¹Ù„Ù…ÙŠ"
    :return: a modified version of the input string. The modified string replaces certain keywords in
    the input string with their corresponding values from the dictionary 'y'. The modified string is
    then returned.
    """
    y = {'Ø±ÙˆØ¶Ø© - 1': 'Ø±1', 'Ø±ÙˆØ¶Ø© - 2': 'Ø±2', 'Ø§Ù„ØµÙ Ø§Ù„Ø£ÙˆÙ„': '1', 'Ø§Ù„ØµÙ Ø§Ù„Ø«Ø§Ù†ÙŠ': '2', 'Ø§Ù„ØµÙ Ø§Ù„Ø«Ø§Ù„Ø«': '3', 'Ø§Ù„ØµÙ Ø§Ù„Ø³Ø§Ø¨Ø¹': '7', 'Ø§Ù„ØµÙ Ø§Ù„Ø«Ø§Ù…Ù†': '8', 'Ø§Ù„ØµÙ Ø§Ù„ØªØ§Ø³Ø¹': '9', 'Ø§Ù„ØµÙ Ø§Ù„Ø±Ø§Ø¨Ø¹': '4', 'Ø§Ù„ØµÙ Ø§Ù„Ø®Ø§Ù…Ø³': '5', 'Ø§Ù„ØµÙ Ø§Ù„Ø³Ø§Ø¯Ø³': '6', 'Ø§Ù„ØµÙ Ø§Ù„Ø¹Ø§Ø´Ø±': '10', 'Ø§Ù„ØµÙ Ø§Ù„Ø­Ø§Ø¯ÙŠ Ø¹Ø´Ø± Ø§Ù„Ø¹Ù„Ù…ÙŠ': '11', 'Ø§Ù„ØµÙ Ø§Ù„Ø«Ø§Ù†ÙŠ Ø¹Ø´Ø± Ø§Ù„Ø¹Ù„Ù…ÙŠ': '12 Ø¹Ù„Ù…ÙŠ', 'Ø§Ù„ØµÙ Ø§Ù„Ø­Ø§Ø¯ÙŠ Ø¹Ø´Ø± Ø§Ù„Ø£Ø¯Ø¨ÙŠ': '11 Ø§Ø¯Ø¨ÙŠ', 'Ø§Ù„ØµÙ Ø§Ù„Ø«Ø§Ù†ÙŠ Ø¹Ø´Ø± Ø§Ù„Ø£Ø¯Ø¨ÙŠ': '12 Ø§Ø¯Ø¨ÙŠ', 'Ø§Ù„ØµÙ Ø§Ù„Ø­Ø§Ø¯ÙŠ Ø¹Ø´Ø± Ø§Ù„Ø´Ø±Ø¹ÙŠ': '11 Ø´Ø±ØºÙŠ', 'Ø§Ù„ØµÙ Ø§Ù„Ø«Ø§Ù†ÙŠ Ø¹Ø´Ø± Ø§Ù„Ø´Ø±Ø¹ÙŠ': '12 Ø´Ø±Ø¹ÙŠ', 'Ø§Ù„ØµÙ Ø§Ù„Ø­Ø§Ø¯ÙŠ Ø¹Ø´Ø± Ø§Ù„ØµØ­ÙŠ': '11 ØµØ­ÙŠ', 'Ø§Ù„ØµÙ Ø§Ù„Ø«Ø§Ù†ÙŠ Ø¹Ø´Ø± Ø§Ù„ØµØ­ÙŠ': '12 ØµØ­ÙŠ', 'Ø§Ù„ØµÙ Ø§Ù„Ø­Ø§Ø¯ÙŠ Ø¹Ø´Ø± - Ø¥Ø¯Ø§Ø±Ø© Ù…Ø¹Ù„ÙˆÙ…Ø§ØªÙŠØ©': '11 Ø§Ø¯Ø§Ø±Ø©', 'Ø§Ù„ØµÙ Ø§Ù„Ø«Ø§Ù†ÙŠ Ø¹Ø´Ø± - Ø¥Ø¯Ø§Ø±Ø© Ù…Ø¹Ù„ÙˆÙ…Ø§ØªÙŠØ©': '12 Ø§Ø¯Ø§Ø±Ø©', 'Ø§Ù„ØµÙ Ø§Ù„Ø­Ø§Ø¯ÙŠ Ø¹Ø´Ø± - Ø§Ù‚ØªØµØ§Ø¯ Ù…Ù†Ø²Ù„ÙŠ': '11 Ø§Ù‚ØªØµØ§Ø¯', 'Ø§Ù„ØµÙ Ø§Ù„Ø«Ø§Ù†ÙŠ Ø¹Ø´Ø± - Ø§Ù‚ØªØµØ§Ø¯ Ù…Ù†Ø²Ù„ÙŠ': '12 Ø§Ù‚ØªØµØ§Ø¯', 'Ø§Ù„ØµÙ Ø§Ù„Ø­Ø§Ø¯ÙŠ Ø¹Ø´Ø±- ÙÙ†Ø¯Ù‚ÙŠ': '11 ÙÙ†Ø¯Ù‚ÙŠ', 'Ø§Ù„ØµÙ Ø§Ù„Ø«Ø§Ù†ÙŠ Ø¹Ø´Ø± - ÙÙ†Ø¯Ù‚ÙŠ': '12 ÙÙ†Ø¯Ù‚ÙŠ', 'Ø§Ù„ØµÙ Ø§Ù„Ø­Ø§Ø¯ÙŠ Ø¹Ø´Ø± - ØµÙ†Ø§Ø¹ÙŠ': '11 ØµÙ†Ø§Ø¹ÙŠ', 'Ø§Ù„ØµÙ Ø§Ù„Ø«Ø§Ù†ÙŠ Ø¹Ø´Ø± - ØµÙ†Ø§Ø¹ÙŠ': '12 ØµÙ†Ø§Ø¹ÙŠ', 'Ø§Ù„ØµÙ Ø§Ù„Ø­Ø§Ø¯ÙŠ Ø¹Ø´Ø± - Ø²Ø±Ø§Ø¹ÙŠ': '11 Ø²Ø±Ø§Ø¹ÙŠ', 'Ø§Ù„ØµÙ Ø§Ù„Ø«Ø§Ù†ÙŠ Ø¹Ø´Ø± - Ø²Ø±Ø§Ø¹ÙŠ': '12 Ø²Ø±Ø§Ø¹ÙŠ'}

    search_str ,class_num = string.split('-')[0] ,string.split('-')[1]

    for key, value in y.items():
        search_key = search_str
        if search_key in key:
            replacement = value
            search_str = search_str.replace(search_key, replacement)

    return f'{search_str}-{class_num}'

def get_basic_info (username , password):
    """
    Retrieves basic information related to a user and institution from an educational management system.

    Parameters:
    - username (str): The username for authentication.
    - password (str): The password for authentication.

    Returns:
    - dict: A dictionary containing the following basic information:
        - 'school_name' (str): The name of the school.
        - 'baldah' (str): The city or region where the school is located.
        - 'grades' (list): List of available education grades.
        - 'modeeriah' (str): The administrative area of the institution.
        - 'melady' (str): The school year in the Gregorian calendar.
        - 'hejri' (str): The school year in the Hijri calendar.
        - 'teacher' (str): The name of the teacher.

    Example:
    - basic_info = get_basic_info('sample_username', 'sample_password')
    """
    auth = get_auth(username ,password )
    user = user_info(auth , username)
    inst_data = inst_name(auth)['data'][0]['Institutions']
    school_name = inst_data['name']
    inst_id= inst_name(auth)['data'][0]['Institutions']['id']
    baldah = make_request(auth=auth , url=FILL_OFFICIAL_MARKS_DOC_WRAPPER_OFFLINE_URL_1.format(inst_id=inst_id))['data'][0]['address'].split('-')[0]
    grades= make_request(auth=auth , url=FILL_OFFICIAL_MARKS_DOC_WRAPPER_OFFLINE_URL_2)
    modeeriah = inst_area(auth)['data'][0]['Areas']['name']
    school_year = get_curr_period(auth)['data']
    melady = str(school_year[0]['end_year'])+' '+str(school_year[0]['start_year'])
    hejri =  str(hijri_converter.convert.Gregorian(school_year[0]['end_year'], 1, 1).to_hijri().year)+' '+str(hijri_converter.convert.Gregorian(school_year[0]['start_year'], 1, 1).to_hijri().year)
    teacher = user['data'][0]['name'].split(' ')[0]+' '+user['data'][0]['name'].split(' ')[-1]

def fill_custom_shape(doc, sheet_name, custom_shape_values, outfile):
    """
    Fills custom shapes in an OpenDocument Spreadsheet (ODS) document.

    Parameters:
    - doc (str): The path to the ODS document.
    - sheet_name (str): The name of the sheet where custom shapes need to be filled.
    - custom_shape_values (dict): A dictionary containing custom shape names as keys
                                 and corresponding values to be filled in the shapes.
    - outfile (str): The path for the output ODS document with filled custom shapes.

    Example:
    - custom_shapes = {
        'modeeriah': f'Ù„ÙˆØ§Ø¡ {modeeriah}',
        'hejri': hejri,
        'melady': melady,
        'baldah': baldah,
        'school': school_name,
        'classes': "7Ø£ ØŒ 7Ø¨",
        'mawad': "Ø§Ù„Ù„ØºØ© Ø§Ù„Ø§Ù†Ø¬Ù„ÙŠØ²ÙŠØ©",
        'teacher' : teacher
    }

    - fill_custom_shape('official_marks_doc_a3_two_face.ods', 'Ø§Ù„ØºÙ„Ø§Ù Ø§Ù„Ø¯Ø§Ø®Ù„ÙŠ', custom_shapes, 'tttttt.ods')
    """
    print(doc)
    # Load the document
    doc = load(str(doc))
    try:
        # Iterate over the sheets in the document
        for sheet in doc.spreadsheet.childNodes[1:-1]:
            # Check if the sheet is the one we want (replace 'Sheet2' with the name of your sheet)
            try :
                if sheet.getAttribute('name') == sheet_name:
                    # Get the custom shapes on the sheet
                    custom_shapes = sheet.getElementsByType(CustomShape)
                    for custom_shape in custom_shapes:     
                        # Check if the custom shape name is in the dictionary of custom shape values
                        if custom_shape.getAttribute('name')  in custom_shape_values:
                            # get the style name
                            p_style = custom_shape.childNodes[0].attributes.get(('urn:oasis:names:tc:opendocument:xmlns:text:1.0', 'style-name'), 'default_style')
                            span_style = custom_shape.childNodes[0].childNodes[0].attributes.get(('urn:oasis:names:tc:opendocument:xmlns:text:1.0', 'style-name'), 'default_style')
                            # clear the text
                            clear_text_custom_shape(custom_shape) 
                            value = custom_shape_values[custom_shape.getAttribute('name')]
                            # Create the text:p element inside the custom shape
                            text_p = P(stylename=str(p_style))
                            text_p.addElement(Span(text=str(value), stylename=str(span_style)))
                            # Add a new text paragraph to the custom shape with the new value
                            custom_shape.addElement(text_p)  
            except ValueError : 
                pass
    except:
        pass
    # Save the modified document
    doc.save(outfile)

def clear_text_custom_shape(shape):
    """Ø¯Ø§Ù„Ø© Ù„Ø­Ø°Ù Ø§Ù„Ù†Øµ Ø§Ù„Ù…ÙˆØ¬ÙˆØ¯ ÙÙŠ Ø§Ù„Ø´ÙƒÙ„ Ø§Ù„Ù…ÙˆØ¬ÙˆØ¯ ÙÙŠ Ø§Ù„ØµÙØ­Ø© ));

    Args:
        shape (str): Ø§Ø³Ù… Ø§Ù„Ø´ÙƒÙ„ 
    """    
    # Remove all child nodes from the shape element
    while len(shape.childNodes) > 0:
        shape.removeChild(shape.childNodes[0])

def get_sheet_custom_shapes(document , sheet_name):
    """Ø¯Ø§Ù„Ø© Ù„Ø§Ø­Ø¶Ø§Ø± Ø§Ø´ÙƒØ§Ù„ ØµÙØ­Ø© Ù…Ù„Ù Ø§Ùˆ Ø¯ÙŠ Ø§Ø³

    Args:
        document (_type_): Ù…Ø³Ø§Ø± Ø§Ù„Ù…Ù„Ù
        sheet_name (_type_): Ø§Ø³Ù… Ø§Ù„ØµÙØ­Ø©

    Returns:
        list: Ù‚Ø§Ø¦Ù…Ø© Ø¨Ø§Ù„Ø§Ø´ÙƒØ§Ù„ Ø§Ù„Ù…ÙˆØ¬ÙˆØ¯Ø© ÙÙŠ Ø§Ù„ØµÙØ­Ø©
    """    
    # Load the document
    doc = load(str(document))
    # Loop through all sheets in the document
    for sheet in doc.spreadsheet.childNodes[1:-1]:      
        # Check if the sheet is the one we want (replace 'Sheet2' with the name of your sheet)
        if sheet.getAttribute('name') == str(sheet_name) :        
            # Get the text boxes on the sheet
            custom_shapes = sheet.getElementsByType(CustomShape)
            return [custom_shape.getAttribute("name") for custom_shape in custom_shapes]

def get_ods_sheets (doc='official_marks_doc_a3_two_face.ods'):
    """Ø¯Ø§Ù„Ø© Ù…Ø®ØªØµØ±Ø© Ù„ÙƒÙŠ Ø§Ø­Ø¶Ø± Ù…Ø¹Ù„ÙˆÙ…Ø§Øª Ø§Ù„ØµÙØ­Ø§Øª ÙÙŠ Ù…Ù„ÙØ§Øª ods =>

    Args:
        doc (str, optional): Ù…Ø³Ø§Ø± Ø§Ù„Ù…Ù„Ù . Defaults to 'official_marks_doc_a3_two_face.ods'.

    Returns:
        list: Ù‚Ø§Ø¦Ù…Ø© Ø¨Ù…Ø¹Ù„ÙˆÙ…Ø§Øª ØµÙØ­Ø§Øª Ø§Ù„Ù…Ù„Ù
    """    
    # Load the ODF document
    doc = load(doc)
    # Get the sheets in the document
    sheets = doc.getElementsByType(Table)
    return [sheet.getAttribute("name") for sheet in sheets]

def page_counter_official_marks_doc_a3_two_face ():
    """Ø¯Ø§Ù„Ø© Ù…Ø³Ø§Ø¹Ø¯Ø© Ù‚Ù…Øª Ø¨Ø§Ù†Ø´Ø§Ø¦Ù‡Ø§ Ù„Ù„Ù…Ø³Ø§Ø¹Ø¯Ø© ÙÙŠ Ø§ÙŠØ¬Ø§Ø¯ ØªØ±ØªÙŠØ¨ Ø§Ù„ØµÙØ­Ø§Øª ÙÙŠ Ø§Ù„Ù…Ù„ÙØ§Øª Ø°Ø§Øª Ø§Ù„ØµÙØ­Ø§Øª Ø°Ø§Øª Ø§Ù„Ù†Ø¸Ø§Ù… Ø§Ù„ÙƒØªÙŠØ¨    """    
    dual_page_dic = {}
    counter = 46
    start_cell = 6
    
    for i in range(3,47,2):
        print ( i , counter)
        print ( counter-1 , i+1 )
        if counter == 46 :
            dual_page_dic.update({ f'{counter}' : f'A{start_cell}:A{start_cell+24}'})
            dual_page_dic.update({ f'{i+1}' : f'A{start_cell+33}:A{start_cell+33+24}'})
            
            dual_page_dic.update({ f'{i}' : f'L{start_cell}:L{start_cell+24}'})
            dual_page_dic.update({ f'{counter-1}' : f'L{start_cell+33}:L{start_cell+33+24}'})
            
            start_cell +=33*2-1      
            counter -= 2
        elif i == 23:
            dual_page_dic.update({ f'{counter}' : f'A{start_cell}:A{start_cell+24}'})
            dual_page_dic.update({ f'{i+1}' : f'A{start_cell+32}:A{start_cell+32+24}'})
            
            dual_page_dic.update({ f'{i}' : f'L{start_cell}:L{start_cell+24}'})
            dual_page_dic.update({ f'{counter-1}' : f'L{start_cell+32}:L{start_cell+32+24}'}) 
                            
            break
        else:
            print(start_cell)
            dual_page_dic.update({ f'{counter}' : f'A{start_cell}:A{start_cell+24}'})
            dual_page_dic.update({ f'{i+1}' : f'A{start_cell+32}:A{start_cell+32+24}'})
            
            dual_page_dic.update({ f'{i}' : f'L{start_cell}:L{start_cell+24}'})
            dual_page_dic.update({ f'{counter-1}' : f'L{start_cell+32}:L{start_cell+32+24}'})            
            start_cell += 32*2
            counter -= 2
        # print(dual_page_dic)
        # input('press anything to continue')        
        
    print(dual_page_dic)

def generate_pdf(doc_path, path , rename_number):
    """Ø§Ù†Ø´Ø§Ø¡ Ù…Ù„Ù Ø¨ÙŠ Ø¯ÙŠ Ø§Ù Ù…Ù† ÙˆØ±Ø¯ Ø¨Ø§Ø³ØªØ®Ø¯Ø§Ù… Ù„ÙŠØ¨Ø±Ø§ Ø§ÙˆÙØ³

    Args:
        doc_path (str): Ù…Ø³Ø§Ø± Ù…Ù„Ù ÙˆØ±Ø¯
        path (str): Ø§Ù„Ù…Ø³Ø§Ø± Ø§Ù„Ù…Ø¬Ù„Ø¯ Ø§Ù„Ù…Ø±Ø§Ø¯ ØªØ­ÙˆÙŠÙ„ Ø§Ù„Ù…Ù„Ù Ø§Ù„ÙŠÙ‡
        rename_number (int): Ø±Ù‚Ù… Ù„Ø¥Ø¹Ø§Ø¯Ø© ØªØ³Ù…ÙŠØ© Ø§Ù„Ù…Ù„Ù Ø§Ù„ÙŠÙ‡
    """
    subprocess.call(['soffice',
                 # '--headless',
                 '--convert-to',
                 'pdf',
                 '--outdir',
                 path,
                 doc_path])
    subprocess.call(['mv',
                    f'{path}/generated.pdf' ,
                    f'{path}/send{rename_number}.pdf'])

def word2pdf(wordFile ,pdfFile):
    """ÙƒÙ…Ø§ ÙŠØ¯Ù„ Ø§Ø³Ù… Ø§Ù„Ù…Ù„Ù ÙÙ‡Ø°Ø§ Ù…Ø§ ÙŠÙ‚ÙˆÙ… Ø¨Ù‡ ØªØ­ÙˆÙŠÙ„ Ù…Ù„Ù ÙˆØ±Ø¯ Ø§Ù„Ù‰ Ù…Ù„Ù Ø¨ÙŠ Ø¯ÙŠ Ø§Ù

    Args:
        wordFile (str): Ù…Ø³Ø§Ø± Ù…Ù„Ù ÙˆØ±Ø¯
        pdfFile (str): Ù…Ø³Ø§Ø± Ù…Ù„Ù Ø¨ÙŠ Ø¯ÙŠ Ø§Ù 
    """    
    convert(wordFile , pdfFile)

def fill_doc(template , context , output):
    """Ø¯Ø§Ù„Ø© Ù„ØªØ¹Ø¨Ø¦Ø© Ù…Ù„Ù ÙˆØ±Ø¯ Ø¨Ø§Ø³ØªØ®Ø¯Ø§Ù… Ø§Ù„Ù…ØªØºÙŠØ±Ø§Øª ÙÙŠ Ø§Ù„Ù…Ù„Ù

    Args:
        template (str): Ù…Ø³Ø§Ø± Ø§Ù„Ù†Ù…ÙˆØ°Ø¬ 
        context (dict): Ù‚Ø§Ù…ÙˆØ³ Ø¨Ø§Ù„Ù…Ø¹Ù„ÙˆÙ…Ø§Øª Ù„ØºØ±Ø¶ ØªØ¹Ø¨Ø¦ØªÙ‡Ø§ ÙÙŠ Ù…Ù„Ù ÙˆØ±Ø¯ ));
        output (str): Ù…Ø³Ø§Ø± Ø§Ù„Ù…Ù„Ù Ø§Ù„Ù…Ø¹Ø¨Ø£ Ø¨Ø§Ù„Ù…Ø¹Ù„ÙˆÙ…Ø§Øª Ø§Ù„Ù…ÙˆØ¬ÙˆØ¯Ø© ÙÙŠ Ø§Ù„Ù‚Ø§Ù…ÙˆØ³
    """    
    doc = DocxTemplate(template)
    context = context
    doc.render(context)
    doc.save(output)

def word_variables(template):
    """Ø¯Ø§Ù„Ø© ØªÙ‚ÙˆÙ… Ø¨Ø·Ø¨Ø§Ø¹Ø© Ø§Ù„Ù…ØªØºÙŠØ±Ø§Øª ÙÙŠ Ù…Ù„Ù ÙˆØ±Ø¯ 

    Args:
        template (str): Ù…Ø³Ø§Ø± Ø§Ù„Ù†Ù…ÙˆØ°Ø¬

    Returns:
        list: Ù‚Ø§Ø¦Ù…Ø© Ø¨Ø§Ù„Ù…ØªØºÙŠØ±Ø§Øª Ø§Ù„Ù…ÙˆØ¬ÙˆØ¯Ø© ÙÙŠ Ù…Ù„Ù ÙˆØ±Ø¯
    """    
    doc = DocxTemplate(template)
    return doc.get_undeclared_template_variables()

def my_jq(data):
    """Ø¯Ø§Ù„Ø© Ù…Ø³Ø§Ø¹Ø¯Ø© ØªÙ‚ÙˆÙ… Ø¨Ø·Ø¨Ø§Ø¹Ø© Ù…Ø¹Ù„ÙˆÙ…Ø§Øª Ø¬Ø§ÙŠØ³ÙˆÙ† Ø¨Ø´ÙƒÙ„ Ø¬Ù…ÙŠÙ„ Ùˆ Ù…Ø±ØªØ¨ ));

    Args:
        data (json): Ù…Ø¹Ù„ÙˆÙ…Ø§Øª Ø¬Ø§Ø³ÙˆÙ† Ø§Ùˆ Ù‚Ø§Ù…ÙˆØ³ 

    Returns:
        json,dict: Ø§Ù„Ù…Ø¹Ù„ÙˆÙ…Ø§Øª Ø¨Ø´ÙƒÙ„ Ù…Ø±ØªØ¨ Ùˆ Ø¬Ù…ÙŠÙ„ Ù„Ù…Ø³Ø§Ø¹Ø¯Ø© Ø§Ù„Ù…Ø·ÙˆØ±
    """    
    json_str = json.dumps(data, indent=4, sort_keys=True, ensure_ascii=False).encode('utf8')
    return highlight(json_str.decode('utf8'), JsonLexer(), TerminalFormatter())

def make_request(url, auth ,session=None,timeout_seconds=500):
    """Ø¯Ø§Ù„Ø© ØªÙ‚ÙˆÙ… Ø¨Ø·Ù„Ø¨Ø§Øª Ø§Ù„Ø§Ù†ØªØ±Ù†Øª Ø¨Ø§Ø³ØªØ®Ø¯Ø§Ù… Ù…ÙƒØªØ¨Ø© Ø±ÙƒÙˆÙŠØ³ØªØ³ Ø¨ØºØ±Ø¶ Ø§Ø³ØªØ¯Ø¹Ø§Ø¡ Ø§Ù„Ù…Ø¹Ù„ÙˆØ§Ù…Ø§Øª

    Args:
        url (_type_): Ø§Ù„Ø±Ø§Ø¨Ø·
        auth (_type_): ØªÙˆÙƒÙŠÙ† Ø§Ù„ØªÙˆØ«ÙŠÙ‚ Ø§Ù„Ø®Ø§Øµ Ø¨Ø§Ù„Ù…Ø³ØªØ®Ø¯Ù…
        session (_type_, optional): Ø¬Ù„Ø³Ø© Ø¨Ø§Ø³ØªØ®Ø¯Ø§Ù… Ù…ÙƒØªØ¨Ø© Ø±ÙƒÙˆÙŠØ³ØªØ³ Ù„ØªØ³Ø±ÙŠØ¹ Ø¹Ù…Ù„ÙŠØ© Ø·Ù„Ø¨ Ø§ÙƒØ«Ø± Ù…Ù† Ø±Ø§Ø¨Ø·. Defaults to None.
        timeout_seconds (int, optional): Ù…Ø¯Ø© Ø²Ù…Ù†ÙŠØ© Ù„Ø§Ù†Ù‡Ø§Ø¡ Ø·Ù„Ø¨ Ø§Ù„ÙˆÙŠØ¨ Ø¨Ø¹Ø¯ ÙØªØ±Ø© Ù…Ø¹ÙŠÙ†Ø©. Defaults to 500 second.

    Returns:
        json : Ø±Ø¯ Ø¨Ø§Ù„Ù…Ø¹Ù„ÙˆÙ…Ø§Øª Ø§Ù„ØªÙŠ Ù‚Ø§Ù… api Ø¨Ø±Ø¯Ù‡Ø§ Ù„Ù„Ø·Ù„Ø¨
    """    
    if 'csrfToken' in auth or 'PHPSESSID' in auth or 'System' in auth:
        headers = {"Cookie": auth }
    else:
        headers = {"Authorization": auth, "ControllerAction": "Results"}
        
    controller_actions = ["Results", "SubjectStudents", "Dashboard", "Staff",'StudentAttendances','SgTree','Students',]
    
    for controller_action in controller_actions:
        headers["ControllerAction"] = controller_action
        if session is None :
            response = requests.request("GET", 
                                        url,
                                        headers=headers,
                                        timeout=timeout_seconds,
                                        verify=False
                                        )
        else :
            response = session.get(url,
                                    headers=headers,
                                    timeout=timeout_seconds,
                                    verify=False
                                    )
        
        if response.status_code != 200 :
            for _ in range(50):
                if session is None :
                    response = requests.request("GET", url,headers=headers,timeout=timeout_seconds,verify=False)
                else :
                    response = session.get(url,headers=headers,timeout=timeout_seconds,verify=False)
                
                if response.status_code == 200:
                    break
        if "error" not in response.url :
            if is_json_response(response):
                return response.json()
        
    print('some thing wrong')
    return ['Some Thing Wrong']

def get_auth(username , password ,proxies=None):
    """
    Log in to the website and retrieve the token for Authorization header.

    Parameters:
        username (str): Username.
        password (str): Password.
        proxies (dict, optional): Proxy settings. Defaults to None.

    Returns:
        str: Token if login successful, False otherwise.
    """
    url = GET_AUTH_URL_1
    payload = {
        "username": username,
        "password": password
    }
    
    proxies = proxies if proxies else None
    
    response = requests.request("POST",
                                url, data=payload ,
                                proxies=proxies,
                                verify=False
                                # verify=False if proxies else True
                                )

    if response.json()['data']['message'] == 'Invalid login creadential':
        return False
    elif response.json()['data']['message'] == 'Api is disabled':
        return get_cookie_as_string(username , password)
    else:
        return response.json()['data']['token']    

def inst_name(auth,session=None):
    """
    Ø§Ø³ØªØ¯Ø¹Ø§Ø¡ Ø§Ø³Ù… Ø§Ù„Ù…Ø¯Ø±Ø³Ø© Ùˆ Ø§Ù„Ø±Ù‚Ù… Ø§Ù„ÙˆØ·Ù†ÙŠ Ùˆ Ø§Ù„Ø±Ù‚Ù… Ø§Ù„ØªØ¹Ø±ÙŠÙÙŠ 
        Ø¹ÙˆØ§Ù…Ù„ Ø§Ù„Ø¯Ø§Ù„Ø© Ø§Ù„Ø±Ø§Ø¨Ø· Ùˆ Ø§Ù„ØªÙˆÙƒÙ†
        ØªØ¹ÙˆØ¯ Ø¨Ø§Ù„Ø±Ù‚Ù… Ø§Ù„ØªØ¹Ø±ÙŠÙÙŠ Ùˆ Ø§Ù„Ø±Ù‚Ù… Ø§Ù„ÙˆØ·Ù†ÙŠ Ùˆ Ø§Ø³Ù… Ø§Ù„Ù…Ø¯Ø±Ø³Ø© 
    """
    url = INST_NAME_URL
    return make_request(url,auth,session=session)   # institution

def inst_area(auth , inst_id = None ,session=None):
    """
    Ø§Ø³ØªØ¯Ø¹Ø§Ø¡ Ù„ÙˆØ§Ø¡ Ø§Ù„Ù…Ø¯Ø±Ø³Ø© Ùˆ Ø§Ù„Ù…Ù†Ø·Ù‚Ø©
    Ø¹ÙˆØ§Ù…Ù„ Ø§Ù„Ø¯Ø§Ù„Ø© Ø§Ù„Ø±Ø§Ø¨Ø· Ùˆ Ø§Ù„ØªÙˆÙƒÙ†
    ØªØ¹ÙˆØ¯ Ø¨Ø§Ø³Ù… Ø§Ù„Ø¨Ù„Ø¯ÙŠØ© Ùˆ Ø§Ø³Ù… Ø§Ù„Ù…Ù†Ø·Ù‚Ø© Ùˆ Ø§Ù„Ù„ÙˆØ§Ø¡ 
    """
    if inst_id is None:
        inst_id = inst_name(auth)['data'][0]['Institutions']['id']
    url = f"https://emis.moe.gov.jo/openemis-core/restful/v2/Institution-Institutions.json?id={inst_id}&_contain=AreaAdministratives,Areas&_fields=AreaAdministratives.name,Areas.name"
    return make_request(url,auth,session=session)

def user_info(auth,username=None , user_id =None,session=None):
    """
        Ø§Ø³ØªØ¯Ø¹Ø§Ø¡ Ù…Ø¹Ù„ÙˆÙ…Ø§Øª Ø¹Ù† Ø§Ù„Ù…Ø¹Ù„Ù… Ø§Ùˆ Ø§Ù„Ù…Ø³ØªØ®Ø¯Ù… 
        Ø¹ÙˆØ§Ù…Ù„ Ø§Ù„Ø¯Ø§Ù„Ø© Ø§Ù„Ø±Ø§Ø¨Ø· Ùˆ Ø§Ù„ØªÙˆÙƒÙ† Ùˆ Ø±Ù‚Ù… Ø§Ù„Ù…Ø³ØªØ®Ø¯Ù…
        ØªØ¹ÙˆØ¯ Ø¨Ø±Ù‚Ù… Ø§Ù„Ù…Ø³ØªØ®Ø¯Ù… Ø§Ù„ÙˆØ·Ù†ÙŠ Ùˆ Ø§Ø³Ù…Ù‡ Ø§Ù„Ø±Ø¨Ø§Ø¹ÙŠ  
    """
    if username:
        url = USER_INFO_URL.format(username=username)
    else:
        url = USER_ID_INFO_URL.format(id=user_id)
    return make_request(url,auth,session=session)

def get_teacher_classes1(auth,ins_id,staff_id,academic_period,session=None):
    """
        Ø§Ø³ØªØ¯Ø¹Ø§Ø¡ Ù…Ø¹Ù„ÙˆÙ…Ø§Øª ØµÙÙˆÙ Ø§Ù„Ù…Ø¹Ù„Ù… 
        Ø¹ÙˆØ§Ù…Ù„ Ø§Ù„Ø¯Ø§Ù„Ø© Ø§Ù„Ø±Ø§Ø¨Ø· Ùˆ Ø§Ù„ØªÙˆÙƒÙ† Ùˆ Ø§Ù„ØªØ¹Ø±ÙŠÙÙŠ Ù„Ù„Ù…Ø¯Ø±Ø³Ø© Ùˆ ØªØ¹Ø±ÙŠÙÙŠ Ø§Ù„ÙØªØ±Ø© Ùˆ staffid 
        ØªØ¹ÙˆØ¯ Ø§Ù„Ø¯Ø§Ù„Ø© Ø¨ØªØ¹Ø±ÙŠÙÙŠ Ø§ÙŠ ØµÙ Ù…Ø¹ Ø§Ù„Ù…Ø¹Ù„Ù… Ùˆ ÙƒÙˆØ¯ Ø§Ù„ØµÙ
    """
    url = f"https://emis.moe.gov.jo/openemis-core/restful/v2/Institution.InstitutionSubjectStaff?institution_id={ins_id}&staff_id={staff_id}&academic_period_id={academic_period}&_contain=InstitutionSubjects"
    return make_request(url,auth,session=session)

def get_teacher_classes2(auth,inst_sub_id,session=None):
    """    Ø§Ø³ØªØ¯Ø¹Ø§Ø¡ Ù…Ø¹Ù„ÙˆÙ…Ø§Øª ØªÙØµÙŠÙ„ÙŠØ© Ø¹Ù† Ø§Ù„ØµÙÙˆÙ 

    Args:
        auth (auth): _description_
        inst_sub_id (int): Ø±Ù‚Ù… Ø§Ù„Ù…Ø§Ø¯Ø© Ø§Ù„ØªØ¹Ø±ÙŠÙÙŠ Ùˆ Ù‡Ø°Ø§ Ø§Ù„Ø±Ù‚Ù… ÙŠØ¯Ù„ Ø¹Ù„Ù‰ Ø§Ù„ØµÙ Ùˆ Ø¹Ù„Ù‰ Ø§Ù„Ù…Ø§Ø¯Ø©
        session (requests.Session(), optional): Ø§Ù„Ø¬Ù„Ø³Ø© Ù„ØªØ³Ø±ÙŠØ¹ Ø¹Ù…Ù„ÙŠØ© Ø§Ø­Ø¶Ø§Ø± Ø§Ù„Ù…Ø¹Ù„ÙˆÙ…Ø§Øª Ø§Ø°Ø§ ØªÙƒØ±Ø± Ø§Ø³ØªØ®Ø¯Ø§Ù… Ø§Ù„Ø¯Ø§Ù„Ø©. Defaults to None.

    Returns:
        _type_: Ù…Ø¹Ù„ÙˆÙ…Ø§Øª Ø¹Ù† Ø§Ù„Ù…Ø§Ø¯Ø© Ùˆ Ø¹Ù† Ø§Ù„ØµÙ Ø¹Ù„Ù‰ Ø´ÙƒÙ„ Ù‚Ø§Ù…ÙˆØ³
    """    
    """
    Ø§Ø³ØªØ¯Ø¹Ø§Ø¡ Ù…Ø¹Ù„ÙˆÙ…Ø§Øª ØªÙØµÙŠÙ„ÙŠØ© Ø¹Ù† Ø§Ù„ØµÙÙˆÙ 
    Ø¹ÙˆØ§Ù…Ù„ Ø§Ù„Ø¯Ø§Ù„Ø© Ø§Ù„Ø±Ø§Ø¨Ø· Ùˆ Ø§Ù„ØªÙˆÙƒÙ† Ùˆ Ø±Ù‚Ù… Ø§Ù„Ù…Ø³ØªØ®Ø¯Ù…
    ØªØ¹ÙˆØ¯ Ø¨Ø§Ø³Ù… Ø§Ù„ØµÙ Ùˆ ØªØ¹Ø±ÙŠÙÙŠ Ø§Ù„ØµÙ Ùˆ Ø¹Ø¯Ø¯ Ø§Ù„Ø·Ù„Ø§Ø¨ ÙÙŠ Ø§Ù„ØµÙ Ùˆ Ø§Ø³Ù… Ø§Ù„Ù…Ø§Ø¯Ø© Ø§Ù„ØªÙŠ ÙŠØ¯Ø±Ø³Ù‡Ø§ Ø§Ù„Ù…Ø¹Ù„Ù… ÙÙŠ Ø§Ù„ØµÙ
    """
    # url = "https://emis.moe.gov.jo/openemis-core/restful/Institution.InstitutionClassSubjects?status=1&_contain=InstitutionSubjects,InstitutionClasses&_limit=0&_orWhere=institution_subject_id:10513896,institution_subject_id:10513912,institution_subject_id:10513928,institution_subject_id:10513944"
    url = GET_TEACHER_CLASSES2_URL.format(inst_sub_id=inst_sub_id)
    
    return make_request(url,auth,session=session)

def get_class_students(auth,academic_period_id,institution_subject_id,institution_class_id,institution_id,education_grade_id=None,session=None):
    """    Ø§Ø³ØªØ¯Ø¹Ø§Ø¡ Ù…Ø¹Ù„ÙˆÙ…Ø§Øª Ø¹Ù† Ø§Ù„Ø·Ù„Ø§Ø¨ ÙÙŠ Ø§Ù„ØµÙ

    Args:
        auth (str): ØªÙˆÙƒÙ† Ø§Ù„Ù…Ø³ØªØ®Ø¯Ù…
        academic_period_id (int): Ø±Ù‚Ù… Ø§Ù„ÙØªØ±Ø© Ø§Ù„Ø§ÙƒØ§Ø¯ÙŠÙ…ÙŠØ© 
        institution_subject_id (int): Ø±Ù‚Ù… Ø§Ù„Ù…Ø§Ø¯Ø© Ø§Ù„ØªØ¹Ø±ÙŠÙÙŠ 
        institution_class_id (int): Ø±Ù‚Ù… Ø§Ù„ØµÙ Ùˆ Ø§Ù„Ø´Ø¹Ø¨Ø© Ø§Ù„ØªØ¹Ø±ÙŠÙÙŠ
        institution_id (int): Ø±Ù‚Ù… Ø§Ù„Ù…Ø¯Ø±Ø³Ø© Ø§Ù„ØªØ¹Ø±ÙŠÙÙŠ
        education_grade_id (int, optional): Ø±Ù‚Ù… Ø§Ù„Ù…Ø±Ø­Ù„Ø© Ø§Ù„ØªØ¹Ù„ÙŠÙ…ÙŠØ© Ø§Ù„ØªØ¹Ø±ÙŠÙÙŠ. Defaults to None.
        session (requests.Session(), optional): Ù…Ø¹Ù„ÙˆÙ…Ø§Øª Ø§Ù„Ø¬Ù„Ø³Ø© Ù„ØªØ³Ø±ÙŠØ¹ Ø§Ø³ØªØ¹Ù…Ø§Ù„Ù‡Ø§ Ø§Ø°Ø§ ØªÙƒØ±Ø±. Defaults to None.

    Raises:
        IndexError: Ø§Ø°Ø§ Ø­ØµÙ„ Ø®Ø·Ø£ Ø¹Ø§Ø¯Ø©Ù‹ ÙŠÙƒÙˆÙ† ØµÙ Ø«Ø§Ù†ÙˆÙŠ ÙØ§Ù‚ÙˆÙ… Ø¨Ø§Ø­Ø¶Ø§Ø± Ù…Ø¹Ù„ÙˆÙ…Ø§Øª Ùˆ Ø§Ø³Ù…Ø§Ø¡ Ø§Ù„Ø·Ù„Ø§Ø¨ Ø­ØªÙ‰ Ø§Ø­ØµÙ„ Ø¹Ù„Ù‰ Ù†ØµØ§Ø¨ Ø§Ù„Ù…Ø¹Ù„Ù… ÙƒØ§Ù…Ù„ Ù…Ø¹ Ø§Ù„ØµÙ Ø§Ù„Ø«Ø§Ù†ÙˆÙŠ 

    Returns:
        list: ØªØ¹ÙˆØ¯ Ø¨Ù…Ø¹Ù„ÙˆÙ…Ø§Øª ØªÙØµÙŠÙ„ÙŠØ© Ø¹Ù† ÙƒÙ„ Ø·Ø§Ù„Ø¨ ÙÙŠ Ø§Ù„ØµÙ Ø¨Ù…Ø§ ÙÙŠ Ø°Ù„Ùƒ Ø§Ø³Ù…Ù‡ Ø§Ù„Ø±Ø¨Ø§Ø¹ÙŠ Ùˆ Ø§Ù„ØªØ¹Ø±ÙŠÙÙŠ Ùˆ Ù…ÙƒØ§Ù† Ø³ÙƒÙ†Ù‡
    """       
    url = GET_CLASS_STUDENTS_URL.format(academic_period_id=academic_period_id,institution_subject_id=institution_subject_id,institution_class_id=institution_class_id,institution_id=institution_id)
    data = make_request(url,auth,session=session)
    if not data['total']:
        try:
            alt_url = f"https://emis.moe.gov.jo/openemis-core/restful/v2/Institution.InstitutionSubjectStudents?_fields=student_id,student_status_id,Users.id,Users.username,Users.openemis_no,Users.first_name,Users.middle_name,Users.third_name,Users.last_name,Users.address,Users.address_area_id,Users.birthplace_area_id,Users.gender_id,Users.date_of_birth,Users.date_of_death,Users.nationality_id,Users.identity_type_id,Users.identity_number,Users.external_reference,Users.status,Users.is_guardian&_limit=0&_finder=StudentResults[institution_id:{institution_id};institution_class_id:{institution_class_id};assessment_id:{get_assessment_id_from_grade_id(auth,education_grade_id)};academic_period_id:{academic_period_id};institution_subject_id:{institution_subject_id};education_grade_id:{education_grade_id}]&_contain=Users"
            data = make_request(alt_url,auth,session=session)
            sorted_list = sorted(data['data'], key=lambda x: x['student_id'])
            grouped_dicts = {k: next(v) for k, v in groupby(sorted_list, key=lambda x: x['student_id'])}
            data['data'] = list(grouped_dicts.values())
            # check if data is empty
            if not data['total']:
                raise IndexError
        except :
            global secondery_students 
            if not len(secondery_students):
                secondery_students =  get_school_students_ids(auth) 
            data = [i for i in secondery_students if i['institution_class_id'] == int(institution_class_id) and i['student_status_id'] ==1]
            data = {'data': data , 'total': len(data)}
            
    enrolled = [i for i in data['data'] if i['student_status_id'] ==1]
    data = {'data': enrolled , 'total': len(enrolled)}
    return data

def enter_mark(auth
                ,marks
                ,assessment_grading_option_id
                ,assessment_id
                ,education_subject_id
                ,education_grade_id
                ,institution_id
                ,academic_period_id
                ,institution_classes_id
                ,student_status_id
                ,student_id
                ,assessment_period_id):
    """
    Ø¯Ø§Ù„Ø© Ù„Ø§Ø¯Ø®Ø§Ù„ Ø¹Ù„Ø§Ù…Ø© Ø§Ù„Ø·Ø§Ù„Ø¨ 
    Ø¹ÙˆØ§Ù…Ù„ Ø§Ù„Ø¯Ø§Ù„Ø© Ø§Ù„Ø¹Ù„Ø§Ù…Ø© Ùˆ Ø±Ù‚Ù… Ø§Ù„Ù…Ø¤Ø³Ø³Ø© Ø§Ù„ØªØ¹Ø±ÙŠÙÙŠ Ùˆ Ø±Ù‚Ù… Ø§Ù„Ø·Ø§Ù„Ø¨ Ùˆ Ø§Ù„Ø±Ù‚Ù… Ø§Ù„ØªØ¹Ø±ÙŠÙÙŠ Ù„Ù„ÙØªØ±Ø© Ø§Ù„Ø§ÙƒØ§Ø¯ÙŠÙ…Ø© Ùˆ Ø±Ù‚Ù… Ø§Ù„Ù…Ø§Ø¯Ø© Ø§Ù„ØªØ¹Ø±ÙŠÙÙŠ
    enter_mark(get_auth() 
                ,marks= 6
                ,assessment_grading_option_id= 8
                ,assessment_id= 188
                ,education_subject_id= 4
                ,education_grade_id= 275
                ,institution_id= 2600
                ,academic_period_id= 13
                ,institution_classes_id= 786120
                ,student_status_id= 1
                ,student_id= 3768676
                ,assessment_period_id= 624)
    Ùˆ ØªØ¹ÙˆØ¯ Ø§Ù„Ø¯Ø§Ù„Ø© Ø¨ÙƒÙˆØ¯ Ø§Ù„Ø§Ø¬Ø§Ø¨Ø© 200 Ùˆ Ø§Ø°Ø§ Ù„Ù… ÙŠØ¹ÙˆØ¯ Ø¨Ù‡ ØªØµØ¯Ø± Ø§Ù„Ø¯Ø§Ù„Ø© Ø®Ø·Ø§
    """
    url = ENTER_MARK_URL
    headers = {"Authorization": auth , "ControllerAction" : "Results" }
    json_data = {
        'marks':marks,
        'assessment_grading_option_id':assessment_grading_option_id,
        'assessment_id':assessment_id,
        'education_subject_id':education_subject_id,
        'education_grade_id':education_grade_id,
        'institution_id':institution_id,
        'academic_period_id':academic_period_id,
        'institution_classes_id':institution_classes_id,
        'student_status_id':student_status_id,
        'student_id':student_id,
        'assessment_period_id':assessment_period_id,
        'action_type': 'default',
    }

    response = requests.post(url,headers=headers,json=json_data)
    if response.status_code != 200:
        raise(Exception("couldn't enter the mark for some reason")) 
    else:
        print(marks , education_subject_id ,student_id , response.status_code)

def get_curr_period(auth,session=None):
    """
    Ø¯Ø§Ù„Ø©  ØªØ³ØªØ¯Ø¹ÙŠ Ù…Ø¹Ù„ÙˆÙ…Ø§Øª Ø§Ù„Ø³Ù†Ø© Ø§Ù„Ø­Ø§Ù„ÙŠØ© Ù…Ù† Ø§Ù„Ø®Ø§Ø¯Ù…
    Ø§Ù„ØªÙˆÙƒÙ† 
    Ùˆ ØªØ¹ÙˆØ¯ Ø¹Ù„Ù‰ Ø§Ù„Ù…Ø³ØªØ®Ø¯Ù… Ø¨Ù…Ø¹Ù„ÙˆÙ…Ø§Øª Ø§Ù„Ø³Ù†Ø© Ø§Ù„Ø¯Ø±Ø§Ø³ÙŠØ© Ø§Ù„Ø­Ø§Ù„ÙŠØ© 
    """
    url = GET_CURR_PERIOD_URL
    return make_request(url,auth,session=session)

def get_assessments(auth,academic_term,assessment_id):
    """Ø¯Ø§Ù„Ø© ØªØ³ØªØ¯Ø¹ÙŠ Ù…Ø¹Ù„ÙˆÙ…Ø§Øª Ø¹Ù† Ø§Ù„Ø§Ù…ØªØ­Ø§Ù†Ø§Øª ÙÙŠ Ø§Ù„ÙØµÙ„
    Ùˆ Ø¹ÙˆØ§Ù…Ù„Ù‡Ø§ Ø§Ø³Ù… Ø§Ù„ÙØµÙ„ Ùˆ ØªØ¹Ø±ÙŠÙÙŠ Ø§Ø®ØªØ¨Ø§Ø± Ø§Ù„Ù…Ø±Ø­Ù„Ø© 
    ØªØ¹ÙˆØ¯ Ø¨Ù…Ø¹Ù„ÙˆÙ…Ø§Øª Ø¹Ù† Ø§Ù„Ø§Ù…ØªØ­Ø§Ù†Ø§Øª Ø§Ù„Ù…ØªÙˆÙØ±Ø© Ø¹Ù„Ù‰ Ø§Ù„Ù…Ù†Ø¸ÙˆÙ…Ø© ÙÙŠ Ø§Ù„ÙØµÙ„

    Args:
        auth (str): Ø¹Ø¨Ø§Ø±Ø© Ø§Ù„ØªÙˆØ«ÙŠÙ‚
        academic_term (str): Ø§Ù„ÙØµÙ„ Ø§Ù„Ø¯Ø±Ø§Ø³ÙŠ Ø§Ù…Ø§ Ø§Ù„ÙØµÙ„ Ø§Ù„Ø§ÙˆÙ„ Ø§Ùˆ Ø§Ù„ÙØµÙ„ Ø§Ù„Ø«Ø§Ù†ÙŠ 
        assessment_id (int): Ø±Ù‚Ù… Ø§Ù„Ù…Ø±Ø­Ù„Ø© Ø§Ù„ØµÙÙŠØ© Ø§Ù„ØªØ¹Ø±ÙŠÙÙŠ

    Returns:
        list: Ù‚Ø§Ø¦Ù…Ø© Ø¨Ù…Ø¹Ù„ÙˆÙ…Ø§Øª Ø§Ù„ØªÙ‚ÙˆÙŠÙ…Ø§Øª Ù„Ù„ØµÙ Ø§Ùˆ Ø§Ù„Ù…Ø±Ø­Ù„Ø© Ø§Ù„ØµÙÙŠØ©
    """    
    url = GET_ASSESSMENTS_URL.format(academic_term=academic_term,assessment_id=assessment_id)
    return make_request(url,auth)

def get_sub_info(auth,class_id,assessment_id,academic_period_id,institution_id):
    """    Ø§Ø³ØªØ¯Ø¹Ø§Ø¡ Ù…Ø¹Ù„ÙˆÙ…Ø§Øª Ø¹Ù† Ù…ÙˆØ§Ø¯ Ø§Ù„ØµÙ
    Ùˆ Ø¹ÙˆØ§Ù…Ù„Ù‡Ø§ Ù‡ÙŠ ØªØ¹Ø±ÙŠÙÙŠ Ø§Ù„ØµÙ Ùˆ ØªØ¹Ø±ÙŠÙÙŠ Ù…Ø±Ø­Ù„Ø© Ø§Ù„Ø§Ø®ØªØ¨Ø§Ø± Ùˆ Ø§Ù„ÙØªØ±Ø© Ø§Ù„Ø§ÙƒØ§Ø¯ÙŠÙ…ÙŠØ© Ùˆ ØªØ¹Ø±ÙŠÙÙŠ Ø§Ù„Ù…Ø¤Ø³Ø³Ø©
    ØªØ¹ÙˆØ¯ Ø¨Ù…Ø¹Ù„ÙˆÙ…Ø§Øª Ø¹Ù† Ù…ÙˆØ§Ø¯ Ø§Ù„ØµÙ Ùˆ Ø§Ù‡Ù…Ù‡Ø§ ØªØ¹Ø±ÙŠÙÙŠ Ø§Ù„Ù…Ø§Ø¯Ø© Ùˆ ÙƒÙˆØ¯ Ø§Ù„Ù…Ø§Ø¯Ø©

    Args:
        auth (str): Ø¬Ø§ÙŠØ³ÙˆÙ† Ø§ÙˆØ« ÙŠØªÙ… Ø§Ø³ØªØ¹Ù…Ø§Ù„Ù‡Ø§ Ù„Ø³Ø­Ø¨ Ø§Ù„Ù…Ø¹Ù„ÙˆÙ…Ø§Øª
        class_id (int): Ø±Ù‚Ù… Ø§Ù„Ø´Ø¹Ø¨Ø© Ø§Ù„ØªØ¹Ø±ÙŠÙÙŠ
        assessment_id (int): Ø±Ù‚Ù… Ø§Ù„Ù…Ø±Ø­Ù„Ø© Ø§Ù„ØªØ¹Ø±ÙŠÙÙŠ
        academic_period_id (int): Ø±Ù‚Ù… Ø§Ù„Ø³Ù†Ø© Ø§Ù„ØªØ¹Ø±ÙŠÙÙŠ
        institution_id (int): Ø±Ù‚Ù… Ø§Ù„Ù…Ø¯Ø±Ø³Ø© Ø§Ù„ØªØ¹Ø±ÙŠÙÙŠ

    Returns:
        list: Ù‚Ø§Ø¦Ù…Ø© Ø¨Ù‚ÙˆØ§Ù…ÙŠØ³ Ù…Ø¹Ù„ÙˆÙ…Ø§Øª Ø§Ù„Ù…ÙˆØ§Ø¯ Ø§Ù„ØªÙŠ ÙŠÙ…ÙƒÙ† Ù„Ù„ØµÙ ØªØ³Ø¬ÙŠÙ„Ù‡Ø§
    """
    url = f"https://emis.moe.gov.jo/openemis-core/restful/v2/Assessment-AssessmentItems.json?_finder=subjectNewTab[class_id:{class_id};assessment_id:{assessment_id};academic_period_id:{academic_period_id};institution_id:{institution_id}]&_limit=0"
    return make_request(url,auth)

def side_marks_document(username , password):
    """Ø§Ù†Ø´Ø§Ø¡ Ù…Ù„Ù ØªÙ‚ÙŠÙŠÙ… Ùˆ Ø§Ø¯Ø§Ø¡ Ø¬Ø§Ù†Ø¨ÙŠ ÙˆØ±Ù‚ÙŠ ÙŠØ­ØªÙˆÙ‰ Ø¹Ù„Ù‰ Ø§Ù„Ø§Ø³Ù…Ø§Ø¡ ÙÙ‚Ø· 

    Args:
        username (str,int): Ø§Ø³Ù… Ø§Ù„Ù…Ø³ØªØ®Ø¯Ù…
        password (str,int): ÙƒÙ„Ù…Ø© Ø§Ù„Ø³Ø±
    """    
    auth = get_auth(username , password)
    period_id = get_curr_period(auth)['data'][0]['id']
    inst_id = inst_name(auth)['data'][0]['Institutions']['id']
    user = user_info(auth , username)['data']
    user_id = user[0]['id']
    years = get_curr_period(auth)
    # Ù…Ø§ Ø¨Ø¹Ø±Ù ÙƒÙŠÙ Ø³ÙˆÙŠØªÙ‡Ø§ Ù„ÙƒÙ† Ø²Ø¨Ø·Øª 
    classes_id_1 = [[value for key , value in i['institution_subject'].items() if key == "id"][0] for i in get_teacher_classes1(auth,inst_id,user_id,period_id)['data']]
    classes_id_2 =[get_teacher_classes2( auth , classes_id_1[i])['data'] for i in range(len(classes_id_1))]
    classes_id_3 = []  
    for class_info in classes_id_2:
        classes_id_3.append([{"institution_class_id": class_info[0]['institution_class_id'] ,"sub_name": class_info[0]['institution_subject']['name'],"class_name": class_info[0]['institution_class']['name']}])
        
    for v in range(len(classes_id_1)):
        # id
        print (classes_id_3[v][0]['institution_class_id'])
        # subject name 
        print (classes_id_3[v][0]['sub_name'])
        # class name
        print (classes_id_3[v][0]['class_name'])
        students = get_class_students(auth
                                    ,period_id
                                    ,classes_id_1[v]
                                    ,classes_id_3[v][0]['institution_class_id']
                                    ,inst_id)
        students_names = sorted([i['user']['name'] for i in students['data']])
        print(students_names)
        
        context={}
        counter = 0
        for name in students_names :
            context[f'name{counter}'] = str(name) 
            counter+=1            
        context[f'sub'] = str(classes_id_3[v][0]['class_name']) 
        context[f'class_name'] = str(classes_id_3[v][0]['sub_name']) 
        context[f'school'] = str(inst_name(auth)['data'][0]['Institutions']['name']) 
        context['y1'] = str(years['data'][0]['start_year'])
        context['y2'] = str(years['data'][0]['end_year'])

        fill_doc('./templet_files/side_marks_note.docx' , context , f'./send_folder/send{v}.docx' )
        context.clear()
        generate_pdf(f'./send_folder/send{v}.docx' , './send_folder' ,v)
        # input("press enter to continue")
        # return students_names

def insert_students_names_in_excel_marks(template , students_id_and_names , outfile):
    """ÙƒØªØ§Ø¨Ø© Ø§Ø³Ù…Ø§Ø¡ Ø§Ù„Ø·Ù„Ø§Ø¨ Ùˆ Ø§Ø¯Ø®Ø§Ù„ Ø±Ù‚Ù…Ù‡Ù… Ø§Ù„ØªØ¹Ø±ÙŠÙÙŠ  ÙÙŠ Ù…Ù„Ù Ø§ÙƒØ³ÙŠÙ„

    Args:
        template (str): Ù…Ø³Ø§Ø± Ø§Ù„Ù†Ù…ÙˆØ°Ø¬
        students_id_and_names (list): Ù‚Ø§Ø¦Ù…Ø© Ø¨Ø§Ø³Ù…Ø§Ø¡ Ø§Ù„Ø·Ù„Ø§Ø¨ Ùˆ Ø§Ø±Ù‚Ø§Ù…Ù‡Ù… Ø§Ù„ØªØ¹Ø±ÙŠÙÙŠØ© 
        outfile (str): Ø§Ù„Ù…Ø³Ø§Ø± Ùˆ Ø§Ø³Ù… Ø§Ù„Ù…Ù„Ù Ø§Ù„Ø° ØªØ±ÙŠØ¯ Ø§Ø®Ø±Ø§Ø¬ Ù…Ù„Ù Ø§ÙƒØ³Ù„ Ø§Ù„ÙŠÙ‡
    """    
    workbook = load_workbook(filename=template)
    sheet = workbook.active
    counter = 2
    for i in students_id_and_names:
        sheet[f'B{counter}'] = i['student_name']
        sheet[f'A{counter}'] = i['student_id']
        counter+=1
    workbook.save( filename = outfile )

def delete_empty_rows(file , outfile):
    """Ø¯Ø§Ù„Ø© ØªÙ‚ÙˆÙ… Ø¨Ø­Ø°Ù Ø§Ù„Ø§Ø³Ø·Ø± Ø§Ù„ÙØ§Ø±ØºØ© Ù…Ù† Ø§Ù„Ø¹Ù„Ø§Ù…Ø§Øª

    Args:
        file (_type_): Ù…Ø³Ø§Ø± Ø§Ù„Ù…Ù„Ù
        outfile (_type_): Ù…Ø³Ø§Ø± Ø§Ù„Ù…Ù„Ù Ø§Ù„Ø°ÙŠ ØªØ±ÙŠØ¯ Ø§Ù† ÙŠØ®Ø±Ø¬ Ø§Ù„ÙŠÙ‡ Ø§Ù„Ù…Ù„Ù Ø§Ù„Ù…ØµÙÙ‰ Ø§Ùˆ Ø¨Ø¯ÙˆÙ† Ø§Ù„Ø§Ø³Ø·Ø± Ø§Ù„ÙØ§Ø±ØºØ©
    """    
    workbook = load_workbook(filename=file)
    sheet = workbook.active
    
    # Find the last row of data in the sheet
    max_row = sheet.max_row
    # breakpoint()
    # Loop through each row in reverse order and check if it is empty
    for row in range(max_row, 1, -1):
        if all([cell.value in (None, '') for cell in sheet[row]][:6]):
            # Remove the row if it is empty
            sheet.delete_rows(row, 1)
            
    # Compute sum for each row
    for row in range(2, sheet.max_row + 1):
        sheet.cell(row=row, column=7).value = f"=SUM(C{row}:F{row})"
    # Set header for sum column
    sheet.cell(row=1, column=7).value = "Ø§Ù„Ù…Ø¬Ù…ÙˆØ¹"
    # Auto-fit column width for sum column
    sheet.column_dimensions['G'].auto_size = True
    
    # Save the updated workbook
    workbook.save(outfile)

def read_excel_marks(file):
    """ÙƒÙ…Ø§ ÙŠØ¯Ù„ Ø§Ø³Ù… Ø§Ù„Ø¯Ø§Ù„Ø©  ØŒ Ù‡Ø°Ù‡ Ø§Ù„Ø¯Ø§Ù„Ø© ØªÙ‚ÙˆÙ… Ø¨Ù‚Ø±Ø£Øª ØµÙØ­Ø§Øª Ø§ÙƒØ³Ù„ Ø§Ù„Ø¹Ù„Ø§Ù…Ø§Øª
    ÙˆÙ‡ÙŠ Ù…Ø­Ø§ÙˆÙ„Ø© ÙØ§Ø´Ù„Ø© Ù‚Ù…Øª Ø¨Ø§Ø­Ø³Ù† Ù…Ù†Ù‡Ø§ ÙÙŠ Ø¯Ø§Ù„Ø© create_e_side_notebook

    Args:
        file (str): Ù…Ø³Ø§Ø± Ø§Ù„Ù…Ù„Ù Ø§Ù„Ø°ÙŠ ØªØ±ÙŠØ¯ Ù‚Ø±Ø§Ø¦ØªÙ‡
    """    
    workbook = load_workbook(filename=file)
    sheet = workbook.active
    counter = 2
    for value in sheet.values:
        if value[0] ==None:
            break
        elif not value[2] == None :
            value = list(value)
            #   Ø§Ù„ØªÙ‚ÙˆÙŠÙ… Ø§Ù„Ø±Ø§Ø¨Ø¹ Ùˆ  Ø§Ù„Ø«Ø§Ù„Ø« Ùˆ  Ø§Ù„Ø«Ø§Ù†ÙŠ Ùˆ   Ø§Ù„Ø§ÙˆÙ„  
            # value[2]+ value[3]+ value[4]+value[5]
            value[6]= value[2]+ value[3]+ value[4]+value[5]
            print(value)                
        else : 
            print(value)

def insert_students_names_and_marks(assessments_json, students_id_and_names , template , outfile):
    """Ø¯Ø§Ù„Ø© Ù…Ø³Ø§Ø¹Ø¯Ø© ØªÙ‚ÙˆÙ… Ø¨Ø§Ù†Ø´Ø§Ø¡ ØµÙØ­Ø§Øª  Ø§ÙƒØ³Ù„ ÙÙŠÙ‡Ø§ Ø§Ø³Ù…Ø§Ø¡ Ùˆ Ø¹Ù„Ø§Ù…Ø§Øª Ø§Ù„Ø·Ù„Ø§Ø¨ 

    Args:
        assessments_json (list): Ù‚Ø§Ø¦Ù…Ø© Ù‚ÙˆØ§Ù…ÙŠØ³ ØªØ­ØªÙˆÙŠ Ø¹Ù„Ø§Ù…Ø§Øª Ø§Ù„Ø·Ù„Ø§Ø¨ Ùˆ Ù…Ø¹Ù„ÙˆÙ…Ø§Øª Ø¹Ù†Ù‡Ø§
        students_id_and_names (list): Ø§Ø³Ù…Ø§Ø¡ Ø§Ù„Ø·Ù„Ø§Ø¨ Ùˆ Ø§Ø³Ù…Ø§Ø¦Ù‡Ù…
        template (str): Ù…Ø³Ø§Ø± Ø§Ù„Ù…Ù„Ù Ø§Ù„Ù†Ù…ÙˆØ°Ø¬
        outfile (str): Ù…Ø³Ø§Ø± Ø§Ù„Ù…Ù„Ù Ø§Ù„Ø°ÙŠ ØªØ±ÙŠØ¯ Ø§Ù†Ø´Ø§Ø¡ Ø§Ù„Ù…Ù„Ù Ø§Ù„ÙŠÙ‡
    """    
    workbook = load_workbook(filename=template)
    sheet = workbook.active
    marks_and_name = []
    dic =  {'Sid':'' ,'Sname': '', 'ass1': '' ,'ass2': '' , 'ass3': '' , 'ass4': '' }
    for i in students_id_and_names:   
    #         print(i['student_id'])
        for v in assessments_json['data']:
            if v['student_id'] == i['student_id'] :  
                dic['Sid'] = i['student_id'] 
                dic['Sname'] = i['student_name'] 
                if v['assessment_period']['name'] == 'Ø§Ù„ØªÙ‚ÙˆÙŠÙ… Ø§Ù„Ø£ÙˆÙ„':
                    dic['ass1'] = v["marks"]     
                elif v['assessment_period']['name'] == 'Ø§Ù„ØªÙ‚ÙˆÙŠÙ… Ø§Ù„Ø«Ø§Ù†ÙŠ':
                    dic['ass2'] = v["marks"]             
                elif v['assessment_period']['name'] == 'Ø§Ù„ØªÙ‚ÙˆÙŠÙ… Ø§Ù„Ø«Ø§Ù„Ø«':
                    dic['ass3'] = v["marks"]           
                elif v['assessment_period']['name'] == 'Ø§Ù„ØªÙ‚ÙˆÙŠÙ… Ø§Ù„Ø±Ø§Ø¨Ø¹':
                    dic['ass4']= v["marks"]
        marks_and_name.append(dic)
        dic =  { 'Sid':'' ,'Sname': '', 'ass1': '' ,'ass2': '' , 'ass3': '' , 'ass4': '' }

        headers = ['id', 'Ø§Ø³Ù… Ø§Ù„Ø·Ø§Ù„Ø¨', 'Ø§Ù„ØªÙ‚ÙˆÙŠÙ… Ø§Ù„Ø§ÙˆÙ„', 'Ø§Ù„ØªÙ‚ÙˆÙŠÙ… Ø§Ù„Ø«Ø§Ù†ÙŠ', 'Ø§Ù„ØªÙ‚ÙˆÙŠÙ… Ø§Ù„Ø«Ø§Ù„Ø«', 'Ø§Ù„ØªÙ‚ÙˆÙŠÙ… Ø§Ù„Ø±Ø§Ø¨Ø¹']
        for col_num, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col_num, value=header)
        # Iterate over the data and insert into rows
        for row_num, row_data in enumerate(marks_and_name, start=2):
            for col_num, cell_data in enumerate(row_data.values(), start=1):
                sheet.cell(row=row_num, column=col_num, value=cell_data)
                
    workbook.save( filename = outfile )
    delete_empty_rows(outfile , outfile)

def create_excel_sheets_marks(username, password ):
    """Ø§ÙˆÙ„ Ù…Ø­Ø§ÙˆÙ„Ø© Ù„ÙŠ Ù„Ø§Ù†Ø´Ø§Ø¡ ÙƒØ´Ù Ø§Ù„ÙƒØªØ±ÙˆÙ†ÙŠ Ø¬Ø§Ù†Ø¨ÙŠ Ø¨Ø¹Ù„Ø§Ù…Ø§Øª Ø§Ù„Ø·Ù„Ø§Ø¨  
    ÙƒØ§Ù†Øª Ù‡Ø°Ù‡ Ø§Ù„Ù…Ø­Ø§ÙˆÙ„Ø© ÙØ§Ø´Ù„Ø© Ùˆ Ø­Ø³Ù†ØªÙ‡Ø§ ÙÙŠ Ø§Ù„Ø¯Ø§Ù„Ø© create_e_side_notebook

    Args:
        username (str,int): _description_
        password (str,int): _description_
    """    
    auth = get_auth(username , password)
    period_id = get_curr_period(auth)['data'][0]['id']
    inst_id = inst_name(auth)['data'][0]['Institutions']['id']
    user_id = user_info(auth , username)['data'][0]['id']
    years = get_curr_period(auth)
    # Ù…Ø§ Ø¨Ø¹Ø±Ù ÙƒÙŠÙ Ø³ÙˆÙŠØªÙ‡Ø§ Ù„ÙƒÙ† Ø²Ø¨Ø·Øª 
    classes_id_1 = [[value for key , value in i['InstitutionSubjects'].items() if key == "id"][0] for i in get_teacher_classes1(auth,inst_id,user_id,period_id)['data']]
    classes_id_2 =[get_teacher_classes2( auth , classes_id_1[i])['data'] for i in range(len(classes_id_1))]
    classes_id_3 = []  
    for class_info in classes_id_2:
        classes_id_3.append([{"institution_class_id": class_info[0]['institution_class_id'] ,"sub_name": class_info[0]['institution_subject']['name'],"class_name": class_info[0]['institution_class']['name']}])

    for v in range(len(classes_id_1)):
        # id
        print (classes_id_3[v][0]['institution_class_id'])
        # subject name 
        print (classes_id_3[v][0]['sub_name'])
        # class name
        print (classes_id_3[v][0]['class_name'])
        students = get_class_students(auth
                                    ,period_id
                                    ,classes_id_1[v]
                                    ,classes_id_3[v][0]['institution_class_id']
                                    ,inst_id)
        # students_names = sorted([i['user']['name'] for i in students['data']])
        # print(students_names)
        students_id_and_names = []
        for IdAndName in students['data']:
            students_id_and_names.append({'student_name': IdAndName['user']['name'] , 'student_id':IdAndName['student_id']})

        assessments_json = make_request(auth=auth , url=f'https://emis.moe.gov.jo/openemis-core/restful/Assessment.AssessmentItemResults?academic_period_id={period_id}&education_subject_id=4&institution_classes_id='+ str(classes_id_3[v][0]['institution_class_id'])+ f'&institution_id={inst_id}&_limit=0&_fields=AssessmentGradingOptions.name,AssessmentGradingOptions.min,AssessmentGradingOptions.max,EducationSubjects.name,EducationSubjects.code,AssessmentPeriods.code,AssessmentPeriods.name,AssessmentPeriods.academic_term,marks,assessment_grading_option_id,student_id,assessment_id,education_subject_id,education_grade_id,assessment_period_id,institution_classes_id&_contain=AssessmentPeriods,AssessmentGradingOptions,EducationSubjects')
        insert_students_names_and_marks(assessments_json , students_id_and_names , './templet_files/excel_marks.xlsx' , './send_folder/' + classes_id_3[v][0]['class_name'] + '.xlsx')

def count_files():
    """Ø¹Ø¯ Ø§Ù„Ù…Ù„ÙØ§Øª ÙÙŠ Ù…Ø¬Ù„Ø¯ Ø§Ù„Ø§Ø±Ø³Ø§Ù„

    Returns:
        list: Ù‚Ø§Ø¦Ù…Ø© Ø¨Ø§Ù„Ù…Ù„ÙØ§Øª Ø§Ù„Ù…ÙˆØ¬ÙˆØ¯Ø© ÙÙŠ Ù…Ø¬Ù„Ø¯ Ø§Ù„Ø§Ø±Ø³Ø§Ù„
    """    
    files = glob.glob('./send_folder/*')
    return files

def delete_send_folder():
    """Ø¯Ø§Ù„Ø© Ø¨Ø³ÙŠØ·Ø© Ù„Ø­Ø°Ù Ù…Ø¬Ù„Ø¯ Ø§Ù„Ø§Ø±Ø³Ø§Ù„ ÙÙŠ Ø§Ù„Ø¨ÙˆØª
    """    
    files = glob.glob('./send_folder/*')
    for f in files:
        os.remove(f)

def get_students_marks(auth,period_id,sub_id,instit_class_id,instit_id):
    """    Ø¯Ø§Ù„Ø© Ù„Ø§Ø³ØªØ¯Ø¹Ø§Ø¡ Ø¹Ù„Ø§Ù…Ø§Øª Ø§Ù„Ø·Ù„Ø§Ø¨ Ùˆ Ø§Ø³Ù…Ø§Ø¦Ù‡Ù… 
    Ùˆ Ø¹ÙˆØ§Ù…Ù„Ù‡Ø§ Ø§Ù„ØªÙˆÙƒÙ† Ø±Ù‚Ù… Ø§Ù„Ø³Ù†Ø© Ø§Ù„ØªØ¹Ø±ÙŠÙÙŠ ÙˆØ±Ù‚Ù… Ø§Ù„Ù…Ø§Ø¯Ø© Ø§Ù„ØªØ¹Ø±ÙŠÙÙŠ Ùˆ Ø±Ù‚Ù… Ø§Ù„Ù…Ø¤Ø³Ø³Ø© Ùˆ  Ø±Ù‚Ù… Ø§Ù„ØµÙ Ø§Ù„ØªØ¹Ø±ÙŠÙÙŠ
    Ùˆ ØªØ¹ÙˆØ¯ Ø¨Ø§Ø³Ù…Ø§Ø¡ Ø§Ù„Ø·Ø§Ù„Ø¨ Ùˆ Ø¹Ù„Ø§Ù…Ø§ØªÙ‡Ù…

    Args:
        auth (str): authentication barriar
        period_id (int): Ø§Ù„ÙØªØ±Ø© Ø§Ù„ØªÙŠ ÙŠÙ†ØªÙ…ÙŠ Ù„Ù‡Ø§ Ù‡Ø°Ø§ Ø§Ù„ÙØµÙ„ 
        sub_id (int): Ø±Ù‚Ù… Ø§Ù„Ù…Ø§Ø¯Ø© Ù…Ø«Ù„Ø§ 7 Ø§Ùˆ 331 
        instit_class_id (int): Ø±Ù‚Ù… Ø§Ù„ØµÙ Ø§Ù„ØªØ¹Ø±ÙŠÙÙŠ 
        instit_id (int): Ø±Ù‚Ù… Ø§Ù„Ù…Ø¤Ø³Ø³Ø© Ø§Ù„ØªØ¹Ø±ÙŠÙÙŠ Ø§Ùˆ Ø§Ù„Ù…Ø¯Ø±Ø³Ø©

    Returns:
        list: Ù‚Ø§Ø¦Ù…Ø© Ø¨Ù‚ÙˆØ§Ù…ÙŠØ³ Ù…Ø¹Ù„ÙˆÙ…Ø§Øª Ø¹Ù„Ø§Ù…Ø§Øª Ø§Ù„Ø·Ù„Ø§Ø¨ 
    """
    
    url = GET_STUDENTS_MARKS_URL.format(period_id=period_id,sub_id=sub_id,instit_class_id=instit_class_id,instit_id=instit_id)
    return make_request(url,auth)

def sort_send_folder_into_two_folders(folder='./send_folder'):
    """Ø¯Ø§Ù„Ø© Ø§Ø³ØªØ¹Ù…Ù„Ù‡Ø§ Ù„ÙØ±Ø² Ù…Ù„ÙØ§Øª Ø§Ù„Ø¨ÙŠ Ø¯ÙŠ Ø§Ù Ø¹Ù† ØºÙŠØ±Ù‡Ø§ ÙÙŠ Ù…Ø¬Ù„Ø¯ send_folder  
    

    Args:
        folder (str, optional): _description_. Defaults to './send_folder'.
    """    
    files = os.listdir(folder)
    pdf_folder = os.path.join(folder, 'PDFs')
    editable_folder = os.path.join(folder, 'Ù‚Ø§Ø¨Ù„ Ù„Ù„ØªØ¹Ø¯ÙŠÙ„')

    os.makedirs(pdf_folder, exist_ok=True)
    os.makedirs(editable_folder, exist_ok=True)

    for file in files:
        if not file.endswith('.json'):
            file_path = os.path.join(folder, file)
            if file.endswith('.pdf'):
                shutil.move(file_path, pdf_folder)
            else:
                shutil.move(file_path, editable_folder)

def extract_primary_and_other_classes(nested_classes):
    """ØªÙ‚ÙˆÙ… Ù‡Ø°Ù‡ Ø§Ù„Ø¯Ø§Ù„Ù‡ Ø¨ÙØµÙ„ Ø§Ù„ØµÙÙˆÙ Ø§Ù„Ø§Ø³Ø§Ø³ÙŠÙ‡ Ø¹Ù† Ø§Ù„ØµÙÙˆÙ Ø§Ù„Ø§Ø¨ØªØ¯Ø§Ø¦ÙŠÙ‡

    Args:
        nested_classes (_type_): Ø§Ø¨Ø¹Ø« Ù„Ù‡Ø§ Ø§Ù„ØµÙÙˆÙ 

    Returns:
        _type_:   ØªØ±Ø¬Ø¹ Ù‚ÙŠÙ…ØªÙŠÙ† Ø§Ù„Ù‚ÙŠÙ…Ù‡ Ø§Ù„Ø§ÙˆÙ„Ù‰ Ø§Ù„ØµÙÙˆÙ Ø§Ù„Ø§Ø¨ØªØ¯Ø§Ø¦ÙŠÙ‡ ÙˆØ§Ù„Ù‚ÙŠÙ…Ù‡ Ø§Ù„Ø«Ø§Ù†ÙŠÙ‡ Ø§Ù„ØµÙÙˆÙ Ø§Ù„Ø§Ø³Ø§Ø³ÙŠÙ‡
    """    
    primary_keywords = ['Ø§Ù„ØµÙ Ø§Ù„Ø£ÙˆÙ„', 'Ø§Ù„ØµÙ Ø§Ù„Ø«Ø§Ù†ÙŠ', 'Ø§Ù„ØµÙ Ø§Ù„Ø«Ø§Ù„Ø«']

    primary_classes = [
        cls for class_list in nested_classes
        for cls in class_list
        if any(keyword in cls['class_name'] for keyword in primary_keywords)
    ]

    other_classes = [
        cls for class_list in nested_classes
        for cls in class_list
        if cls not in primary_classes
    ]

    return primary_classes, other_classes

def main():
    print('starting script')
    setup_logging("access.log")
    # create_coloured_certs_wrapper('9991014194','Zzaid#079079' ,term2=True)
    # create_certs_wrapper('9991014194','Zzaid#079079','2002186139' , term2=True)

    passwords= '''9831046736
Emad1989/eE@123456
9961021489
113859/9841050562
9842053654/654321
9882048821/123456
9842001456
9911018064
9842048442/123456'''

    passwords = '''9771043526/9771043526#Hh  Ø¬  Ø±'''
    passwords = '2000220556/Aa@2000220556'
    # diverse_bulk_work(passwords)
    bulk_e_side_note_marks(passwords)

    # create_e_side_marks_doc(username , password)
    
    # convert_to_marks_offline_from_send_folder(A4_context = False , A3_context = True )
    playsound()
    print('finished')

if __name__ == "__main__":
    main()